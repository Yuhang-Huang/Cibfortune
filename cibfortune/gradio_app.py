#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Qwen3-VL-8B-Instruct Gradio Web界面
提供友好的Web界面来使用Qwen3-VL模型
"""

import os
import gradio as gr
import torch
from transformers import Qwen3VLForConditionalGeneration, AutoProcessor
from PIL import Image
import requests
from io import BytesIO
import time
import json
import csv
from datetime import datetime

# 设置环境变量
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'

class Qwen3VLGradioApp:
    """Qwen3-VL Gradio应用类"""
    
    def __init__(self):
        self.model = None
        self.processor = None
        self.model_path = "D:\cibfortune\Cibfortune\cibfortune\models\qwen3-vl-2b-instruct"
        self.is_loaded = False
        self.chat_history = []
        self.chat_messages = []
        self.last_image = None
        self.last_ocr_markdown = None

    def _sanitize_markdown(self, text: str) -> str:
        if not text:
            return ""
        s = text.strip()
        lines = s.splitlines()
        out = []
        in_fence = False
        for line in lines:
            stripped = line.strip()
            if stripped.startswith("```"):
                in_fence = not in_fence
                continue
            out.append(line) if not in_fence else None
        cleaned = "\n".join(out).strip()
        return cleaned if cleaned else s
        
    def load_model(self, progress=gr.Progress()):
        """加载模型"""
        if self.is_loaded:
            return "✅ 模型已经加载完成！"
        
        try:
            progress(0.1, desc="检查模型路径...")
            if not os.path.exists(self.model_path):
                return f"❌ 模型路径不存在: {self.model_path}"
            
            progress(0.3, desc="加载模型...")
            self.model = Qwen3VLForConditionalGeneration.from_pretrained(
                self.model_path,
                dtype="auto",
                device_map="cuda",
                load_in_4bit=True,
            )
            
            progress(0.7, desc="加载处理器...")
            self.processor = AutoProcessor.from_pretrained(self.model_path)
            
            progress(1.0, desc="完成！")
            self.is_loaded = True
            
            return "✅ 模型加载成功！可以开始使用了。"
            
        except Exception as e:
            return f"❌ 模型加载失败: {str(e)}"
    
    def _prepare_user_message(self, image, prompt):
        prompt_clean = (prompt or "").strip()
        resolved_image = image if image is not None else self.last_image
        if resolved_image is None:
            raise ValueError("❌ 请上传图像！")
        if not prompt_clean:
            raise ValueError("❌ 请输入问题！")
        if image is not None:
            self.last_image = image
        content = [
            {"type": "image", "image": resolved_image},
            {"type": "text", "text": prompt_clean},
        ]
        return prompt_clean, {"role": "user", "content": content}

    def _run_inference(self, image, prompt, max_tokens, temperature, prepared=None):
        if prepared is None:
            prompt_clean, user_message = self._prepare_user_message(image, prompt)
        else:
            prompt_clean, user_message = prepared
        messages = self.chat_messages + [user_message]
        inputs = self.processor.apply_chat_template(
            messages,
            tokenize=True,
            add_generation_prompt=True,
            return_dict=True,
            return_tensors="pt"
        ).to(self.model.device)

        generation_kwargs = {
            "max_new_tokens": max_tokens,
            "temperature": temperature,
            "do_sample": True if temperature > 0 else False
        }

        start_time = time.time()
        with torch.no_grad():
            generated_ids = self.model.generate(**inputs, **generation_kwargs)
        generation_time = time.time() - start_time

        generated_ids_trimmed = [
            out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text = self.processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )
        response = output_text[0]

        assistant_message = {"role": "assistant", "content": [{"type": "text", "text": response}]}
        self.chat_messages.extend([user_message, assistant_message])
        return prompt_clean, response, generation_time

    def _clone_history(self, history):
        return [[turn[0], turn[1]] for turn in history]

    def _chunk_response(self, text, chunk_size=80):
        if not text:
            return []
        return [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]

    @staticmethod
    def _parse_markdown_sections(markdown_text):
        sections = []
        lines = markdown_text.splitlines()
        i = 0

        while i < len(lines):
            line = lines[i]
            stripped = line.strip()
            is_table = (
                stripped.startswith("|")
                and stripped.count("|") >= 2
                and i + 1 < len(lines)
                and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
            )

            if is_table:
                header = [cell.strip() for cell in stripped.strip("|").split("|")]
                i += 2
                rows = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if not (row_line.startswith("|") and row_line.count("|") >= 2):
                        break
                    row = [cell.strip() for cell in row_line.strip("|").split("|")]
                    rows.append(row)
                    i += 1
                sections.append({"type": "table", "header": header, "rows": rows})
                continue

            text_block = []
            while i < len(lines):
                current = lines[i]
                stripped_current = current.strip()
                next_is_table = (
                    stripped_current.startswith("|")
                    and stripped_current.count("|") >= 2
                    and i + 1 < len(lines)
                    and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                )
                if next_is_table:
                    break
                text_block.append(current)
                i += 1
                if i < len(lines) and lines[i] == "":
                    text_block.append(lines[i])
            text_content = "\n".join(text_block).strip("\n")
            if text_content:
                sections.append({"type": "text", "text": text_content})

        return sections

    def chat_with_image(self, image, text, history, max_tokens, temperature):
        """与图像对话（流式反馈）"""
        original_text = text

        if not self.is_loaded:
            yield history, original_text
            return

        try:
            prepared = self._prepare_user_message(image, text)
        except ValueError as exc:
            yield history, original_text
            return

        prompt_clean, _ = prepared
        history_copy = self._clone_history(history)
        history_copy.append([f"👤 {prompt_clean}", "🤖 正在思考..."])
        yield self._clone_history(history_copy), original_text

        try:
            _, response, _ = self._run_inference(image, text, max_tokens, temperature, prepared=prepared)
        except Exception as e:
            history_copy[-1][1] = f"❌ 生成失败: {str(e)}"
            self.chat_history = self._clone_history(history_copy)
            yield self._clone_history(history_copy), original_text
            return

        assembled = ""
        chunks = self._chunk_response(response)
        if not chunks:
            chunks = [""]
        for chunk in chunks:
            assembled += chunk
            history_copy[-1][1] = f"🤖 {assembled}▌"
            yield self._clone_history(history_copy), original_text

        history_copy[-1][1] = f"🤖 {response}"
        final_history = self._clone_history(history_copy)
        self.chat_history = final_history
        yield final_history, original_text
    
    def ocr_analysis(self, image, prompt: str = None):
        """OCR文字识别"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        
        default_prompt = "请识别并提取这张图片中的所有文字内容。如果图片中有多种语言，请分别标注语言类型。"
        effective_prompt = (prompt or "").strip() or default_prompt
        
        try:
            prompt_clean, response, _ = self._run_inference(image, effective_prompt, max_tokens=1024, temperature=0.7)
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## OCR识别结果\n\n{cleaned}"
            return f"📝 OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"
    
    def spatial_analysis(self, image, prompt: str = None):
        """空间感知分析"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        
        default_prompt = """请分析这张图片中的空间关系，包括：
            1. 物体的相对位置关系
            2. 视角和观察角度
            3. 物体的遮挡关系
            4. 深度和距离感
            5. 空间布局的整体描述"""
        effective_prompt = (prompt or "").strip() or default_prompt
        
        try:
            prompt_clean, response, _ = self._run_inference(image, effective_prompt, max_tokens=768, temperature=0.7)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {response}"])
            return f"🔍 空间分析结果:\n\n{response}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 空间分析失败: {str(e)}"
    
    def visual_coding(self, image, output_format, prompt: str = None):
        """视觉编程"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"

        try:
            format_prompts = {
                "HTML": "请根据图片生成对应的HTML结构代码，包含必要的语义标签。",
                "CSS": "请为该图片对应的界面生成合理的CSS样式代码，包括布局与颜色。",
                "JavaScript": "请根据图片交互生成JavaScript代码示例，包含必要的事件与逻辑。",
                "Python": "请生成能复现该界面/布局的Python示例代码（如使用streamlit或flask的伪代码）。",
            }
            base_prompt = format_prompts.get(output_format, format_prompts["HTML"]) + " 请只输出代码，不要额外说明。"
            effective_prompt = (prompt or "").strip() or base_prompt

            prompt_clean, response, _ = self._run_inference(image, effective_prompt, max_tokens=2048, temperature=0.4)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {response}"])
            return f"💻 {output_format}代码:\n\n```{output_format.lower()}\n{response}\n```"
            
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 代码生成失败: {str(e)}"
    
    def clear_history(self):
        """清空对话历史"""
        self.chat_history = []
        self.chat_messages = []
        self.last_image = None
        self.last_ocr_markdown = None
        return []

    def export_last_ocr(self):
        if not getattr(self, "last_ocr_markdown", None):
            return "❌ 没有可保存的文本样式，请先执行一次OCR识别！"

        export_dir = os.path.join("ocr_exports")
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        sections = self._parse_markdown_sections(self.last_ocr_markdown)

        excel_path = os.path.join(export_dir, f"ocr_{timestamp}.xlsx")
        excel_note = ""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "表格1" if sections else "OCR文本"
            table_idx = 0
            for section in sections:
                if section["type"] == "table":
                    table_idx += 1
                    if table_idx > 1:
                        ws = wb.create_sheet(title=f"表格{table_idx}")
                    ws.append(section["header"])
                    for row in section["rows"]:
                        ws.append(row)
                elif section["type"] == "text" and section["text"]:
                    if table_idx > 0:
                        ws = wb.create_sheet(title=f"文本{table_idx}")
                    for line in section["text"].splitlines():
                        ws.append([line])
            if not sections:
                for line in self.last_ocr_markdown.splitlines():
                    ws.append([line])
            wb.save(excel_path)
        except Exception as exc:
            excel_path = os.path.join(export_dir, f"ocr_{timestamp}.csv")
            with open(excel_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["OCR Result"])
                for line in self.last_ocr_markdown.splitlines():
                    writer.writerow([line])
            excel_note = f"⚠️ Excel导出失败({exc})，已保存为CSV"

        json_path = os.path.join(export_dir, f"ocr_{timestamp}.json")
        json_content = {
            "markdown": self.last_ocr_markdown,
            "sections": sections,
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(json_content, f, ensure_ascii=False, indent=2)

        message_lines = [
            "✅ 文本样式已保存：",
            f"- Excel: {excel_path}" + (f" ({excel_note})" if excel_note else ""),
            f"- JSON: {json_path}",
        ]
        return "\n".join(message_lines)

# 创建应用实例
# 创建应用实例
app = Qwen3VLGradioApp()

def create_interface():
    """创建Gradio界面"""
    
    with gr.Blocks(
        title="Qwen3-VL-8B-Instruct Web界面",
        theme=gr.themes.Soft(),
        css="""
        .gradio-container {
            max-width: 1600px !important;
        }
        .chat-message {
            padding: 10px;
            margin: 5px 0;
            border-radius: 10px;
        }
        #ocr-md {
            max-height: 560px;
            overflow: auto;
            border: 1px solid #eee;
            padding: 10px;
            border-radius: 6px;
            background: #fff;
        }
        """
    ) as interface:
        
        gr.Markdown("""
        # 🤖 Qwen3-VL-8B-Instruct Web界面
        
        欢迎使用Qwen3-VL-8B-Instruct多模态大语言模型！这个界面提供了友好的Web交互方式。
        
        **主要功能：**
        - 🖼️ 图像理解和对话
        - 📝 OCR文字识别
        - 🔍 空间感知分析
        - 💻 视觉编程（生成代码）
        """)
        
        with gr.Tab("🚀 模型管理"):
            gr.Markdown("### 模型加载")
            with gr.Row():
                load_btn = gr.Button("🔄 加载模型", variant="primary", size="lg")
                status_text = gr.Textbox(
                    label="状态", 
                    value="⏳ 模型未加载，请点击加载模型按钮",
                    interactive=False
                )
            
            load_btn.click(
                app.load_model,
                outputs=[status_text]
            )
        
        with gr.Tab("💬 图像对话"):
            gr.Markdown("### 与图像进行对话")
            
            with gr.Row():
                with gr.Column(scale=1):
                    image_input = gr.Image(
                        label="上传图像",
                        type="pil",
                        height=400
                    )
                    
                    with gr.Row():
                        max_tokens = gr.Slider(
                            minimum=50, maximum=1024, value=256,
                            label="最大生成长度"
                        )
                        temperature = gr.Slider(
                            minimum=0.1, maximum=2.0, value=0.7,
                            label="创造性 (Temperature)"
                        )
                
                with gr.Column(scale=2):
                    chatbot = gr.Chatbot(
                        label="对话历史",
                        height=400,
                        show_label=True,
                        render_markdown=True
                    )
                    
                    with gr.Row():
                        text_input = gr.Textbox(
                            label="输入问题",
                            placeholder="请描述这张图片...",
                            lines=2
                        )
                        send_btn = gr.Button("发送", variant="primary")
                    
                    with gr.Row():
                        clear_btn = gr.Button("🗑️ 清空历史")
            
            # 事件绑定
            send_btn.click(
                app.chat_with_image,
                inputs=[image_input, text_input, chatbot, max_tokens, temperature],
                outputs=[chatbot, text_input]
            )
            
            text_input.submit(
                app.chat_with_image,
                inputs=[image_input, text_input, chatbot, max_tokens, temperature],
                outputs=[chatbot, text_input]
            )
            
        
        with gr.Tab("📝 OCR识别"):
            gr.Markdown("### 文字识别")
            
            with gr.Row():
                with gr.Column():
                    ocr_image = gr.Image(
                        label="上传图像进行OCR识别",
                        type="pil",
                        height=300
                    )
                    ocr_btn = gr.Button("🔍 开始识别", variant="primary")
                
                with gr.Column(scale=2):
                    ocr_md = gr.Markdown(
                        value="",
                        elem_id="ocr-md"
                    )
                    save_style_btn = gr.Button("💾 导出样式", variant="secondary", interactive=False)
                    ocr_export_status = gr.Textbox(
                        label="导出状态",
                        interactive=False,
                        lines=4
                    )

            def _run_ocr(image):
                result = app.ocr_analysis(image)
                can_save = bool(app.last_ocr_markdown) and not result.startswith("❌")
                display_md = app.last_ocr_markdown if can_save else ""
                status = "" if can_save else result
                return display_md, gr.update(interactive=can_save), status

            ocr_btn.click(
                _run_ocr,
                inputs=[ocr_image],
                outputs=[ocr_md, save_style_btn, ocr_export_status]
            )

            save_style_btn.click(
                app.export_last_ocr,
                outputs=[ocr_export_status]
            )

        def _clear_all():
            app.clear_history()
            return [], gr.update(interactive=False), ""

        clear_btn.click(
            _clear_all,
            outputs=[chatbot, save_style_btn, ocr_export_status]
        )
        
        with gr.Tab("🔍 空间分析"):
            gr.Markdown("### 空间感知分析")
            
            with gr.Row():
                with gr.Column():
                    spatial_image = gr.Image(
                        label="上传图像进行空间分析",
                        type="pil",
                        height=300
                    )
                    spatial_btn = gr.Button("🔍 开始分析", variant="primary")
                
                with gr.Column():
                    spatial_result = gr.Textbox(
                        label="分析结果",
                        lines=15,
                        max_lines=20
                    )
            
            spatial_btn.click(
                app.spatial_analysis,
                inputs=[spatial_image],
                outputs=[spatial_result]
            )
        
        with gr.Tab("💻 视觉编程"):
            gr.Markdown("### 从图像生成代码")
            
            with gr.Row():
                with gr.Column():
                    code_image = gr.Image(
                        label="上传图像生成代码",
                        type="pil",
                        height=300
                    )
                    
                    code_format = gr.Dropdown(
                        choices=["HTML", "CSS", "JavaScript", "Python"],
                        value="HTML",
                        label="选择代码类型"
                    )
                    
                    code_btn = gr.Button("💻 生成代码", variant="primary")
                
                with gr.Column():
                    code_result = gr.Textbox(
                        label="生成的代码",
                        lines=15,
                        max_lines=20
                    )
            
            code_btn.click(
                app.visual_coding,
                inputs=[code_image, code_format],
                outputs=[code_result]
            )
        
        with gr.Tab("ℹ️ 使用说明"):
            gr.Markdown("""
            ## 使用说明
            
            ### 1. 模型管理
            - 首次使用需要点击"加载模型"按钮
            - 模型加载可能需要几分钟时间
            - 加载完成后可以开始使用所有功能
            
            ### 2. 图像对话
            - 上传图像后可以与其进行对话
            - 支持多轮对话，保持上下文
            - 可以调整生成参数（长度、创造性）
            
            ### 3. OCR识别
            - 上传包含文字的图像
            - 自动识别并提取所有文字内容
            - 支持32种语言识别
            
            ### 4. 空间分析
            - 分析图像中的空间关系
            - 包括物体位置、视角、遮挡关系等
            - 适用于3D场景理解
            
            ### 5. 视觉编程
            - 从图像生成各种类型的代码
            - 支持HTML、CSS、JavaScript、Python
            - 适用于UI设计和原型开发
            
            ### 注意事项
            - 确保模型路径正确：`/data/storage1/wulin/models/qwen3-vl-8b-instruct`
            - 需要足够的内存（建议16GB+）
            - 支持GPU加速（自动检测）
            """)
    
    return interface

def main():
    """主函数"""
    print("🚀 启动Qwen3-VL-8B-Instruct Web界面...")
    
    # 创建界面
    interface = create_interface()
    
    interface.queue()

    # 启动服务
    interface.launch(
        server_name="0.0.0.0",  # 允许外部访问
        server_port=7860,       # 端口
        share=False,            # 不创建公共链接
        debug=True,             # 调试模式
        show_error=True         # 显示错误
    )

if __name__ == "__main__":
    main()
