#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Qwen3-VL-8B-Instruct 统一Gradio界面
支持在同一界面内切换「通用版」与「专业版」，并提供触屏友好样式
"""

import os
import json
import inspect
import io
import hashlib
import time
import csv
import html
import numpy as np
from datetime import datetime
import shutil
import atexit
import gc

import gradio as gr
from transformers import Qwen3VLForConditionalGeneration, AutoProcessor
from ocr_card_rag_api import CardOCRWithRAG

try:
    import torch
except Exception:
    torch = None

# 统一环境变量
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'

class AdvancedQwen3VLApp:
    """高级Qwen3-VL应用类"""

    def __init__(self):
        self.model = None
        self.processor = None
        self.model_path = "\D:\cibfortune\Cibfortune\cibfortune\models\qwen3-vl-2b-instruct"
        self.is_loaded = False
        self.chat_history = []
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.chat_messages = []
        self.last_image = None
        self.last_saved_image_path = None
        self.last_image_digest = None
        self.last_ocr_markdown = None
        self.last_ocr_html = None
        # 卡证OCR多模态RAG组件
        self.card_rag_store = None
        self.card_rag_ready = False
        self.card_rag_dir = "rag_cards"
        # API 卡证OCR（RAG + Qwen API）
        self.card_api = None
        # API 票据OCR（使用qwen-vl-max模型）
        self.bill_api = None
        # 字段模板文件
        self.field_templates_file = "card_field_templates.md"
        # 当前识别的卡证类型和字段
        self.current_card_type = None
        self.current_default_fields = []
        self.current_custom_fields = []
        self.current_field_template_html = None  # 存储HTML表格结构
        self.current_final_fields_html = None  # 存储最终字段列表的HTML（包含自定义字段）

    def _ensure_card_rag_loaded(self):
        """懒加载卡证RAG图片库（若存在 rag_cards 目录），支持多种RAG实现方式。"""
        if self.card_rag_ready:
            return
        try:
            if not os.path.isdir(self.card_rag_dir):
                self.card_rag_ready = True  # 标记为已尝试，避免重复检查
                return
            
            # 优先尝试使用 multimodal_rag 模块
            try:
                from multimodal_rag import MultiModalDocumentLoader, MultiModalVectorStore
                loader = MultiModalDocumentLoader()
                docs = loader.load_images_from_folder(self.card_rag_dir)
                if not docs:
                    self.card_rag_ready = True
                    return
                store = MultiModalVectorStore(persist_directory="./multimodal_chroma_card")
                store.create_vector_store(docs)
                self.card_rag_store = store
                self.card_rag_ready = True
                print("✅ 使用multimodal_rag加载RAG图片库成功")
                return
            except Exception as e:
                print(f"⚠️ 使用multimodal_rag加载失败: {e}，尝试使用简化版RAG")
            
            # 如果multimodal_rag不可用，尝试使用SimpleRAGStore（从ocr_card_rag_api导入）
            try:
                from ocr_card_rag_api import SimpleRAGStore
                print("使用简化版RAG功能（基于卡面样式特征）...")
                store = SimpleRAGStore(use_style_features=True)
                store.load_images_from_folder(self.card_rag_dir)
                
                if not store.image_embeddings:
                    print("⚠️ RAG图片库为空")
                    self.card_rag_ready = True
                    return False
                
                self.card_rag_store = store
                self.card_rag_ready = True
                print(f"✅ 使用简化版RAG加载成功，共 {len(store.image_embeddings)} 张图片")
                return
            except Exception as e:
                print(f"⚠️ 使用简化版RAG加载失败: {e}")
            
            # 如果都失败了，标记为已尝试
            self.card_rag_ready = True
        except Exception as e:
            print(f"加载RAG图片库失败: {e}")
            self.card_rag_ready = True

    def _ensure_card_api_loaded(self):
        """懒加载卡证OCR API（RAG增强 + Qwen API 客户端）"""
        if self.card_api is not None:
            return
        try:
            api = CardOCRWithRAG(
                api_key=None,
                model="qwen3-vl-plus",
                rag_image_dir=self.card_rag_dir,
                persist_directory="./multimodal_chroma_card",
            )
            api.load_model()
            api.load_rag_library()
            self.card_api = api
        except Exception:
            self.card_api = None
        except Exception:
            # RAG 初始化失败时忽略，走纯模型路径
            self.card_rag_store = None
            self.card_rag_ready = True

    def _rag_search_card(self, image, top_k: int = 3):
        """
        对输入图片进行RAG检索，返回相似图片信息（与ocr_card_rag_api.py中的逻辑一致）
        
        Args:
            image: 输入图片（PIL Image）
            top_k: 返回最相似的k张图片
            
        Returns:
            相似图片列表，每个元素包含 {filename, similarity, metadata}
        """
        if not self.card_rag_store or not hasattr(self.card_rag_store, "image_embeddings"):
            return []
            
        try:
            # 生成查询图片的嵌入向量
            # 兼容两种实现：MultiModalVectorStore 使用 .embeddings.embed_image，SimpleRAGStore 直接使用 .embed_image
            if hasattr(self.card_rag_store, "embeddings") and hasattr(self.card_rag_store.embeddings, "embed_image"):
                # 使用 MultiModalVectorStore
                query_emb = self.card_rag_store.embeddings.embed_image(image)
            elif hasattr(self.card_rag_store, "embed_image"):
                # 使用 SimpleRAGStore
                query_emb = self.card_rag_store.embed_image(image)
            else:
                print("⚠️ RAG存储不支持embed_image方法")
                return []
            
            # 计算与图片库中所有图片的相似度
            similarities = []
            # 如果SimpleRAGStore有compute_similarity方法，使用它（支持样式相似度）
            use_compute_similarity = hasattr(self.card_rag_store, "compute_similarity")
            
            # 确保查询向量的维度
            query_dim = len(query_emb) if hasattr(query_emb, '__len__') else query_emb.shape[0] if hasattr(query_emb, 'shape') else 0
            
            for idx, emb in enumerate(self.card_rag_store.image_embeddings):
                try:
                    # 检查维度是否匹配
                    emb_dim = len(emb) if hasattr(emb, '__len__') else emb.shape[0] if hasattr(emb, 'shape') else 0
                    
                    if query_dim != emb_dim:
                        # 维度不匹配，跳过或使用默认相似度
                        print(f"⚠️ 特征维度不匹配: 查询向量={query_dim}, 图片库向量={emb_dim}，跳过该图片")
                        continue
                    
                    if use_compute_similarity:
                        # 使用样式相似度或CLIP相似度（根据SimpleRAGStore的配置）
                        similarity = self.card_rag_store.compute_similarity(query_emb, emb)
                    else:
                        # 使用余弦相似度（MultiModalVectorStore）
                        dot_product = np.dot(query_emb, emb)
                        norm_query = np.linalg.norm(query_emb)
                        norm_emb = np.linalg.norm(emb)
                        denom = norm_query * norm_emb + 1e-8
                        similarity = float(dot_product / denom) if denom > 0 else 0.0
                    similarities.append((similarity, idx))
                except Exception as e:
                    # 如果计算相似度时出错，跳过该图片
                    print(f"⚠️ 计算相似度失败（图片{idx}）: {str(e)}")
                    continue
            
            # 排序并取Top-K
            similarities.sort(key=lambda x: x[0], reverse=True)
            top_results = []
            
            for sim, idx in similarities[:top_k]:
                if idx < len(self.card_rag_store.image_metadatas):
                    meta = self.card_rag_store.image_metadatas[idx]
                    filename = meta.get("filename") or os.path.basename(meta.get("source", "")) or f"图片{idx+1}"
                    top_results.append({
                        "filename": filename,
                        "similarity": sim,
                        "metadata": meta
                    })
                    
            return top_results
            
        except Exception as e:
            print(f"⚠️ RAG检索失败: {str(e)}")
            return []

    def _build_enhanced_prompt_card(self, base_prompt: str, rag_results: list, custom_prompt: str = None):
        """
        构建增强后的提示词（包含RAG检索结果，与ocr_card_rag_api.py中的逻辑一致）
        
        Args:
            base_prompt: 基础提示词
            rag_results: RAG检索结果
            custom_prompt: 用户自定义提示词
            
        Returns:
            增强后的完整提示词
        """
        if custom_prompt:
            prompt = custom_prompt
        else:
            prompt = base_prompt
            
        # 如果有RAG检索结果，添加到提示词中
        if rag_results:
            rag_context = "\n基于图片库检索到的相似卡证：\n"
            for rank, result in enumerate(rag_results, 1):
                filename = result["filename"]
                similarity = result["similarity"]
                rag_context += f"- 卡面{rank}: {filename} | 相似度={similarity:.3f}\n"
            rag_context += "\n"
            filenames = [result["filename"].split(".")[0] for result in rag_results]
            banks = [filename.split("_")[0] for filename in filenames]
            prompt = rag_context + prompt
            prompt = prompt + (
                f"6. 如果是银行卡且字段列表包含'卡面类型'，则按照以下规则填充：\n"
                f"  - 基于图片库检索到的相似卡证结果{filenames}，填充\"卡面类型\"字段。字段值规则如下：\n"
                f"       -**禁止**自定义、生成、猜测或编造新的卡面类型值。\n"
                f"       -当出现任何不确定、模糊或不匹配情况时，\"卡面类型\"字段的值**必须且只能为\"其他\"**。\n"
                f"       -若识别出的\"发卡行\"字段的值存在与{banks}中银行名称相同的情况，"
                f"则\"卡面类型\"字段的值只能从{filenames}中**严格选择一个**。\n"
            )
            
        return prompt

    def _ensure_bill_api_loaded(self):
        """懒加载票据OCR API（使用qwen-vl-max模型）"""
        if self.bill_api is not None:
            return
        try:
            api = CardOCRWithRAG(
                api_key=None,
                model="qwen3-vl-plus",  # 票据OCR使用qwen-vl-max模型
                rag_image_dir=None,  # 票据OCR不使用RAG
                persist_directory=None,
            )
            api.load_model()
            # 票据OCR不使用RAG，跳过RAG库加载
            self.bill_api = api
        except Exception:
            self.bill_api = None

    def _load_field_templates(self):
        """从card_field_templates目录下的md文件加载字段模板"""
        templates = {}
        html_templates = {}  # 存储HTML表格内容
        templates_dir = "card_field_templates"
        
        def parse_html_table(content):
            """解析HTML表格，提取字段名称，正确处理rowspan和子字段组合"""
            try:
                from bs4 import BeautifulSoup
            except ImportError:
                print("⚠️ 需要安装beautifulsoup4来解析HTML表格")
                return []
            
            fields = []
            try:
                soup = BeautifulSoup(content, 'html.parser')
                table = soup.find('table')
                if not table:
                    return []
                
                rows = table.find_all('tr')
                if not rows:
                    return []
                
                # 子字段列表（需要与父类别组合）
                sub_fields = ['全称', '账号', '开户银行', '开户行行号', '开户行名称', '出票人', '承兑人']
                
                # 用于跟踪每个列位置的活跃rowspan类别
                # 格式: {列位置: {'name': '类别名', 'remaining_rows': 剩余行数}}
                active_rowspans = {}
                
                # 遍历每一行
                for row_idx, row in enumerate(rows):
                    cells = row.find_all(['th', 'td'])
                    if not cells:
                        continue
                    
                    # 第一步：计算每个单元格的实际列位置（考虑colspan和rowspan）
                    current_col = 0
                    row_cells_info = []
                    
                    for cell in cells:
                        text = cell.get_text(strip=True)
                        colspan = int(cell.get('colspan', 1))
                        rowspan = int(cell.get('rowspan', 1))
                        
                        # 跳过被rowspan占用的列
                        while current_col in active_rowspans:
                            current_col += 1
                        
                        cell_info = {
                            'text': text,
                            'col': current_col,
                            'colspan': colspan,
                            'rowspan': rowspan
                        }
                        row_cells_info.append(cell_info)
                        
                        current_col += colspan
                    
                    # 第三步：设置新的rowspan类别（在同一行处理字段提取之前）
                    for cell_info in row_cells_info:
                        text = cell_info['text']
                        col = cell_info['col']
                        colspan = cell_info['colspan']
                        rowspan = cell_info['rowspan']
                        
                        # 如果有rowspan，记录活跃的类别
                        # 注意：即使文本在sub_fields中，如果有rowspan，也应该作为类别处理
                        if rowspan > 1 and text:
                            for c in range(col, col + colspan):
                                active_rowspans[c] = {
                                    'name': text,
                                    'remaining_rows': rowspan - 1
                                }
                    
                    # 第四步：处理当前行的字段提取
                    for cell_info in row_cells_info:
                        text = cell_info['text']
                        col = cell_info['col']
                        colspan = cell_info['colspan']
                        rowspan = cell_info['rowspan']
                        
                        if not text:
                            continue
                        
                        # 如果该单元格有rowspan，说明它是类别，已经在上面设置了active_rowspans，跳过
                        if rowspan > 1:
                            continue
                        
                        # 检查是否是子字段
                        if text in sub_fields:
                            # 查找该列位置的活跃rowspan类别
                            parent_category = None
                            # 检查当前列及其左侧列是否有活跃的rowspan
                            for check_col in range(col, -1, -1):
                                if check_col in active_rowspans:
                                    parent_category = active_rowspans[check_col]['name']
                                    break
                            
                            if parent_category:
                                # 组合字段名：父类别 + 子字段
                                full_field = f"{parent_category}{text}"
                                if full_field not in fields:
                                    fields.append(full_field)
                            else:
                                # 没有父类别，作为独立字段（如单独的"出票人"、"承兑人"）
                                if text not in fields:
                                    fields.append(text)
                        else:
                            # 独立字段（如"出票日期"、"汇票到期日"、"票据状态"等）
                            if colspan > 1:
                                # 跨列字段，直接添加
                                if text not in fields:
                                    fields.append(text)
                            else:
                                # 单列字段，检查该列是否有活跃的rowspan（且不是当前单元格）
                                if col not in active_rowspans or active_rowspans[col]['name'] != text:
                                    if text not in fields:
                                        fields.append(text)
                    
                    # 第五步：更新rowspan剩余行数，移除已结束的（在字段提取之后）
                    for col in list(active_rowspans.keys()):
                        active_rowspans[col]['remaining_rows'] -= 1
                        if active_rowspans[col]['remaining_rows'] < 0:
                            del active_rowspans[col]
                
                # 去重并保持顺序
                fields = list(dict.fromkeys(fields))
                
                return fields
                
            except Exception as e:
                print(f"⚠️ 解析HTML表格失败: {e}")
                import traceback
                traceback.print_exc()
                return []
        
        def parse_markdown_table(content):
            """解析Markdown表格，提取字段名称列"""
            fields = []
            lines = content.split('\n')
            in_table = False
            header_found = False
            field_name_col_idx = None
            header_col_count = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # 检测表格开始（包含 | 的行）
                if '|' in line:
                    if not in_table:
                        in_table = True
                        header_found = False
                    
                    # 分割表格行（保留空字符串以保持列索引）
                    all_cells = [cell.strip() for cell in line.split('|')]
                    # 移除首尾的空字符串（Markdown表格格式：| col1 | col2 |）
                    cells = [c for c in all_cells[1:-1] if c.strip()] if len(all_cells) > 2 else [c.strip() for c in all_cells if c.strip()]
                    
                    # 处理表头
                    if not header_found and len(cells) > 0:
                        # 查找"字段名称"列的索引
                        for idx, cell in enumerate(cells):
                            if '字段名称' in cell or '字段名' in cell:
                                field_name_col_idx = idx
                                header_col_count = len(cells)
                                break
                        header_found = True
                        continue
                    
                    # 跳过分隔行（包含---的行）
                    if '---' in line or all(c in '-: ' for c in line):
                        continue
                    
                    # 提取字段名称
                    if field_name_col_idx is not None and len(cells) > 0:
                        field_name = None
                        # 如果列数与表头相同（3列），使用表头确定的列索引
                        if header_col_count and len(cells) == header_col_count:
                            if len(cells) > field_name_col_idx:
                                # 检查第一列是否是字段类别（如"出票信息"、"收款信息"等）
                                first_col = cells[0].strip() if len(cells) > 0 else ""
                                category_keywords = ['出票信息', '收款信息', '承兑信息', '承兑信息（续）', '保证信息', '保证信息（续）']
                                # 如果第一列是类别，则字段名称在第二列（索引1）
                                if first_col in category_keywords:
                                    if len(cells) > 1:
                                        field_name = cells[1].strip()
                                else:
                                    # 如果第一列不是类别，可能是字段名称在指定列
                                    field_name = cells[field_name_col_idx].strip()
                        # 如果列数不同（通常是2列），假设第一列是字段名称
                        elif len(cells) == 2:
                            field_name = cells[0].strip()
                        
                        # 添加字段名称（排除空值和类别名）
                        if field_name and field_name not in ['出票信息', '收款信息', '承兑信息', '承兑信息（续）', '保证信息', '保证信息（续）']:
                            if field_name not in fields:
                                fields.append(field_name)
                else:
                    # 如果不在表格中，尝试解析列表格式
                    if line.startswith('- '):
                        field = line.replace('- ', '').strip()
                        if field and field not in fields:
                            fields.append(field)
            
            return fields
        
        try:
            if os.path.isdir(templates_dir):
                # 从目录中的md文件加载
                for filename in os.listdir(templates_dir):
                    if filename.endswith('.md'):
                        card_type = filename.replace('.md', '')
                        filepath = os.path.join(templates_dir, filename)
                        fields = []
                        try:
                            with open(filepath, 'r', encoding='utf-8') as f:
                                content = f.read()
                                # 检查是否是HTML表格格式
                                # 兼容带属性的<table ...>，采用更宽松的检测
                                is_html = '<table' in content.lower()
                                
                                # 保存HTML内容（如果是HTML格式）
                                if is_html:
                                    html_templates[card_type] = content
                                
                                # 先尝试解析HTML表格格式
                                if is_html:
                                    fields = parse_html_table(content)
                                else:
                                    # 尝试解析Markdown表格格式
                                    fields = parse_markdown_table(content)
                                
                                # 如果表格解析没有结果，再尝试列表格式
                                if not fields:
                                    for line in content.split('\n'):
                                        line = line.strip()
                                        if line.startswith('- '):
                                            field = line.replace('- ', '').strip()
                                            if field:
                                                fields.append(field)
                            
                            if fields:
                                # 确保第一个字段是"卡证类型"（如果还没有的话）
                                if not fields or fields[0] != "卡证类型":
                                    fields.insert(0, "卡证类型")
                                templates[card_type] = fields
                                print(f"✅ 成功加载 {card_type} 字段模板，共 {len(fields)} 个字段")
                            else:
                                print(f"⚠️ {card_type} 字段模板解析结果为空，将使用默认模板")
                        except Exception as e:
                            print(f"⚠️ 加载模板文件 {filename} 失败: {e}")
                            continue
            else:
                # 如果目录不存在，使用默认模板
                templates = {
                    "身份证": ["卡证类型", "姓名", "性别", "民族", "出生日期", "住址", "公民身份号码", "签发机关", "有效期限"],
                    "银行卡": ["卡证类型", "发卡行", "卡号", "有效期", "姓名", "卡面类型"],
                    "驾驶证": ["卡证类型", "姓名", "性别", "国籍", "住址", "出生日期", "初次领证日期", "准驾车型", "有效期限", "档案编号", "证号"],
                    "护照": ["卡证类型", "姓名", "性别", "出生日期", "出生地点", "护照号码", "签发日期", "有效期至", "签发机关"],
                    "工牌": ["卡证类型", "姓名", "工号", "部门", "职位", "公司名称", "有效期"],
                    "银行承兑汇票": ["卡证类型", "出票人名称", "出票人账号", "出票人开户行", "出票人保证人姓名", "票据金额（大写）", "票据金额（小写）", "收款人名称", "收款人账号", "收款人开户行", "保证人地址", "保证日期", "承兑人名称", "承兑人账号", "承兑人开户行行号", "承兑人开户行名称", "承兑人承诺", "本汇票已承兑，到期无条件付款", "承兑日期", "交易合同号", "能否转让", "保证人姓名", "信用等级", "审查意见"],
                    "其他": ["卡证类型", "姓名", "证件号码", "有效期"]
                }
        except Exception as e:
            print(f"⚠️ 加载字段模板失败: {e}")
            # 使用默认模板
            templates = {
                "身份证": ["卡证类型", "姓名", "性别", "民族", "出生日期", "住址", "公民身份号码", "签发机关", "有效期限"],
                "银行卡": ["卡证类型", "发卡行", "卡号", "有效期", "姓名", "卡面类型"],
                "驾驶证": ["卡证类型", "姓名", "性别", "国籍", "住址", "出生日期", "初次领证日期", "准驾车型", "有效期限", "档案编号", "证号"],
                "护照": ["卡证类型", "姓名", "性别", "出生日期", "出生地点", "护照号码", "签发日期", "有效期至", "签发机关"],
                "工牌": ["卡证类型", "姓名", "工号", "部门", "职位", "公司名称", "有效期"],
                "银行承兑汇票": ["卡证类型", "出票人名称", "出票人账号", "出票人开户行", "出票人保证人姓名", "票据金额（大写）", "票据金额（小写）", "收款人名称", "收款人账号", "收款人开户行", "保证人地址", "保证日期", "承兑人名称", "承兑人账号", "承兑人开户行行号", "承兑人开户行名称", "承兑人承诺", "本汇票已承兑，到期无条件付款", "承兑日期", "交易合同号", "能否转让", "保证人姓名", "信用等级", "审查意见"],
                "其他": ["卡证类型", "姓名", "证件号码", "有效期"]
            }
        # 将HTML模板存储到实例变量中
        self.field_template_htmls = html_templates
        return templates

    def detect_card_type(self, image):
        """第一步：识别卡证类型并加载默认字段模板"""
        if image is None:
            return None, [], "❌ 请先上传图片"
        
        try:
            self._ensure_card_api_loaded()
            if self.card_api is None:
                return None, [], "❌ 卡证OCR API未初始化"
            
            # 使用简化的提示词只识别卡证类型（不包含银行承兑汇票）
            type_prompt = (
                "请识别这张图片中的卡证类型。\n"
                "只允许从以下类别中选择一种：身份证、银行卡、驾驶证、护照、工牌、其他。\n"
                "只输出卡证类型，不要输出其他内容。"
            )
            
            result = self.card_api.recognize_card(
                image,
                custom_prompt=type_prompt,
                use_rag=False,
                max_tokens=50,
                temperature=0.1
            )
            
            if not result.get("success"):
                return None, [], None, f"❌ 识别失败: {result.get('error', '未知错误')}"
            
            # 从结果中提取卡证类型（不包含银行承兑汇票）
            result_text = result.get("result", "").strip()
            card_types = ["身份证", "银行卡", "驾驶证", "护照", "工牌", "其他"]
            detected_type = None
            
            for ct in card_types:
                if ct in result_text:
                    detected_type = ct
                    break
            
            if not detected_type:
                detected_type = "其他"
            
            # 加载对应的默认字段模板（卡证OCR不使用HTML模板）
            templates = self._load_field_templates()
            default_fields = templates.get(detected_type, templates.get("其他", []))
            
            # 卡证OCR不使用HTML模板，强制设置为None
            html_template = None
            
            # 保存当前状态
            self.current_card_type = detected_type
            self.current_default_fields = default_fields.copy()
            self.current_custom_fields = []
            self.current_field_template_html = None  # 卡证OCR不使用HTML模板
            
            return detected_type, default_fields, html_template, f"✅ 识别成功：{detected_type}"
            
        except Exception as e:
            return None, [], None, f"❌ 识别失败: {str(e)}"

    def detect_bill_type(self, image):
        """票据识别第一步：识别票据类型并加载默认字段模板（使用HTML模板）"""
        supported_bill_type = ["银行承兑汇票", "商业承兑汇票", "转账支票", "现金支票", "普通支票", "本票", "付款回单", "收款回单"]

        if image is None:
            return None, [], None, "❌ 请先上传图片"
        
        try:
            self._ensure_bill_api_loaded()
            if self.bill_api is None:
                return None, [], None, "❌ 票据OCR API未初始化"
            
            # 票据OCR只识别银行承兑汇票
            type_prompt = (
                "请识别这张图片中的票据类型。\n"
                f"只允许从以下类别中选择一种：{supported_bill_type}。\n"
                "转账支票类型必须有\"转账支票\"关键词，现金支票类型必须有\"现金支票\"关键词，其他支票为普通支票\n"
                "只输出票据类型，不要输出其他内容。"
            )
            
            result = self.bill_api.recognize_card(
                image,
                custom_prompt=type_prompt,
                use_rag=False,
                max_tokens=50,
                temperature=0.1
            )
            
            if not result.get("success"):
                return None, [], None, f"❌ 识别失败: {result.get('error', '未知错误')}"
            
            # 从结果中提取票据类型
            result_text = result.get("result", "").strip()
            detected_type = None
            
            for bt in supported_bill_type:
                if bt in result_text:
                    detected_type = bt
                    break
            
            # No need to set default
            # if not detected_type:
            #     detected_type = "银行承兑汇票"  # 默认使用银行承兑汇票
            
            # 加载对应的默认字段模板（票据OCR使用HTML模板）
            templates = self._load_field_templates()
            #todo: add template of other bills
            default_fields = templates.get(detected_type, templates.get("其他票据", [])) 
            
            # 获取HTML表格内容（票据OCR必须使用HTML模板）
            html_template = getattr(self, 'field_template_htmls', {}).get(detected_type, None)
            
            # 保存当前状态
            self.current_card_type = detected_type
            self.current_default_fields = default_fields.copy()
            self.current_custom_fields = []
            self.current_field_template_html = html_template
            
            return detected_type, default_fields, html_template, f"✅ 识别成功：{detected_type}"
            
        except Exception as e:
            return None, [], None, f"❌ 识别失败: {str(e)}"

    def update_fields(self, card_type, default_fields, custom_fields_text):
        """第二步：合并默认字段和自定义字段"""
        try:
            # 解析自定义字段（每行一个字段）
            custom_fields = []
            if custom_fields_text:
                for line in custom_fields_text.strip().split('\n'):
                    field = line.strip()
                    if field and field not in default_fields:
                        custom_fields.append(field)
            
            # 合并字段
            all_fields = default_fields + custom_fields
            
            # 保存当前状态
            self.current_card_type = card_type
            self.current_default_fields = default_fields
            self.current_custom_fields = custom_fields
            
            return all_fields, f"✅ 字段已更新，共 {len(all_fields)} 个字段"
            
        except Exception as e:
            return [], f"❌ 更新字段失败: {str(e)}"

    def ocr_card_with_fields(self, image, fields_to_extract):
        """第三步：使用指定字段进行OCR识别"""
        if image is None:
            return "❌ 请先上传图片"
        
        if not fields_to_extract:
            return "❌ 请先设置要提取的字段"
        
        try:
            self._ensure_card_api_loaded()
            if self.card_api is None:
                return "❌ 卡证OCR API未初始化"
            
            # 构建包含字段列表的提示词
            fields_list = "、".join(fields_to_extract)
            
            # 卡证OCR不使用HTML模板，只使用Markdown格式
            has_html_template = False
            
            if False:  # 卡证OCR不使用HTML模板
                # 如果有HTML模板，要求大模型返回填充后的HTML表格
                custom_prompt = (
                    f"你是专业的票据/卡证OCR引擎。请阅读并识别输入图片内容，并在下面提供的HTML表格模板中填充对应字段的值。\n"
                    f"\n"
                    f"【卡证类型】{self.current_card_type or '未知'}\n"
                    f"【字段列表（必须全部覆盖，缺失填写'无'）】{fields_list}\n"
                    f"\n"
                    f"【HTML表格模板】\n"
                    f"{html_template}\n"
                    f"\n"
                    f"要求：\n"
                    f"- 只返回填充后的HTML表格（保持原有结构、行列、合并单元格和样式/属性），不要返回任何其他说明文字。\n"
                    f"- 不新增或删除字段，不改变表头文案；未识别到的填写'无'。\n"
                    f"- 仅在需要填写值的单元格写入文本，避免修改字段名单元格。\n"
                    f"- 禁止输出任何猜测或编造的内容。\n"
                    f"- 禁止输出未在字段列表中的字段和字段值。\n"
                    f"- 不要使用代码块标记符号（例如 ``` ）。"
                )
            else:
                # 如果没有HTML模板，使用原来的Markdown表格格式
                custom_prompt = (
                    f"你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                    f"\n"
                    f"任务要求：\n"
                    f"1. 识别卡证类型：{self.current_card_type or '未知'}\n"
                    f"2. 提取以下字段（必须全部提取，如果图片中没有该字段则填写'无'）：{fields_list}，禁止提取该列表以外的字段和字段值\n"
                    f"3. 以Markdown表格形式输出，表格包含两列：字段名、字段值\n"
                    f"4. 不要使用代码块标记符号（例如 ``` ）\n"
                    f"5. 输出限制：\n"
                    f"   - 最终输出只包含Markdown表格。\n"
                    f"   - 禁止输出任何猜测或编造的内容。\n"
                    f"   - 禁止输出任何其他文字或解释性内容。\n"
                    f"   - 禁止输出未在字段列表中的字段和字段值。"
                )
            
            # 只有银行卡类型才使用RAG
            use_rag = (self.current_card_type == "银行卡")
            
            result = self.card_api.recognize_card(
                image,
                custom_prompt=custom_prompt,
                use_rag=use_rag,
            )
            
            if not result.get("success"):
                return f"❌ OCR识别失败: {result.get('error', '未知错误')}"
            
            # 在终端输出RAG相似度匹配结果
            rag_info = result.get("rag_info")
            if rag_info and rag_info.get("enabled") and rag_info.get("results"):
                print("\n" + "=" * 60)
                print("📊 RAG相似度匹配结果")
                print("=" * 60)
                print(f"找到 {len(rag_info['results'])} 张相似图片：\n")
                for i, r in enumerate(rag_info["results"], 1):
                    filename = r.get("filename", "未知")
                    similarity = r.get("similarity", 0.0)
                    print(f"  {i}. {filename}")
                    print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
                print("=" * 60 + "\n")
            
            raw_result = (result.get("result") or "").strip()
            
            # 如果模型按要求直接返回HTML表格，则优先使用HTML（注入可编辑样式）
            if has_html_template and "<table" in raw_result.lower():
                try:
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(raw_result, 'html.parser')
                    table = soup.find('table')
                    if table:
                        # 添加样式使表格更美观且可编辑
                        table['class'] = (table.get('class', []) or []) + ['ocr-result-table']
                        # 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                        for tr in table.find_all('tr'):
                            if tr.get('height'):
                                del tr['height']
                            if tr.get('width'):
                                del tr['width']
                        for td in table.find_all('td'):
                            if td.get('height'):
                                del td['height']
                            if td.get('width'):
                                del td['width']
                        # 移除table的固定width属性
                        if table.get('width'):
                            del table['width']
                        if table.get('style'):
                            # 移除style中的width和height（使用Python的re模块）
                            import re
                            style = table.get('style', '')
                            style = re.sub(r'width\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                            style = re.sub(r'height\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                            style = style.strip()
                            if style:
                                table['style'] = style
                            else:
                                del table['style']
                        # 移除colgroup中的固定宽度设置
                        for colgroup in table.find_all('colgroup'):
                            for col in colgroup.find_all('col'):
                                if col.get('width'):
                                    del col['width']
                                if col.get('style'):
                                    import re
                                    style = col.get('style', '')
                                    style = re.sub(r'width\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                                    style = style.strip()
                                    if style:
                                        col['style'] = style
                                    else:
                                        del col['style']
                        # 获取所有字段名（用于识别哪些单元格是字段名，哪些是值）
                        field_names = set(fields_to_extract)
                        for td in table.find_all('td'):
                            cell_text = td.get_text(strip=True)
                            # 如果单元格文本不是字段名，且不是空，则设置为可编辑（这是值单元格）
                            if cell_text and cell_text not in field_names:
                                td['contenteditable'] = 'true'
                            # 如果单元格为空，也设置为可编辑（可能是待填充的值单元格）
                            elif not cell_text:
                                td['contenteditable'] = 'true'
                        
                        # 优化的表格样式：可调整大小的容器，表格随容器大小变化
                        styled_html = f"""
                        <style>
                        /* 可调整大小的表格容器 */
                        .ocr-result-table-container {{
                            position: relative;
                            display: inline-block;
                            min-width: 500px;
                            min-height: 300px;
                            max-width: 95vw;
                            max-height: 90vh;
                            width: 100%;
                            height: 600px;
                            resize: both;
                            overflow: auto;  /* 允许滚动，确保表格不超出容器 */
                            border: 2px solid #e0e0e0;
                            border-radius: 8px;
                            padding: 10px;
                            background-color: #f8f9fa;
                            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
                            margin: 20px 0;
                        }}
                        /* 调整大小手柄样式 */
                        .ocr-result-table-container::-webkit-resizer {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            border-radius: 0 0 8px 0;
                            width: 20px;
                            height: 20px;
                        }}
                        /* 调整大小提示 */
                        .ocr-result-table-container::before {{
                            content: '↘ 拖拽调整大小';
                            position: absolute;
                            top: 5px;
                            right: 5px;
                            font-size: 11px;
                            color: #667eea;
                            background: rgba(255, 255, 255, 0.9);
                            padding: 2px 6px;
                            border-radius: 4px;
                            pointer-events: none;
                            opacity: 0.7;
                            z-index: 5;
                            transition: opacity 0.3s ease;
                        }}
                        .ocr-result-table-container:hover::before {{
                            opacity: 1;
                        }}
                        /* 调整大小时的边框高亮 */
                        .ocr-result-table-container:active {{
                            border-color: #667eea;
                            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
                        }}
                        .ocr-result-table {{
                            width: auto;  /* 表格宽度根据内容自适应 */
                            min-width: 100%;  /* 最小宽度为容器宽度 */
                            max-width: 100%;  /* 最大宽度不超过容器 */
                            border-collapse: collapse;
                            margin: 0;
                            font-size: 14px;
                            table-layout: auto;  /* 使用auto，让列宽根据内容自动调整 */
                            box-shadow: none;
                            border-radius: 8px;
                            overflow: visible;  /* 允许内容溢出，不裁剪 */
                            background-color: #ffffff;
                        }}
                        .ocr-result-table th,
                        .ocr-result-table td {{
                            border: 1px solid #e0e0e0;
                            padding: 12px 16px;
                            text-align: left;
                            vertical-align: top;
                            word-break: break-word;
                            word-wrap: break-word;
                            transition: all 0.2s ease;
                            line-height: 1.6;
                            height: auto !important;  /* 行高根据内容自动调整，覆盖HTML中的固定height */
                            min-height: auto !important;
                            overflow: visible;  /* 允许内容显示，不裁剪 */
                            width: auto;  /* 列宽根据内容自动调整 */
                            max-width: none;  /* 不限制最大宽度 */
                        }}
                        /* 字段名列：根据内容自适应宽度 */
                        .ocr-result-table td:not([contenteditable="true"]) {{
                            background-color: #f8f9fa;
                            font-weight: 600;
                            color: #374151;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 120px;  /* 最小宽度 */
                            max-width: 300px;  /* 最大宽度限制，避免过宽 */
                            white-space: nowrap;  /* 字段名不换行 */
                            font-size: 14px;
                            border-right: 2px solid #d1d5db;
                            height: auto !important;
                            overflow: visible;
                        }}
                        /* 值列：根据内容自适应宽度 */
                        .ocr-result-table td[contenteditable="true"] {{
                            background-color: #ffffff;
                            cursor: text;
                            min-height: 20px;
                            height: auto !important;  /* 行高根据内容自动调整 */
                            position: relative;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 200px;  /* 最小宽度 */
                            max-width: none;  /* 不限制最大宽度，允许长文本 */
                            overflow: visible;  /* 允许内容显示 */
                            word-break: break-word;  /* 长文本自动换行 */
                        }}
                        /* 根据文本长度动态调整样式（保持列宽比例） */
                        .ocr-result-table td[contenteditable="true"][data-length="short"] {{
                            font-size: 15px;
                            padding: 10px 14px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="medium"] {{
                            font-size: 14px;
                            padding: 12px 16px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="long"] {{
                            font-size: 13px;
                            padding: 14px 18px;
                            line-height: 1.7;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="very-long"] {{
                            font-size: 12px;
                            padding: 16px 20px;
                            line-height: 1.8;
                            height: auto !important;
                        }}
                        .ocr-result-table th {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: #ffffff;
                            font-weight: 600;
                            font-size: 15px;
                            text-transform: uppercase;
                            letter-spacing: 0.5px;
                            border-color: #5568d3;
                        }}
                        .ocr-result-table tr:nth-child(even) {{
                            background-color: #f8f9fa;
                        }}
                        .ocr-result-table tr:nth-child(odd) {{
                            background-color: #ffffff;
                        }}
                        .ocr-result-table tr:hover {{
                            background-color: #f0f4ff;
                        }}
                        .ocr-result-table tr:hover td:not([contenteditable="true"]) {{
                            background-color: #e5e7eb;
                        }}
                        .ocr-result-table td[contenteditable="true"]:hover {{
                            background-color: #f8f9ff;
                            box-shadow: inset 0 0 0 1px #667eea;
                        }}
                        .ocr-result-table td[contenteditable="true"]:focus {{
                            outline: none;
                            background-color: #eef5ff;
                            box-shadow: inset 0 0 0 2px #667eea, 0 0 0 3px rgba(102, 126, 234, 0.1);
                            border-radius: 4px;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:before {{
                            content: "点击编辑...";
                            color: #999;
                            font-style: italic;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:focus:before {{
                            content: "";
                        }}
                        /* 优化长文本显示 */
                        .ocr-result-table td[contenteditable="true"] {{
                            overflow-wrap: break-word;
                            hyphens: auto;
                        }}
                        /* 响应式设计 */
                        @media (max-width: 768px) {{
                            .ocr-result-table-container {{
                                min-width: 300px;
                                min-height: 200px;
                            }}
                            .ocr-result-table {{
                                font-size: 12px;
                                table-layout: fixed;
                            }}
                            .ocr-result-table th,
                            .ocr-result-table td {{
                                padding: 8px 12px;
                            }}
                            .ocr-result-table td:not([contenteditable="true"]) {{
                                width: 30%;
                                font-size: 12px;
                            }}
                            .ocr-result-table td[contenteditable="true"] {{
                                width: 70%;
                            }}
                        }}
                        </style>
                        <script>
                        (function() {{
                            // 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                            function removeFixedHeights() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    // 移除table的width属性
                                    if (table.hasAttribute('width')) {{
                                        table.removeAttribute('width');
                                    }}
                                    if (table.style.width) {{
                                        table.style.width = '';
                                    }}
                                    
                                    // 移除tr的height和width属性
                                    var rows = table.querySelectorAll('tr');
                                    rows.forEach(function(row) {{
                                        if (row.hasAttribute('height')) {{
                                            row.removeAttribute('height');
                                        }}
                                        if (row.hasAttribute('width')) {{
                                            row.removeAttribute('width');
                                        }}
                                    }});
                                    
                                    // 移除td和th的height和width属性
                                    var cells = table.querySelectorAll('td, th');
                                    cells.forEach(function(cell) {{
                                        if (cell.hasAttribute('height')) {{
                                            cell.removeAttribute('height');
                                        }}
                                        if (cell.hasAttribute('width')) {{
                                            cell.removeAttribute('width');
                                        }}
                                        // 移除内联样式中的height和width
                                        if (cell.style.height) {{
                                            cell.style.height = '';
                                        }}
                                        if (cell.style.width) {{
                                            cell.style.width = '';
                                        }}
                                    }});
                                    
                                    // 移除colgroup中的width属性
                                    var colgroups = table.querySelectorAll('colgroup');
                                    colgroups.forEach(function(colgroup) {{
                                        var cols = colgroup.querySelectorAll('col');
                                        cols.forEach(function(col) {{
                                            if (col.hasAttribute('width')) {{
                                                col.removeAttribute('width');
                                            }}
                                            if (col.style.width) {{
                                                col.style.width = '';
                                            }}
                                        }});
                                    }});
                                }}
                            }}
                            
                            // 根据文本长度动态设置data-length属性
                            function updateCellLength() {{
                                var cells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                cells.forEach(function(cell) {{
                                    var text = cell.textContent || cell.innerText || '';
                                    var length = text.length;
                                    cell.removeAttribute('data-length');
                                    if (length > 0) {{
                                        if (length <= 20) {{
                                            cell.setAttribute('data-length', 'short');
                                        }} else if (length <= 50) {{
                                            cell.setAttribute('data-length', 'medium');
                                        }} else if (length <= 100) {{
                                            cell.setAttribute('data-length', 'long');
                                        }} else {{
                                            cell.setAttribute('data-length', 'very-long');
                                        }}
                                    }}
                                }});
                            }}
                            
                            // 页面加载后执行
                            setTimeout(function() {{
                                removeFixedHeights();
                                updateCellLength();
                            }}, 100);
                            
                            // 监听内容变化
                            var observer = new MutationObserver(function(mutations) {{
                                removeFixedHeights();
                                updateCellLength();
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true,
                                        attributes: true,
                                        attributeFilter: ['height', 'style']
                                    }});
                                }}
                            }}, 200);
                        }})();
                        </script>
                        <div class="ocr-result-table-container">
                            {str(table)}
                        </div>
                        """
                        self.last_ocr_html = styled_html
                        self.last_ocr_markdown = ""  # HTML模式下不生成Markdown
                        return styled_html
                    else:
                        # 如果解析失败，回退到Markdown处理
                        cleaned = self._sanitize_markdown(raw_result)
                        self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
                        self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                        return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
                except Exception as e:
                    print(f"⚠️ HTML表格解析失败，回退到Markdown格式: {e}")
                    # 解析失败，回退到Markdown处理
                    cleaned = self._sanitize_markdown(raw_result)
                    self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
                    self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                    return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
            else:
                # 否则按Markdown处理
                cleaned = self._sanitize_markdown(raw_result)
            self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
            
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def ocr_bill_with_fields(self, image, fields_to_extract):
        """票据OCR第三步：使用指定字段进行OCR识别（使用HTML模板）"""
        if image is None:
            return "❌ 请先上传图片"
        
        if not fields_to_extract:
            return "❌ 请先设置要提取的字段"
        
        try:
            self._ensure_bill_api_loaded()
            if self.bill_api is None:
                return "❌ 票据OCR API未初始化"
            
            # 构建包含字段列表的提示词
            fields_list = "、".join(fields_to_extract)
            
            # 票据OCR使用HTML模板
            html_template = getattr(self, 'current_final_fields_html', None)
            if not html_template:
                html_template = getattr(self, 'current_field_template_html', None)
            has_html_template = html_template is not None and html_template.strip()
            
            if has_html_template:
                # 如果有HTML模板，要求大模型返回填充后的HTML表格
                # 将字段列表格式化为更清晰的格式，确保模型不会遗漏
                fields_list_formatted = "\n".join([f"  {i+1}. {field}" for i, field in enumerate(fields_to_extract)])
                
                custom_prompt = (
                    f"你是专业的票据OCR引擎。请仔细阅读并识别输入图片中的所有内容，并在下面提供的HTML表格模板中填充对应字段的值。\n"
                    f"\n"
                    f"【票据类型】{self.current_card_type or '未知'}\n"
                    f"\n"
                    f"【必须识别的字段列表（共{len(fields_to_extract)}个字段，必须全部识别，一个都不能遗漏）】\n"
                    f"{fields_list_formatted}\n"
                    f"\n"
                    f"【重要要求】\n"
                    f"- **必须识别上述所有{len(fields_to_extract)}个字段，一个都不能遗漏**\n"
                    f"- 如果图片中没有某个字段的值，该字段的值必须填写'无'，但不能跳过该字段\n"
                    f"- 请仔细检查图片中的每一个位置，确保所有字段都被识别和填充\n"
                    f"- 对于组合字段（如'出票人全称'、'出票人账号'等），需要分别识别每个子字段\n"
                    f"\n"
                    f"【HTML表格模板】\n"
                    f"{html_template}\n"
                    f"\n"
                    f"【输出要求】\n"
                    f"- 只返回填充后的HTML表格（保持原有结构、行列、合并单元格和样式/属性），不要返回任何其他说明文字\n"
                    f"- 不新增或删除字段，不改变表头文案；未识别到的填写'无'\n"
                    f"- 仅在需要填写值的单元格写入文本，避免修改字段名单元格\n"
                    f"- 禁止输出任何猜测或编造的内容\n"
                    f"- 禁止输出未在字段列表中的字段和字段值\n"
                    f"- 不要使用代码块标记符号（例如 ``` ）\n"
                )
            else:
                # 如果没有HTML模板，使用Markdown表格格式（不应该发生，但作为兜底）
                custom_prompt = (
                    f"你是专业的票据OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                    f"\n"
                    f"任务要求：\n"
                    f"1. 识别票据类型：{self.current_card_type or '未知'}\n"
                    f"2. 提取以下字段（必须全部提取，如果图片中没有该字段则填写'无'）：{fields_list}，禁止提取该列表以外的字段和字段值\n"
                    f"3. 以Markdown表格形式输出，表格包含两列：字段名、字段值\n"
                    f"4. 不要使用代码块标记符号（例如 ``` ）\n"
                    f"5. 输出限制：\n"
                    f"   - 最终输出只包含Markdown表格。\n"
                    f"   - 禁止输出任何猜测或编造的内容。\n"
                    f"   - 禁止输出任何其他文字或解释性内容。\n"
                    f"   - 禁止输出未在字段列表中的字段和字段值。"
                )
            
            # 票据OCR不使用RAG
            use_rag = False
            
            # 票据OCR使用更大的max_tokens，确保能输出完整的HTML表格
            # 根据字段数量动态调整max_tokens（每个字段大约需要50-100 tokens）
            estimated_tokens = len(fields_to_extract) * 100 + 2000  # 基础2000 + 每个字段100
            max_tokens = max(2048, min(estimated_tokens, 8192))  # 最小2048，最大8192
            
            result = self.bill_api.recognize_card(
                image,
                custom_prompt=custom_prompt,
                use_rag=use_rag,
                max_tokens=max_tokens,
                temperature=0.1,  # 降低温度，提高准确性
            )
            
            if not result.get("success"):
                return f"❌ OCR识别失败: {result.get('error', '未知错误')}"
            
            raw_result = (result.get("result") or "").strip()
            
            # 如果模型按要求直接返回HTML表格，则优先使用HTML（注入可编辑样式）
            if has_html_template and "<table" in raw_result.lower():
                try:
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(raw_result, 'html.parser')
                    table = soup.find('table')
                    if table:
                        # 添加样式使表格更美观且可编辑
                        table['class'] = (table.get('class', []) or []) + ['ocr-result-table']
                        # 获取所有字段名（用于识别哪些单元格是字段名，哪些是值）
                        field_names = set(fields_to_extract)
                        for td in table.find_all('td'):
                            cell_text = td.get_text(strip=True)
                            # 如果单元格文本不是字段名，且不是空，则设置为可编辑（这是值单元格）
                            if cell_text and cell_text not in field_names:
                                td['contenteditable'] = 'true'
                            # 如果单元格为空，也设置为可编辑（可能是待填充的值单元格）
                            elif not cell_text:
                                td['contenteditable'] = 'true'
                        
                        # 优化的表格样式：可调整大小的容器，表格随容器大小变化
                        # 添加JavaScript代码，监听编辑事件并更新隐藏的Textbox
                        styled_html = f"""
                        <style>
                        /* 可调整大小的表格容器 */
                        .ocr-result-table-container {{
                            position: relative;
                            display: inline-block;
                            min-width: 500px;
                            min-height: 300px;
                            max-width: 95vw;
                            max-height: 90vh;
                            width: 100%;
                            height: 600px;
                            resize: both;
                            overflow: auto;  /* 允许滚动，确保表格不超出容器 */
                            border: 2px solid #e0e0e0;
                            border-radius: 8px;
                            padding: 10px;
                            background-color: #f8f9fa;
                            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
                            margin: 20px 0;
                        }}
                        /* 调整大小手柄样式 */
                        .ocr-result-table-container::-webkit-resizer {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            border-radius: 0 0 8px 0;
                            width: 20px;
                            height: 20px;
                        }}
                        /* 调整大小提示 */
                        .ocr-result-table-container::before {{
                            content: '↘ 拖拽调整大小';
                            position: absolute;
                            top: 5px;
                            right: 5px;
                            font-size: 11px;
                            color: #667eea;
                            background: rgba(255, 255, 255, 0.9);
                            padding: 2px 6px;
                            border-radius: 4px;
                            pointer-events: none;
                            opacity: 0.7;
                            z-index: 5;
                            transition: opacity 0.3s ease;
                        }}
                        .ocr-result-table-container:hover::before {{
                            opacity: 1;
                        }}
                        /* 调整大小时的边框高亮 */
                        .ocr-result-table-container:active {{
                            border-color: #667eea;
                            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
                        }}
                        .ocr-result-table {{
                            width: auto;  /* 表格宽度根据内容自适应 */
                            min-width: 100%;  /* 最小宽度为容器宽度 */
                            max-width: 100%;  /* 最大宽度不超过容器 */
                            border-collapse: collapse;
                            margin: 0;
                            font-size: 14px;
                            table-layout: auto;  /* 使用auto，让列宽根据内容自动调整 */
                            box-shadow: none;
                            border-radius: 8px;
                            overflow: visible;  /* 允许内容溢出，不裁剪 */
                            background-color: #ffffff;
                        }}
                        .ocr-result-table th,
                        .ocr-result-table td {{
                            border: 1px solid #e0e0e0;
                            padding: 12px 16px;
                            text-align: left;
                            vertical-align: top;
                            word-break: break-word;
                            word-wrap: break-word;
                            transition: all 0.2s ease;
                            line-height: 1.6;
                            height: auto !important;  /* 行高根据内容自动调整，覆盖HTML中的固定height */
                            min-height: auto !important;
                            overflow: visible;  /* 允许内容显示，不裁剪 */
                            width: auto;  /* 列宽根据内容自动调整 */
                            max-width: none;  /* 不限制最大宽度 */
                        }}
                        /* 字段名列：根据内容自适应宽度 */
                        .ocr-result-table td:not([contenteditable="true"]) {{
                            background-color: #f8f9fa;
                            font-weight: 600;
                            color: #374151;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 120px;  /* 最小宽度 */
                            max-width: 300px;  /* 最大宽度限制，避免过宽 */
                            white-space: nowrap;  /* 字段名不换行 */
                            font-size: 14px;
                            border-right: 2px solid #d1d5db;
                            height: auto !important;
                            overflow: visible;
                        }}
                        /* 值列：根据内容自适应宽度 */
                        .ocr-result-table td[contenteditable="true"] {{
                            background-color: #ffffff;
                            cursor: text;
                            min-height: 20px;
                            height: auto !important;  /* 行高根据内容自动调整 */
                            position: relative;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 200px;  /* 最小宽度 */
                            max-width: none;  /* 不限制最大宽度，允许长文本 */
                            overflow: visible;  /* 允许内容显示 */
                            word-break: break-word;  /* 长文本自动换行 */
                        }}
                        /* 根据文本长度动态调整样式（保持列宽比例） */
                        .ocr-result-table td[contenteditable="true"][data-length="short"] {{
                            font-size: 15px;
                            padding: 10px 14px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="medium"] {{
                            font-size: 14px;
                            padding: 12px 16px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="long"] {{
                            font-size: 13px;
                            padding: 14px 18px;
                            line-height: 1.7;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="very-long"] {{
                            font-size: 12px;
                            padding: 16px 20px;
                            line-height: 1.8;
                            height: auto !important;
                        }}
                        .ocr-result-table th {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: #ffffff;
                            font-weight: 600;
                            font-size: 15px;
                            text-transform: uppercase;
                            letter-spacing: 0.5px;
                            border-color: #5568d3;
                        }}
                        .ocr-result-table tr:nth-child(even) {{
                            background-color: #f8f9fa;
                        }}
                        .ocr-result-table tr:nth-child(odd) {{
                            background-color: #ffffff;
                        }}
                        .ocr-result-table tr:hover {{
                            background-color: #f0f4ff;
                        }}
                        .ocr-result-table tr:hover td:not([contenteditable="true"]) {{
                            background-color: #e5e7eb;
                        }}
                        .ocr-result-table td[contenteditable="true"]:hover {{
                            background-color: #f8f9ff;
                            box-shadow: inset 0 0 0 1px #667eea;
                        }}
                        .ocr-result-table td[contenteditable="true"]:focus {{
                            outline: none;
                            background-color: #eef5ff;
                            box-shadow: inset 0 0 0 2px #667eea, 0 0 0 3px rgba(102, 126, 234, 0.1);
                            border-radius: 4px;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:before {{
                            content: "点击编辑...";
                            color: #999;
                            font-style: italic;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:focus:before {{
                            content: "";
                        }}
                        /* 优化长文本显示 */
                        .ocr-result-table td[contenteditable="true"] {{
                            overflow-wrap: break-word;
                            hyphens: auto;
                        }}
                        /* 响应式设计 */
                        @media (max-width: 768px) {{
                            .ocr-result-table-container {{
                                min-width: 300px;
                                min-height: 200px;
                            }}
                            .ocr-result-table {{
                                font-size: 12px;
                                table-layout: fixed;
                            }}
                            .ocr-result-table th,
                            .ocr-result-table td {{
                                padding: 8px 12px;
                            }}
                            .ocr-result-table td:not([contenteditable="true"]) {{
                                width: 30%;
                                font-size: 12px;
                            }}
                            .ocr-result-table td[contenteditable="true"] {{
                                width: 70%;
                            }}
                        }}
                        </style>
                        <script>
                        (function() {{
                            // 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                            function removeFixedHeights() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    // 移除table的width属性
                                    if (table.hasAttribute('width')) {{
                                        table.removeAttribute('width');
                                    }}
                                    if (table.style.width) {{
                                        table.style.width = '';
                                    }}
                                    
                                    // 移除tr的height和width属性
                                    var rows = table.querySelectorAll('tr');
                                    rows.forEach(function(row) {{
                                        if (row.hasAttribute('height')) {{
                                            row.removeAttribute('height');
                                        }}
                                        if (row.hasAttribute('width')) {{
                                            row.removeAttribute('width');
                                        }}
                                    }});
                                    
                                    // 移除td和th的height和width属性
                                    var cells = table.querySelectorAll('td, th');
                                    cells.forEach(function(cell) {{
                                        if (cell.hasAttribute('height')) {{
                                            cell.removeAttribute('height');
                                        }}
                                        if (cell.hasAttribute('width')) {{
                                            cell.removeAttribute('width');
                                        }}
                                        // 移除内联样式中的height和width
                                        if (cell.style.height) {{
                                            cell.style.height = '';
                                        }}
                                        if (cell.style.width) {{
                                            cell.style.width = '';
                                        }}
                                    }});
                                    
                                    // 移除colgroup中的width属性
                                    var colgroups = table.querySelectorAll('colgroup');
                                    colgroups.forEach(function(colgroup) {{
                                        var cols = colgroup.querySelectorAll('col');
                                        cols.forEach(function(col) {{
                                            if (col.hasAttribute('width')) {{
                                                col.removeAttribute('width');
                                            }}
                                            if (col.style.width) {{
                                                col.style.width = '';
                                            }}
                                        }});
                                    }});
                                }}
                            }}
                            
                            // 根据文本长度动态设置data-length属性
                            function updateCellLength() {{
                                var cells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                cells.forEach(function(cell) {{
                                    var text = cell.textContent || cell.innerText || '';
                                    var length = text.length;
                                    cell.removeAttribute('data-length');
                                    if (length > 0) {{
                                        if (length <= 20) {{
                                            cell.setAttribute('data-length', 'short');
                                        }} else if (length <= 50) {{
                                            cell.setAttribute('data-length', 'medium');
                                        }} else if (length <= 100) {{
                                            cell.setAttribute('data-length', 'long');
                                        }} else {{
                                            cell.setAttribute('data-length', 'very-long');
                                        }}
                                    }}
                                }});
                            }}
                            
                            // 页面加载后执行
                            setTimeout(function() {{
                                removeFixedHeights();
                                updateCellLength();
                            }}, 100);
                            
                            // 监听内容变化
                            var observer = new MutationObserver(function(mutations) {{
                                removeFixedHeights();
                                updateCellLength();
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true,
                                        attributes: true,
                                        attributeFilter: ['height', 'style']
                                    }});
                                }}
                            }}, 200);
                        }})();
                        </script>
                        <div class="ocr-result-table-container">
                            {str(table)}
                        </div>
                        <script>
                        (function() {{
                            var updateTimeout = null;
                            
                            function updateEditedContent() {{
                                // 清除之前的定时器
                                if (updateTimeout) {{
                                    clearTimeout(updateTimeout);
                                }}
                                
                                // 延迟更新，避免频繁触发
                                updateTimeout = setTimeout(function() {{
                                    var table = document.querySelector('.ocr-result-table');
                                    if (!table) return;
                                    
                                    // 获取完整的HTML（包括样式）
                                    var fullHtml = document.querySelector('#bill-ocr-result-html, [id*="bill-ocr-result-html"]');
                                    var htmlContent = '';
                                    
                                    if (fullHtml) {{
                                        // 获取包含表格的完整HTML
                                        var container = fullHtml.querySelector('.ocr-result-table') || fullHtml;
                                        htmlContent = container.innerHTML;
                                    }} else {{
                                        // 如果没有找到容器，直接获取表格的outerHTML
                                        htmlContent = table.outerHTML;
                                    }}
                                    
                                    // 查找隐藏的Textbox - 使用多种方法
                                    var hiddenInput = null;
                                    
                                    // 方法1: 直接通过ID查找
                                    hiddenInput = document.getElementById('bill-ocr-result-html-edited');
                                    
                                    // 方法2: 通过ID包含关键字查找
                                    if (!hiddenInput) {{
                                        var inputs = document.querySelectorAll('input, textarea');
                                        for (var i = 0; i < inputs.length; i++) {{
                                            if (inputs[i].id && inputs[i].id.includes('bill-ocr-result-html-edited')) {{
                                                hiddenInput = inputs[i];
                                                break;
                                            }}
                                        }}
                                    }}
                                    
                                    // 方法3: 通过name属性查找
                                    if (!hiddenInput) {{
                                        hiddenInput = document.querySelector('input[name*="bill-ocr-result-html-edited"], textarea[name*="bill-ocr-result-html-edited"]');
                                    }}
                                    
                                    // 方法4: 通过data属性或class查找
                                    if (!hiddenInput) {{
                                        var allInputs = document.querySelectorAll('input[type="text"], textarea');
                                        for (var i = 0; i < allInputs.length; i++) {{
                                            var input = allInputs[i];
                                            // 检查是否在Gradio的隐藏组件区域
                                            if (input.style.display === 'none' || input.hidden || input.offsetParent === null) {{
                                                // 尝试设置值，看是否能找到正确的输入框
                                                var testValue = input.value;
                                                input.value = 'TEST_' + Date.now();
                                                if (input.value === 'TEST_' + Date.now()) {{
                                                    input.value = testValue; // 恢复原值
                                                    // 这可能是我们要找的输入框，但需要更精确的匹配
                                                }}
                                            }}
                                        }}
                                    }}
                                    
                                    if (hiddenInput) {{
                                        // 获取完整的HTML内容（包括样式）
                                        var styleTag = document.querySelector('style');
                                        var styleContent = styleTag ? styleTag.outerHTML : '';
                                        var fullContent = styleContent + '\\n' + table.outerHTML;
                                        
                                        hiddenInput.value = fullContent;
                                        
                                        // 触发多种事件，确保Gradio捕获到变化
                                        var events = ['input', 'change', 'blur', 'keyup'];
                                        events.forEach(function(eventType) {{
                                            var event = new Event(eventType, {{ bubbles: true, cancelable: true }});
                                            hiddenInput.dispatchEvent(event);
                                        }});
                                        
                                        // 也尝试直接设置属性
                                        if (hiddenInput.setAttribute) {{
                                            hiddenInput.setAttribute('value', fullContent);
                                        }}
                                        
                                        console.log('[DEBUG] 已更新隐藏Textbox，内容长度:', fullContent.length);
                                    }} else {{
                                        console.warn('[DEBUG] 未找到隐藏的Textbox组件');
                                        // 如果找不到，尝试通过window对象存储
                                        if (window.gradioEditedContent === undefined) {{
                                            window.gradioEditedContent = {{}};
                                        }}
                                        window.gradioEditedContent['bill-ocr-result-html-edited'] = htmlContent;
                                    }}
                                }}, 300);
                            }}
                            
                            // 监听所有可编辑单元格的输入事件
                            function attachListeners() {{
                                var editableCells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                editableCells.forEach(function(cell) {{
                                    // 移除旧的监听器（如果存在）
                                    var newCell = cell.cloneNode(true);
                                    cell.parentNode.replaceChild(newCell, cell);
                                    
                                    // 添加新的监听器
                                    newCell.addEventListener('input', updateEditedContent);
                                    newCell.addEventListener('blur', updateEditedContent);
                                    newCell.addEventListener('keyup', updateEditedContent);
                                    newCell.addEventListener('paste', function() {{
                                        setTimeout(updateEditedContent, 100);
                                    }});
                                }});
                                
                                // 初始更新
                                updateEditedContent();
                            }}
                            
                            // 延迟执行，确保DOM已加载
                            setTimeout(attachListeners, 500);
                            
                            // 使用MutationObserver监听表格变化（动态添加的单元格）
                            var observer = new MutationObserver(function(mutations) {{
                                var shouldReattach = false;
                                mutations.forEach(function(mutation) {{
                                    if (mutation.type === 'childList' && mutation.addedNodes.length > 0) {{
                                        shouldReattach = true;
                                    }}
                                }});
                                if (shouldReattach) {{
                                    setTimeout(attachListeners, 100);
                                }}
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true
                                    }});
                                }}
                            }}, 500);
                            
                            // 页面卸载前保存
                            window.addEventListener('beforeunload', updateEditedContent);
                            
                            // 监听导出按钮点击事件，在导出前强制更新内容
                            function setupExportButton() {{
                                var exportBtn = document.getElementById('bill-ocr-export-btn') || 
                                               document.querySelector('button[id*="bill-ocr-export-btn"]') ||
                                               document.querySelector('button:contains("导出结果")');
                                
                                if (exportBtn) {{
                                    exportBtn.addEventListener('click', function(e) {{
                                        console.log('[DEBUG] 导出按钮被点击，强制更新内容...');
                                        // 立即更新内容，不延迟
                                        var table = document.querySelector('.ocr-result-table');
                                        if (table) {{
                                            var styleTag = document.querySelector('style');
                                            var styleContent = styleTag ? styleTag.outerHTML : '';
                                            // 获取编辑后的表格HTML（包含所有用户编辑的内容）
                                            var tableHtml = table.outerHTML;
                                            var fullContent = styleContent + '\\n' + tableHtml;
                                            
                                            console.log('[DEBUG] 获取到的表格HTML长度:', tableHtml.length);
                                            console.log('[DEBUG] 表格内容预览:', tableHtml.substring(0, 200));
                                            
                                            // 查找隐藏的Textbox - 使用多种方法
                                            var hiddenInput = null;
                                            
                                            // 方法1: 直接通过ID查找
                                            hiddenInput = document.getElementById('bill-ocr-result-html-edited');
                                            
                                            // 方法2: 通过ID包含关键字查找
                                            if (!hiddenInput) {{
                                                var inputs = document.querySelectorAll('input, textarea');
                                                for (var i = 0; i < inputs.length; i++) {{
                                                    if (inputs[i].id && inputs[i].id.includes('bill-ocr-result-html-edited')) {{
                                                        hiddenInput = inputs[i];
                                                        break;
                                                    }}
                                                }}
                                            }}
                                            
                                            // 方法3: 查找所有隐藏的输入框
                                            if (!hiddenInput) {{
                                                var allInputs = document.querySelectorAll('input[type="text"], textarea');
                                                for (var i = 0; i < allInputs.length; i++) {{
                                                    var input = allInputs[i];
                                                    // 检查是否是隐藏的组件
                                                    if ((input.style.display === 'none' || input.hidden || input.offsetParent === null) &&
                                                        input.id && input.id.includes('bill')) {{
                                                        hiddenInput = input;
                                                        break;
                                                    }}
                                                }}
                                            }}
                                            
                                            if (hiddenInput) {{
                                                console.log('[DEBUG] 找到隐藏Textbox，ID:', hiddenInput.id);
                                                hiddenInput.value = fullContent;
                                                
                                                // 触发所有可能的事件，确保Gradio捕获到变化
                                                var events = ['input', 'change', 'blur', 'keyup', 'focus'];
                                                events.forEach(function(eventType) {{
                                                    try {{
                                                        var event = new Event(eventType, {{ bubbles: true, cancelable: true }});
                                                        hiddenInput.dispatchEvent(event);
                                                    }} catch(err) {{
                                                        console.error('触发事件失败:', eventType, err);
                                                    }}
                                                }});
                                                
                                                // 也尝试直接设置属性
                                                if (hiddenInput.setAttribute) {{
                                                    hiddenInput.setAttribute('value', fullContent);
                                                }}
                                                
                                                console.log('[DEBUG] 导出前已强制更新，内容长度:', fullContent.length);
                                                console.log('[DEBUG] Textbox当前值长度:', hiddenInput.value.length);
                                            }} else {{
                                                console.error('[DEBUG] 导出前未找到隐藏Textbox，尝试所有输入框...');
                                                var allInputs = document.querySelectorAll('input, textarea');
                                                console.log('[DEBUG] 找到', allInputs.length, '个输入框');
                                                for (var i = 0; i < Math.min(allInputs.length, 10); i++) {{
                                                    console.log('  输入框', i, ':', allInputs[i].id, allInputs[i].name, allInputs[i].className);
                                                }}
                                            }}
                                        }} else {{
                                            console.error('[DEBUG] 未找到表格元素');
                                        }}
                                    }}, true); // 使用捕获阶段，确保先执行
                                }} else {{
                                    // 如果按钮还没加载，延迟重试
                                    setTimeout(setupExportButton, 500);
                                }}
                            }}
                            
                            // 延迟设置导出按钮监听器
                            setTimeout(setupExportButton, 1000);
                        }})();
                        </script>
                        """
                        self.last_ocr_html = styled_html
                        self.last_ocr_markdown = ""  # HTML模式下不生成Markdown
                        return styled_html
                    else:
                        # 如果解析失败，回退到Markdown处理
                        cleaned = self._sanitize_markdown(raw_result)
                        self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                        self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                        return f"🪪 票据OCR识别结果:\n\n{cleaned}"
                except Exception as e:
                    print(f"⚠️ HTML表格解析失败，回退到Markdown格式: {e}")
                    # 解析失败，回退到Markdown处理
                    cleaned = self._sanitize_markdown(raw_result)
                    self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                    self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                    return f"🪪 票据OCR识别结果:\n\n{cleaned}"
            else:
                # 否则按Markdown处理
                cleaned = self._sanitize_markdown(raw_result)
                self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                return f"🪪 票据OCR识别结果:\n\n{cleaned}"
            
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def load_model(self, progress=gr.Progress()):
        """加载模型"""
        if self.is_loaded:
            return "✅ 模型已经加载完成！", gr.update(interactive=True)

        if torch is None:
            return "❌ 模型加载失败: 未检测到PyTorch，请先安装。", gr.update(interactive=False)

        try:
            progress(0.1, desc="检查模型路径...")
            if not os.path.exists(self.model_path):
                return f"❌ 模型路径不存在: {self.model_path}", gr.update(interactive=False)

            progress(0.3, desc="加载模型...")
            self.model = Qwen3VLForConditionalGeneration.from_pretrained(
                self.model_path,
                dtype="auto",
                device_map="cuda",
                load_in_4bit=True,
            )

            progress(0.7, desc="加载处理器...")
            self.processor = AutoProcessor.from_pretrained(self.model_path)
            print("加载处理器")
            progress(1.0, desc="完成！")
            self.is_loaded = True

            return "✅ 模型加载成功！可以开始使用了。", gr.update(interactive=True)

        except Exception as e:
            return f"❌ 模型加载失败: {str(e)}", gr.update(interactive=False)

    def _prepare_user_message(self, image, prompt):
        prompt_clean = (prompt or "").strip()
        resolved_image = image if image is not None else self.last_image
        if resolved_image is None:
            raise ValueError("❌ 请上传图像！")
        if not prompt_clean:
            raise ValueError("❌ 请输入问题！")
        if image is not None:
            self.last_image = image
        content = [
            {"type": "image", "image": resolved_image},
            {"type": "text", "text": prompt_clean},
        ]
        return prompt_clean, {"role": "user", "content": content}

    def _run_inference(self,
                       image,
                       prompt,
                       max_tokens,
                       temperature,
                       top_p,
                       top_k,
                       repetition_penalty,
                       prepared=None):
        if prepared is None:
            prompt_clean, user_message = self._prepare_user_message(image, prompt)
        else:
            prompt_clean, user_message = prepared
        messages = self.chat_messages + [user_message]

        inputs = self.processor.apply_chat_template(
            messages,
            tokenize=True,
            add_generation_prompt=True,
            return_dict=True,
            return_tensors="pt"
        ).to(self.model.device)

        generation_kwargs = {
            "max_new_tokens": max_tokens,
            "temperature": temperature,
            "top_p": top_p,
            "top_k": top_k,
            "do_sample": True if temperature > 0 else False,
            "repetition_penalty": repetition_penalty
        }

        start_time = time.time()
        with torch.no_grad():
            generated_ids = self.model.generate(**inputs, **generation_kwargs)
        generation_time = time.time() - start_time

        generated_ids_trimmed = [
            out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text = self.processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )
        response = output_text[0]

        assistant_message = {"role": "assistant", "content": [{"type": "text", "text": response}]}
        self.chat_messages.extend([user_message, assistant_message])
        return prompt_clean, response, generation_time

    def _clone_history(self, history):
        return [[turn[0], turn[1]] for turn in history]

    def _chunk_response(self, text, chunk_size=80):
        if not text:
            return []
        return [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]

    def _parse_markdown_sections(self, markdown_text):
        """
        将 Markdown 文本拆分为 table/text 段，支持：
        - 管道表格（| a | b |）
        - HTML <table>（若存在）
        并在解析前对围栏代码块进行去围栏清洗，确保导出与渲染一致。
        """
        sections = []
        if not markdown_text:
            return sections

        # 1) 先去掉围栏，使得“代码块中的表格”也能被识别为可导出的内容
        cleaned_md = self._sanitize_markdown(markdown_text)

        # 2) 先尝试解析 HTML 表格（若模型输出了 <table>）
        html_tables = []
        try:
            from bs4 import BeautifulSoup  # 可选依赖
            soup = BeautifulSoup(cleaned_md, "html.parser")
            for t in soup.find_all("table"):
                headers = []
                header_row = t.find("tr")
                if header_row:
                    # 如果有 <th> 用 th；否则用首行的 td 作为 header
                    ths = header_row.find_all("th")
                    if ths:
                        headers = [th.get_text(strip=True) for th in ths]
                        data_rows = header_row.find_next_siblings("tr")
                    else:
                        tds = header_row.find_all("td")
                        headers = [td.get_text(strip=True) for td in tds]
                        data_rows = header_row.find_next_siblings("tr")
                rows = []
                for r in (data_rows or []):
                    cols = r.find_all(["td", "th"])
                    rows.append([c.get_text(strip=True) for c in cols])
                if headers or rows:
                    html_tables.append({"type": "table", "header": headers, "rows": rows})
        except Exception:
            # 如果 bs4 不在环境中，则略过 HTML 解析
            pass

        # 3) 解析管道表格
        lines = cleaned_md.splitlines()
        i = 0
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()

            # 管道表格判定：当前行和下一行构成 header + 分隔
            is_table = (
                stripped.startswith("|")
                and stripped.count("|") >= 2
                and i + 1 < len(lines)
                and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                and lines[i + 1].strip().startswith("|")
            )

            if is_table:
                header = [cell.strip() for cell in stripped.strip("|").split("|")]
                i += 2  # 跳过 header 与分隔线
                rows = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if not (row_line.startswith("|") and row_line.count("|") >= 2):
                        break
                    row = [cell.strip() for cell in row_line.strip("|").split("|")]
                    rows.append(row)
                    i += 1
                sections.append({"type": "table", "header": header, "rows": rows})
                continue

            # 普通文本块（直到遇到下一个表格或文件结束）
            text_block = []
            while i < len(lines):
                current = lines[i]
                stripped_current = current.strip()
                next_is_table = (
                    stripped_current.startswith("|")
                    and stripped_current.count("|") >= 2
                    and i + 1 < len(lines)
                    and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                    and lines[i + 1].strip().startswith("|")
                )
                if next_is_table:
                    break
                text_block.append(current)
                i += 1
                # 保留空行，改善段落分隔的可读性
                if i < len(lines) and lines[i] == "":
                    text_block.append(lines[i])

            text_content = "\n".join(text_block).strip("\n")
            if text_content:
                sections.append({"type": "text", "text": text_content})

        # 4) 若存在 HTML 表，优先把 HTML 表也加入（放在解析结果前面，避免遗漏）
        if html_tables:
            # 将 HTML 表插在最前面（也可根据需要合并/去重）
            sections = html_tables + sections

        return sections

    @staticmethod
    def _text_to_html_block(text: str) -> str:
        if not text:
            return ""
        escaped = html.escape(text)
        replaced = escaped.replace("\n", "<br>")
        return f'<div class="ocr-text">{replaced}</div>'

    @staticmethod
    def _table_to_html_block(header, rows) -> str:
        header = header or []
        rows = rows or []
        thead = ""
        if header:
            header_cells = "".join(f"<th>{html.escape(str(cell))}</th>" for cell in header)
            thead = f"<thead><tr>{header_cells}</tr></thead>"
        body_rows = []
        for row in rows:
            row_cells = "".join(f"<td>{html.escape(str(cell))}</td>" for cell in row)
            body_rows.append(f"<tr>{row_cells}</tr>")
        tbody = f"<tbody>{''.join(body_rows)}</tbody>" if body_rows else "<tbody></tbody>"
        return f'<table class="ocr-table">{thead}{tbody}</table>'

    def _render_sections_as_html(self, markdown_text: str) -> str:
        if not markdown_text:
            return ""
        sections = self._parse_markdown_sections(markdown_text)
        if not sections:
            escaped = html.escape(markdown_text.strip())
            return f"<pre>{escaped}</pre>" if escaped else ""
        blocks = []
        for section in sections:
            if section.get("type") == "table":
                blocks.append(self._table_to_html_block(section.get("header"), section.get("rows")))
            elif section.get("type") == "text":
                blocks.append(self._text_to_html_block(section.get("text", "")))
        return '<div class="ocr-preview">' + "".join(blocks) + "</div>"

    def chat_with_image(self, image, text, history, max_tokens, temperature, top_p, top_k, repetition_penalty: float = 1.0, presence_penalty: float = 1.5):
        """与图像对话（流式反馈）"""
        original_text = text

        if not self.is_loaded:
            yield history, original_text, "❌ 请先加载模型！"
            return

        try:
            prepared = self._prepare_user_message(image, text)
        except ValueError as exc:
            yield history, original_text, str(exc)
            return

        prompt_clean, _ = prepared
        history_copy = self._clone_history(history)
        history_copy.append([f"👤 {prompt_clean}", "🤖 正在思考..."])
        yield self._clone_history(history_copy), original_text, "🤖 正在思考..."

        try:
            _, response, generation_time = self._run_inference(
                image,
                text,
                max_tokens,
                temperature,
                top_p,
                top_k,
                repetition_penalty,
                prepared=prepared
            )
        except Exception as e:
            history_copy[-1][1] = f"❌ 生成失败: {str(e)}"
            self.chat_history = self._clone_history(history_copy)
            yield self._clone_history(history_copy), original_text, f"❌ 错误: {str(e)}"
            return

        assembled = ""
        chunks = self._chunk_response(response)
        if not chunks:
            chunks = [""]
        for chunk in chunks:
            assembled += chunk
            history_copy[-1][1] = f"🤖 {assembled}▌"
            yield self._clone_history(history_copy), original_text, f"🤖 {assembled}▌"

        stats = (
            f"⏱️ 生成时间: {generation_time:.2f}秒 | 📝 生成长度: {len(response)}字符"
            f" | ⚙️ 最大长度: {max_tokens}"
        )
        if max_tokens > 1024:
            stats += " | ⏳ 提示: 较大的最大长度可能延长生成时间"
        history_copy[-1][1] = f"🤖 {response}"
        self.chat_history = self._clone_history(history_copy)
        yield self._clone_history(history_copy), original_text, stats

    def _sanitize_markdown(self, text: str) -> str:
        if not text:
            return ""
        s = text.strip()
        lines = s.splitlines()
        out = []
        in_fence = False
        for line in lines:
            stripped = line.strip()
            if stripped.startswith("```"):
                in_fence = not in_fence
                continue
            if not in_fence:
                out.append(line)
        cleaned = "\n".join(out).strip()
        return cleaned if cleaned else s

    def ocr_analysis(self, image, prompt: str = None):
        """OCR文字识别，可选自定义提示词"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "请识别并提取这张图片中的所有文字内容，尽量还原原本样式，并标注语言类型。"
            " 请确保所有带样式或表格内容使用Markdown表格表示。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1024,
                temperature=0.7,
                top_p=0.8,
                top_k=50,
                repetition_penalty=1.0
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"📝 OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def ocr_card(self, image, prompt: str = None):
        """卡证OCR识别：身份证/银行卡/驾驶证等结构化提取（使用本地模型，流程与API版本一致）"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        
        # 使用与ocr_card_api相同的默认提示词
        default_prompt = (
                "你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                "\n"
                "任务要求如下：\n"
                "\n"
            "1. 识别卡证类型：只允许从以下类别中选择一种：\n"
            "   - 身份证 / 银行卡 / 驾驶证 / 护照 / 工牌 / 其他。\n"
            "   Markdown表格中添加“卡证类型”字段，并用类别选择赋值。\n"
            "   **重要**：如果识别为银行卡，必须严格遵守第3条银行卡特殊要求！\n"
            "\n"
            "2. 输出格式：\n"
            "   - 以Markdown表格形式输出所有识别出的关键字段及其对应的值。\n"
            "   - 若字段中包含“卡号”，请确保该字段的值仅包含数字。\n"
            "   - 不要使用代码块标记符号（例如 ``` ）。\n"
            "\n"
            "3. 银行卡特殊要求（必须严格遵守）：\n"
            "   如果识别的卡证类型是银行卡，必须在Markdown表格的最后额外添加一个字段：\n"
            "   - 字段名：卡面类型（必须添加，不可省略）。\n"
            "   - 基于图片库检索到的相似卡证结果，填充“卡面类型”字段。字段值规则如下：\n"
            "       ① 当出现任何不确定、模糊或不匹配情况时，“卡面类型”字段的值**必须且只能为“其他”**，不得填写相似图片名或其他文本。\n"
            "       ② 若识别出的“发卡行”字段的值与这些相似卡证文件名中`_`前面的银行名称相同，"
            "则“卡面类型”字段的值只能从相似卡证文件名中**严格选择一个**，格式为`银行名称_卡面类型`，去掉文件后缀名，如`中国银行_visa卡`。\n"
            "       ③ 禁止自定义、生成、猜测或编造新的卡面类型值。任何不存在基于图片库检索到的相似卡证文件名的值都视为错误。\n"
            "   **重要提醒**：银行卡的Markdown表格必须包含“卡面类型”字段，这是强制要求，不能省略！\n"
            "   - 如果不是银行卡，则不添加“卡面类型”字段。\n"
            "\n"
            "4. 输出限制：\n"
            "   - 最终输出只包含Markdown表格。\n"
            "   - 禁止输出任何其他文字或解释性内容。\n"
            "   - 如果是银行卡，表格中必须包含“卡面类型”字段，否则输出不完整。\n"
        )

        effective_prompt = (prompt or "").strip() or default_prompt

        # RAG检索（使用与API版本相同的逻辑）
        rag_results = []
        try:
            self._ensure_card_rag_loaded()
            if self.card_rag_store and getattr(self.card_rag_store, "image_embeddings", None):
                rag_results = self._rag_search_card(image, top_k=3)
        except Exception as e:
            print(f"⚠️ RAG检索失败: {str(e)}")
            rag_results = []

        # 在终端输出RAG相似度匹配结果（与API版本一致）
        if rag_results:
            print("\n" + "=" * 60)
            print("📊 RAG相似度匹配结果")
            print("=" * 60)
            print(f"找到 {len(rag_results)} 张相似图片：\n")
            for i, r in enumerate(rag_results, 1):
                filename = r.get("filename", "未知")
                similarity = r.get("similarity", 0.0)
                print(f"  {i}. {filename}")
                print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
            print("=" * 60 + "\n")
        else:
            print("\n⚠️ 未找到相似图片\n")

        # 构建增强提示词（使用与API版本相同的逻辑）
        enhanced_prompt = self._build_enhanced_prompt_card(
            base_prompt=default_prompt,
            rag_results=rag_results,
            custom_prompt=effective_prompt if (prompt or "").strip() else None
        )

        # 在终端输出发送给模型的完整prompt（与API版本一致）
        print("\n" + "=" * 80)
        print("📝 发送给模型的完整Prompt")
        print("=" * 80)
        print(enhanced_prompt)
        print("=" * 80 + "\n")

        try:
            # 使用本地模型进行推理
            prompt_clean, response, _ = self._run_inference(
                image,
                enhanced_prompt,
                max_tokens=1024,
                temperature=0.3,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 卡证OCR识别失败: {str(e)}"

    def ocr_card_api(self, image, prompt: str = None):
        """卡证OCR识别（API调用 + RAG增强）"""
        # 注：如无需强制本地模型加载，可移除此判断
        try:
            self._ensure_card_api_loaded()
            if self.card_api is None:
                return "�?卡证OCR API初始化失败"
            default_prompt = (
                "你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                "\n"
                "任务要求如下：\n"
                "\n"
            "1. 识别卡证类型：只允许从以下类别中选择一种：\n"
            "   - 身份证 / 银行卡 / 驾驶证 / 护照 / 工牌 / 其他。\n"
            "   Markdown表格中添加“卡证类型”字段，并用类别选择赋值。\n"
            "   **重要**：如果识别为银行卡，必须严格遵守第3条银行卡特殊要求！\n"
            "\n"
            "2. 输出格式：\n"
            "   - 以Markdown表格形式输出所有识别出的关键字段及其对应的值。\n"
            "   - 若字段中包含“卡号”，请确保该字段的值仅包含数字。\n"
            "   - 不要使用代码块标记符号（例如 ``` ）。\n"
            "\n"
            "3. 银行卡特殊要求（必须严格遵守）：\n"
            "   如果识别的卡证类型是银行卡，必须在Markdown表格的最后额外添加一个字段：\n"
            "   - 字段名：卡面类型（必须添加，不可省略）。\n"
            "   - 基于图片库检索到的相似卡证结果，填充“卡面类型”字段。字段值规则如下：\n"
            "       ① 当出现任何不确定、模糊或不匹配情况时，“卡面类型”字段的值**必须且只能为“其他”**，不得填写相似图片名或其他文本。\n"
            "       ② 若识别出的“发卡行”字段的值与这些相似卡证文件名中`_`前面的银行名称相同，"
            "则“卡面类型”字段的值只能从相似卡证文件名中**严格选择一个**，格式为`银行名称_卡面类型`，去掉文件后缀名，如`中国银行_visa卡`。\n"
            "       ③ 禁止自定义、生成、猜测或编造新的卡面类型值。任何不存在基于图片库检索到的相似卡证文件名的值都视为错误。\n"
            "   **重要提醒**：银行卡的Markdown表格必须包含“卡面类型”字段，这是强制要求，不能省略！\n"
            "   - 如果不是银行卡，则不添加“卡面类型”字段。\n"
            "\n"
            "4. 输出限制：\n"
            "   - 最终输出只包含Markdown表格。\n"
            "   - 禁止输出任何其他文字或解释性内容。\n"
            "   - 如果是银行卡，表格中必须包含“卡面类型”字段，否则输出不完整。\n"
            )

            effective_prompt = (prompt or "").strip() or default_prompt
            result = self.card_api.recognize_card(
                image,
                custom_prompt=effective_prompt,
                use_rag=True,
            )
            if not result.get("success"):
                return f"�?卡证OCR API调用失败: {result.get('error') or '未知错误'}"

            # 在终端输出RAG相似度匹配结果
            rag_info = result.get("rag_info")
            if rag_info and rag_info.get("enabled") and rag_info.get("results"):
                print("\n" + "=" * 60)
                print("📊 RAG相似度匹配结果")
                print("=" * 60)
                print(f"找到 {len(rag_info['results'])} 张相似图片：\n")
                for i, r in enumerate(rag_info["results"], 1):
                    filename = r.get("filename", "未知")
                    similarity = r.get("similarity", 0.0)
                    print(f"  {i}. {filename}")
                    print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
                print("=" * 60 + "\n")
            elif rag_info and not rag_info.get("enabled"):
                print(f"\n⚠️ RAG未启用: {rag_info.get('reason', '未知原因')}\n")
            else:
                print("\n⚠️ 未找到相似图片\n")

            cleaned = self._sanitize_markdown(result.get("result") or "")
            self.last_ocr_markdown = f"## 卡证OCR识别（API）结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别（API）结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别（API）结果:\n\n{cleaned}"
        except Exception as e:
            return f"�?卡证OCR API识别失败: {str(e)}"

    def ocr_receipt(self, image, prompt: str = None):
        """票据OCR识别：发票/小票等表格与关键项解析"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "你是发票/小票OCR专家。请解析图片中的票据并输出：\n"
            "- 以Markdown表格给出关键信息：票据类型、开票日期、发票代码、发票号码、校验码、购买方、销售方、税号、项目、数量、单价、金额、税率、税额、合计金额(含税/不含税)；\n"
            "- 若检测到多行项目，请以表格形式逐行列出；\n"
            "- 表格下方给出识别置信度与可疑项提示；\n"
            "- 不要使用围栏代码块，保持Markdown可渲染。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1536,
                temperature=0.2,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🧾 票据OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 票据OCR识别失败: {str(e)}"

    def ocr_agreement(self, image, prompt: str = None):
        """协议OCR识别：合同/协议段落与条款解析"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "你是合同/协议OCR与条款解析助手。请完成：\n"
            "1) 识别全文，保持段落结构；\n"
            "2) 以Markdown表格提炼关键信息：合同名称、甲方、乙方、签署日期、生效日期、终止日期、金额/币种、违约条款、争议解决、签章情况；\n"
            "3) 如有编号的条款，保留编号并逐条列出；\n"
            "4) 在末尾给出“风险提示”列表（如空白处、涂改处、关键要素缺失等）；\n"
            "5) 不要输出围栏代码块。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=2048,
                temperature=0.3,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 协议OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>协议OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"📄 协议OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 协议OCR识别失败: {str(e)}"

    def spatial_analysis(self, image, prompt: str = None):
        """空间感知分析，可选自定义提示词"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "请分析这张图片中的空间关系，包括相对位置、视角、遮挡、深度与距离感，并给出整体布局描述。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=768,
                temperature=0.7,
                top_p=0.8,
                top_k=50,
                repetition_penalty=1.0
            )
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {response}"])
            return f"📐 空间分析结果:\n\n{response}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 空间分析失败: {str(e)}"

    def visual_coding(self, image, output_format: str = "HTML", prompt: str = None):
        """视觉编程生成代码，可选自定义提示词"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        base_prompts = {
            "HTML": "请根据图片生成对应的HTML结构代码，包含必要的语义标签。",
            "CSS": "请为该图片对应的界面生成合理的CSS样式代码，包括布局与颜色。",
            "JavaScript": "请根据图片交互生成JavaScript代码示例，包含必要的事件与逻辑。",
            "Python": "请生成能复现该界面/布局的Python示例代码（如使用streamlit或flask的伪代码）。",
        }
        default_prompt = base_prompts.get(output_format, base_prompts["HTML"]) + " 请只输出代码，不要额外说明。"
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1024,
                temperature=0.4,
                top_p=0.8,
                top_k=50,
                repetition_penalty=1.0
            )
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {response}"])
            return response
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 视觉编程失败: {str(e)}"

    def batch_analysis(self, images, analysis_type):
        """批量分析"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"

        if not images:
            return "❌ 请上传图像！"

        results = []

        for i, image in enumerate(images):
            try:
                if analysis_type == "描述":
                    prompt = "请真实、详细、客观地描述这张图片的内容。"
                elif analysis_type == "OCR":
                    prompt = "请识别并提取这张图片中的所有文字内容，尽量还原原本样式，并标注语言类型。"
                elif analysis_type == "空间分析":
                    prompt = "请分析这张图片中的空间关系和物体位置，包括相对位置、视角、遮挡、深度与距离感，并给出整体布局描述。"
                elif analysis_type == "情感分析":
                    prompt = "请分析这张图片传达的情感或氛围。"
                else:
                    prompt = "请分析这张图片。"

                messages = [
                    {
                        "role": "user",
                        "content": [
                            {"type": "image", "image": image},
                            {"type": "text", "text": prompt},
                        ],
                    }
                ]

                inputs = self.processor.apply_chat_template(
                    messages,
                    tokenize=True,
                    add_generation_prompt=True,
                    return_dict=True,
                    return_tensors="pt"
                )
                inputs = inputs.to(self.model.device)

                with torch.no_grad():
                    generated_ids = self.model.generate(**inputs, max_new_tokens=512)

                generated_ids_trimmed = [
                    out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
                ]
                output_text = self.processor.batch_decode(
                    generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
                )

                results.append(f"📷 图像 {i+1}:\n{output_text[0]}\n" + "="*50 + "\n")

            except Exception as e:
                results.append(f"📷 图像 {i+1}: ❌ 分析失败 - {str(e)}\n" + "="*50 + "\n")

        return "".join(results)

    def compare_images(self, image1, image2, comparison_type):
        """图像对比"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"

        if image1 is None or image2 is None:
            return "❌ 请上传两张图像进行对比！"

        try:
            if comparison_type == "相似性":
                prompt = "请对比这两张图片，分析它们的相似之处和不同之处。"
            elif comparison_type == "风格":
                prompt = "请对比这两张图片的艺术风格、色彩搭配和构图特点。"
            elif comparison_type == "内容":
                prompt = "请对比这两张图片的内容，分析它们描述的场景或主题。"
            else:
                prompt = "请对比这两张图片，提供详细的对比分析。"

            messages = [
                {
                    "role": "user",
                    "content": [
                        {"type": "image", "image": image1},
                        {"type": "image", "image": image2},
                        {"type": "text", "text": prompt},
                    ],
                }
            ]

            inputs = self.processor.apply_chat_template(
                messages,
                tokenize=True,
                add_generation_prompt=True,
                return_dict=True,
                return_tensors="pt"
            )
            inputs = inputs.to(self.model.device)

            with torch.no_grad():
                generated_ids = self.model.generate(**inputs, max_new_tokens=1024)

            generated_ids_trimmed = [
                out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
            ]
            output_text = self.processor.batch_decode(
                generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
            )

            return f"🔍 对比分析结果:\n\n{output_text[0]}"

        except Exception as e:
            return f"❌ 对比分析失败: {str(e)}"

    def export_chat_history(self):
        """导出对话历史"""
        if not self.chat_history:
            return "❌ 没有对话历史可导出！"

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"chat_history_{timestamp}.json"

            # 保存为JSON格式
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.chat_history, f, ensure_ascii=False, indent=2)

            return f"✅ 对话历史已导出到: {filename}"

        except Exception as e:
            return f"❌ 导出失败: {str(e)}"

    def clear_history(self):
        """清空对话历史"""
        self.chat_history = []
        self.chat_messages = []
        self.last_image = None
        self.last_saved_image_path = None
        self.last_image_digest = None
        self.last_ocr_markdown = None
        self.last_ocr_html = None
        if hasattr(self, "session_turn_image_paths"):
            self.session_turn_image_paths.clear()
        return []

    def export_last_ocr(self):
        if not self.last_ocr_markdown:
            return "❌ 没有可保存的文本样式，请先执行一次OCR识别！"

        export_dir = os.path.join("ocr_exports")
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        sections = self._parse_markdown_sections(self.last_ocr_markdown)

        excel_path = os.path.join(export_dir, f"ocr_{timestamp}.xlsx")
        excel_note = ""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "表格1" if sections else "OCR文本"
            table_idx = 0
            for section in sections:
                if section["type"] == "table":
                    table_idx += 1
                    if table_idx > 1:
                        ws = wb.create_sheet(title=f"表格{table_idx}")
                    ws.append(section["header"])
                    for row in section["rows"]:
                        ws.append(row)
                elif section["type"] == "text" and section["text"]:
                    if table_idx > 0:
                        ws = wb.create_sheet(title=f"文本{table_idx}")
                    for line in section["text"].splitlines():
                        ws.append([line])
            if not sections:
                for line in self.last_ocr_markdown.splitlines():
                    ws.append([line])
            wb.save(excel_path)
        except Exception as exc:
            excel_path = os.path.join(export_dir, f"ocr_{timestamp}.csv")
            with open(excel_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["OCR Result"])
                for line in self.last_ocr_markdown.splitlines():
                    writer.writerow([line])
            excel_note = f"⚠️ Excel导出失败({exc})，已保存为CSV"

        json_path = os.path.join(export_dir, f"ocr_{timestamp}.json")
        json_content = {
            "markdown": self.last_ocr_markdown,
            "sections": sections,
        }
        with open(json_path, "w", encoding='utf-8') as f:
            json.dump(json_content, f, ensure_ascii=False, indent=2)

        message_lines = [
            "✅ 文本样式已保存：",
            f"- Excel: {excel_path}" + (f" ({excel_note})" if excel_note else ""),
            f"- JSON: {json_path}",
        ]
        return "\n".join(message_lines)


DEFAULT_TASK_PROMPTS = {
    "任务问答": "请根据图片完成指定任务。",
    "OCR识别": "请识别并提取这张图片中的所有文字内容，并标注语言类型。请确保所有带样式或表格内容使用Markdown表格表示。",
    "卡证OCR识别": "请进行卡证类识别并以Markdown表格输出关键字段（如姓名、证件号、有效期、卡号等）",
    "票据OCR识别": "请解析发票/小票等票据，输出关键信息和多行项目表格，并在下方给出置信度与可疑项。",
    "协议OCR识别": "请提取合同/协议关键信息（甲乙方、日期、金额、条款等），保留段落与条款编号，并在末尾给出风险提示。",
    "空间分析": "请分析这张图片中的空间关系，包括相对位置、视角、遮挡、深度与距离感，并给出整体布局描述。",
    "情感分析": "请分析这张图片传达的情感或氛围，并说明理由。",
}

VISUAL_CODING_PROMPTS = {
    "HTML": "请根据图片生成对应的HTML结构代码，包含必要的语义标签。请只输出代码，不要额外说明。",
    "CSS": "请为该图片对应的界面生成合理的CSS样式代码，包括布局与颜色。请只输出代码，不要额外说明。",
    "JavaScript": "请根据图片交互生成JavaScript代码示例，包含必要的事件与逻辑。请只输出代码，不要额外说明。",
    "Python": "请生成能复现该界面/布局的Python示例代码（如使用streamlit或flask的伪代码）。请只输出代码，不要额外说明。",
}


def _plain_text_to_html(text: str) -> str:
    if not text:
        return ""
    escaped = html.escape(str(text))
    replaced = escaped.replace("\n", "<br>")
    return f'<div class="stats-text">{replaced}</div>'


def _get_default_prompt(task: str, code_format: str = None) -> str:
    if task == "视觉编程":
        fmt = code_format or "HTML"
        return VISUAL_CODING_PROMPTS.get(fmt, VISUAL_CODING_PROMPTS["HTML"])
    return DEFAULT_TASK_PROMPTS.get(task, DEFAULT_TASK_PROMPTS["任务问答"])


# 单例应用
app = AdvancedQwen3VLApp()

# 会话级图片保存目录与轨迹
IMAGE_SAVE_ROOT = "chat_history/images"
SESSION_IMAGE_DIR = os.path.join(IMAGE_SAVE_ROOT, getattr(app, "session_id", datetime.now().strftime("%Y%m%d_%H%M%S")))
os.makedirs(SESSION_IMAGE_DIR, exist_ok=True)
app.session_turn_image_paths = []  # 与对话轮次对齐的图片路径（无图则为 None）


def _toggle_mode(mode, current_task, current_code_format):
    """根据模式切换组件可见性，并预填充默认提示。"""
    is_pro = (mode == "专业版")
    task_value = current_task if is_pro else "任务问答"
    code_visible = is_pro and task_value == "视觉编程"
    text_value = _get_default_prompt(task_value, current_code_format) if is_pro else ""
    return (
        gr.update(visible=is_pro),                       # adv_params_box
        gr.update(visible=is_pro),                       # stats_output
        gr.update(visible=is_pro),                       # tab_batch
        gr.update(visible=is_pro),                       # tab_compare
        gr.update(visible=is_pro, value=task_value),     # pro_task dropdown
        gr.update(visible=code_visible),                 # code_format dropdown
        gr.update(value=text_value),                     # text_input prompt
    )


def _toggle_task(task, code_format):
    """任务切换时调整代码下拉可见性并预填提示。"""
    is_visual = (task == "视觉编程")
    prompt = _get_default_prompt(task, code_format)
    code_kwargs = {"visible": is_visual}
    if is_visual and not code_format:
        code_kwargs["value"] = "HTML"
    return gr.update(**code_kwargs), gr.update(value=prompt)


def _update_code_prompt(task, code_format):
    if task != "视觉编程":
        return gr.update()
    return gr.update(value=_get_default_prompt(task, code_format))


def handle_unified_chat(image,
                        text,
                        history,
                        max_tokens,
                        temperature,
                        top_p,
                        top_k,
                        mode,
                        pro_task,
                        code_format,
                        repetition_penalty,
                        presence_penalty):
    """统一的发送处理：
    - 通用版：普通问答
    - 专业版：按任务分派到不同方法
    返回: history, cleared_text, stats
    """
    user_text = (text or "").strip()
    saved_image_path = None
    image_digest = None
    if image is not None:
        try:
            buffer = io.BytesIO()
            image.save(buffer, format="PNG")
            image_digest = hashlib.md5(buffer.getvalue()).hexdigest()
        except Exception:
            image_digest = None

        should_save = image_digest is None or image_digest != getattr(app, "last_image_digest", None)
        if should_save:
            try:
                ts = datetime.now().strftime("%H%M%S%f")
                saved_image_path = os.path.join(SESSION_IMAGE_DIR, f"img_{ts}.png")
                image.save(saved_image_path)
                if image_digest is not None:
                    app.last_image_digest = image_digest
            except Exception:
                saved_image_path = None

    prev_turns = len(history)
    image_recorded = False

    def record_image_path():
        nonlocal image_recorded
        if image_recorded:
            return
        if saved_image_path and saved_image_path != getattr(app, "last_saved_image_path", None):
            app.session_turn_image_paths.append(saved_image_path)
            app.last_saved_image_path = saved_image_path
            if image_digest is not None:
                app.last_image_digest = image_digest
        else:
            existing = getattr(app, "last_saved_image_path", None)
            app.session_turn_image_paths.append(existing)
        image_recorded = True

    try:
        if mode == "通用版":
            effective_prompt = user_text
            chat_result = app.chat_with_image(
                image,
                effective_prompt,
                history,
                max_tokens,
                temperature,
                top_p,
                top_k,
                repetition_penalty,
                presence_penalty
            )

            if inspect.isgenerator(chat_result):
                for out_history, cleared, stats in chat_result:
                    if not image_recorded and len(out_history) > prev_turns:
                        record_image_path()
                    app.chat_history = out_history
                    button_update = gr.update(interactive=bool(app.last_ocr_markdown))
                    stats_update = gr.update(value=_plain_text_to_html(stats), visible=True)
                    yield out_history, cleared, stats_update, button_update, gr.update(value="", visible=True)
            else:
                out_history, cleared, stats = chat_result
                if not image_recorded and len(out_history) > prev_turns:
                    record_image_path()
                app.chat_history = out_history
                button_update = gr.update(interactive=bool(app.last_ocr_markdown))
                stats_update = gr.update(value=_plain_text_to_html(stats), visible=True)
                yield out_history, cleared, stats_update, button_update, gr.update(value="", visible=True)

        else:
            task = pro_task or "任务问答"
            if task == "OCR识别":
                if image is None:
                    stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图像！"), visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), "❌ 请上传图像！"
                    return

                result = app.ocr_analysis(image)

                if result.startswith("❌"):
                    stats_update = gr.update(value="", visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), result
                    return

                prompt_text = user_text if user_text else _get_default_prompt(task, code_format)
                updated_history = history + [[f"👤 {prompt_text}", result]]
                app.chat_history = updated_history
                if not image_recorded:
                    record_image_path()
                ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
                stats_update = gr.update(value=ocr_preview, visible=True)
                status_update = "✅ OCR识别完成，可导出样式"
                yield updated_history, "", stats_update, gr.update(interactive=bool(app.last_ocr_markdown)), status_update
                return

            if task == "卡证OCR识别（API）":
                if image is None:
                    stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图像！"), visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), "❌ 请上传图像！"
                    return
                result = app.ocr_card_api(image)
                if result.startswith("❌"):
                    stats_update = gr.update(value="", visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), result
                    return
                prompt_text = user_text if user_text else _get_default_prompt("卡证OCR识别", code_format)
                updated_history = history + [[f"👤 {prompt_text}", result]]
                app.chat_history = updated_history
                if not image_recorded:
                    record_image_path()
                ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
                stats_update = gr.update(value=ocr_preview, visible=True)
                yield updated_history, "", stats_update, gr.update(interactive=bool(app.last_ocr_markdown)), "✅ 卡证OCR识别(API)完成，可导出样式"
                return

            if task == "卡证OCR识别":
                if image is None:
                    stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图像！"), visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), "❌ 请上传图像！"
                    return
                result = app.ocr_card(image)
                if result.startswith("❌"):
                    stats_update = gr.update(value="", visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), result
                    return
                prompt_text = user_text if user_text else _get_default_prompt(task, code_format)
                updated_history = history + [[f"👤 {prompt_text}", result]]
                app.chat_history = updated_history
                if not image_recorded:
                    record_image_path()
                ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
                stats_update = gr.update(value=ocr_preview, visible=True)
                yield updated_history, "", stats_update, gr.update(interactive=bool(app.last_ocr_markdown)), "✅ 卡证OCR识别完成，可导出样式"
                return

            if task == "票据OCR识别":
                if image is None:
                    stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图像！"), visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), "❌ 请上传图像！"
                    return
                result = app.ocr_receipt(image)
                if result.startswith("❌"):
                    stats_update = gr.update(value="", visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), result
                    return
                prompt_text = user_text if user_text else _get_default_prompt(task, code_format)
                updated_history = history + [[f"👤 {prompt_text}", result]]
                app.chat_history = updated_history
                if not image_recorded:
                    record_image_path()
                ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
                stats_update = gr.update(value=ocr_preview, visible=True)
                yield updated_history, "", stats_update, gr.update(interactive=bool(app.last_ocr_markdown)), "✅ 票据OCR识别完成，可导出样式"
                return

            if task == "协议OCR识别":
                if image is None:
                    stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图像！"), visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), "❌ 请上传图像！"
                    return
                result = app.ocr_agreement(image)
                if result.startswith("❌"):
                    stats_update = gr.update(value="", visible=True)
                    yield history, text, stats_update, gr.update(interactive=False), result
                    return
                prompt_text = user_text if user_text else _get_default_prompt(task, code_format)
                updated_history = history + [[f"👤 {prompt_text}", result]]
                app.chat_history = updated_history
                if not image_recorded:
                    record_image_path()
                ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
                stats_update = gr.update(value=ocr_preview, visible=True)
                yield updated_history, "", stats_update, gr.update(interactive=bool(app.last_ocr_markdown)), "✅ 协议OCR识别完成，可导出样式"
                return

            effective_prompt = user_text if user_text else _get_default_prompt(task, code_format)
            chat_result = app.chat_with_image(
                image,
                effective_prompt,
                history,
                max_tokens,
                temperature,
                top_p,
                top_k,
                repetition_penalty,
                presence_penalty
            )

            if inspect.isgenerator(chat_result):
                for out_history, cleared, stats in chat_result:
                    if not image_recorded and len(out_history) > prev_turns:
                        record_image_path()
                    app.chat_history = out_history
                    button_update = gr.update(interactive=bool(app.last_ocr_markdown))
                    stats_update = gr.update(value=_plain_text_to_html(stats), visible=True)
                    yield out_history, cleared, stats_update, button_update, gr.update()
            else:
                out_history, cleared, stats = chat_result
                if not image_recorded and len(out_history) > prev_turns:
                    record_image_path()
                app.chat_history = out_history
                button_update = gr.update(interactive=bool(app.last_ocr_markdown))
                stats_update = gr.update(value=_plain_text_to_html(stats), visible=True)
                yield out_history, cleared, stats_update, button_update, gr.update()

        if not image_recorded and len(app.chat_history) > prev_turns:
            record_image_path()

    except Exception as e:
        history.append(["👤", f"❌ 错误: {str(e)}"])
        app.chat_history = history
        if not image_recorded and len(history) > prev_turns:
            record_image_path()
        button_update = gr.update(interactive=bool(app.last_ocr_markdown))
        stats_update = gr.update(value=_plain_text_to_html(f"❌ 错误: {str(e)}"), visible=True)
        yield history, text, stats_update, button_update, f"❌ 错误: {str(e)}"


def save_chat_to_folder(save_dir, history):
    """将当前聊天历史保存到指定文件夹（JSON）。"""
    try:
        if not save_dir:
            return "❌ 保存失败：未指定保存目录"
        os.makedirs(save_dir, exist_ok=True)
        # 每次保存使用独立导出子目录，避免图片累积到同一目录
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = os.path.join(save_dir, timestamp)
        os.makedirs(export_dir, exist_ok=True)
        images_target_dir = os.path.join(export_dir, "images")
        os.makedirs(images_target_dir, exist_ok=True)

        image_paths = getattr(app, "session_turn_image_paths", [])
        copied_rel_paths = []
        seen_images = set()
        for p in image_paths:
            abs_path = os.path.abspath(p) if p else None
            if not p or not os.path.exists(p) or abs_path in seen_images:
                copied_rel_paths.append(None)
                continue
            try:
                basename = os.path.basename(p)
                target = os.path.join(images_target_dir, basename)
                if os.path.abspath(p) != os.path.abspath(target):
                    shutil.copy2(p, target)
                seen_images.add(abs_path)
                copied_rel_paths.append(os.path.join("images", basename))
            except Exception:
                copied_rel_paths.append(None)
        filename = os.path.join(export_dir, f"chat_history_{timestamp}.json")
        # history 是 [(user, bot), ...]
        data = []
        turns = history or []
        for idx, pair in enumerate(turns):
            try:
                u, b = pair
            except Exception:
                u, b = pair, ""
            img_rel = copied_rel_paths[idx] if idx < len(copied_rel_paths) else None
            data.append({"user": u, "assistant": b, "image_path": img_rel})
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return f"✅ 已保存到: {filename}"
    except Exception as e:
        return f"❌ 保存失败: {str(e)}"


def create_unified_interface():
    """创建统一Gradio界面。"""

    touch_css = """
    :root {
        --radius-lg: 22px;
        --radius-md: 14px;
        --surface: #ffffff;
        --surface-muted: #f5f7fb;
        --surface-border: #e2e8f0;
        --text-primary: #0f172a;
        --text-secondary: #64748b;
        --accent: #2563eb;
        --accent-soft: rgba(37, 99, 235, 0.12);
    }
    body {
        background: linear-gradient(135deg, #eef2ff 0%, #f9fafc 55%, #ffffff 100%);
        color: var(--text-primary);
    }
    .gradio-container {
        max-width: 1650px !important;
        margin: 0 auto;
        padding: 20px 24px 48px;
        font-size: 16px;
        color: var(--text-primary);
    }
    .gradio-container .gr-markdown {
        color: var(--text-primary);
    }
    #unified-header {
        background: linear-gradient(130deg, rgba(37, 99, 235, 0.12), rgba(59, 130, 246, 0.1));
        border: 1px solid rgba(37, 99, 235, 0.18);
        padding: 24px 28px;
        border-radius: 28px;
        box-shadow: 0 18px 36px rgba(15, 23, 42, 0.08);
        margin-bottom: 22px;
    }
    #unified-header h1 {
        margin: 0 0 6px;
        font-size: 26px;
        font-weight: 600;
        letter-spacing: 0.2px;
    }
    #unified-header p {
        margin: 0;
        color: var(--text-secondary);
    }
    #unified-mode-bar {
        background: var(--surface);
        border-radius: 24px;
        box-shadow: 0 20px 40px rgba(15, 23, 42, 0.06);
        padding: 20px 22px;
        gap: 18px;
        margin-bottom: 20px;
        border: 1px solid var(--surface-border);
    }
    #unified-mode-bar .gradio-button,
    #unified-mode-bar button {
        font-size: 16px !important;
        padding: 12px 18px !important;
        border-radius: 14px !important;
    }
    #unified-mode-bar textarea,
    #unified-mode-bar input[type="text"] {
        background: var(--surface);
        border: 1px solid rgba(148, 163, 184, 0.35);
        border-radius: 14px;
        color: var(--text-primary);
        box-shadow: inset 0 1px 3px rgba(15, 23, 42, 0.04);
    }
    #unified-mode-bar textarea:focus,
    #unified-mode-bar input[type="text"]:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
    }
    .gradio-container .tabs {
        background: transparent;
        border: none;
    }
    .gradio-container .tabitem {
        border-radius: var(--radius-md);
        background: #f8fafc;
        border: 1px solid transparent;
        color: var(--text-secondary);
    }
    .gradio-container .tabitem.selected {
        border-color: rgba(37, 99, 235, 0.25);
        color: var(--text-primary);
        background: #ffffff;
        box-shadow: 0 10px 20px rgba(37, 99, 235, 0.08);
    }
    #unified-input-panel,
    #unified-chat-panel,
    #unified-batch-panel,
    #unified-compare-panel {
        background: var(--surface);
        border-radius: 24px;
        padding: 22px 24px;
        box-shadow: 0 22px 44px rgba(15, 23, 42, 0.06);
        border: 1px solid var(--surface-border);
    }
    #unified-input-panel .gradio-slider > label,
    #unified-input-panel .gradio-dropdown > label {
        color: var(--text-secondary);
    }
    #unified-chat-panel {
        display: flex;
        flex-direction: column;
        gap: 16px;
    }
    #unified-chatbot > .wrap {
        background: #f8fafc;
        border-radius: 20px;
        border: 1px solid rgba(148, 163, 184, 0.25);
        padding: 8px 10px;
    }
    #unified-chatbot .message {
        border-radius: 16px !important;
        padding: 12px 14px !important;
        line-height: 1.6;
        font-size: 15px;
        color: var(--text-primary);
    }
    #unified-chatbot .message.user {
        background: linear-gradient(138deg, rgba(37, 99, 235, 0.16), rgba(96, 165, 250, 0.12));
        border: 1px solid rgba(37, 99, 235, 0.22);
        color: var(--text-primary);
        align-self: flex-end;
    }
    #unified-chatbot .message.bot {
        background: #ffffff;
        border: 1px solid rgba(203, 213, 225, 0.9);
        color: var(--text-primary);
        align-self: flex-start;
    }
    #unified-query textarea {
        border-radius: 16px;
        border: 1px solid rgba(148, 163, 184, 0.35);
        background: var(--surface);
        color: var(--text-primary);
        box-shadow: inset 0 1px 3px rgba(15, 23, 42, 0.05);
    }
    #unified-query textarea:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
    }
    #unified-stats .stats-text {
        background: var(--accent-soft);
        border-radius: 16px;
        border: 1px solid rgba(37, 99, 235, 0.2);
        color: var(--text-primary);
        font-weight: 500;
        padding: 12px 14px;
        line-height: 1.6;
        margin-bottom: 12px;
        word-break: break-word;
    }
    .gradio-container .gradio-button.primary {
        background: linear-gradient(135deg, #2563eb, #1d4ed8);
        border: none;
        color: #ffffff;
        font-weight: 600;
        box-shadow: 0 18px 30px rgba(37, 99, 235, 0.22);
    }
    .gradio-container .gradio-button.primary:hover {
        filter: brightness(1.03);
    }
    .gradio-container .gradio-button.secondary {
        background: rgba(37, 99, 235, 0.1);
        border: 1px solid rgba(37, 99, 235, 0.18);
        color: var(--text-primary);
    }
    .gradio-container textarea,
    .gradio-container input[type="text"],
    .gradio-container input[type="number"] {
        background: var(--surface);
        border: 1px solid rgba(148, 163, 184, 0.35);
        color: var(--text-primary);
        border-radius: 16px;
    }
    .gradio-container textarea:focus,
    .gradio-container input[type="text"]:focus,
    .gradio-container input[type="number"]:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
    }
    #unified-batch-panel textarea,
    #unified-compare-panel textarea {
        min-height: 320px;
    }
    .gradio-container .dropdown span.label,
    .gradio-container .slider > label,
    .gradio-container .dropdown label {
        color: var(--text-secondary);
    }
    .gradio-container .gradio-dropdown .wrap select,
    .gradio-container .gradio-dropdown .wrap button {
        background: var(--surface);
        color: var(--text-primary);
        border-color: rgba(148, 163, 184, 0.4);
    }
    .gradio-container .gradio-dropdown .wrap select:focus,
    .gradio-container .gradio-dropdown .wrap button:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
    }
    /*
    Bigger markdown preview area for unified stats (OCR/table preview)
    */
    #unified-stats {
        max-height: 560px;
        overflow: auto;
        border: 1px solid rgba(148, 163, 184, 0.35);
        padding: 12px 14px;
        border-radius: 14px;
        background: #ffffff;
    }
    #unified-stats table {
        width: 100%;
        border-collapse: collapse;
        margin: 8px 0 14px;
    }
    #unified-stats th,
    #unified-stats td {
        border: 1px solid #e5e7eb;
        padding: 8px 10px;
        text-align: left;
        vertical-align: top;
        font-size: 14px;
        line-height: 1.55;
    }
    #unified-stats thead th {
        background: #f8fafc;
        font-weight: 600;
    }
    #unified-stats code {
        background: #f8fafc;
        border: 1px solid #e5e7eb;
        padding: 1px 4px;
        border-radius: 6px;
        font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
    }
    /* 字段表格样式 */
    .gradio-container .dataframe {
        border-radius: 14px;
        border: 1px solid rgba(148, 163, 184, 0.35);
        overflow: hidden;
    }
    .gradio-container .dataframe table {
        width: 100%;
        border-collapse: collapse;
    }
    .gradio-container .dataframe th {
        background: linear-gradient(135deg, rgba(37, 99, 235, 0.1), rgba(59, 130, 246, 0.08));
        color: var(--text-primary);
        font-weight: 600;
        padding: 10px 12px;
        border-bottom: 2px solid rgba(37, 99, 235, 0.2);
    }
    .gradio-container .dataframe td {
        padding: 8px 12px;
        border-bottom: 1px solid rgba(148, 163, 184, 0.2);
    }
    .gradio-container .dataframe tr:hover {
        background: rgba(37, 99, 235, 0.04);
    }
    .gradio-container .dataframe input[type="text"] {
        border: 1px solid rgba(148, 163, 184, 0.3);
        border-radius: 6px;
        padding: 4px 8px;
        width: 100%;
    }
    .gradio-container .dataframe input[type="text"]:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.1);
    }
    """

    with gr.Blocks(
        title="多模态大模型智能助手",
        theme=gr.themes.Soft(),
        css=touch_css
    ) as interface:

        gr.HTML("""
        <section id="unified-header">
          <h1>🤖 多模态大模型智能助手</h1>
          <p>全新布局与对话样式，通用 / 专业双模式随心切换，支持任务分派与对话保存。</p>
        </section>
        """)

        with gr.Row(elem_id="unified-mode-bar"):
            with gr.Column(scale=1, min_width=240):
                mode = gr.Radio(
                    choices=["通用版", "专业版"], value="通用版", label="界面模式"
                )
                pro_task = gr.Dropdown(
                    choices=["任务问答", "OCR识别", "卡证OCR识别", "卡证OCR识别（API）", "票据OCR识别", "协议OCR识别", "空间分析", "视觉编程", "情感分析"],
                    value="任务问答",
                    label="专业任务",
                    visible=False,
                )
            with gr.Column(scale=1, min_width=240):
                load_btn = gr.Button("🔄 加载模型", variant="primary")
                status_text = gr.Textbox(
                    label="运行状态",
                    value="⏳ 模型未加载，请点击加载模型按钮",
                    interactive=False,
                    lines=3,
                )
            with gr.Column(scale=1, min_width=240):
                save_btn = gr.Button("💾 保存当前对话", variant="secondary")
                save_dir = gr.Textbox(value="chat_history", label="保存目录", interactive=False)

        load_btn.click(app.load_model, outputs=[status_text, load_btn])

        # 样式在 Blocks 实例化时应用，无需运行时切换

        with gr.Tab("💬 图像对话"):
            with gr.Row(equal_height=True):
                with gr.Column(scale=1):
                    with gr.Group(elem_id="unified-input-panel"):
                        gr.Markdown("### 图像与参数设置")
                        image_input = gr.Image(label="上传图像", type="pil", height=400)

                        # 通用参数
                        with gr.Row(equal_height=True):
                            max_tokens = gr.Slider(
                                minimum=512, maximum=16384, value=8192, label="最大生成长度 (out_seq_length)"
                            )
                            temperature = gr.Slider(
                                minimum=0.0, maximum=2.0, value=0.7, label="创造性 (temperature)"
                            )

                        # 专业参数容器（默认隐藏）
                        with gr.Accordion("🎛️ 高级参数", open=False, visible=False) as adv_params_box:
                            top_p = gr.Slider(
                                minimum=0.0, maximum=1.0, value=0.8, label="top_p"
                            )
                            top_k = gr.Slider(
                                minimum=0, maximum=100, value=20, label="top_k"
                            )
                            repetition_penalty = gr.Slider(
                                minimum=0.8, maximum=2.0, value=1.0, step=0.05, label="repetition_penalty"
                            )
                            presence_penalty = gr.Slider(
                                minimum=0.0, maximum=3.0, value=1.5, step=0.1, label="presence_penalty (占位)"
                            )

                with gr.Column(scale=2):
                    with gr.Group(elem_id="unified-chat-panel"):
                        gr.Markdown("### 对话与输出")
                        chatbot = gr.Chatbot(
                            label=None,
                            height=500,
                            show_label=False,
                            type="tuples",
                            elem_id="unified-chatbot",
                            render_markdown=True
                        )
                        text_input = gr.Textbox(
                            label=None,
                            placeholder="输入想了解的内容，按 Enter 或点击发送。",
                            lines=3,
                            elem_id="unified-query"
                        )
                        with gr.Row():
                            send_btn = gr.Button("发送", variant="primary", scale=1)
                            clear_btn = gr.Button("🗑️ 清空历史", variant="secondary", scale=1)
                        with gr.Row():
                            with gr.Column(scale=4):
                                stats_output = gr.HTML(
                                    value="",
                                    visible=False,
                                    elem_id="unified-stats"
                                )
                            with gr.Column(scale=1, min_width=220):
                                ocr_export_btn = gr.Button("💾 导出样式", variant="secondary", interactive=False)
                                ocr_export_status = gr.Textbox(
                                    label="导出状态",
                                    interactive=False,
                                    lines=4
                                )
                code_format = gr.Dropdown(
                    choices=["HTML", "CSS", "JavaScript", "Python"],
                    value="HTML",
                    label="代码类型",
                    visible=False,
                )

            # 通用/专业两种调用路径（利用同一高级应用，专业多两个参数与统计输出）
            send_btn.click(
                handle_unified_chat,
                inputs=[image_input, text_input, chatbot, max_tokens, temperature, top_p, top_k, mode, pro_task, code_format, repetition_penalty, presence_penalty],
                outputs=[chatbot, text_input, stats_output, ocr_export_btn, ocr_export_status],
            )
            text_input.submit(
                handle_unified_chat,
                inputs=[image_input, text_input, chatbot, max_tokens, temperature, top_p, top_k, mode, pro_task, code_format, repetition_penalty, presence_penalty],
                outputs=[chatbot, text_input, stats_output, ocr_export_btn, ocr_export_status],
            )
            def _clear_session():
                app.clear_history()
                return [], "", gr.update(value="", visible=False), gr.update(interactive=False), ""

            clear_btn.click(
                _clear_session,
                outputs=[chatbot, text_input, stats_output, ocr_export_btn, ocr_export_status],
            )
        save_btn.click(save_chat_to_folder, inputs=[save_dir, chatbot], outputs=[status_text])

        ocr_export_btn.click(
            app.export_last_ocr,
            outputs=[ocr_export_status],
        )

        # 高级功能Tab（默认隐藏，通过模式切换显示）
        with gr.Tab("📊 批量分析", visible=False) as tab_batch:
            with gr.Group(elem_id="unified-batch-panel"):
                with gr.Row():
                    with gr.Column():
                        batch_images = gr.File(
                            label="上传多个图像", file_count="multiple", file_types=["image"]
                        )
                        analysis_type = gr.Dropdown(
                            choices=["描述", "OCR", "空间分析", "情感分析"], value="描述", label="分析类型"
                        )
                        batch_btn = gr.Button("🔍 开始批量分析", variant="primary")
                    with gr.Column():
                        batch_result = gr.Markdown()
            batch_btn.click(app.batch_analysis, inputs=[batch_images, analysis_type], outputs=[batch_result])

        with gr.Tab("🔄 图像对比", visible=False) as tab_compare:
            with gr.Group(elem_id="unified-compare-panel"):
                with gr.Row():
                    with gr.Column():
                        compare_image1 = gr.Image(label="图像1", type="pil", height=220)
                        compare_image2 = gr.Image(label="图像2", type="pil", height=220)
                        comparison_type = gr.Dropdown(
                            choices=["相似性", "风格", "内容", "综合"], value="相似性", label="对比类型"
                        )
                        compare_btn = gr.Button("🔄 开始对比", variant="primary")
                    with gr.Column():
                        compare_result = gr.Markdown()
            compare_btn.click(
                app.compare_images,
                inputs=[compare_image1, compare_image2, comparison_type],
                outputs=[compare_result],
            )

        with gr.Tab("🪪 卡证OCR（三步流程）"):
            gr.Markdown("### 三步流程：识别类型 → 自定义字段 → OCR识别")
            
            with gr.Row():
                with gr.Column(scale=1):
                    card_image = gr.Image(
                        label="上传卡证图片",
                        type="pil",
                        height=400
                    )
                    
                    with gr.Row():
                        detect_type_btn = gr.Button("🔍 第一步：识别卡证类型", variant="primary")
                    
                    card_type_output = gr.Textbox(
                        label="识别的卡证类型",
                        interactive=False,
                        visible=False
                    )
                    
                    default_fields_title = gr.Markdown("### 📋 默认字段模板", visible=False)
                    # HTML表格展示（用于HTML格式的模板）
                    default_fields_html = gr.HTML(
                        label="默认字段模板（HTML表格）",
                        visible=False,
                        elem_id="default-fields-html"
                    )
                    # Dataframe展示（用于非HTML格式的模板）
                    default_fields_output = gr.Dataframe(
                        label="默认字段",
                        headers=["序号", "字段名"],
                        datatype=["number", "str"],
                        interactive=True,
                        visible=False,
                        wrap=True,
                        type="array"  # 明确指定返回格式为2D数组
                    )
                    
                    custom_fields_title = gr.Markdown("### ➕ 自定义字段", visible=False)
                    custom_fields_input = gr.Dataframe(
                        label="添加自定义字段（每行一个字段名）",
                        headers=["字段名"],
                        datatype=["str"],
                        interactive=True,
                        visible=False,
                        wrap=True,
                        row_count=(1, "dynamic"),
                        col_count=(1, "fixed"),
                        type="array",  # 明确指定返回格式为2D数组
                        value=[[""]]  # 初始值：一个空行
                    )
                    
                    with gr.Row():
                        add_custom_field_btn = gr.Button("➕ 添加自定义字段", variant="secondary", visible=False, size="sm")
                    
                    with gr.Row():
                        update_fields_btn = gr.Button("🔗 第二步：合并字段", variant="secondary", visible=False)
                    
                    all_fields_title = gr.Markdown("### ✅ 最终字段列表（将用于OCR识别）", visible=False)
                    # HTML表格展示（用于HTML格式的模板）
                    all_fields_html = gr.HTML(
                        label="最终字段列表（HTML表格）",
                        visible=False,
                        elem_id="all-fields-html"
                    )
                    # Dataframe展示（用于非HTML格式的模板）
                    all_fields_output = gr.Dataframe(
                        label="最终字段列表",
                        headers=["序号", "字段名", "来源"],
                        datatype=["number", "str", "str"],
                        interactive=False,
                        visible=False,
                        wrap=True,
                        type="array"  # 明确指定返回格式为2D数组
                    )
                    
                    fields_status = gr.Textbox(
                        label="状态",
                        interactive=False,
                        visible=False
                    )
                    
                    with gr.Row():
                        ocr_with_fields_btn = gr.Button("🚀 第三步：开始OCR识别", variant="primary", visible=False)
                
                with gr.Column(scale=2):
                    with gr.Row():
                        gr.Markdown("### 📊 OCR识别结果")
                        with gr.Column(scale=1, min_width=200):
                            ocr_export_format = gr.Dropdown(
                                choices=["Markdown (.md)", "Excel (.xlsx)", "CSV (.csv)", "JSON (.json)"],
                                value="Markdown (.md)",
                                label="导出格式",
                                visible=False
                            )
                        ocr_export_btn_3step = gr.Button("💾 导出结果", variant="secondary", visible=False, size="sm")
                    
                    # HTML表格展示（用于HTML格式的模板）
                    ocr_result_html = gr.HTML(
                        label="OCR识别结果（HTML表格）",
                        visible=False,
                        elem_id="ocr-result-html"
                    )
                    # Dataframe展示（用于非HTML格式的模板）
                    ocr_result = gr.Dataframe(
                        label="OCR识别结果（可编辑表格）",
                        headers=["字段名", "字段值"],
                        datatype=["str", "str"],
                        interactive=True,
                        visible=False,
                        wrap=True,
                        type="array"
                    )
                    
                    ocr_export_status_3step = gr.Textbox(
                        label="导出状态",
                        interactive=False,
                        visible=False,
                        lines=3
                    )
            
            # 辅助函数：确保值是标量（非可迭代）
            def ensure_scalar(value):
                """确保值是标量，如果是可迭代对象则转换为字符串"""
                if value is None:
                    return ""
                elif isinstance(value, str):
                    return value
                elif isinstance(value, (list, tuple)):
                    return "".join(str(x) for x in value) if value else ""
                elif hasattr(value, '__iter__'):
                    try:
                        return "".join(str(x) for x in value)
                    except:
                        return str(value)
                else:
                    return str(value)
            
            # 第一步：识别卡证类型
            def step1_detect_type(image):
                if image is None:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        "❌ 请先上传图片"
                    )
                
                result = app.detect_card_type(image)
                if len(result) == 4:
                    card_type, default_fields, html_template, status_msg = result
                else:
                    # 兼容旧版本（没有HTML模板）
                    card_type, default_fields, status_msg = result
                    html_template = None
                
                if card_type:
                    # 卡证OCR不使用HTML模板，只使用DataFrame展示
                    # 转换为DataFrame格式：[[序号, 字段名], ...]
                    default_fields_df = []
                    for i, field in enumerate(default_fields, 1):
                        field_str = ensure_scalar(field).strip()
                        if field_str:
                            default_fields_df.append([int(i), field_str])
                    # 清空自定义字段
                    custom_fields_df = [[""]]
                    return (
                        gr.update(value=card_type, visible=True),
                        gr.update(visible=False),  # HTML表格隐藏
                        gr.update(value=default_fields_df, visible=True),  # Dataframe
                        gr.update(visible=True),  # default_fields_title
                        gr.update(value=custom_fields_df, visible=True),  # custom_fields_input
                        gr.update(visible=True),  # custom_fields_title
                        gr.update(visible=True),  # add_custom_field_btn
                        gr.update(visible=True),  # update_fields_btn
                        gr.update(visible=False),  # all_fields_title (初始隐藏)
                        gr.update(value=[], visible=False),  # all_fields_output (初始为空)
                        gr.update(value=status_msg, visible=True),  # fields_status
                    )
                else:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(value=status_msg, visible=True),  # fields_status
                    )
            
            # 第二步：更新字段
            def step2_update_fields(card_type, default_fields_df, custom_fields_df):
                # 检查第一步是否完成：需要card_type存在
                if not card_type:
                    return (
                        gr.update(visible=False),
                        gr.update(value=[], visible=False),
                        gr.update(value="❌ 请先完成第一步：识别卡证类型", visible=True)
                    )
                
                # 优先从应用状态获取字段列表（适用于HTML表格情况）
                default_fields = []
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    default_fields = app.current_default_fields.copy()
                    print(f"[DEBUG] 从app状态获取默认字段: {default_fields}")
                
                # 如果应用状态中没有，则从DataFrame提取字段名
                if not default_fields and default_fields_df is not None:
                    print(f"[DEBUG] default_fields_df原始数据: {default_fields_df}, 类型: {type(default_fields_df)}")
                    
                    # 处理不同的数据格式
                    rows = []
                    if hasattr(default_fields_df, 'values'):
                        # 如果是pandas DataFrame
                        try:
                            rows = default_fields_df.values.tolist()
                        except:
                            rows = list(default_fields_df.values) if hasattr(default_fields_df, 'values') else []
                    elif isinstance(default_fields_df, (list, tuple)):
                        # 如果是列表或元组
                        if len(default_fields_df) == 0:
                            rows = []
                        elif len(default_fields_df) > 0 and isinstance(default_fields_df[0], str):
                            # 第一个元素是字符串，可能是列名列表，跳过
                            print(f"[DEBUG] 警告：default_fields_df似乎是列名列表，跳过: {default_fields_df}")
                            rows = []
                        elif len(default_fields_df) > 0 and isinstance(default_fields_df[0], (list, tuple)):
                            # 第一个元素是列表/元组，这是行数据列表
                            rows = default_fields_df
                        else:
                            # 其他情况，尝试作为行数据处理
                            rows = default_fields_df
                    elif isinstance(default_fields_df, dict):
                        # 如果是字典，尝试提取数据
                        if 'data' in default_fields_df:
                            rows = default_fields_df['data']
                        else:
                            rows = []
                    else:
                        rows = []
                    
                    print(f"[DEBUG] 处理后的rows: {rows}, 类型: {type(rows)}, 长度: {len(rows) if hasattr(rows, '__len__') else 'N/A'}")
                    
                    # 遍历行数据
                    for i, row in enumerate(rows):
                        if not row:
                            continue
                        print(f"[DEBUG] 处理行{i}: {row}, 类型: {type(row)}")
                        
                        # 跳过列名（字符串）
                        if isinstance(row, str):
                            print(f"[DEBUG] 跳过列名: {row}")
                            continue
                        
                        # 处理行格式：应该是列表或元组 [序号, 字段名]
                        if isinstance(row, (list, tuple)):
                            if len(row) >= 2:
                                # 取第二个元素（索引1，字段名列）
                                field_value = row[1]
                            else:
                                continue
                        elif isinstance(row, dict):
                            # 如果是字典，尝试从'字段名'键获取
                            field_value = row.get('字段名') or row.get(1)
                        else:
                            continue
                        
                        print(f"[DEBUG] 提取的field_value: {field_value!r}, 类型: {type(field_value)}")
                        # 使用辅助函数确保字段值是标量
                        field_str = ensure_scalar(field_value).strip()
                        print(f"[DEBUG] 提取后: field_str={field_str!r}, type={type(field_str)}")
                        if field_str:
                            default_fields.append(field_str)
                
                # 如果仍然没有字段，说明第一步未完成
                if not default_fields:
                    return (
                        gr.update(visible=False),
                        gr.update(value=[], visible=False),
                        gr.update(value="❌ 请先完成第一步：识别卡证类型", visible=True)
                    )
                
                # 从自定义字段DataFrame提取
                custom_fields_list = []
                if custom_fields_df is not None:
                    print(f"[DEBUG] custom_fields_df原始数据: {custom_fields_df}, 类型: {type(custom_fields_df)}")
                    
                    # 处理不同的数据格式（与默认字段相同的逻辑）
                    if isinstance(custom_fields_df, dict):
                        if 'data' in custom_fields_df:
                            rows = custom_fields_df['data']
                        elif 'values' in custom_fields_df:
                            rows = custom_fields_df['values']
                        else:
                            rows = [v for v in custom_fields_df.values() if isinstance(v, (list, tuple)) and len(v) > 0]
                            if rows:
                                rows = rows[0]
                            else:
                                rows = []
                    elif hasattr(custom_fields_df, 'values'):
                        try:
                            rows = custom_fields_df.values.tolist()
                        except:
                            rows = list(custom_fields_df.values) if hasattr(custom_fields_df, 'values') else []
                    elif isinstance(custom_fields_df, (list, tuple)):
                        rows = custom_fields_df
                    else:
                        try:
                            rows = list(custom_fields_df)
                        except:
                            rows = []
                    
                    print(f"[DEBUG] 处理后的custom rows: {rows}, 类型: {type(rows)}")
                    
                    # 遍历行数据
                    for row in rows:
                        if not row:
                            continue
                        print(f"[DEBUG] 处理自定义行: {row}, 类型: {type(row)}")
                        
                        # 处理不同的行格式
                        if isinstance(row, dict):
                            # 如果是字典，尝试从'字段名'键获取
                            field_value = row.get('字段名') or row.get(0) or (row.get(list(row.keys())[0]) if len(row) > 0 else None)
                        elif isinstance(row, (list, tuple)):
                            # 如果是列表或元组，取第一个元素（索引0）
                            if len(row) > 0:
                                field_value = row[0]
                            else:
                                continue
                        else:
                            continue
                        
                        print(f"[DEBUG] 提取的自定义field_value: {field_value!r}, 类型: {type(field_value)}")
                        # 使用辅助函数确保字段值是标量
                        field_str = ensure_scalar(field_value).strip()
                        if field_str and field_str not in default_fields:
                            custom_fields_list.append(field_str)
                
                # 转换为DataFrame格式：[[序号, 字段名, 来源], ...]
                all_fields_df = []
                default_count = 0
                custom_count = 0
                
                # 添加默认字段（过滤空字段）
                idx = 1
                for field in default_fields:
                    # 使用辅助函数确保字段名是标量
                    field_str = ensure_scalar(field).strip()
                    if field_str:
                        # 确保每个元素都是标量值，不是可迭代对象
                        # 显式转换为字符串，确保不是其他类型
                        field_name = str(field_str)
                        print(f"[DEBUG] 添加默认字段: idx={idx}, field_str={field_str!r}, field_name={field_name!r}, type={type(field_name)}")
                        all_fields_df.append([int(idx), field_name, "默认"])
                        idx += 1
                        default_count += 1
                
                # 添加自定义字段，序号从当前idx开始
                for field in custom_fields_list:
                    # 使用辅助函数确保字段名是标量
                    field_str = ensure_scalar(field).strip()
                    if field_str:
                        # 确保每个元素都是标量值，不是可迭代对象
                        # 显式转换为字符串，确保不是其他类型
                        field_name = str(field_str)
                        print(f"[DEBUG] 添加自定义字段: idx={idx}, field_str={field_str!r}, field_name={field_name!r}, type={type(field_name)}")
                        all_fields_df.append([int(idx), field_name, "自定义"])
                        idx += 1
                        custom_count += 1
                
                # 调试输出
                print(f"\n[DEBUG] 最终字段列表数据:")
                print(f"  all_fields_df类型: {type(all_fields_df)}")
                print(f"  all_fields_df内容: {all_fields_df}")
                for i, row in enumerate(all_fields_df):
                    print(f"  行{i}: {row}, 类型: {type(row)}, 字段名类型: {type(row[1]) if len(row) > 1 else 'N/A'}")
                
                # 生成状态消息
                total_count = len(all_fields_df)
                if total_count == 0:
                    status_msg = "⚠️ 警告：没有有效字段，请至少添加一个字段"
                else:
                    status_msg = f"✅ 字段已更新，共 {total_count} 个字段（默认：{default_count}，自定义：{custom_count}）"
                
                # 保存到app状态
                app.current_card_type = card_type
                app.current_default_fields = default_fields.copy()
                app.current_custom_fields = custom_fields_list.copy()
                
                # 卡证OCR不使用HTML模板，只使用DataFrame
                app.current_final_fields_html = None
                
                # 直接使用DataFrame展示（卡证OCR不使用HTML模板）
                return (
                    gr.update(visible=True),  # all_fields_title
                    gr.update(visible=False, value=""),  # HTML表格隐藏
                    gr.update(value=all_fields_df, visible=True),  # Dataframe
                    gr.update(value=status_msg, visible=True)
                )
            
            # 辅助函数：将Markdown表格转换为Dataframe格式
            def markdown_table_to_dataframe(markdown_text):
                """将Markdown表格转换为Dataframe格式（2D数组）"""
                if not markdown_text:
                    return []
                
                # 解析Markdown表格
                sections = app._parse_markdown_sections(markdown_text)
                dataframe_data = []
                
                for section in sections:
                    if section["type"] == "table":
                        header = section.get("header", [])
                        rows = section.get("rows", [])
                        
                        # 如果header为空，使用第一行作为header
                        if not header and rows:
                            header = rows[0] if rows else ["字段名", "字段值"]
                            rows = rows[1:] if len(rows) > 1 else []
                        
                        # 确保header至少有两列
                        if len(header) < 2:
                            header = ["字段名", "字段值"]
                        
                        # 转换为Dataframe格式：每行是[字段名, 字段值]
                        for row in rows:
                            if len(row) >= 2:
                                dataframe_data.append([str(row[0]).strip(), str(row[1]).strip()])
                            elif len(row) == 1:
                                dataframe_data.append([str(row[0]).strip(), ""])
                
                # 如果没有找到表格，返回空列表
                return dataframe_data if dataframe_data else []
            
            # 第三步：OCR识别
            def step3_ocr(image, all_fields_df):
                if image is None:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False)
                    )
                
                # 优先从应用状态获取字段列表（适用于HTML表格情况）
                fields_list = []
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    fields_list = app.current_default_fields.copy()
                if hasattr(app, 'current_custom_fields') and app.current_custom_fields:
                    fields_list.extend(app.current_custom_fields)
                
                # 如果应用状态中没有，则从DataFrame提取字段名
                if not fields_list:
                    # 安全地检查DataFrame是否为空
                    has_fields = all_fields_df is not None and (
                        (isinstance(all_fields_df, list) and len(all_fields_df) > 0) or
                        (hasattr(all_fields_df, '__len__') and len(all_fields_df) > 0)
                    )
                    
                    if not has_fields:
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False)
                        )
                    
                    # 从DataFrame提取字段名（排除"来源"列）
                    if all_fields_df is not None:
                        # 确保是列表格式
                        if not isinstance(all_fields_df, list):
                            try:
                                all_fields_df = all_fields_df.tolist() if hasattr(all_fields_df, 'tolist') else list(all_fields_df)
                            except:
                                all_fields_df = []
                        
                        for row in all_fields_df:
                            if row and len(row) >= 2 and row[1] and str(row[1]).strip():
                                fields_list.append(str(row[1]).strip())
                
                if not fields_list:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False)
                    )
                
                result = app.ocr_card_with_fields(image, fields_list)
                
                # 卡证OCR不使用HTML模板，只使用Markdown/DataFrame
                # 提取Markdown文本（去掉可能的图标前缀）
                if result.startswith("🪪"):
                    markdown_text = result.split(":", 1)[1].strip() if ":" in result else result
                else:
                    markdown_text = result
                
                # 保存到app状态以便导出
                app.last_ocr_markdown = f"## 卡证OCR识别结果（三步流程）\n\n{markdown_text}"
                
                # 解析OCR结果，提取字段值字典
                ocr_data = {}
                sections = app._parse_markdown_sections(markdown_text)
                for section in sections:
                    if section["type"] == "table":
                        rows = section.get("rows", [])
                        for row in rows:
                            if len(row) >= 2:
                                field_name = str(row[0]).strip()
                                field_value = str(row[1]).strip()
                                if field_name:
                                    ocr_data[field_name] = field_value
                
                # 卡证OCR不使用HTML模板，只使用DataFrame
                # 将OCR结果转换为DataFrame格式
                ocr_dataframe = []
                for field_name, field_value in ocr_data.items():
                    ocr_dataframe.append([field_name, field_value])
                
                # 如果没有数据，返回空DataFrame
                if not ocr_dataframe:
                    ocr_dataframe = [["", ""]]
                
                return (
                    gr.update(visible=False),  # HTML表格隐藏
                    gr.update(value=ocr_dataframe, visible=True),  # Dataframe
                    gr.update(visible=True),  # 导出格式
                    gr.update(visible=True),  # 导出按钮
                    gr.update(visible=False, value="")  # 导出状态
                )
                
                # 以下代码不再使用（卡证OCR不使用HTML模板）
                if False:
                    # 使用HTML表格展示OCR结果
                    try:
                        from bs4 import BeautifulSoup
                        soup = BeautifulSoup(html_template, 'html.parser')
                        table = soup.find('table')
                        
                        if table:
                            # 填充OCR结果到表格中
                            # 策略：遍历所有行，查找包含字段名的单元格，然后在同一行或下一行填充值
                            for row in table.find_all('tr'):
                                cells = row.find_all(['td', 'th'])
                                for i, cell in enumerate(cells):
                                    cell_text = cell.get_text(strip=True)
                                    # 检查是否是字段名（在OCR结果中存在）
                                    if cell_text and cell_text in ocr_data:
                                        # 查找同一行中的下一个空单元格来填充值
                                        found = False
                                        for j in range(i + 1, len(cells)):
                                            next_cell = cells[j]
                                            next_text = next_cell.get_text(strip=True)
                                            # 如果下一个单元格为空，填充OCR结果
                                            if not next_text:
                                                next_cell.string = ocr_data[cell_text]
                                                found = True
                                                break
                                            # 如果下一个单元格不是字段名，也填充（可能是值单元格）
                                            elif next_text not in ocr_data or next_text == '':
                                                next_cell.string = ocr_data[cell_text]
                                                found = True
                                                break
                                        
                                        # 如果同一行没有找到合适的单元格，在当前单元格后插入
                                        if not found:
                                            value_cell = soup.new_tag('td')
                                            value_cell.string = ocr_data[cell_text]
                                            cell.insert_after(value_cell)
                                        
                                        # 标记已处理，避免重复填充
                                        ocr_data.pop(cell_text, None)
                            
                            # 添加自定义字段（如果有）
                            custom_fields = getattr(app, 'current_custom_fields', [])
                            for custom_field in custom_fields:
                                if custom_field in ocr_data:
                                    new_row = soup.new_tag('tr')
                                    new_row['class'] = 'custom-field-row'
                                    field_cell = soup.new_tag('td')
                                    field_cell.string = custom_field
                                    field_cell['colspan'] = '2'
                                    value_cell = soup.new_tag('td')
                                    value_cell.string = ocr_data.get(custom_field, '')
                                    value_cell['colspan'] = '3'
                                    new_row.append(field_cell)
                                    new_row.append(value_cell)
                                    table.append(new_row)
                            
                            # 添加样式
                            styled_html = f"""
                            <style>
                            .ocr-result-table {{
                                width: 100%;
                                border-collapse: collapse;
                                margin: 10px 0;
                                font-size: 14px;
                            }}
                            .ocr-result-table th,
                            .ocr-result-table td {{
                                border: 1px solid #ddd;
                                padding: 8px;
                                text-align: left;
                            }}
                            .ocr-result-table th {{
                                background-color: #f2f2f2;
                                font-weight: bold;
                            }}
                            .ocr-result-table tr:nth-child(even) {{
                                background-color: #f9f9f9;
                            }}
                            .ocr-result-table .custom-field-row {{
                                background-color: #fff3cd !important;
                            }}
                            .ocr-result-table td[contenteditable="true"] {{
                                background-color: #e7f3ff;
                                cursor: text;
                            }}
                            </style>
                            {str(table)}
                            """
                            
                            return (
                                gr.update(value=styled_html, visible=True),  # HTML表格
                                gr.update(value=[], visible=False),  # Dataframe隐藏，传递空列表避免验证错误
                                gr.update(visible=True),  # ocr_export_format
                                gr.update(visible=True),  # ocr_export_btn_3step
                                gr.update(visible=False, value="")  # ocr_export_status_3step
                            )
                    except Exception as e:
                        print(f"⚠️ 生成OCR结果HTML表格失败: {e}")
                        import traceback
                        traceback.print_exc()
                        # 降级到Dataframe展示
                        dataframe_data = markdown_table_to_dataframe(markdown_text)
                        return (
                            gr.update(visible=False, value=""),  # HTML表格隐藏
                            gr.update(value=dataframe_data, visible=True),  # Dataframe
                            gr.update(visible=True),  # ocr_export_format
                            gr.update(visible=True),  # ocr_export_btn_3step
                            gr.update(visible=False, value="")  # ocr_export_status_3step
                        )
                else:
                    # 没有HTML模板，使用Dataframe展示
                    dataframe_data = markdown_table_to_dataframe(markdown_text)
                    return (
                        gr.update(visible=False, value=""),  # HTML表格隐藏
                        gr.update(value=dataframe_data, visible=True),  # Dataframe
                        gr.update(visible=True),  # ocr_export_format
                        gr.update(visible=True),  # ocr_export_btn_3step
                        gr.update(visible=False, value="")  # ocr_export_status_3step
                    )
            
            detect_type_btn.click(
                step1_detect_type,
                inputs=[card_image],
                outputs=[
                    card_type_output,
                    default_fields_html,
                    default_fields_output,
                    default_fields_title,
                    custom_fields_input,
                    custom_fields_title,
                    add_custom_field_btn,
                    update_fields_btn,
                    all_fields_title,
                    all_fields_output,
                    fields_status
                ]
            )
            
            # 添加自定义字段按钮的功能
            def add_custom_field(current_data):
                """在自定义字段Dataframe中添加一个新行"""
                if current_data is None:
                    return [[""]]
                # 确保是列表格式
                if not isinstance(current_data, list):
                    try:
                        current_data = current_data.tolist() if hasattr(current_data, 'tolist') else list(current_data)
                    except:
                        current_data = [[""]]
                # 添加一个新行
                new_data = list(current_data) if current_data else []
                new_data.append([""])
                return new_data
            
            add_custom_field_btn.click(
                add_custom_field,
                inputs=[custom_fields_input],
                outputs=[custom_fields_input]
            )
            
            update_fields_btn.click(
                step2_update_fields,
                inputs=[card_type_output, default_fields_output, custom_fields_input],
                outputs=[all_fields_title, all_fields_html, all_fields_output, fields_status]
            )
            
            # 当字段更新后，显示OCR按钮（用于Dataframe）
            def show_ocr_btn_from_dataframe(all_fields_df):
                # 优先检查应用状态（适用于HTML表格情况）
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    return gr.update(visible=True)
                
                # 安全地检查DataFrame
                if all_fields_df is None:
                    return gr.update(visible=False)
                
                # 确保是列表格式
                if not isinstance(all_fields_df, list):
                    try:
                        all_fields_df = all_fields_df.tolist() if hasattr(all_fields_df, 'tolist') else list(all_fields_df)
                    except:
                        return gr.update(visible=False)
                
                if len(all_fields_df) > 0:
                    # 检查是否有有效字段
                    has_fields = any(
                        row and len(row) >= 2 and row[1] and str(row[1]).strip()
                        for row in all_fields_df
                    )
                    return gr.update(visible=has_fields)
                return gr.update(visible=False)
            
            # 当HTML表格更新后，显示OCR按钮（用于HTML）
            def show_ocr_btn_from_html(html_content):
                # HTML组件变化时，直接检查应用状态
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    return gr.update(visible=True)
                return gr.update(visible=False)
            
            # 监听all_fields_output和all_fields_html的变化
            all_fields_output.change(
                show_ocr_btn_from_dataframe,
                inputs=[all_fields_output],
                outputs=[ocr_with_fields_btn]
            )
            all_fields_html.change(
                show_ocr_btn_from_html,
                inputs=[all_fields_html],
                outputs=[ocr_with_fields_btn]
            )
            
            ocr_with_fields_btn.click(
                step3_ocr,
                inputs=[card_image, all_fields_output],  # all_fields_output可能为空（HTML表格情况），但会从app状态获取
                outputs=[ocr_result_html, ocr_result, ocr_export_format, ocr_export_btn_3step, ocr_export_status_3step]
            )
            
            # 辅助函数：将Dataframe转换为Markdown表格
            def dataframe_to_markdown_table(dataframe_data):
                """将Dataframe数据转换为Markdown表格格式"""
                if not dataframe_data:
                    return ""
                
                # 确保是列表格式
                if not isinstance(dataframe_data, list):
                    try:
                        dataframe_data = dataframe_data.tolist() if hasattr(dataframe_data, 'tolist') else list(dataframe_data)
                    except:
                        return ""
                
                # 构建Markdown表格
                lines = ["| 字段名 | 字段值 |", "|--------|--------|"]
                for row in dataframe_data:
                    if row and len(row) >= 2:
                        field_name = str(row[0]).strip() if row[0] else ""
                        field_value = str(row[1]).strip() if row[1] else ""
                        lines.append(f"| {field_name} | {field_value} |")
                
                return "\n".join(lines)
            
            # 导出OCR结果（从Dataframe读取当前编辑后的内容）
            def export_ocr_result_3step(dataframe_data, export_format):
                """导出三步流程的OCR结果（支持多种格式）"""
                if not dataframe_data or (isinstance(dataframe_data, list) and len(dataframe_data) == 0):
                    return gr.update(visible=True, value="❌ 没有可保存的OCR结果，请先执行OCR识别！")
                
                export_dir = os.path.join("ocr_exports")
                os.makedirs(export_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # 确保是列表格式
                if not isinstance(dataframe_data, list):
                    try:
                        dataframe_data = dataframe_data.tolist() if hasattr(dataframe_data, 'tolist') else list(dataframe_data)
                    except:
                        dataframe_data = []
                
                try:
                    # 根据选择的格式导出
                    if export_format == "Markdown (.md)":
                        # 转换为Markdown表格
                        markdown_table = dataframe_to_markdown_table(dataframe_data)
                        markdown_content = f"## 卡证OCR识别结果（三步流程）\n\n{markdown_table}"
                        
                        file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.md")
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write(markdown_content)
                        
                        app.last_ocr_markdown = markdown_content
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 Markdown文件: {file_path}\n\n已保存当前编辑后的内容。")
                    
                    elif export_format == "Excel (.xlsx)":
                        try:
                            from openpyxl import Workbook
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "OCR结果"
                            
                            # 添加表头
                            ws.append(["字段名", "字段值"])
                            
                            # 添加数据行
                            for row in dataframe_data:
                                if row and len(row) >= 2:
                                    ws.append([str(row[0]).strip() if row[0] else "", str(row[1]).strip() if row[1] else ""])
                            
                            file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.xlsx")
                            wb.save(file_path)
                            return gr.update(visible=True, value=f"✅ 导出成功！\n📊 Excel文件: {file_path}\n\n已保存当前编辑后的内容。")
                        except Exception as e:
                            # 如果Excel导出失败，尝试CSV
                            file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.csv")
                            with open(file_path, "w", encoding="utf-8", newline="") as f:
                                writer = csv.writer(f)
                                writer.writerow(["字段名", "字段值"])
                                for row in dataframe_data:
                                    if row and len(row) >= 2:
                                        writer.writerow([str(row[0]).strip() if row[0] else "", str(row[1]).strip() if row[1] else ""])
                            return gr.update(visible=True, value=f"⚠️ Excel导出失败，已保存为CSV\n📄 CSV文件: {file_path}\n\n错误: {str(e)}")
                    
                    elif export_format == "CSV (.csv)":
                        file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.csv")
                        with open(file_path, "w", encoding="utf-8", newline="") as f:
                            writer = csv.writer(f)
                            writer.writerow(["字段名", "字段值"])
                            for row in dataframe_data:
                                if row and len(row) >= 2:
                                    writer.writerow([str(row[0]).strip() if row[0] else "", str(row[1]).strip() if row[1] else ""])
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 CSV文件: {file_path}\n\n已保存当前编辑后的内容。")
                    
                    elif export_format == "JSON (.json)":
                        # 转换为字典列表
                        json_data = []
                        for row in dataframe_data:
                            if row and len(row) >= 2:
                                json_data.append({
                                    "字段名": str(row[0]).strip() if row[0] else "",
                                    "字段值": str(row[1]).strip() if row[1] else ""
                                })
                        
                        file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.json")
                        with open(file_path, "w", encoding="utf-8") as f:
                            json.dump(json_data, f, ensure_ascii=False, indent=2)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 JSON文件: {file_path}\n\n已保存当前编辑后的内容。")
                    
                    else:
                        return gr.update(visible=True, value=f"❌ 不支持的导出格式: {export_format}")
                        
                except Exception as e:
                    return gr.update(visible=True, value=f"❌ 导出失败: {str(e)}")
            
            ocr_export_btn_3step.click(
                export_ocr_result_3step,
                inputs=[ocr_result, ocr_export_format],
                outputs=[ocr_export_status_3step]
            )

        with gr.Tab("📄 单据OCR（三步流程）"):
            gr.Markdown("### 三步流程：识别类型 → 自定义字段 → OCR识别（使用HTML表格模板）")
            
            with gr.Row():
                with gr.Column(scale=1):
                    bill_image = gr.Image(
                        label="上传票据图片",
                        type="pil",
                        height=400
                    )
                    
                    with gr.Row():
                        detect_bill_type_btn = gr.Button("🔍 第一步：识别票据类型", variant="primary")
                    
                    bill_type_output = gr.Textbox(
                        label="识别的票据类型",
                        interactive=False,
                        visible=False
                    )
                    
                    bill_default_fields_title = gr.Markdown("### 📋 默认字段模板", visible=False)
                    # HTML表格展示（票据OCR使用HTML模板）
                    bill_default_fields_html = gr.HTML(
                        label="默认字段模板（HTML表格）",
                        visible=False,
                        elem_id="bill-default-fields-html"
                    )
                    
                    bill_custom_fields_title = gr.Markdown("### ➕ 自定义字段", visible=False)
                    bill_custom_fields_input = gr.Dataframe(
                        label="添加自定义字段（每行一个字段名）",
                        headers=["字段名"],
                        datatype=["str"],
                        interactive=True,
                        visible=False,
                        wrap=True,
                        row_count=(1, "dynamic"),
                        col_count=(1, "fixed"),
                        type="array",
                        value=[[""]]
                    )
                    
                    with gr.Row():
                        bill_add_custom_field_btn = gr.Button("➕ 添加自定义字段", variant="secondary", visible=False, size="sm")
                    
                    with gr.Row():
                        bill_update_fields_btn = gr.Button("🔗 第二步：合并字段", variant="secondary", visible=False)
                    
                    bill_all_fields_title = gr.Markdown("### ✅ 最终字段列表（将用于OCR识别）", visible=False)
                    # HTML表格展示（票据OCR使用HTML模板）
                    bill_all_fields_html = gr.HTML(
                        label="最终字段列表（HTML表格）",
                        visible=False,
                        elem_id="bill-all-fields-html"
                    )
                    
                    bill_fields_status = gr.Textbox(
                        label="状态",
                        interactive=False,
                        visible=False
                    )
                    
                    with gr.Row():
                        bill_ocr_with_fields_btn = gr.Button("🚀 第三步：开始OCR识别", variant="primary", visible=False)
                
                with gr.Column(scale=2):
                    with gr.Row():
                        gr.Markdown("### 📊 OCR识别结果")
                        with gr.Column(scale=1, min_width=200):
                            bill_ocr_export_format = gr.Dropdown(
                                choices=["Markdown (.md)", "Excel (.xlsx)", "CSV (.csv)", "JSON (.json)"],
                                value="Markdown (.md)",
                                label="导出格式",
                                visible=False
                            )
                        bill_ocr_export_btn_3step = gr.Button("💾 导出结果", variant="secondary", visible=False, size="sm", elem_id="bill-ocr-export-btn")
                    
                    # HTML表格展示（票据OCR使用HTML模板）
                    bill_ocr_result_html = gr.HTML(
                        label="OCR识别结果（HTML表格）",
                        visible=False,
                        elem_id="bill-ocr-result-html"
                    )
                    
                    # 隐藏的Textbox，用于存储编辑后的HTML内容
                    # 注意：不使用elem_id，让Gradio自动生成ID，然后通过返回值更新
                    bill_ocr_result_html_edited = gr.Textbox(
                        label="编辑后的HTML内容",
                        visible=False
                    )
                    
                    bill_ocr_export_status_3step = gr.Textbox(
                        label="导出状态",
                        interactive=False,
                        visible=False,
                        lines=3
                    )
            
            # 第一步：识别票据类型
            def bill_step1_detect_type(image):
                if image is None:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        "❌ 请先上传图片"
                    )
                
                result = app.detect_bill_type(image)
                if len(result) == 4:
                    bill_type, default_fields, html_template, status_msg = result
                else:
                    bill_type, default_fields, status_msg = result
                    html_template = None
                
                if bill_type:
                    has_html_template = html_template is not None and html_template.strip()
                    
                    if has_html_template:
                        styled_html = f"""
                        <style>
                        .field-template-table {{
                            width: 100%;
                            border-collapse: collapse;
                            margin: 10px 0;
                            font-size: 14px;
                        }}
                        .field-template-table th,
                        .field-template-table td {{
                            border: 1px solid #ddd;
                            padding: 8px;
                            text-align: left;
                        }}
                        .field-template-table th {{
                            background-color: #f2f2f2;
                            font-weight: bold;
                        }}
                        .field-template-table tr:nth-child(even) {{
                            background-color: #f9f9f9;
                        }}
                        </style>
                        {html_template}
                        """
                        return (
                            gr.update(value=bill_type, visible=True),
                            gr.update(value=styled_html, visible=True),
                            gr.update(visible=True),
                            gr.update(value=[[""]], visible=True),
                            gr.update(visible=True),
                            gr.update(visible=True),
                            gr.update(visible=True),
                            gr.update(value=status_msg, visible=True)
                        )
                    else:
                        return (
                            gr.update(value=bill_type, visible=True),
                            gr.update(visible=False),
                            gr.update(visible=True),
                            gr.update(value=[[""]], visible=True),
                            gr.update(visible=True),
                            gr.update(visible=True),
                            gr.update(visible=True),
                            gr.update(value=status_msg, visible=True)
                        )
                else:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(value=status_msg, visible=True)
                    )
            
            # 第二步：合并字段（票据OCR使用HTML模板）
            def bill_step2_update_fields(card_type, custom_fields_df):
                if not card_type:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(value="❌ 请先完成第一步：识别票据类型", visible=True)
                    )
                
                default_fields = []
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    default_fields = app.current_default_fields.copy()
                
                custom_fields_list = []
                if custom_fields_df is not None:
                    if isinstance(custom_fields_df, (list, tuple)):
                        rows = custom_fields_df
                    else:
                        rows = []
                    
                    for row in rows:
                        if not row:
                            continue
                        if isinstance(row, (list, tuple)) and len(row) > 0:
                            field_value = row[0]
                        else:
                            continue
                        field_str = str(field_value).strip()
                        if field_str and field_str not in default_fields:
                            custom_fields_list.append(field_str)
                
                app.current_custom_fields = custom_fields_list.copy()
                
                html_template = getattr(app, 'current_field_template_html', None)
                has_html_template = html_template is not None and html_template.strip()
                
                final_fields_html = None
                
                if has_html_template and custom_fields_list:
                    try:
                        from bs4 import BeautifulSoup
                        soup = BeautifulSoup(html_template, 'html.parser')
                        table = soup.find('table')
                        
                        if table:
                            for custom_field in custom_fields_list:
                                new_row = soup.new_tag('tr')
                                field_cell = soup.new_tag('td')
                                field_cell.string = custom_field
                                field_cell['colspan'] = '2'
                                value_cell = soup.new_tag('td')
                                value_cell.string = ''
                                value_cell['colspan'] = '3'
                                new_row.append(field_cell)
                                new_row.append(value_cell)
                                table.append(new_row)
                            
                            styled_html = f"""
                            <style>
                            .all-fields-table {{
                                width: 100%;
                                border-collapse: collapse;
                                margin: 10px 0;
                                font-size: 14px;
                            }}
                            .all-fields-table th,
                            .all-fields-table td {{
                                border: 1px solid #ddd;
                                padding: 8px;
                                text-align: left;
                            }}
                            .all-fields-table th {{
                                background-color: #f2f2f2;
                                font-weight: bold;
                            }}
                            .all-fields-table tr:nth-child(even) {{
                                background-color: #f9f9f9;
                            }}
                            .custom-field-row {{
                                background-color: #fff3cd !important;
                            }}
                            </style>
                            {str(table)}
                            """
                            
                            final_fields_html = str(table)
                            app.current_final_fields_html = final_fields_html
                            
                            total_count = len(default_fields) + len(custom_fields_list)
                            status_msg = f"✅ 字段已更新，共 {total_count} 个字段（默认：{len(default_fields)}，自定义：{len(custom_fields_list)}）"
                            
                            return (
                                gr.update(visible=True),
                                gr.update(value=styled_html, visible=True),
                                gr.update(value=status_msg, visible=True)
                            )
                    except Exception as e:
                        print(f"⚠️ 生成HTML表格失败: {e}")
                        status_msg = f"⚠️ 生成HTML表格失败: {e}"
                        return (
                            gr.update(visible=True),
                            gr.update(visible=False),
                            gr.update(value=status_msg, visible=True)
                        )
                elif has_html_template:
                    try:
                        from bs4 import BeautifulSoup
                        soup = BeautifulSoup(html_template, 'html.parser')
                        table = soup.find('table')
                        if table:
                            final_fields_html = str(table)
                        else:
                            final_fields_html = html_template
                    except:
                        final_fields_html = html_template
                    
                    app.current_final_fields_html = final_fields_html
                    
                    styled_html = f"""
                    <style>
                    .all-fields-table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin: 10px 0;
                        font-size: 14px;
                    }}
                    .all-fields-table th,
                    .all-fields-table td {{
                        border: 1px solid #ddd;
                        padding: 8px;
                        text-align: left;
                    }}
                    .all-fields-table th {{
                        background-color: #f2f2f2;
                        font-weight: bold;
                    }}
                    .all-fields-table tr:nth-child(even) {{
                        background-color: #f9f9f9;
                    }}
                    </style>
                    {html_template}
                    """
                    status_msg = f"✅ 字段已更新，共 {len(default_fields)} 个字段"
                    return (
                        gr.update(visible=True),
                        gr.update(value=styled_html, visible=True),
                        gr.update(value=status_msg, visible=True)
                    )
                else:
                    app.current_final_fields_html = None
                    status_msg = "⚠️ 未找到HTML模板"
                    return (
                        gr.update(visible=True),
                        gr.update(visible=False),
                        gr.update(value=status_msg, visible=True)
                    )
            
            # 第三步：OCR识别
            def bill_step3_ocr(image):
                if image is None:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False)
                    )
                
                fields_list = []
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    fields_list = app.current_default_fields.copy()
                if hasattr(app, 'current_custom_fields') and app.current_custom_fields:
                    fields_list.extend(app.current_custom_fields)
                
                if not fields_list:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False)
                    )
                
                result = app.ocr_bill_with_fields(image, fields_list)
                
                html_template = getattr(app, 'current_field_template_html', None)
                has_html_template = html_template is not None and html_template.strip()
                
                if has_html_template and "<table" in result.lower():
                    app.last_ocr_html = result
                    app.last_ocr_markdown = ""
                    return (
                        gr.update(value=result, visible=True),
                        gr.update(value=result, visible=False),  # 同时更新隐藏的Textbox
                        gr.update(visible=True),
                        gr.update(visible=True),
                        gr.update(visible=False, value="")
                    )
                else:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False, value=""),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False)
                    )
            
            # 导出票据OCR结果
            def bill_export_ocr_result_3step(html_content, export_format):
                if not html_content or not html_content.strip():
                    return gr.update(visible=True, value="❌ 没有可保存的OCR结果，请先执行OCR识别！")
                
                # 如果接收到的内容看起来像是完整的HTML（包含style或script标签），尝试提取表格
                # 否则直接返回错误
                if '<style>' in html_content or '<script>' in html_content:
                    # 这是完整的HTML，需要提取表格部分
                    pass
                elif '<table' not in html_content.lower():
                    return gr.update(visible=True, value="❌ 未找到表格数据，无法导出！")
                
                export_dir = os.path.join("ocr_exports")
                os.makedirs(export_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                try:
                    from bs4 import BeautifulSoup
                    # 解析HTML内容，提取表格数据
                    soup = BeautifulSoup(html_content, 'html.parser')
                    # 移除script和style标签，只保留表格
                    for script in soup(["script", "style"]):
                        script.decompose()
                    
                    table = soup.find('table', class_='ocr-result-table') or soup.find('table')
                    
                    if not table:
                        return gr.update(visible=True, value="❌ 未找到表格数据，无法导出！")
                    
                    # 提取表格数据：处理复杂的表格结构（包含rowspan和colspan）
                    def extract_table_data(table):
                        """提取表格数据，处理rowspan和colspan"""
                        data = []
                        rows = table.find_all('tr')
                        
                        for row in rows:
                            cells = row.find_all(['td', 'th'])
                            if not cells:
                                continue
                            
                            # 提取所有单元格的文本
                            row_data = []
                            for cell in cells:
                                cell_text = cell.get_text(strip=True)
                                # 跳过空单元格或样式类名
                                if cell_text and cell_text not in ['et2', 'et9', 'et11']:
                                    row_data.append(cell_text)
                            
                            # 如果行中有数据，尝试配对字段名和值
                            if len(row_data) >= 2:
                                # 尝试配对：第一个是字段名，后续是值
                                for i in range(0, len(row_data) - 1, 2):
                                    if i + 1 < len(row_data):
                                        field = row_data[i]
                                        value = row_data[i + 1]
                                        # 跳过明显的样式类名
                                        if field not in ['et2', 'et9', 'et11', ''] and value not in ['et2', 'et9', 'et11', '']:
                                            data.append([field, value])
                            elif len(row_data) == 1:
                                # 单列数据，可能是字段名或值（需要与上一行配对）
                                pass  # 暂时跳过单列数据
                        
                        return data
                    
                    # 改进的提取方法：处理复杂的表格结构（rowspan和colspan）
                    def extract_simple_table_data(table):
                        """提取表格数据，处理rowspan和colspan"""
                        data = []
                        rows = table.find_all('tr')
                        
                        # 已知的字段名列表（用于识别字段名）
                        known_fields = [
                            '出票日期', '票据状态', '汇票到期日', '票号', '出票人', '收票人',
                            '全称', '账号', '开户银行', '出票保证信息', '票据金额', '承兑人信息',
                            '开户行行号', '开户行名称', '交易合同号', '能否转让', '承兑信息',
                            '承兑保证信息', '评级信息', '备注'
                        ]
                        
                        for row in rows:
                            cells = row.find_all(['td', 'th'])
                            if not cells:
                                continue
                            
                            # 提取所有非空单元格文本
                            cell_texts = []
                            for cell in cells:
                                text = cell.get_text(strip=True)
                                # 跳过样式类名、空文本和纯数字
                                if text and text not in ['et2', 'et9', 'et11', 'et3', '']:
                                    # 检查是否是样式类名（通常是短字符串且全小写或全大写）
                                    if not (len(text) <= 3 and text.isalnum() and text.islower()):
                                        cell_texts.append(text)
                            
                            if len(cell_texts) < 2:
                                continue
                            
                            # 识别字段名和值
                            # 字段名通常是：1) 在known_fields中 2) 较短且不包含大量数字
                            # 值通常是：1) 较长 2) 包含数字或特殊字符
                            field = None
                            values = []
                            
                            for text in cell_texts:
                                # 检查是否是已知字段名
                                is_field = False
                                for known_field in known_fields:
                                    if known_field in text or text in known_field:
                                        if not field:  # 如果还没有找到字段名
                                            field = text
                                            is_field = True
                                            break
                                
                                if not is_field:
                                    # 判断是否是字段名（较短且不包含大量数字）
                                    if not field and len(text) < 15 and text.count('0') + text.count('1') + text.count('2') + text.count('3') + text.count('4') + text.count('5') + text.count('6') + text.count('7') + text.count('8') + text.count('9') < len(text) * 0.3:
                                        field = text
                                    else:
                                        values.append(text)
                            
                            # 如果有字段名和值，添加到数据中
                            if field and values:
                                # 合并多个值为一个（用空格分隔）
                                value = ' '.join(values)
                                # 避免重复添加相同的字段
                                if not any(d[0] == field for d in data):
                                    data.append([field, value])
                            elif field and not values:
                                # 只有字段名没有值，可能是rowspan的情况，跳过或标记为空
                                pass
                        
                        return data
                    
                    # 使用简单方法提取数据
                    table_data = extract_simple_table_data(table)
                    
                    # 调试信息
                    print(f"[DEBUG] 提取到的表格数据: {len(table_data)} 条")
                    for i, (field, value) in enumerate(table_data[:5]):  # 只打印前5条
                        print(f"  {i+1}. {field}: {value[:50]}...")
                    
                    if not table_data:
                        # 如果提取失败，尝试更简单的方法
                        print("[DEBUG] 简单提取失败，尝试备用方法...")
                        table_data = []
                        rows = table.find_all('tr')
                        for row in rows:
                            cells = row.find_all(['td', 'th'])
                            texts = [cell.get_text(strip=True) for cell in cells]
                            texts = [t for t in texts if t and t not in ['et2', 'et9', 'et11', 'et3'] and len(t) > 1]
                            if len(texts) >= 2:
                                # 简单配对：第一个是字段名，其余是值
                                field = texts[0]
                                value = ' '.join(texts[1:])
                                if field and value:
                                    table_data.append([field, value])
                        
                        if not table_data:
                            return gr.update(visible=True, value="❌ 表格数据为空，无法导出！请检查表格格式。")
                    
                    if export_format == "Markdown (.md)":
                        markdown_lines = ["## 票据OCR识别结果\n\n| 字段名 | 字段值 |"]
                        markdown_lines.append("|--------|--------|")
                        for field, value in table_data:
                            # 转义Markdown特殊字符
                            field_escaped = field.replace('|', '\\|')
                            value_escaped = value.replace('|', '\\|').replace('\n', ' ')
                            markdown_lines.append(f"| {field_escaped} | {value_escaped} |")
                        markdown_content = "\n".join(markdown_lines)
                        
                        file_name = f"bill_ocr_{timestamp}.md"
                        file_path = os.path.join(export_dir, file_name)
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write(markdown_content)
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 Markdown文件已保存到:\n{abs_file_path}")
                    elif export_format == "Excel (.xlsx)":
                        import pandas as pd
                        df = pd.DataFrame(table_data, columns=["字段名", "字段值"])
                        file_name = f"bill_ocr_{timestamp}.xlsx"
                        file_path = os.path.join(export_dir, file_name)
                        df.to_excel(file_path, index=False)
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 Excel文件已保存到:\n{abs_file_path}")
                    elif export_format == "CSV (.csv)":
                        import pandas as pd
                        df = pd.DataFrame(table_data, columns=["字段名", "字段值"])
                        file_name = f"bill_ocr_{timestamp}.csv"
                        file_path = os.path.join(export_dir, file_name)
                        df.to_csv(file_path, index=False, encoding='utf-8-sig')
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 CSV文件已保存到:\n{abs_file_path}")
                    elif export_format == "JSON (.json)":
                        import json
                        data = {field: value for field, value in table_data}
                        file_name = f"bill_ocr_{timestamp}.json"
                        file_path = os.path.join(export_dir, file_name)
                        with open(file_path, "w", encoding="utf-8") as f:
                            json.dump(data, f, ensure_ascii=False, indent=2)
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 JSON文件已保存到:\n{abs_file_path}")
                    else:
                        return gr.update(visible=True, value=f"❌ 不支持的导出格式: {export_format}")
                except Exception as e:
                    import traceback
                    error_msg = f"❌ 导出失败: {str(e)}\n{traceback.format_exc()}"
                    print(error_msg)
                    return gr.update(visible=True, value=f"❌ 导出失败: {str(e)}")
            
            # 绑定事件
            detect_bill_type_btn.click(
                bill_step1_detect_type,
                inputs=[bill_image],
                outputs=[bill_type_output, bill_default_fields_html, bill_default_fields_title, 
                        bill_custom_fields_input, bill_custom_fields_title, bill_add_custom_field_btn,
                        bill_update_fields_btn, bill_fields_status]
            )
            
            def bill_add_custom_field(current_data):
                if current_data is None:
                    current_data = [[""]]
                elif not isinstance(current_data, list):
                    try:
                        current_data = current_data.tolist() if hasattr(current_data, 'tolist') else list(current_data)
                    except:
                        current_data = [[""]]
                new_data = list(current_data) if current_data else []
                new_data.append([""])
                return new_data
            
            bill_add_custom_field_btn.click(
                bill_add_custom_field,
                inputs=[bill_custom_fields_input],
                outputs=[bill_custom_fields_input]
            )
            
            bill_update_fields_btn.click(
                bill_step2_update_fields,
                inputs=[bill_type_output, bill_custom_fields_input],
                outputs=[bill_all_fields_title, bill_all_fields_html, bill_fields_status]
            )
            
            def bill_show_ocr_btn_from_html(all_fields_html):
                if all_fields_html and all_fields_html.strip():
                    return gr.update(visible=True)
                return gr.update(visible=False)
            
            bill_all_fields_html.change(
                bill_show_ocr_btn_from_html,
                inputs=[bill_all_fields_html],
                outputs=[bill_ocr_with_fields_btn]
            )
            
            bill_ocr_with_fields_btn.click(
                bill_step3_ocr,
                inputs=[bill_image],
                outputs=[bill_ocr_result_html, bill_ocr_result_html_edited, bill_ocr_export_format, bill_ocr_export_btn_3step, bill_ocr_export_status_3step]
            )
            
            # 监听HTML组件的change事件，同步更新隐藏的Textbox
            def sync_edited_html(html_content):
                if html_content:
                    return html_content
                return ""
            
            bill_ocr_result_html.change(
                sync_edited_html,
                inputs=[bill_ocr_result_html],
                outputs=[bill_ocr_result_html_edited]
            )
            
            # 导出函数：使用JavaScript更新隐藏的Textbox，然后从Textbox读取
            def export_with_js_content(html_edited, export_format):
                """导出函数：使用JavaScript更新后的内容"""
                print(f"[DEBUG] export_with_js_content接收到内容:")
                print(f"  - html_edited类型: {type(html_edited)}")
                print(f"  - html_edited长度: {len(html_edited) if html_edited else 0}")
                if html_edited:
                    print(f"  - html_edited预览: {html_edited[:200]}...")
                
                if not html_edited or not html_edited.strip():
                    return gr.update(visible=True, value="❌ 没有可保存的OCR结果，请先执行OCR识别！")
                
                # 调用导出函数
                return bill_export_ocr_result_3step(html_edited, export_format)
            
            # JavaScript函数：在导出前从DOM读取编辑后的表格内容并更新隐藏的Textbox
            js_code = """
            function() {
                var table = document.querySelector('.ocr-result-table');
                if (!table) {
                    console.error('[DEBUG] 未找到表格元素');
                    return [null];
                }
                
                // 获取编辑后的表格HTML（包含所有用户编辑的内容）
                var styleTag = document.querySelector('style');
                var styleContent = styleTag ? styleTag.outerHTML : '';
                var tableHtml = table.outerHTML;
                var fullContent = styleContent + '\\n' + tableHtml;
                
                console.log('[DEBUG] 从DOM获取的表格HTML长度:', tableHtml.length);
                console.log('[DEBUG] 表格内容预览:', tableHtml.substring(0, 200));
                console.log('[DEBUG] 准备返回编辑后的内容，长度:', fullContent.length);
                
                // 返回编辑后的内容，Gradio会自动更新bill_ocr_result_html_edited组件
                return [fullContent];
            }
            """
            
            # 使用JavaScript更新隐藏的Textbox，然后导出
            # 第一步：JavaScript更新bill_ocr_result_html_edited组件
            # 第二步：从bill_ocr_result_html_edited读取内容并导出
            bill_ocr_export_btn_3step.click(
                fn=None,  # 不使用Python函数，只执行JavaScript
                inputs=None,
                outputs=[bill_ocr_result_html_edited],  # JavaScript返回的值更新这个组件
                js=js_code
            ).then(
                export_with_js_content,
                inputs=[bill_ocr_result_html_edited, bill_ocr_export_format],
                outputs=[bill_ocr_export_status_3step]
            )

        with gr.Tab("ℹ️ 使用说明"):
            gr.Markdown(
                """
                - 先点击「加载模型」后再使用各项功能。
                - 「通用版」适合快速上手与日常使用；「专业版」提供更细粒度的生成参数与高级功能（批量分析、图像对比）。
                - 已默认优化为更易触摸点击的界面尺寸。
                """
            )

        # 绑定模式切换：控制高级组件可见性
        mode.change(
            _toggle_mode,
            inputs=[mode, pro_task, code_format],
            outputs=[adv_params_box, stats_output, tab_batch, tab_compare, pro_task, code_format, text_input],
        )

        pro_task.change(
            _toggle_task,
            inputs=[pro_task, code_format],
            outputs=[code_format, text_input],
        )

        code_format.change(
            _update_code_prompt,
            inputs=[pro_task, code_format],
            outputs=[text_input],
        )

    return interface


def main():
    print("🚀 启动Qwen3-VL-8B-Instruct 统一Web界面...")
    interface = create_unified_interface()

    def _cleanup():
        # 清理模型与显存
        try:
            app.model = None
            app.processor = None
            app.is_loaded = False
        except Exception:
            pass
        try:
            if torch is not None and hasattr(torch, "cuda") and torch.cuda.is_available():
                torch.cuda.empty_cache()
                torch.cuda.ipc_collect()
        except Exception:
            pass
        try:
            gc.collect()
        except Exception:
            pass

    # 注册进程退出清理
    atexit.register(_cleanup)
    interface.queue()
    interface.launch(
        server_name="127.0.0.1",
        server_port=None,  # 自动选择可用端口，避免端口占用错误
        share=False,
        debug=True,
        show_error=True,
    )


if __name__ == "__main__":
    main()
