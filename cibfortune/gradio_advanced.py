#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Qwen3-VL-8B-Instruct é«˜çº§Gradioç•Œé¢
åŒ…å«æ›´å¤šé«˜çº§åŠŸèƒ½å’Œæ›´å¥½çš„ç”¨æˆ·ä½“éªŒ
"""

import os
import gradio as gr
import torch
from transformers import Qwen3VLForConditionalGeneration, AutoProcessor
from PIL import Image
import requests
from io import BytesIO
import time
import json
import base64
import csv
from datetime import datetime

# è®¾ç½®ç¯å¢ƒå˜é‡
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'

class AdvancedQwen3VLApp:
    """é«˜çº§Qwen3-VLåº”ç”¨ç±»"""
    
    def __init__(self):
        self.model = None
        self.processor = None
        self.model_path = "/data/storage1/wulin/models/qwen3-vl-8b-instruct"
        self.is_loaded = False
        self.chat_history = []
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.chat_messages = []
        self.last_image = None
        self.last_saved_image_path = None
        self.last_image_digest = None
        self.last_ocr_markdown = None
        
    def load_model(self, progress=gr.Progress()):
        """åŠ è½½æ¨¡å‹"""
        if self.is_loaded:
            return "âœ… æ¨¡å‹å·²ç»åŠ è½½å®Œæˆï¼", gr.update(interactive=True)
        
        try:
            progress(0.1, desc="æ£€æŸ¥æ¨¡å‹è·¯å¾„...")
            if not os.path.exists(self.model_path):
                return f"âŒ æ¨¡å‹è·¯å¾„ä¸å­˜åœ¨: {self.model_path}", gr.update(interactive=False)
            
            progress(0.3, desc="åŠ è½½æ¨¡å‹...")
            self.model = Qwen3VLForConditionalGeneration.from_pretrained(
                self.model_path,
                dtype="auto",
                device_map="auto"
            )
            
            progress(0.7, desc="åŠ è½½å¤„ç†å™¨...")
            self.processor = AutoProcessor.from_pretrained(self.model_path)
            
            progress(1.0, desc="å®Œæˆï¼")
            self.is_loaded = True
            
            return "âœ… æ¨¡å‹åŠ è½½æˆåŠŸï¼å¯ä»¥å¼€å§‹ä½¿ç”¨äº†ã€‚", gr.update(interactive=True)
            
        except Exception as e:
            return f"âŒ æ¨¡å‹åŠ è½½å¤±è´¥: {str(e)}", gr.update(interactive=False)
    
    def _prepare_user_message(self, image, prompt):
        prompt_clean = (prompt or "").strip()
        resolved_image = image if image is not None else self.last_image
        if resolved_image is None:
            raise ValueError("âŒ è¯·ä¸Šä¼ å›¾åƒï¼")
        if not prompt_clean:
            raise ValueError("âŒ è¯·è¾“å…¥é—®é¢˜ï¼")
        if image is not None:
            self.last_image = image
        content = [
            {"type": "image", "image": resolved_image},
            {"type": "text", "text": prompt_clean},
        ]
        return prompt_clean, {"role": "user", "content": content}

    def _run_inference(self,
                       image,
                       prompt,
                       max_tokens,
                       temperature,
                       top_p,
                       top_k,
                       repetition_penalty,
                       prepared=None):
        if prepared is None:
            prompt_clean, user_message = self._prepare_user_message(image, prompt)
        else:
            prompt_clean, user_message = prepared
        messages = self.chat_messages + [user_message]

        inputs = self.processor.apply_chat_template(
            messages,
            tokenize=True,
            add_generation_prompt=True,
            return_dict=True,
            return_tensors="pt"
        ).to(self.model.device)

        generation_kwargs = {
            "max_new_tokens": max_tokens,
            "temperature": temperature,
            "top_p": top_p,
            "top_k": top_k,
            "do_sample": True if temperature > 0 else False,
            "repetition_penalty": repetition_penalty
        }

        start_time = time.time()
        with torch.no_grad():
            generated_ids = self.model.generate(**inputs, **generation_kwargs)
        generation_time = time.time() - start_time

        generated_ids_trimmed = [
            out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text = self.processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )
        response = output_text[0]

        assistant_message = {"role": "assistant", "content": [{"type": "text", "text": response}]}
        self.chat_messages.extend([user_message, assistant_message])
        return prompt_clean, response, generation_time

    def _clone_history(self, history):
        return [[turn[0], turn[1]] for turn in history]

    def _chunk_response(self, text, chunk_size=80):
        if not text:
            return []
        return [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]

    def _parse_markdown_sections(self, markdown_text):
        """
        å°† Markdown æ–‡æœ¬æ‹†åˆ†ä¸º table/text æ®µï¼Œæ”¯æŒï¼š
        - ç®¡é“è¡¨æ ¼ï¼ˆ| a | b |ï¼‰
        - HTML <table>ï¼ˆè‹¥å­˜åœ¨ï¼‰
        å¹¶åœ¨è§£æå‰å¯¹å›´æ ä»£ç å—è¿›è¡Œå»å›´æ æ¸…æ´—ï¼Œç¡®ä¿å¯¼å‡ºä¸æ¸²æŸ“ä¸€è‡´ã€‚
        """
        sections = []
        if not markdown_text:
            return sections
        
        # 1) å…ˆå»æ‰å›´æ ï¼Œä½¿å¾—â€œä»£ç å—ä¸­çš„è¡¨æ ¼â€ä¹Ÿèƒ½è¢«è¯†åˆ«ä¸ºå¯å¯¼å‡ºçš„å†…å®¹
        cleaned_md = self._sanitize_markdown(markdown_text)
        
        # 2) å…ˆå°è¯•è§£æ HTML è¡¨æ ¼ï¼ˆè‹¥æ¨¡å‹è¾“å‡ºäº† <table>ï¼‰
        html_tables = []
        try:
            from bs4 import BeautifulSoup  # å¯é€‰ä¾èµ–
            soup = BeautifulSoup(cleaned_md, "html.parser")
            for t in soup.find_all("table"):
                headers = []
                header_row = t.find("tr")
                if header_row:
                    # å¦‚æœæœ‰ <th> ç”¨ thï¼›å¦åˆ™ç”¨é¦–è¡Œçš„ td ä½œä¸º header
                    ths = header_row.find_all("th")
                    if ths:
                        headers = [th.get_text(strip=True) for th in ths]
                        data_rows = header_row.find_next_siblings("tr")
                    else:
                        tds = header_row.find_all("td")
                        headers = [td.get_text(strip=True) for td in tds]
                        data_rows = header_row.find_next_siblings("tr")
                rows = []
                for r in (data_rows or []):
                    cols = r.find_all(["td", "th"])
                    rows.append([c.get_text(strip=True) for c in cols])
                if headers or rows:
                    html_tables.append({"type": "table", "header": headers, "rows": rows})
        except Exception:
            # å¦‚æœ bs4 ä¸åœ¨ç¯å¢ƒä¸­ï¼Œåˆ™ç•¥è¿‡ HTML è§£æ
            pass
        
        # 3) è§£æç®¡é“è¡¨æ ¼
        lines = cleaned_md.splitlines()
        i = 0
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()
            
            # ç®¡é“è¡¨æ ¼åˆ¤å®šï¼šå½“å‰è¡Œå’Œä¸‹ä¸€è¡Œæ„æˆ header + åˆ†éš”
            is_table = (
                stripped.startswith("|")
                and stripped.count("|") >= 2
                and i + 1 < len(lines)
                and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                and lines[i + 1].strip().startswith("|")
            )
            
            if is_table:
                header = [cell.strip() for cell in stripped.strip("|").split("|")]
                i += 2  # è·³è¿‡ header ä¸åˆ†éš”çº¿
                rows = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if not (row_line.startswith("|") and row_line.count("|") >= 2):
                        break
                    row = [cell.strip() for cell in row_line.strip("|").split("|")]
                    rows.append(row)
                    i += 1
                sections.append({"type": "table", "header": header, "rows": rows})
                continue
            
            # æ™®é€šæ–‡æœ¬å—ï¼ˆç›´åˆ°é‡åˆ°ä¸‹ä¸€ä¸ªè¡¨æ ¼æˆ–æ–‡ä»¶ç»“æŸï¼‰
            text_block = []
            while i < len(lines):
                current = lines[i]
                stripped_current = current.strip()
                next_is_table = (
                    stripped_current.startswith("|")
                    and stripped_current.count("|") >= 2
                    and i + 1 < len(lines)
                    and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                    and lines[i + 1].strip().startswith("|")
                )
                if next_is_table:
                    break
                text_block.append(current)
                i += 1
                # ä¿ç•™ç©ºè¡Œï¼Œæ”¹å–„æ®µè½åˆ†éš”çš„å¯è¯»æ€§
                if i < len(lines) and lines[i] == "":
                    text_block.append(lines[i])
            
            text_content = "\n".join(text_block).strip("\n")
            if text_content:
                sections.append({"type": "text", "text": text_content})
        
        # 4) è‹¥å­˜åœ¨ HTML è¡¨ï¼Œä¼˜å…ˆæŠŠ HTML è¡¨ä¹ŸåŠ å…¥ï¼ˆæ”¾åœ¨è§£æç»“æœå‰é¢ï¼Œé¿å…é—æ¼ï¼‰
        if html_tables:
            # å°† HTML è¡¨æ’åœ¨æœ€å‰é¢ï¼ˆä¹Ÿå¯æ ¹æ®éœ€è¦åˆå¹¶/å»é‡ï¼‰
            sections = html_tables + sections
        
        return sections

    def chat_with_image(self, image, text, history, max_tokens, temperature, top_p, top_k, repetition_penalty: float = 1.0, presence_penalty: float = 1.5):
        """ä¸å›¾åƒå¯¹è¯ï¼ˆæµå¼åé¦ˆï¼‰"""
        original_text = text

        if not self.is_loaded:
            yield history, original_text, gr.update(value="âŒ è¯·å…ˆåŠ è½½æ¨¡å‹ï¼", visible=True)
            return

        try:
            prepared = self._prepare_user_message(image, text)
        except ValueError as exc:
            yield history, original_text, gr.update(value=str(exc), visible=True)
            return

        prompt_clean, _ = prepared
        history_copy = self._clone_history(history)
        history_copy.append([f"ğŸ‘¤ {prompt_clean}", "ğŸ¤– æ­£åœ¨æ€è€ƒ..."])
        yield self._clone_history(history_copy), original_text, gr.update(value="ğŸ¤– æ­£åœ¨æ€è€ƒ...", visible=True)

        try:
            _, response, generation_time = self._run_inference(
                image,
                text,
                max_tokens,
                temperature,
                top_p,
                top_k,
                repetition_penalty,
                prepared=prepared
            )
        except Exception as e:
            history_copy[-1][1] = f"âŒ ç”Ÿæˆå¤±è´¥: {str(e)}"
            self.chat_history = self._clone_history(history_copy)
            yield self._clone_history(history_copy), original_text, gr.update(value=f"âŒ é”™è¯¯: {str(e)}", visible=True)
            return

        assembled = ""
        chunks = self._chunk_response(response)
        if not chunks:
            chunks = [""]
        for chunk in chunks:
            assembled += chunk
            history_copy[-1][1] = f"ğŸ¤– {assembled}â–Œ"
            yield self._clone_history(history_copy), original_text, gr.update(value=f"ğŸ¤– {assembled}â–Œ", visible=True)

        stats = (
            f"â±ï¸ ç”Ÿæˆæ—¶é—´: {generation_time:.2f}ç§’ | ğŸ“ ç”Ÿæˆé•¿åº¦: {len(response)}å­—ç¬¦"
            f" | âš™ï¸ æœ€å¤§é•¿åº¦: {max_tokens}"
        )
        if max_tokens > 1024:
            stats += " | â³ æç¤º: è¾ƒå¤§çš„æœ€å¤§é•¿åº¦å¯èƒ½å»¶é•¿ç”Ÿæˆæ—¶é—´"
        history_copy[-1][1] = f"ğŸ¤– {response}"
        self.chat_history = self._clone_history(history_copy)
        yield self._clone_history(history_copy), original_text, gr.update(value=stats, visible=True)

    def _sanitize_markdown(self, text: str) -> str:
        if not text:
            return ""
        s = text.strip()
        lines = s.splitlines()
        out = []
        in_fence = False
        for line in lines:
            stripped = line.strip()
            if stripped.startswith("```"):
                in_fence = not in_fence
                continue
            if not in_fence:
                out.append(line)
        cleaned = "\n".join(out).strip()
        return cleaned if cleaned else s

    def ocr_analysis(self, image, prompt: str = None):
        """OCRæ–‡å­—è¯†åˆ«ï¼Œå¯é€‰è‡ªå®šä¹‰æç¤ºè¯"""
        if not self.is_loaded:
            return "âŒ è¯·å…ˆåŠ è½½æ¨¡å‹ï¼"
        default_prompt = (
            "è¯·è¯†åˆ«å¹¶æå–è¿™å¼ å›¾ç‰‡ä¸­çš„æ‰€æœ‰æ–‡å­—å†…å®¹ï¼Œå°½é‡è¿˜åŸåŸæœ¬æ ·å¼ï¼Œå¹¶æ ‡æ³¨è¯­è¨€ç±»å‹ã€‚"
            " è¯·ç¡®ä¿æ‰€æœ‰å¸¦æ ·å¼æˆ–è¡¨æ ¼å†…å®¹ä½¿ç”¨Markdownè¡¨æ ¼è¡¨ç¤ºã€‚"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1024,
                temperature=0.7,
                top_p=0.8,
                top_k=50,
                repetition_penalty=1.0
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"ğŸ‘¤ {prompt_clean}", f"ğŸ¤– {cleaned}"])
            self.last_ocr_markdown = f"## OCRè¯†åˆ«ç»“æœ\n\n{cleaned}"
            return f"ğŸ“ OCRè¯†åˆ«ç»“æœ:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"âŒ OCRè¯†åˆ«å¤±è´¥: {str(e)}"

    def spatial_analysis(self, image, prompt: str = None):
        """ç©ºé—´æ„ŸçŸ¥åˆ†æï¼Œå¯é€‰è‡ªå®šä¹‰æç¤ºè¯"""
        if not self.is_loaded:
            return "âŒ è¯·å…ˆåŠ è½½æ¨¡å‹ï¼"
        default_prompt = (
            "è¯·åˆ†æè¿™å¼ å›¾ç‰‡ä¸­çš„ç©ºé—´å…³ç³»ï¼ŒåŒ…æ‹¬ç›¸å¯¹ä½ç½®ã€è§†è§’ã€é®æŒ¡ã€æ·±åº¦ä¸è·ç¦»æ„Ÿï¼Œå¹¶ç»™å‡ºæ•´ä½“å¸ƒå±€æè¿°ã€‚"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=768,
                temperature=0.7,
                top_p=0.8,
                top_k=50,
                repetition_penalty=1.0
            )
            self.chat_history.append([f"ğŸ‘¤ {prompt_clean}", f"ğŸ¤– {response}"])
            return f"ğŸ“ ç©ºé—´åˆ†æç»“æœ:\n\n{response}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"âŒ ç©ºé—´åˆ†æå¤±è´¥: {str(e)}"

    def visual_coding(self, image, output_format: str = "HTML", prompt: str = None):
        """è§†è§‰ç¼–ç¨‹ç”Ÿæˆä»£ç ï¼Œå¯é€‰è‡ªå®šä¹‰æç¤ºè¯"""
        if not self.is_loaded:
            return "âŒ è¯·å…ˆåŠ è½½æ¨¡å‹ï¼"
        base_prompts = {
            "HTML": "è¯·æ ¹æ®å›¾ç‰‡ç”Ÿæˆå¯¹åº”çš„HTMLç»“æ„ä»£ç ï¼ŒåŒ…å«å¿…è¦çš„è¯­ä¹‰æ ‡ç­¾ã€‚",
            "CSS": "è¯·ä¸ºè¯¥å›¾ç‰‡å¯¹åº”çš„ç•Œé¢ç”Ÿæˆåˆç†çš„CSSæ ·å¼ä»£ç ï¼ŒåŒ…æ‹¬å¸ƒå±€ä¸é¢œè‰²ã€‚",
            "JavaScript": "è¯·æ ¹æ®å›¾ç‰‡äº¤äº’ç”ŸæˆJavaScriptä»£ç ç¤ºä¾‹ï¼ŒåŒ…å«å¿…è¦çš„äº‹ä»¶ä¸é€»è¾‘ã€‚",
            "Python": "è¯·ç”Ÿæˆèƒ½å¤ç°è¯¥ç•Œé¢/å¸ƒå±€çš„Pythonç¤ºä¾‹ä»£ç ï¼ˆå¦‚ä½¿ç”¨streamlitæˆ–flaskçš„ä¼ªä»£ç ï¼‰ã€‚",
        }
        default_prompt = base_prompts.get(output_format, base_prompts["HTML"]) + " è¯·åªè¾“å‡ºä»£ç ï¼Œä¸è¦é¢å¤–è¯´æ˜ã€‚"
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1024,
                temperature=0.4,
                top_p=0.8,
                top_k=50,
                repetition_penalty=1.0
            )
            self.chat_history.append([f"ğŸ‘¤ {prompt_clean}", f"ğŸ¤– {response}"])
            return response
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"âŒ è§†è§‰ç¼–ç¨‹å¤±è´¥: {str(e)}"
    
    def batch_analysis(self, images, analysis_type):
        """æ‰¹é‡åˆ†æ"""
        if not self.is_loaded:
            return "âŒ è¯·å…ˆåŠ è½½æ¨¡å‹ï¼"
        
        if not images:
            return "âŒ è¯·ä¸Šä¼ å›¾åƒï¼"
        
        results = []
        
        for i, image in enumerate(images):
            try:
                if analysis_type == "æè¿°":
                    prompt = "è¯·çœŸå®ã€è¯¦ç»†ã€å®¢è§‚åœ°æè¿°è¿™å¼ å›¾ç‰‡çš„å†…å®¹ã€‚"
                elif analysis_type == "OCR":
                    prompt = "è¯·è¯†åˆ«å¹¶æå–è¿™å¼ å›¾ç‰‡ä¸­çš„æ‰€æœ‰æ–‡å­—å†…å®¹ï¼Œå°½é‡è¿˜åŸåŸæœ¬æ ·å¼ï¼Œå¹¶æ ‡æ³¨è¯­è¨€ç±»å‹ã€‚"
                elif analysis_type == "ç©ºé—´åˆ†æ":
                    prompt = "è¯·åˆ†æè¿™å¼ å›¾ç‰‡ä¸­çš„ç©ºé—´å…³ç³»å’Œç‰©ä½“ä½ç½®ï¼ŒåŒ…æ‹¬ç›¸å¯¹ä½ç½®ã€è§†è§’ã€é®æŒ¡ã€æ·±åº¦ä¸è·ç¦»æ„Ÿï¼Œå¹¶ç»™å‡ºæ•´ä½“å¸ƒå±€æè¿°ã€‚"
                elif analysis_type == "æƒ…æ„Ÿåˆ†æ":
                    prompt = "è¯·åˆ†æè¿™å¼ å›¾ç‰‡ä¼ è¾¾çš„æƒ…æ„Ÿæˆ–æ°›å›´ã€‚"
                else:
                    prompt = "è¯·åˆ†æè¿™å¼ å›¾ç‰‡ã€‚"
                
                messages = [
                    {
                        "role": "user",
                        "content": [
                            {"type": "image", "image": image},
                            {"type": "text", "text": prompt},
                        ],
                    }
                ]
                
                inputs = self.processor.apply_chat_template(
                    messages,
                    tokenize=True,
                    add_generation_prompt=True,
                    return_dict=True,
                    return_tensors="pt"
                )
                inputs = inputs.to(self.model.device)
                
                with torch.no_grad():
                    generated_ids = self.model.generate(**inputs, max_new_tokens=512)
                
                generated_ids_trimmed = [
                    out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
                ]
                output_text = self.processor.batch_decode(
                    generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
                )
                
                results.append(f"ğŸ“· å›¾åƒ {i+1}:\n{output_text[0]}\n" + "="*50 + "\n")
                
            except Exception as e:
                results.append(f"ğŸ“· å›¾åƒ {i+1}: âŒ åˆ†æå¤±è´¥ - {str(e)}\n" + "="*50 + "\n")
        
        return "".join(results)
    
    def compare_images(self, image1, image2, comparison_type):
        """å›¾åƒå¯¹æ¯”"""
        if not self.is_loaded:
            return "âŒ è¯·å…ˆåŠ è½½æ¨¡å‹ï¼"
        
        if image1 is None or image2 is None:
            return "âŒ è¯·ä¸Šä¼ ä¸¤å¼ å›¾åƒè¿›è¡Œå¯¹æ¯”ï¼"
        
        try:
            if comparison_type == "ç›¸ä¼¼æ€§":
                prompt = "è¯·å¯¹æ¯”è¿™ä¸¤å¼ å›¾ç‰‡ï¼Œåˆ†æå®ƒä»¬çš„ç›¸ä¼¼ä¹‹å¤„å’Œä¸åŒä¹‹å¤„ã€‚"
            elif comparison_type == "é£æ ¼":
                prompt = "è¯·å¯¹æ¯”è¿™ä¸¤å¼ å›¾ç‰‡çš„è‰ºæœ¯é£æ ¼ã€è‰²å½©æ­é…å’Œæ„å›¾ç‰¹ç‚¹ã€‚"
            elif comparison_type == "å†…å®¹":
                prompt = "è¯·å¯¹æ¯”è¿™ä¸¤å¼ å›¾ç‰‡çš„å†…å®¹ï¼Œåˆ†æå®ƒä»¬æè¿°çš„åœºæ™¯æˆ–ä¸»é¢˜ã€‚"
            else:
                prompt = "è¯·å¯¹æ¯”è¿™ä¸¤å¼ å›¾ç‰‡ï¼Œæä¾›è¯¦ç»†çš„å¯¹æ¯”åˆ†æã€‚"
            
            messages = [
                {
                    "role": "user",
                    "content": [
                        {"type": "image", "image": image1},
                        {"type": "image", "image": image2},
                        {"type": "text", "text": prompt},
                    ],
                }
            ]
            
            inputs = self.processor.apply_chat_template(
                messages,
                tokenize=True,
                add_generation_prompt=True,
                return_dict=True,
                return_tensors="pt"
            )
            inputs = inputs.to(self.model.device)
            
            with torch.no_grad():
                generated_ids = self.model.generate(**inputs, max_new_tokens=1024)
            
            generated_ids_trimmed = [
                out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
            ]
            output_text = self.processor.batch_decode(
                generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
            )
            
            return f"ğŸ” å¯¹æ¯”åˆ†æç»“æœ:\n\n{output_text[0]}"
            
        except Exception as e:
            return f"âŒ å¯¹æ¯”åˆ†æå¤±è´¥: {str(e)}"
    
    def export_chat_history(self):
        """å¯¼å‡ºå¯¹è¯å†å²"""
        if not self.chat_history:
            return "âŒ æ²¡æœ‰å¯¹è¯å†å²å¯å¯¼å‡ºï¼"
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"chat_history_{timestamp}.json"
            
            # ä¿å­˜ä¸ºJSONæ ¼å¼
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.chat_history, f, ensure_ascii=False, indent=2)
            
            return f"âœ… å¯¹è¯å†å²å·²å¯¼å‡ºåˆ°: {filename}"
            
        except Exception as e:
            return f"âŒ å¯¼å‡ºå¤±è´¥: {str(e)}"
    
    def clear_history(self):
        """æ¸…ç©ºå¯¹è¯å†å²"""
        self.chat_history = []
        self.chat_messages = []
        self.last_image = None
        self.last_saved_image_path = None
        self.last_image_digest = None
        self.last_ocr_markdown = None
        if hasattr(self, "session_turn_image_paths"):
            self.session_turn_image_paths.clear()
        return []

    def export_last_ocr(self):
        if not self.last_ocr_markdown:
            return "âŒ æ²¡æœ‰å¯ä¿å­˜çš„æ–‡æœ¬æ ·å¼ï¼Œè¯·å…ˆæ‰§è¡Œä¸€æ¬¡OCRè¯†åˆ«ï¼"

        export_dir = os.path.join("ocr_exports")
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        sections = self._parse_markdown_sections(self.last_ocr_markdown)

        excel_path = os.path.join(export_dir, f"ocr_{timestamp}.xlsx")
        excel_note = ""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "è¡¨æ ¼1" if sections else "OCRæ–‡æœ¬"
            table_idx = 0
            for section in sections:
                if section["type"] == "table":
                    table_idx += 1
                    if table_idx > 1:
                        ws = wb.create_sheet(title=f"è¡¨æ ¼{table_idx}")
                    ws.append(section["header"])
                    for row in section["rows"]:
                        ws.append(row)
                elif section["type"] == "text" and section["text"]:
                    if table_idx > 0:
                        ws = wb.create_sheet(title=f"æ–‡æœ¬{table_idx}")
                    for line in section["text"].splitlines():
                        ws.append([line])
            if not sections:
                for line in self.last_ocr_markdown.splitlines():
                    ws.append([line])
            wb.save(excel_path)
        except Exception as exc:
            excel_path = os.path.join(export_dir, f"ocr_{timestamp}.csv")
            with open(excel_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["OCR Result"])
                for line in self.last_ocr_markdown.splitlines():
                    writer.writerow([line])
            excel_note = f"âš ï¸ Excelå¯¼å‡ºå¤±è´¥({exc})ï¼Œå·²ä¿å­˜ä¸ºCSV"

        json_path = os.path.join(export_dir, f"ocr_{timestamp}.json")
        json_content = {
            "markdown": self.last_ocr_markdown,
            "sections": sections,
        }
        with open(json_path, "w", encoding='utf-8') as f:
            json.dump(json_content, f, ensure_ascii=False, indent=2)

        message_lines = [
            "âœ… æ–‡æœ¬æ ·å¼å·²ä¿å­˜ï¼š",
            f"- Excel: {excel_path}" + (f" ({excel_note})" if excel_note else ""),
            f"- JSON: {json_path}",
        ]
        return "\n".join(message_lines)

# åˆ›å»ºåº”ç”¨å®ä¾‹
app = AdvancedQwen3VLApp()

def create_advanced_interface():
    """åˆ›å»ºé«˜çº§Gradioç•Œé¢"""
    
    with gr.Blocks(
        title="Qwen3-VL-8B-Instruct é«˜çº§ç•Œé¢",
        theme=gr.themes.Soft(),
        css="""
        :root {
            --radius-lg: 18px;
            --radius-md: 12px;
            --surface: #ffffff;
            --surface-muted: #f4f6fb;
            --surface-border: #e2e8f0;
            --text-primary: #0f172a;
            --text-secondary: #64748b;
            --accent: #2563eb;
            --accent-soft: rgba(37, 99, 235, 0.12);
        }
        body {
            background: linear-gradient(140deg, #eef2ff 0%, #f8fafc 45%, #ffffff 100%);
            color: var(--text-primary);
        }
        .gradio-container {
            max-width: 1600px !important;
            margin: 0 auto;
            padding: 18px 22px 48px;
            color: var(--text-primary);
        }
        #advanced-header {
            background: linear-gradient(135deg, rgba(37, 99, 235, 0.12), rgba(96, 165, 250, 0.1));
            border: 1px solid rgba(37, 99, 235, 0.18);
            padding: 22px 26px;
            border-radius: 24px;
            box-shadow: 0 16px 32px rgba(15, 23, 42, 0.08);
            margin-bottom: 20px;
        }
        #advanced-header h1 {
            margin: 0 0 6px;
            font-size: 26px;
            font-weight: 600;
            letter-spacing: 0.2px;
            color: var(--text-primary);
        }
        #advanced-header p {
            margin: 0;
            font-size: 15px;
            color: var(--text-secondary);
        }
        .gradio-container .tabs {
            background: transparent;
            border: none;
        }
        .gradio-container .tabitem {
            border-radius: var(--radius-md);
            background: #f8fafc;
            border: 1px solid transparent;
            color: var(--text-secondary);
        }
        .gradio-container .tabitem.selected {
            border-color: rgba(37, 99, 235, 0.25);
            color: var(--text-primary);
            background: #ffffff;
            box-shadow: 0 8px 18px rgba(37, 99, 235, 0.08);
        }
        #advanced-input-panel, #advanced-chat-panel, #advanced-secondary-panel {
            background: var(--surface);
            border-radius: 22px;
            padding: 20px 22px;
            border: 1px solid var(--surface-border);
            box-shadow: 0 20px 40px rgba(15, 23, 42, 0.06);
        }
        #advanced-chat-panel {
            display: flex;
            flex-direction: column;
            gap: 16px;
        }
        #advanced-chatbot > .wrap {
            background: #f8fafc;
            border-radius: 18px;
            border: 1px solid rgba(148, 163, 184, 0.25);
            padding: 8px 10px;
        }
        #advanced-chatbot .message {
            border-radius: 14px !important;
            padding: 12px 14px !important;
            font-size: 15px;
            line-height: 1.6;
            color: var(--text-primary);
        }
        #advanced-chatbot .message.user {
            background: linear-gradient(135deg, rgba(37, 99, 235, 0.18), rgba(59, 130, 246, 0.12));
            border: 1px solid rgba(37, 99, 235, 0.25);
            color: var(--text-primary);
            align-self: flex-end;
        }
        #advanced-chatbot .message.bot {
            background: #ffffff;
            border: 1px solid rgba(203, 213, 225, 0.9);
            color: var(--text-primary);
            align-self: flex-start;
        }
        #advanced-chatbot .message.bot .markdown ul {
            padding-left: 22px;
        }
        #advanced-query textarea {
            border-radius: 14px;
            border: 1px solid rgba(148, 163, 184, 0.35);
            background: var(--surface);
            color: var(--text-primary);
            box-shadow: inset 0 1px 3px rgba(15, 23, 42, 0.05);
        }
        #advanced-query textarea:focus {
            border-color: var(--accent);
            box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
        }
        #advanced-params .slider {
            padding: 6px 0;
        }
        #advanced-params .slider input[type="range"]::-webkit-slider-thumb {
            background: var(--accent);
        }
        #advanced-params .slider input[type="range"]::-moz-range-thumb {
            background: var(--accent);
        }
        #advanced-stats textarea {
            background: var(--accent-soft);
            border: 1px solid rgba(37, 99, 235, 0.2);
            border-radius: 14px;
            color: var(--text-primary);
            font-weight: 500;
        }
        .gradio-container .gradio-button.primary {
            background: linear-gradient(135deg, #2563eb, #1d4ed8);
            border: none;
            color: #ffffff;
            font-weight: 600;
            box-shadow: 0 16px 28px rgba(37, 99, 235, 0.22);
        }
        .gradio-container .gradio-button.primary:hover {
            filter: brightness(1.03);
        }
        .gradio-container .gradio-button.secondary {
            background: rgba(37, 99, 235, 0.1);
            border: 1px solid rgba(37, 99, 235, 0.18);
            color: var(--text-primary);
        }
        .gradio-container textarea,
        .gradio-container input[type="text"],
        .gradio-container input[type="number"] {
            background: var(--surface);
            border: 1px solid rgba(148, 163, 184, 0.35);
            color: var(--text-primary);
            border-radius: 14px;
        }
        .gradio-container textarea:focus,
        .gradio-container input[type="text"]:focus,
        .gradio-container input[type="number"]:focus {
            border-color: var(--accent);
            box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
        }
        .gradio-container .slider > label,
        .gradio-container .checkbox-group > label,
        .gradio-container .radio-group > label {
            color: var(--text-secondary);
        }
        #ocr-md {
            max-height: 560px;
            overflow: auto;
            border: 1px solid rgba(148, 163, 184, 0.35);
            padding: 12px;
            border-radius: 14px;
            background: #ffffff;
        }
        """
    ) as interface:
        
        gr.HTML("""
        <section id="advanced-header">
            <h1>ğŸ¤– å¤šæ¨¡æ€å¤§è¯­è¨€æ¨¡å‹æ™ºèƒ½åˆ†æåŠ©æ‰‹</h1>
            <p>å‡çº§åçš„é¡µé¢å¸ƒå±€ä¸å¯¹è¯æ¡†æ ·å¼ï¼Œè®©å›¾åƒé—®ç­”ä¸é«˜çº§åˆ†æä½“éªŒæ›´æ²‰æµ¸ã€æ›´é«˜æ•ˆã€‚</p>
        </section>
        """)
        
        with gr.Tab("ğŸš€ æ¨¡å‹ç®¡ç†"):
            gr.Markdown("### æ¨¡å‹åŠ è½½ä¸ç®¡ç†")
            with gr.Row():
                with gr.Column():
                    load_btn = gr.Button("ğŸ”„ åŠ è½½æ¨¡å‹", variant="primary", size="lg")
                    status_text = gr.Textbox(
                        label="çŠ¶æ€", 
                        value="â³ æ¨¡å‹æœªåŠ è½½ï¼Œè¯·ç‚¹å‡»åŠ è½½æ¨¡å‹æŒ‰é’®",
                        interactive=False
                    )
                with gr.Column():
                    model_info = gr.Textbox(
                        label="æ¨¡å‹ä¿¡æ¯",
                        value=f"æ¨¡å‹è·¯å¾„: {app.model_path}",
                        interactive=False
                    )
            
            load_btn.click(
                app.load_model,
                outputs=[status_text, load_btn]
            )
        
        with gr.Tab("ğŸ’¬ æ™ºèƒ½å¯¹è¯"):
            with gr.Row(equal_height=True):
                with gr.Column(scale=1):
                    with gr.Group(elem_id="advanced-input-panel"):
                        gr.Markdown("### å›¾åƒä¸ç”Ÿæˆè®¾ç½®")
                        image_input = gr.Image(
                            label="ä¸Šä¼ å›¾åƒ",
                            type="pil",
                            height=390
                        )

                        with gr.Accordion("ğŸ›ï¸ ç”Ÿæˆå‚æ•°", open=False, elem_id="advanced-params"):
                            max_tokens = gr.Slider(
                                minimum=50, maximum=2048, value=256,
                                label="æœ€å¤§ç”Ÿæˆé•¿åº¦"
                            )
                            temperature = gr.Slider(
                                minimum=0.1, maximum=2.0, value=0.7,
                                label="åˆ›é€ æ€§ (Temperature)"
                            )
                            top_p = gr.Slider(
                                minimum=0.1, maximum=1.0, value=0.8,
                                label="Top-p"
                            )
                            top_k = gr.Slider(
                                minimum=1, maximum=100, value=20,
                                label="Top-k"
                            )

                with gr.Column(scale=2):
                    with gr.Group(elem_id="advanced-chat-panel"):
                        gr.Markdown("### å¯¹è¯ä¸è¾“å‡º")
                        chatbot = gr.Chatbot(
                            label=None,
                            height=600,
                            show_label=False,
                            type="tuples",
                            elem_id="advanced-chatbot",
                            render_markdown=True
                        )
                        text_input = gr.Textbox(
                            label=None,
                            placeholder="è¾“å…¥æƒ³äº†è§£çš„å†…å®¹ï¼ŒæŒ‰ Enter æˆ–ç‚¹å‡»å‘é€ã€‚",
                            lines=3,
                            elem_id="advanced-query"
                        )
                        send_btn = gr.Button("å‘é€", variant="primary")

                        stats_output = gr.Markdown(
                            value="",
                            elem_id="advanced-stats"
                        )

                        with gr.Row():
                            clear_btn = gr.Button("ğŸ—‘ï¸ æ¸…ç©ºå†å²", variant="secondary")
                            export_btn = gr.Button("ğŸ“ å¯¼å‡ºå¯¹è¯", variant="secondary")

        with gr.Tab("ğŸ“ OCRè¯†åˆ«"):
            gr.Markdown("### æ–‡å­—è¯†åˆ«")
            with gr.Row():
                with gr.Column(scale=1):
                    ocr_image = gr.Image(
                        label="ä¸Šä¼ å›¾åƒè¿›è¡ŒOCRè¯†åˆ«",
                        type="pil",
                        height=320
                    )
                    ocr_btn = gr.Button("ğŸ” å¼€å§‹è¯†åˆ«", variant="primary")
                with gr.Column(scale=1):
                    with gr.Row():
                        with gr.Column(scale=4):
                            ocr_md = gr.Markdown(
                                value="ï¼ˆè¯†åˆ«ç»“æœä¼šä»¥ Markdown æ¸²æŸ“ï¼ŒåŒ…æ‹¬è¡¨æ ¼ï¼‰",
                                elem_id="ocr-md"
                            )
                        with gr.Column(scale=1):
                            save_style_btn = gr.Button("ğŸ’¾ å¯¼å‡ºæ ·å¼", variant="secondary", interactive=False)
                            ocr_export_status = gr.Textbox(
                                label="å¯¼å‡ºçŠ¶æ€",
                                interactive=False,
                                lines=4
                            )

        # äº‹ä»¶ç»‘å®š
        def _run_ocr(image):
            result = app.ocr_analysis(image)
            can_save = bool(app.last_ocr_markdown)
            # å¯¹æ˜¾ç¤ºå†…å®¹ï¼šå¦‚æœæˆåŠŸï¼Œå±•ç¤ºæ¸²æŸ“åçš„ Markdownï¼›å¦‚æœå¤±è´¥ï¼Œå°†é”™è¯¯æç¤ºæ”¾åˆ°å³ä¾§çŠ¶æ€æ¡†
            display_md = app.last_ocr_markdown if can_save else ""
            status = "" if can_save else result
            return display_md, gr.update(interactive=can_save), status

        def _clear_all():
            app.clear_history()
            return (
                [],
                "",
                gr.update(value="", visible=False),
                gr.update(value="ï¼ˆè¯†åˆ«ç»“æœä¼šä»¥ Markdown æ¸²æŸ“ï¼ŒåŒ…æ‹¬è¡¨æ ¼ï¼‰", visible=True),
                gr.update(interactive=False),
                "",
            )

        send_btn.click(
            app.chat_with_image,
            inputs=[image_input, text_input, chatbot, max_tokens, temperature, top_p, top_k],
            outputs=[chatbot, text_input, stats_output]
        )

        text_input.submit(
            app.chat_with_image,
            inputs=[image_input, text_input, chatbot, max_tokens, temperature, top_p, top_k],
            outputs=[chatbot, text_input, stats_output]
        )

        clear_btn.click(
            _clear_all,
            outputs=[chatbot, text_input, stats_output, ocr_md, save_style_btn, ocr_export_status]
        )

        export_btn.click(
            app.export_chat_history,
            outputs=[stats_output]
        )

        ocr_btn.click(
            _run_ocr,
            inputs=[ocr_image],
            outputs=[ocr_md, save_style_btn, ocr_export_status]
        )

        save_style_btn.click(
            app.export_last_ocr,
            outputs=[ocr_export_status]
        )

        with gr.Tab("ğŸ“Š æ‰¹é‡åˆ†æ"):
            gr.Markdown("### æ‰¹é‡å›¾åƒåˆ†æ")
            
            with gr.Row():
                with gr.Column():
                    batch_images = gr.File(
                        label="ä¸Šä¼ å¤šä¸ªå›¾åƒ",
                        file_count="multiple",
                        file_types=["image"]
                    )
                    
                    analysis_type = gr.Dropdown(
                        choices=["æè¿°", "OCR", "ç©ºé—´åˆ†æ", "æƒ…æ„Ÿåˆ†æ"],
                        value="æè¿°",
                        label="åˆ†æç±»å‹"
                    )
                    
                    batch_btn = gr.Button("ğŸ” å¼€å§‹æ‰¹é‡åˆ†æ", variant="primary")
                
                with gr.Column():
                    batch_result = gr.Markdown()
            
            batch_btn.click(
                app.batch_analysis,
                inputs=[batch_images, analysis_type],
                outputs=[batch_result]
            )
        
        with gr.Tab("ğŸ”„ å›¾åƒå¯¹æ¯”"):
            gr.Markdown("### å›¾åƒå¯¹æ¯”åˆ†æ")
            
            with gr.Row():
                with gr.Column():
                    compare_image1 = gr.Image(
                        label="å›¾åƒ1",
                        type="pil",
                        height=200
                    )
                    compare_image2 = gr.Image(
                        label="å›¾åƒ2", 
                        type="pil",
                        height=200
                    )
                    
                    comparison_type = gr.Dropdown(
                        choices=["ç›¸ä¼¼æ€§", "é£æ ¼", "å†…å®¹", "ç»¼åˆ"],
                        value="ç›¸ä¼¼æ€§",
                        label="å¯¹æ¯”ç±»å‹"
                    )
                    
                    compare_btn = gr.Button("ğŸ”„ å¼€å§‹å¯¹æ¯”", variant="primary")
                
                with gr.Column():
                    compare_result = gr.Markdown()
            
            compare_btn.click(
                app.compare_images,
                inputs=[compare_image1, compare_image2, comparison_type],
                outputs=[compare_result]
            )
        
        with gr.Tab("â„¹ï¸ ä½¿ç”¨è¯´æ˜"):
            gr.Markdown("""
            ## ğŸ“– è¯¦ç»†ä½¿ç”¨è¯´æ˜
            
            ### ğŸš€ æ¨¡å‹ç®¡ç†
            - **åŠ è½½æ¨¡å‹**: é¦–æ¬¡ä½¿ç”¨å¿…é¡»ç‚¹å‡»"åŠ è½½æ¨¡å‹"æŒ‰é’®
            - **æ¨¡å‹è·¯å¾„**: `/data/storage1/wulin/models/qwen3-vl-8b-instruct`
            - **åŠ è½½æ—¶é—´**: é€šå¸¸éœ€è¦10ç§’ï¼Œè¯·è€å¿ƒç­‰å¾…
            
            ### ğŸ’¬ æ™ºèƒ½å¯¹è¯
            - **å›¾åƒä¸Šä¼ **: æ”¯æŒJPGã€PNGç­‰å¸¸è§æ ¼å¼
            - **å‚æ•°è°ƒèŠ‚**: 
              - æœ€å¤§ç”Ÿæˆé•¿åº¦: æ§åˆ¶å›ç­”çš„è¯¦ç»†ç¨‹åº¦
              - åˆ›é€ æ€§: æ•°å€¼è¶Šé«˜å›ç­”è¶Šæœ‰åˆ›æ„
              - Top-p/Top-k: æ§åˆ¶ç”Ÿæˆçš„éšæœºæ€§
            - **å¤šè½®å¯¹è¯**: æ”¯æŒåŸºäºå›¾åƒçš„è¿ç»­å¯¹è¯
            - **å†å²ç®¡ç†**: å¯æ¸…ç©ºæˆ–å¯¼å‡ºå¯¹è¯å†å²
            
            ### ğŸ“Š æ‰¹é‡åˆ†æ
            - **å¤šå›¾åƒä¸Šä¼ **: ä¸€æ¬¡å¯ä¸Šä¼ å¤šå¼ å›¾åƒ
            - **åˆ†æç±»å‹**:
              - æè¿°: è¯¦ç»†æè¿°å›¾åƒå†…å®¹
              - OCR: æå–å›¾åƒä¸­çš„æ–‡å­—
              - ç©ºé—´åˆ†æ: åˆ†æç©ºé—´å…³ç³»
              - æƒ…æ„Ÿåˆ†æ: åˆ†æå›¾åƒæƒ…æ„Ÿæ°›å›´
            
            ### ğŸ”„ å›¾åƒå¯¹æ¯”
            - **å¯¹æ¯”ç±»å‹**:
              - ç›¸ä¼¼æ€§: åˆ†æå›¾åƒçš„ç›¸ä¼¼å’Œå·®å¼‚
              - é£æ ¼: å¯¹æ¯”è‰ºæœ¯é£æ ¼å’Œè‰²å½©
              - å†…å®¹: å¯¹æ¯”åœºæ™¯å’Œä¸»é¢˜
              - ç»¼åˆ: å…¨é¢çš„å¯¹æ¯”åˆ†æ
            
            ### âš ï¸ æ³¨æ„äº‹é¡¹
            - ç¡®ä¿æœ‰è¶³å¤Ÿçš„å†…å­˜ï¼ˆå»ºè®®16GB+ï¼‰
            - æ”¯æŒGPUåŠ é€Ÿï¼ˆè‡ªåŠ¨æ£€æµ‹ï¼‰
            - å¤§å›¾åƒå¯èƒ½éœ€è¦æ›´é•¿çš„å¤„ç†æ—¶é—´
            - å»ºè®®ä¸€æ¬¡å¤„ç†ä¸è¶…è¿‡10å¼ å›¾åƒ
            """)
    
    return interface

def main():
    """ä¸»å‡½æ•°"""
    print("ğŸš€ å¯åŠ¨Qwen3-VL-8B-Instruct é«˜çº§Webç•Œé¢...")
    
    # åˆ›å»ºç•Œé¢
    interface = create_advanced_interface()
    
    interface.queue()

    # å¯åŠ¨æœåŠ¡
    interface.launch(
        server_name="0.0.0.0",
        server_port=7861,  # ä½¿ç”¨ä¸åŒç«¯å£é¿å…å†²çª
        share=False,
        debug=True,
        show_error=True
    )

if __name__ == "__main__":
    main()
