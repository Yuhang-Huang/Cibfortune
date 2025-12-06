#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Qwen3-VL-8B-Instruct 统一Gradio界面
单一界面覆盖图文问答与各类OCR工具，提供触屏友好样式
"""

import os
# 禁用 Gradio analytics 以避免网络连接错误（不影响功能）
os.environ['GRADIO_ANALYTICS_ENABLED'] = 'False'
import json
import inspect
import io
import hashlib
import time
import csv
import html
import re
import numpy as np
from datetime import datetime
import shutil
import atexit
import gc
import tempfile
import cv2

import gradio as gr
from transformers import Qwen3VLForConditionalGeneration, AutoProcessor
from ocr_card_rag_api import CardOCRWithRAG
from PIL import Image
from image import paddleocr_super_resolution

# 尝试导入文档知识库
try:
    from doc_knowledge_base import DocumentKnowledgeBase
    DOC_KB_AVAILABLE = True
except ImportError:
    DOC_KB_AVAILABLE = False
    print("⚠️ 文档知识库模块未找到，知识库功能将不可用")


# 尝试导入PDF处理库
PDF_AVAILABLE = False
PDF2IMAGE_AVAILABLE = False
PYMUPDF_AVAILABLE = False

# 尝试导入 pdf2image
try:
    from pdf2image import convert_from_path, convert_from_bytes
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    pass

# 尝试导入 PyMuPDF
try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    pass

# 确定使用哪个库（优先使用 PyMuPDF，因为它不需要外部依赖）
if PYMUPDF_AVAILABLE:
    PDF_AVAILABLE = True
    PDF_LIB = "pymupdf"
elif PDF2IMAGE_AVAILABLE:
    PDF_AVAILABLE = True
    PDF_LIB = "pdf2image"
else:
    PDF_AVAILABLE = False
    PDF_LIB = None
    print("⚠️ PDF处理库未安装，文档OCR将不支持PDF格式。可安装: pip install PyMuPDF 或 pip install pdf2image")

try:
    import torch
except Exception:
    torch = None

# 统一环境变量
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'

class AdvancedQwen3VLApp:
    """高级Qwen3-VL应用类"""

    def __init__(self):
        self.model = None
        self.processor = None
        # self.model_path = "D:\cibfortune\Cibfortune\cibfortune\models\qwen3-vl-2b-instruct"
        self.model_path = "/data/storage1/wulin/models/qwen3-vl-8b-instruct"
        self.is_loaded = False
        self.chat_history = []
        self.session_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.chat_messages = []
        self.last_image = None
        self.last_saved_image_path = None
        self.last_image_digest = None
        self.last_ocr_markdown = None
        self.last_ocr_html = None
        # 卡证OCR多模态RAG组件
        self.card_rag_store = None
        self.card_rag_ready = False
        self.card_rag_dir = "rag_cards"
        # API 卡证OCR（RAG + Qwen API）
        self.card_api = None
        # API 票据OCR（使用qwen-vl-max模型）
        self.bill_api = None
        # API 文档OCR（使用qwen3-vl-plus模型）
        self.doc_api = None
        # 文档OCR（使用 PaddleOCR API，不需要本地 PaddleOCR）
        self.last_ocr_text = None
        self.last_ocr_page_texts = []  # 每页的文本列表
        self.last_ocr_text_chunks = []  # 文本切片列表
        self.last_ocr_output_dir = None
        self.last_ocr_files = {}
        # 文档知识库
        self.doc_knowledge_base = None
        self.last_doc_id = None  # 最后添加的文档ID
        self.doc_kb_use_qwen3vl = True  # 使用qwen3-vl的文本编码器
        # 文档知识库将在模型加载后初始化，使用qwen3-vl的文本编码器
        # 字段模板文件
        self.field_templates_file = "card_field_templates.md"
        # 当前识别的卡证类型和字段
        self.current_card_type = None
        self.current_default_fields = []
        self.current_custom_fields = []
        self.current_field_template_html = None  # 存储HTML表格结构
        self.current_final_fields_html = None  # 存储最终字段列表的HTML（包含自定义字段）
        self.current_parsed_dict = None

    def _super_resolve_image_for_ocr(self, image):
        """
        使用 image.py 中的双三次插值函数对图像进行超分辨率处理，
        在提供给大模型前尽可能提升清晰度。
        兼容 PIL.Image 和 numpy.ndarray 输入，失败时回退为原图。
        """
        if image is None:
            return None

        try:
            # 统一转换为 PIL.Image
            if isinstance(image, Image.Image):
                pil_img = image
            else:
                try:
                    pil_img = Image.fromarray(np.array(image))
                except Exception:
                    # 无法转换时直接返回原始输入
                    return image

            # 将图像暂存为临时文件，复用 image.py 中的 paddleocr_super_resolution 逻辑
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
                tmp_in = f.name
                pil_img.save(tmp_in, format="PNG")

            tmp_out = tmp_in.replace(".png", "_sr.png")

            try:
                upscaled = paddleocr_super_resolution(
                    tmp_in,
                    output_path=tmp_out,
                )
            except Exception as e:
                print(f"⚠️ 超分辨率处理失败，使用原图: {e}")
                return pil_img
            finally:
                # 清理输入临时文件（输出文件可能供排查使用，先不强制删除）
                try:
                    if os.path.exists(tmp_in):
                        os.remove(tmp_in)
                except Exception:
                    pass

            if upscaled is None:
                # 处理失败时直接返回原图
                return pil_img

            # image.py 返回的是 OpenCV BGR 格式的 ndarray
            if isinstance(upscaled, np.ndarray):
                try:
                    upscaled_rgb = cv2.cvtColor(upscaled, cv2.COLOR_BGR2RGB)
                    return Image.fromarray(upscaled_rgb)
                except Exception as e:
                    print(f"⚠️ 转换超分图像为PIL失败，使用原图: {e}")
                    return pil_img

            # 如果未来 paddleocr_super_resolution 直接返回 PIL.Image，也能兼容
            if isinstance(upscaled, Image.Image):
                return upscaled

            # 其他未知类型，回退为原图
            return pil_img

        except Exception as e:
            print(f"⚠️ 超分辨率预处理异常，使用原图: {e}")
            return image

    def _ensure_card_rag_loaded(self):
        """懒加载卡证RAG图片库（若存在 rag_cards 目录），支持多种RAG实现方式。"""
        if self.card_rag_ready:
            return
        try:
            if not os.path.isdir(self.card_rag_dir):
                self.card_rag_ready = True  # 标记为已尝试，避免重复检查
                return
            
            # 优先尝试使用 multimodal_rag 模块
            try:
                from multimodal_rag import MultiModalDocumentLoader, MultiModalVectorStore
                loader = MultiModalDocumentLoader()
                docs = loader.load_images_from_folder(self.card_rag_dir)
                if not docs:
                    self.card_rag_ready = True
                    return
                store = MultiModalVectorStore(persist_directory="./multimodal_chroma_card")
                store.create_vector_store(docs)
                self.card_rag_store = store
                self.card_rag_ready = True
                print("✅ 使用multimodal_rag加载RAG图片库成功")
                return
            except Exception as e:
                print(f"⚠️ 使用multimodal_rag加载失败: {e}，尝试使用简化版RAG")
            
            # 如果multimodal_rag不可用，尝试使用SimpleRAGStore（从ocr_card_rag_api导入）
            try:
                from ocr_card_rag_api import SimpleRAGStore
                print("使用简化版RAG功能（基于卡面样式特征）...")
                store = SimpleRAGStore(use_style_features=True)
                store.load_images_from_folder(self.card_rag_dir)
                
                if not store.image_embeddings:
                    print("⚠️ RAG图片库为空")
                    self.card_rag_ready = True
                    return False
                
                self.card_rag_store = store
                self.card_rag_ready = True
                print(f"✅ 使用简化版RAG加载成功，共 {len(store.image_embeddings)} 张图片")
                return
            except Exception as e:
                print(f"⚠️ 使用简化版RAG加载失败: {e}")
            
            # 如果都失败了，标记为已尝试
            self.card_rag_ready = True
        except Exception as e:
            print(f"加载RAG图片库失败: {e}")
            self.card_rag_ready = True

    def _ensure_card_api_loaded(self):
        """懒加载卡证OCR API（RAG增强 + Qwen API 客户端）"""
        if self.card_api is not None:
            return
        try:
            api = CardOCRWithRAG(
                api_key=None,
                model="qwen3-vl-plus",
                rag_image_dir=self.card_rag_dir,
                persist_directory="./multimodal_chroma_card",
            )
            api.load_model()
            api.load_rag_library()
            self.card_api = api
        except Exception:
            self.card_api = None
        except Exception:
            # RAG 初始化失败时忽略，走纯模型路径
            self.card_rag_store = None
            self.card_rag_ready = True

    def _rag_search_card(self, image, top_k: int = 3):
        """
        对输入图片进行RAG检索，返回相似图片信息（与ocr_card_rag_api.py中的逻辑一致）
        
        Args:
            image: 输入图片（PIL Image）
            top_k: 返回最相似的k张图片
            
        Returns:
            相似图片列表，每个元素包含 {filename, similarity, metadata}
        """
        if not self.card_rag_store or not hasattr(self.card_rag_store, "image_embeddings"):
            return []
            
        try:
            # 生成查询图片的嵌入向量
            # 兼容两种实现：MultiModalVectorStore 使用 .embeddings.embed_image，SimpleRAGStore 直接使用 .embed_image
            if hasattr(self.card_rag_store, "embeddings") and hasattr(self.card_rag_store.embeddings, "embed_image"):
                # 使用 MultiModalVectorStore
                query_emb = self.card_rag_store.embeddings.embed_image(image)
            elif hasattr(self.card_rag_store, "embed_image"):
                # 使用 SimpleRAGStore
                query_emb = self.card_rag_store.embed_image(image)
            else:
                print("⚠️ RAG存储不支持embed_image方法")
                return []
            
            # 计算与图片库中所有图片的相似度
            similarities = []
            # 如果SimpleRAGStore有compute_similarity方法，使用它（支持样式相似度）
            use_compute_similarity = hasattr(self.card_rag_store, "compute_similarity")
            
            # 确保查询向量的维度
            query_dim = len(query_emb) if hasattr(query_emb, '__len__') else query_emb.shape[0] if hasattr(query_emb, 'shape') else 0
            
            for idx, emb in enumerate(self.card_rag_store.image_embeddings):
                try:
                    # 检查维度是否匹配
                    emb_dim = len(emb) if hasattr(emb, '__len__') else emb.shape[0] if hasattr(emb, 'shape') else 0
                    
                    if query_dim != emb_dim:
                        # 维度不匹配，跳过或使用默认相似度
                        print(f"⚠️ 特征维度不匹配: 查询向量={query_dim}, 图片库向量={emb_dim}，跳过该图片")
                        continue
                    
                    if use_compute_similarity:
                        # 使用样式相似度或CLIP相似度（根据SimpleRAGStore的配置）
                        similarity = self.card_rag_store.compute_similarity(query_emb, emb)
                    else:
                        # 使用余弦相似度（MultiModalVectorStore）
                        dot_product = np.dot(query_emb, emb)
                        norm_query = np.linalg.norm(query_emb)
                        norm_emb = np.linalg.norm(emb)
                        denom = norm_query * norm_emb + 1e-8
                        similarity = float(dot_product / denom) if denom > 0 else 0.0
                    similarities.append((similarity, idx))
                except Exception as e:
                    # 如果计算相似度时出错，跳过该图片
                    print(f"⚠️ 计算相似度失败（图片{idx}）: {str(e)}")
                    continue
            
            # 排序并取Top-K
            similarities.sort(key=lambda x: x[0], reverse=True)
            top_results = []
            
            for sim, idx in similarities[:top_k]:
                if idx < len(self.card_rag_store.image_metadatas):
                    meta = self.card_rag_store.image_metadatas[idx]
                    filename = meta.get("filename") or os.path.basename(meta.get("source", "")) or f"图片{idx+1}"
                    top_results.append({
                        "filename": filename,
                        "similarity": sim,
                        "metadata": meta
                    })
                    
            return top_results
            
        except Exception as e:
            print(f"⚠️ RAG检索失败: {str(e)}")
            return []

    def _build_enhanced_prompt_card(self, base_prompt: str, rag_results: list, custom_prompt: str = None):
        """
        构建增强后的提示词（包含RAG检索结果，与ocr_card_rag_api.py中的逻辑一致）
        
        Args:
            base_prompt: 基础提示词
            rag_results: RAG检索结果
            custom_prompt: 用户自定义提示词
            
        Returns:
            增强后的完整提示词
        """
        if custom_prompt:
            prompt = custom_prompt
        else:
            prompt = base_prompt
            
        # 如果有RAG检索结果，添加到提示词中
        if rag_results:
            rag_context = "\n基于图片库检索到的相似卡证：\n"
            for rank, result in enumerate(rag_results, 1):
                filename = result["filename"]
                similarity = result["similarity"]
                rag_context += f"- 卡面{rank}: {filename} | 相似度={similarity:.3f}\n"
            rag_context += "\n"
            filenames = [result["filename"].split(".")[0] for result in rag_results]
            banks = [filename.split("_")[0] for filename in filenames]
            prompt = rag_context + prompt
            prompt = prompt + (
                f"6. 如果是银行卡且字段列表包含'卡面类型'，则按照以下规则填充：\n"
                f"  - 基于图片库检索到的相似卡证结果{filenames}，填充\"卡面类型\"字段。字段值规则如下：\n"
                f"       -**禁止**自定义、生成、猜测或编造新的卡面类型值。\n"
                f"       -当出现任何不确定、模糊或不匹配情况时，\"卡面类型\"字段的值**必须且只能为\"其他\"**。\n"
                f"       -若识别出的\"发卡行\"字段的值存在与{banks}中银行名称相同的情况，"
                f"则\"卡面类型\"字段的值只能从{filenames}中**严格选择一个**。\n"
            )
            
        return prompt

    def _ensure_bill_api_loaded(self):
        """懒加载票据OCR API（使用qwen-vl-max模型）"""
        if self.bill_api is not None:
            return
        try:
            api = CardOCRWithRAG(
                api_key=None,
                # model="qwen3-vl-plus",
                model="qwen-vl-max", # 都试试
                rag_image_dir=None,  # 票据OCR不使用RAG
                persist_directory=None,
            )
            api.load_model()
            # 票据OCR不使用RAG，跳过RAG库加载
            self.bill_api = api
        except Exception:
            self.bill_api = None

    def _ensure_doc_api_loaded(self):
        """懒加载文档OCR API（使用qwen3-vl-plus模型）"""
        if self.doc_api is not None:
            return
        try:
            api = CardOCRWithRAG(
                api_key=None,
                model="qwen3-vl-plus",  # 文档OCR使用qwen3-vl-plus模型
                rag_image_dir=None,  # 文档OCR不使用RAG
                persist_directory=None,
            )
            api.load_model()
            # 文档OCR不使用RAG，跳过RAG库加载
            self.doc_api = api
        except Exception as e:
            print(f"⚠️ 文档OCR 加载失败: {e}")
            self.doc_api = None

    def _load_field_templates(self):
        """从card_field_templates目录下的md文件加载字段模板"""
        templates = {}
        html_templates = {}  # 存储HTML表格内容
        templates_dir = "card_field_templates"
        
        def parse_html_table(content):
            """解析HTML表格，提取字段名称，正确处理rowspan和子字段组合"""
            try:
                from bs4 import BeautifulSoup
            except ImportError:
                print("⚠️ 需要安装beautifulsoup4来解析HTML表格")
                return []
            
            fields = []
            try:
                soup = BeautifulSoup(content, 'html.parser')
                table = soup.find('table')
                if not table:
                    return []
                
                rows = table.find_all('tr')
                if not rows:
                    return []
                
                # 子字段列表（需要与父类别组合）
                sub_fields = ['名称', '统一社会信用代码/纳税人识别号', '全称', '账号', '开户银行', '开户行行号', '开户行名称', '出票人', '承兑人']
                
                # 用于跟踪每个列位置的活跃rowspan类别
                # 格式: {列位置: {'name': '类别名', 'remaining_rows': 剩余行数}}
                active_rowspans = {}
                
                # 遍历每一行
                for row_idx, row in enumerate(rows):
                    cells = row.find_all(['th', 'td'])
                    if not cells:
                        continue
                    
                    # 第一步：计算每个单元格的实际列位置（考虑colspan和rowspan）
                    current_col = 0
                    row_cells_info = []
                    
                    for cell in cells:
                        text = cell.get_text(strip=True)
                        colspan = int(cell.get('colspan', 1))
                        rowspan = int(cell.get('rowspan', 1))
                        
                        # 跳过被rowspan占用的列
                        while current_col in active_rowspans:
                            current_col += 1
                        
                        cell_info = {
                            'text': text,
                            'col': current_col,
                            'colspan': colspan,
                            'rowspan': rowspan
                        }
                        row_cells_info.append(cell_info)
                        
                        current_col += colspan
                    
                    # 第三步：设置新的rowspan类别（在同一行处理字段提取之前）
                    for cell_info in row_cells_info:
                        text = cell_info['text']
                        col = cell_info['col']
                        colspan = cell_info['colspan']
                        rowspan = cell_info['rowspan']
                        
                        # 如果有rowspan，记录活跃的类别
                        # 注意：即使文本在sub_fields中，如果有rowspan，也应该作为类别处理
                        if rowspan > 1 and text:
                            for c in range(col, col + colspan):
                                active_rowspans[c] = {
                                    'name': text,
                                    'remaining_rows': rowspan - 1
                                }
                    
                    # 第四步：处理当前行的字段提取
                    for cell_info in row_cells_info:
                        text = cell_info['text']
                        col = cell_info['col']
                        colspan = cell_info['colspan']
                        rowspan = cell_info['rowspan']
                        
                        if not text:
                            continue
                        
                        # 如果该单元格有rowspan，说明它是类别，已经在上面设置了active_rowspans，跳过
                        if rowspan > 1:
                            continue
                        
                        # 检查是否是子字段
                        if text in sub_fields:
                            # 查找该列位置的活跃rowspan类别
                            parent_category = None
                            # 检查当前列及其左侧列是否有活跃的rowspan
                            for check_col in range(col, -1, -1):
                                if check_col in active_rowspans:
                                    parent_category = active_rowspans[check_col]['name']
                                    break
                            
                            if parent_category:
                                # 组合字段名：父类别 + 子字段
                                full_field = f"{parent_category}_{text}"
                                if full_field not in fields:
                                    fields.append(full_field)
                            else:
                                # 没有父类别，作为独立字段（如单独的"出票人"、"承兑人"）
                                if text not in fields:
                                    fields.append(text)
                        else:
                            # 独立字段（如"出票日期"、"汇票到期日"、"票据状态"等）
                            if colspan > 1:
                                # 跨列字段，直接添加
                                if text not in fields:
                                    fields.append(text)
                            else:
                                # 单列字段，检查该列是否有活跃的rowspan（且不是当前单元格）
                                if col not in active_rowspans or active_rowspans[col]['name'] != text:
                                    if text not in fields:
                                        fields.append(text)
                    
                    # 第五步：更新rowspan剩余行数，移除已结束的（在字段提取之后）
                    for col in list(active_rowspans.keys()):
                        active_rowspans[col]['remaining_rows'] -= 1
                        if active_rowspans[col]['remaining_rows'] < 0:
                            del active_rowspans[col]
                
                # 去重并保持顺序
                fields = list(dict.fromkeys(fields))
                
                return fields
                
            except Exception as e:
                print(f"⚠️ 解析HTML表格失败: {e}")
                import traceback
                traceback.print_exc()
                return []
        
        def parse_markdown_table(content):
            """解析Markdown表格，提取字段名称列"""
            fields = []
            lines = content.split('\n')
            in_table = False
            header_found = False
            field_name_col_idx = None
            header_col_count = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # 检测表格开始（包含 | 的行）
                if '|' in line:
                    if not in_table:
                        in_table = True
                        header_found = False
                    
                    # 分割表格行（保留空字符串以保持列索引）
                    all_cells = [cell.strip() for cell in line.split('|')]
                    # 移除首尾的空字符串（Markdown表格格式：| col1 | col2 |）
                    cells = [c for c in all_cells[1:-1] if c.strip()] if len(all_cells) > 2 else [c.strip() for c in all_cells if c.strip()]
                    
                    # 处理表头
                    if not header_found and len(cells) > 0:
                        # 查找"字段名称"列的索引
                        for idx, cell in enumerate(cells):
                            if '字段名称' in cell or '字段名' in cell:
                                field_name_col_idx = idx
                                header_col_count = len(cells)
                                break
                        header_found = True
                        continue
                    
                    # 跳过分隔行（包含---的行）
                    if '---' in line or all(c in '-: ' for c in line):
                        continue
                    
                    # 提取字段名称
                    if field_name_col_idx is not None and len(cells) > 0:
                        field_name = None
                        # 如果列数与表头相同（3列），使用表头确定的列索引
                        if header_col_count and len(cells) == header_col_count:
                            if len(cells) > field_name_col_idx:
                                # 检查第一列是否是字段类别（如"出票信息"、"收款信息"等）
                                first_col = cells[0].strip() if len(cells) > 0 else ""
                                category_keywords = ['出票信息', '收款信息', '承兑信息', '承兑信息（续）', '保证信息', '保证信息（续）']
                                # 如果第一列是类别，则字段名称在第二列（索引1）
                                if first_col in category_keywords:
                                    if len(cells) > 1:
                                        field_name = cells[1].strip()
                                else:
                                    # 如果第一列不是类别，可能是字段名称在指定列
                                    field_name = cells[field_name_col_idx].strip()
                        # 如果列数不同（通常是2列），假设第一列是字段名称
                        elif len(cells) == 2:
                            field_name = cells[0].strip()
                        
                        # 添加字段名称（排除空值和类别名）
                        if field_name and field_name not in ['出票信息', '收款信息', '承兑信息', '承兑信息（续）', '保证信息', '保证信息（续）']:
                            if field_name not in fields:
                                fields.append(field_name)
                else:
                    # 如果不在表格中，尝试解析列表格式
                    if line.startswith('- '):
                        field = line.replace('- ', '').strip()
                        if field and field not in fields:
                            fields.append(field)
            
            return fields
        
        try:
            if os.path.isdir(templates_dir):
                # 从目录中的md文件加载
                for filename in os.listdir(templates_dir):
                    if filename.endswith('.md'):
                        card_type = filename.replace('.md', '')
                        filepath = os.path.join(templates_dir, filename)
                        fields = []
                        try:
                            with open(filepath, 'r', encoding='utf-8') as f:
                                content = f.read()
                                # 检查是否是HTML表格格式
                                # 兼容带属性的<table ...>，采用更宽松的检测
                                is_html = '<table' in content.lower()
                                
                                # 保存HTML内容（如果是HTML格式）
                                if is_html:
                                    html_templates[card_type] = content
                                
                                # 先尝试解析HTML表格格式
                                if is_html:
                                    fields = parse_html_table(content)
                                else:
                                    # 尝试解析Markdown表格格式
                                    fields = parse_markdown_table(content)
                                
                                # 如果表格解析没有结果，再尝试列表格式
                                if not fields:
                                    for line in content.split('\n'):
                                        line = line.strip()
                                        if line.startswith('- '):
                                            field = line.replace('- ', '').strip()
                                            if field:
                                                fields.append(field)
                            
                            if fields:
                                # 确保第一个字段是"卡证类型"（如果还没有的话）
                                if not fields or fields[0] != "卡证类型":
                                    fields.insert(0, "卡证类型")
                                templates[card_type] = fields
                                print(f"✅ 成功加载 {card_type} 字段模板，共 {len(fields)} 个字段")
                            else:
                                print(f"⚠️ {card_type} 字段模板解析结果为空，将使用默认模板")
                        except Exception as e:
                            print(f"⚠️ 加载模板文件 {filename} 失败: {e}")
                            continue
            else:
                # 如果目录不存在，使用默认模板
                templates = {
                    "身份证": ["卡证类型", "姓名", "性别", "民族", "出生日期", "住址", "公民身份号码", "签发机关", "有效期限"],
                    "银行卡": ["卡证类型", "发卡行", "卡号", "有效期", "姓名", "卡面类型"],
                    "驾驶证": ["卡证类型", "姓名", "性别", "国籍", "住址", "出生日期", "初次领证日期", "准驾车型", "有效期限", "档案编号", "证号"],
                    "护照": ["卡证类型", "姓名", "性别", "出生日期", "出生地点", "护照号码", "签发日期", "有效期至", "签发机关"],
                    "工牌": ["卡证类型", "姓名", "工号", "部门", "职位", "公司名称", "有效期"],
                    "银行承兑汇票": ["卡证类型", "出票人名称", "出票人账号", "出票人开户行", "出票人保证人姓名", "票据金额（大写）", "票据金额（小写）", "收款人名称", "收款人账号", "收款人开户行", "保证人地址", "保证日期", "承兑人名称", "承兑人账号", "承兑人开户行行号", "承兑人开户行名称", "承兑人承诺", "本汇票已承兑，到期无条件付款", "承兑日期", "交易合同号", "能否转让", "保证人姓名", "信用等级", "审查意见"],
                    "其他": ["卡证类型", "姓名", "证件号码", "有效期"]
                }
        except Exception as e:
            print(f"⚠️ 加载字段模板失败: {e}")
            # 使用默认模板
            templates = {
                "身份证": ["卡证类型", "姓名", "性别", "民族", "出生日期", "住址", "公民身份号码", "签发机关", "有效期限"],
                "银行卡": ["卡证类型", "发卡行", "卡号", "有效期", "姓名", "卡面类型"],
                "驾驶证": ["卡证类型", "姓名", "性别", "国籍", "住址", "出生日期", "初次领证日期", "准驾车型", "有效期限", "档案编号", "证号"],
                "护照": ["卡证类型", "姓名", "性别", "出生日期", "出生地点", "护照号码", "签发日期", "有效期至", "签发机关"],
                "工牌": ["卡证类型", "姓名", "工号", "部门", "职位", "公司名称", "有效期"],
                "银行承兑汇票": ["卡证类型", "出票人名称", "出票人账号", "出票人开户行", "出票人保证人姓名", "票据金额（大写）", "票据金额（小写）", "收款人名称", "收款人账号", "收款人开户行", "保证人地址", "保证日期", "承兑人名称", "承兑人账号", "承兑人开户行行号", "承兑人开户行名称", "承兑人承诺", "本汇票已承兑，到期无条件付款", "承兑日期", "交易合同号", "能否转让", "保证人姓名", "信用等级", "审查意见"],
                "其他": ["卡证类型", "姓名", "证件号码", "有效期"]
            }
        # 将HTML模板存储到实例变量中
        self.field_template_htmls = html_templates
        return templates

    def detect_card_type(self, image):
        """第一步：识别卡证类型并加载默认字段模板"""
        if image is None:
            return None, [], "❌ 请先上传图片"
        
        try:
            # 在提供给大模型前先做一次超分辨率预处理
            image_sr = self._super_resolve_image_for_ocr(image)

            self._ensure_card_api_loaded()
            if self.card_api is None:
                return None, [], "❌ 卡证OCR 未初始化"
            
            # 使用简化的提示词只识别卡证类型（不包含银行承兑汇票）
            type_prompt = (
                "请识别这张图片中的卡证类型。\n"
                "只允许从以下类别中选择一种：身份证、银行卡、驾驶证、护照、工牌、其他。\n"
                "只输出卡证类型，不要输出其他内容。"
            )
            
            result = self.card_api.recognize_card(
                image_sr,
                custom_prompt=type_prompt,
                use_rag=False,
                max_tokens=50,
                temperature=0.1
            )
            
            if not result.get("success"):
                return None, [], None, f"❌ 识别失败: {result.get('error', '未知错误')}"
            
            # 从结果中提取卡证类型（不包含银行承兑汇票）
            result_text = result.get("result", "").strip()
            card_types = ["身份证", "银行卡", "驾驶证", "护照", "工牌", "其他"]
            detected_type = None
            
            for ct in card_types:
                if ct in result_text:
                    detected_type = ct
                    break
            
            if not detected_type:
                detected_type = "其他"
            
            # 加载对应的默认字段模板（卡证OCR不使用HTML模板）
            templates = self._load_field_templates()
            default_fields = templates.get(detected_type, templates.get("其他", []))
            
            # 卡证OCR不使用HTML模板，强制设置为None
            html_template = None
            
            # 保存当前状态
            self.current_card_type = detected_type
            self.current_default_fields = default_fields.copy()
            self.current_custom_fields = []
            self.current_field_template_html = None  # 卡证OCR不使用HTML模板
            
            return detected_type, default_fields, html_template, f"✅ 识别成功：{detected_type}"
            
        except Exception as e:
            return None, [], None, f"❌ 识别失败: {str(e)}"

    def detect_bill_type(self, image):
        """票据识别第一步：识别票据类型并加载默认字段模板（使用HTML模板）"""
        fixed_bill_type = ["银行承兑汇票", "商业承兑汇票", "转账支票", "现金支票", "普通支票", "本票", "付款回单", "收款回单", "代发业务回单", "电子发票（铁路电子客票）",]
        extendable_bill_type = ["代发业务清单", "单位活期明细对账单", "电子发票（增值税专用发票）", "电子发票（普通发票）", "中央非税收入统一票据" ]

        supported_bill_type = fixed_bill_type + extendable_bill_type
        if image is None:
            return None, [], None, "❌ 请先上传图片"
        
        try:
            # 在提供给大模型前先做一次超分辨率预处理
            image_sr = self._super_resolve_image_for_ocr(image)

            self._ensure_bill_api_loaded()
            if self.bill_api is None:
                return None, [], None, "❌ 票据OCR 未初始化"
            
            # 票据OCR只识别银行承兑汇票
            type_prompt = (
                "请识别这张图片中的票据类型。\n"
                f"只允许从以下类别中选择一种：{supported_bill_type}。\n"
                "转账支票类型必须有\"转账支票\"关键词，现金支票类型必须有\"现金支票\"关键词，其他支票为普通支票\n"
                "只输出票据类型，不要输出其他内容。"
            )
            
            result = self.bill_api.recognize_card(
                image_sr,
                custom_prompt=type_prompt,
                use_rag=False,
                max_tokens=50,
                temperature=0.1
            )
            
            if not result.get("success"):
                return None, [], None, f"❌ 识别失败: {result.get('error', '未知错误')}"
            
            # 从结果中提取票据类型
            result_text = result.get("result", "").strip()
            detected_type = None
            
            for bt in supported_bill_type:
                if bt in result_text:
                    detected_type = bt
                    break
            
            # No need to set default
            # if not detected_type:
            #     detected_type = "银行承兑汇票"  # 默认使用银行承兑汇票
            
            # 加载对应的默认字段模板（票据OCR使用HTML模板）
            templates = self._load_field_templates()
            #todo: add template of other bills
            default_fields = templates.get(detected_type, templates.get("其他票据", [])) 
            
            # 获取HTML表格内容（票据OCR必须使用HTML模板）
            html_template = getattr(self, 'field_template_htmls', {}).get(detected_type, None)
            
            # 保存当前状态
            self.current_card_type = detected_type
            self.current_default_fields = default_fields.copy()
            self.current_custom_fields = []
            self.current_field_template_html = html_template
            self.current_parsed_dict = None # 清除过去解析结果
            
            return detected_type, default_fields, html_template, f"✅ 识别成功：{detected_type}"
            
        except Exception as e:
            return None, [], None, f"❌ 识别失败: {str(e)}"

    def update_fields(self, card_type, default_fields, custom_fields_text):
        """第二步：合并默认字段和自定义字段"""
        try:
            # 解析自定义字段（每行一个字段）
            custom_fields = []
            if custom_fields_text:
                for line in custom_fields_text.strip().split('\n'):
                    field = line.strip()
                    if field and field not in default_fields:
                        custom_fields.append(field)
            
            # 合并字段
            all_fields = default_fields + custom_fields
            
            # 保存当前状态
            self.current_card_type = card_type
            self.current_default_fields = default_fields
            self.current_custom_fields = custom_fields
            
            return all_fields, f"✅ 字段已更新，共 {len(all_fields)} 个字段"
            
        except Exception as e:
            return [], f"❌ 更新字段失败: {str(e)}"

    def ocr_card_with_fields(self, image, fields_to_extract):
        """第三步：使用指定字段进行OCR识别"""
        if image is None:
            return "❌ 请先上传图片"
        
        if not fields_to_extract:
            return "❌ 请先设置要提取的字段"
        
        try:
            # 在提供给大模型前先做一次超分辨率预处理
            image_sr = self._super_resolve_image_for_ocr(image)

            self._ensure_card_api_loaded()
            if self.card_api is None:
                return "❌ 卡证OCR 未初始化"
            
            # 构建包含字段列表的提示词
            fields_list = "、".join(fields_to_extract)
            
            # 卡证OCR不使用HTML模板，只使用Markdown格式
            has_html_template = False
            
            if False:  # 卡证OCR不使用HTML模板
                # 如果有HTML模板，要求大模型返回填充后的HTML表格
                custom_prompt = (
                    f"你是专业的票据/卡证OCR引擎。请阅读并识别输入图片内容，并在下面提供的HTML表格模板中填充对应字段的值。\n"
                    f"\n"
                    f"【卡证类型】{self.current_card_type or '未知'}\n"
                    f"【字段列表（必须全部覆盖，缺失填写'无'）】{fields_list}\n"
                    f"\n"
                    f"【HTML表格模板】\n"
                    f"{html_template}\n"
                    f"\n"
                    f"要求：\n"
                    f"- 只返回填充后的HTML表格（保持原有结构、行列、合并单元格和样式/属性），不要返回任何其他说明文字。\n"
                    f"- 不新增或删除字段，不改变表头文案；未识别到的填写'无'。\n"
                    f"- 仅在需要填写值的单元格写入文本，避免修改字段名单元格。\n"
                    f"- 禁止输出任何猜测或编造的内容。\n"
                    f"- 禁止输出未在字段列表中的字段和字段值。\n"
                    f"- 不要使用代码块标记符号（例如 ``` ）。"
                )
            else:
                # 如果没有HTML模板，使用原来的Markdown表格格式
                custom_prompt = (
                    f"你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                    f"\n"
                    f"任务要求：\n"
                    f"1. 识别卡证类型：{self.current_card_type or '未知'}\n"
                    f"2. 提取以下字段（必须全部提取，如果图片中没有该字段则填写'无'）：{fields_list}，禁止提取该列表以外的字段和字段值\n"
                    f"3. 以Markdown表格形式输出，表格包含两列：字段名、字段值\n"
                    f"4. 不要使用代码块标记符号（例如 ``` ）\n"
                    f"5. 输出限制：\n"
                    f"   - 最终输出只包含Markdown表格。\n"
                    f"   - 禁止输出任何猜测或编造的内容。\n"
                    f"   - 禁止输出任何其他文字或解释性内容。\n"
                    f"   - 禁止输出未在字段列表中的字段和字段值。"
                )
            
            # 只有银行卡类型才使用RAG
            use_rag = (self.current_card_type == "银行卡")
            
            result = self.card_api.recognize_card(
                image_sr,
                custom_prompt=custom_prompt,
                use_rag=use_rag,
            )
            
            if not result.get("success"):
                return f"❌ OCR识别失败: {result.get('error', '未知错误')}"
            
            # 在终端输出RAG相似度匹配结果
            rag_info = result.get("rag_info")
            if rag_info and rag_info.get("enabled") and rag_info.get("results"):
                print("\n" + "=" * 60)
                print("📊 RAG相似度匹配结果")
                print("=" * 60)
                print(f"找到 {len(rag_info['results'])} 张相似图片：\n")
                for i, r in enumerate(rag_info["results"], 1):
                    filename = r.get("filename", "未知")
                    similarity = r.get("similarity", 0.0)
                    print(f"  {i}. {filename}")
                    print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
                print("=" * 60 + "\n")
            
            raw_result = (result.get("result") or "").strip()
            
            # 如果模型按要求直接返回HTML表格，则优先使用HTML（注入可编辑样式）
            if has_html_template and "<table" in raw_result.lower():
                try:
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(raw_result, 'html.parser')
                    table = soup.find('table')
                    if table:
                        # 添加样式使表格更美观且可编辑
                        table['class'] = (table.get('class', []) or []) + ['ocr-result-table']
                        # 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                        for tr in table.find_all('tr'):
                            if tr.get('height'):
                                del tr['height']
                            if tr.get('width'):
                                del tr['width']
                        for td in table.find_all('td'):
                            if td.get('height'):
                                del td['height']
                            if td.get('width'):
                                del td['width']
                        # 移除table的固定width属性
                        if table.get('width'):
                            del table['width']
                        if table.get('style'):
                            # 移除style中的width和height（使用Python的re模块）
                            import re
                            style = table.get('style', '')
                            style = re.sub(r'width\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                            style = re.sub(r'height\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                            style = style.strip()
                            if style:
                                table['style'] = style
                            else:
                                del table['style']
                        # 移除colgroup中的固定宽度设置
                        for colgroup in table.find_all('colgroup'):
                            for col in colgroup.find_all('col'):
                                if col.get('width'):
                                    del col['width']
                                if col.get('style'):
                                    import re
                                    style = col.get('style', '')
                                    style = re.sub(r'width\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                                    style = style.strip()
                                    if style:
                                        col['style'] = style
                                    else:
                                        del col['style']
                        # 获取所有字段名（用于识别哪些单元格是字段名，哪些是值）
                        field_names = set(fields_to_extract)
                        for td in table.find_all('td'):
                            cell_text = td.get_text(strip=True)
                            # 如果单元格文本不是字段名，且不是空，则设置为可编辑（这是值单元格）
                            if cell_text and cell_text not in field_names:
                                td['contenteditable'] = 'true'
                            # 如果单元格为空，也设置为可编辑（可能是待填充的值单元格）
                            elif not cell_text:
                                td['contenteditable'] = 'true'
                        
                        # 优化的表格样式：可调整大小的容器，表格随容器大小变化
                        styled_html = f"""
                        <style>
                        /* 可调整大小的表格容器 */
                        .ocr-result-table-container {{
                            position: relative;
                            display: inline-block;
                            min-width: 500px;
                            min-height: 300px;
                            max-width: 95vw;
                            max-height: 90vh;
                            width: 100%;
                            height: 600px;
                            resize: both;
                            overflow: auto;  /* 允许滚动，确保表格不超出容器 */
                            border: 2px solid #e0e0e0;
                            border-radius: 8px;
                            padding: 10px;
                            background-color: #f8f9fa;
                            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
                            margin: 20px 0;
                        }}
                        /* 调整大小手柄样式 */
                        .ocr-result-table-container::-webkit-resizer {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            border-radius: 0 0 8px 0;
                            width: 20px;
                            height: 20px;
                        }}
                        /* 调整大小提示 */
                        .ocr-result-table-container::before {{
                            content: '↘ 拖拽调整大小';
                            position: absolute;
                            top: 5px;
                            right: 5px;
                            font-size: 11px;
                            color: #667eea;
                            background: rgba(255, 255, 255, 0.9);
                            padding: 2px 6px;
                            border-radius: 4px;
                            pointer-events: none;
                            opacity: 0.7;
                            z-index: 5;
                            transition: opacity 0.3s ease;
                        }}
                        .ocr-result-table-container:hover::before {{
                            opacity: 1;
                        }}
                        /* 调整大小时的边框高亮 */
                        .ocr-result-table-container:active {{
                            border-color: #667eea;
                            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
                        }}
                        .ocr-result-table {{
                            width: auto;  /* 表格宽度根据内容自适应 */
                            min-width: 100%;  /* 最小宽度为容器宽度 */
                            max-width: 100%;  /* 最大宽度不超过容器 */
                            border-collapse: collapse;
                            margin: 0;
                            font-size: 14px;
                            table-layout: auto;  /* 使用auto，让列宽根据内容自动调整 */
                            box-shadow: none;
                            border-radius: 8px;
                            overflow: visible;  /* 允许内容溢出，不裁剪 */
                            background-color: #ffffff;
                        }}
                        .ocr-result-table th,
                        .ocr-result-table td {{
                            border: 1px solid #e0e0e0;
                            padding: 12px 16px;
                            text-align: left;
                            vertical-align: top;
                            word-break: break-word;
                            word-wrap: break-word;
                            transition: all 0.2s ease;
                            line-height: 1.6;
                            height: auto !important;  /* 行高根据内容自动调整，覆盖HTML中的固定height */
                            min-height: auto !important;
                            overflow: visible;  /* 允许内容显示，不裁剪 */
                            width: auto;  /* 列宽根据内容自动调整 */
                            max-width: none;  /* 不限制最大宽度 */
                        }}
                        /* 字段名列：根据内容自适应宽度 */
                        .ocr-result-table td:not([contenteditable="true"]) {{
                            background-color: #f8f9fa;
                            font-weight: 600;
                            color: #374151;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 120px;  /* 最小宽度 */
                            max-width: 300px;  /* 最大宽度限制，避免过宽 */
                            white-space: nowrap;  /* 字段名不换行 */
                            font-size: 14px;
                            border-right: 2px solid #d1d5db;
                            height: auto !important;
                            overflow: visible;
                        }}
                        /* 值列：根据内容自适应宽度 */
                        .ocr-result-table td[contenteditable="true"] {{
                            background-color: #ffffff;
                            cursor: text;
                            min-height: 20px;
                            height: auto !important;  /* 行高根据内容自动调整 */
                            position: relative;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 200px;  /* 最小宽度 */
                            max-width: none;  /* 不限制最大宽度，允许长文本 */
                            overflow: visible;  /* 允许内容显示 */
                            word-break: break-word;  /* 长文本自动换行 */
                        }}
                        /* 根据文本长度动态调整样式（保持列宽比例） */
                        .ocr-result-table td[contenteditable="true"][data-length="short"] {{
                            font-size: 15px;
                            padding: 10px 14px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="medium"] {{
                            font-size: 14px;
                            padding: 12px 16px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="long"] {{
                            font-size: 13px;
                            padding: 14px 18px;
                            line-height: 1.7;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="very-long"] {{
                            font-size: 12px;
                            padding: 16px 20px;
                            line-height: 1.8;
                            height: auto !important;
                        }}
                        .ocr-result-table th {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: #ffffff;
                            font-weight: 600;
                            font-size: 15px;
                            text-transform: uppercase;
                            letter-spacing: 0.5px;
                            border-color: #5568d3;
                        }}
                        .ocr-result-table tr:nth-child(even) {{
                            background-color: #f8f9fa;
                        }}
                        .ocr-result-table tr:nth-child(odd) {{
                            background-color: #ffffff;
                        }}
                        .ocr-result-table tr:hover {{
                            background-color: #f0f4ff;
                        }}
                        .ocr-result-table tr:hover td:not([contenteditable="true"]) {{
                            background-color: #e5e7eb;
                        }}
                        .ocr-result-table td[contenteditable="true"]:hover {{
                            background-color: #f8f9ff;
                            box-shadow: inset 0 0 0 1px #667eea;
                        }}
                        .ocr-result-table td[contenteditable="true"]:focus {{
                            outline: none;
                            background-color: #eef5ff;
                            box-shadow: inset 0 0 0 2px #667eea, 0 0 0 3px rgba(102, 126, 234, 0.1);
                            border-radius: 4px;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:before {{
                            content: "点击编辑...";
                            color: #999;
                            font-style: italic;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:focus:before {{
                            content: "";
                        }}
                        /* 优化长文本显示 */
                        .ocr-result-table td[contenteditable="true"] {{
                            overflow-wrap: break-word;
                            hyphens: auto;
                        }}
                        /* 响应式设计 */
                        @media (max-width: 768px) {{
                            .ocr-result-table-container {{
                                min-width: 300px;
                                min-height: 200px;
                            }}
                            .ocr-result-table {{
                                font-size: 12px;
                                table-layout: fixed;
                            }}
                            .ocr-result-table th,
                            .ocr-result-table td {{
                                padding: 8px 12px;
                            }}
                            .ocr-result-table td:not([contenteditable="true"]) {{
                                width: 30%;
                                font-size: 12px;
                            }}
                            .ocr-result-table td[contenteditable="true"] {{
                                width: 70%;
                            }}
                        }}
                        </style>
                        <script>
                        (function() {{
                            // 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                            function removeFixedHeights() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    // 移除table的width属性
                                    if (table.hasAttribute('width')) {{
                                        table.removeAttribute('width');
                                    }}
                                    if (table.style.width) {{
                                        table.style.width = '';
                                    }}
                                    
                                    // 移除tr的height和width属性
                                    var rows = table.querySelectorAll('tr');
                                    rows.forEach(function(row) {{
                                        if (row.hasAttribute('height')) {{
                                            row.removeAttribute('height');
                                        }}
                                        if (row.hasAttribute('width')) {{
                                            row.removeAttribute('width');
                                        }}
                                    }});
                                    
                                    // 移除td和th的height和width属性
                                    var cells = table.querySelectorAll('td, th');
                                    cells.forEach(function(cell) {{
                                        if (cell.hasAttribute('height')) {{
                                            cell.removeAttribute('height');
                                        }}
                                        if (cell.hasAttribute('width')) {{
                                            cell.removeAttribute('width');
                                        }}
                                        // 移除内联样式中的height和width
                                        if (cell.style.height) {{
                                            cell.style.height = '';
                                        }}
                                        if (cell.style.width) {{
                                            cell.style.width = '';
                                        }}
                                    }});
                                    
                                    // 移除colgroup中的width属性
                                    var colgroups = table.querySelectorAll('colgroup');
                                    colgroups.forEach(function(colgroup) {{
                                        var cols = colgroup.querySelectorAll('col');
                                        cols.forEach(function(col) {{
                                            if (col.hasAttribute('width')) {{
                                                col.removeAttribute('width');
                                            }}
                                            if (col.style.width) {{
                                                col.style.width = '';
                                            }}
                                        }});
                                    }});
                                }}
                            }}
                            
                            // 根据文本长度动态设置data-length属性
                            function updateCellLength() {{
                                var cells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                cells.forEach(function(cell) {{
                                    var text = cell.textContent || cell.innerText || '';
                                    var length = text.length;
                                    cell.removeAttribute('data-length');
                                    if (length > 0) {{
                                        if (length <= 20) {{
                                            cell.setAttribute('data-length', 'short');
                                        }} else if (length <= 50) {{
                                            cell.setAttribute('data-length', 'medium');
                                        }} else if (length <= 100) {{
                                            cell.setAttribute('data-length', 'long');
                                        }} else {{
                                            cell.setAttribute('data-length', 'very-long');
                                        }}
                                    }}
                                }});
                            }}
                            
                            // 页面加载后执行
                            setTimeout(function() {{
                                removeFixedHeights();
                                updateCellLength();
                            }}, 100);
                            
                            // 监听内容变化
                            var observer = new MutationObserver(function(mutations) {{
                                removeFixedHeights();
                                updateCellLength();
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true,
                                        attributes: true,
                                        attributeFilter: ['height', 'style']
                                    }});
                                }}
                            }}, 200);
                        }})();
                        </script>
                        <div class="ocr-result-table-container">
                            {str(table)}
                        </div>
                        """
                        self.last_ocr_html = styled_html
                        self.last_ocr_markdown = ""  # HTML模式下不生成Markdown
                        return styled_html
                    else:
                        # 如果解析失败，回退到Markdown处理
                        cleaned = self._sanitize_markdown(raw_result)
                        self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
                        self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                        return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
                except Exception as e:
                    print(f"⚠️ HTML表格解析失败，回退到Markdown格式: {e}")
                    # 解析失败，回退到Markdown处理
                    cleaned = self._sanitize_markdown(raw_result)
                    self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
                    self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                    return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
            else:
                # 否则按Markdown处理
                cleaned = self._sanitize_markdown(raw_result)
            self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
            
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def ocr_bill_with_fields(self, image, fields_to_extract):
        """票据OCR第三步：使用指定字段进行OCR识别（使用HTML模板）"""
        if image is None:
            return "❌ 请先上传图片"
        
        if not fields_to_extract:
            return "❌ 请先设置要提取的字段"
        
        try:
            # 在提供给大模型前先做一次超分辨率预处理
            image_sr = self._super_resolve_image_for_ocr(image)

            self._ensure_bill_api_loaded()
            if self.bill_api is None:
                return "❌ 票据OCR 未初始化"
            
            # 构建包含字段列表的提示词
            fields_list = "、".join(fields_to_extract)

            # 票据OCR使用HTML模板
            html_template = getattr(self, 'current_final_fields_html', None)
            if not html_template:
                html_template = getattr(self, 'current_field_template_html', None)
            has_html_template = html_template is not None and html_template.strip()
            
            if has_html_template:
                # 如果有HTML模板，要求大模型返回填充后的HTML表格
                # 将字段列表格式化为更清晰的格式，确保模型不会遗漏
                fields_list_formatted = "\n".join([f"  {i+1}. {field}" for i, field in enumerate(fields_to_extract)])
                
                custom_prompt = (
                    f"你是专业的票据OCR引擎。请仔细阅读并识别输入图片中的所有内容，并在下面提供的HTML表格模板中填充对应字段的值。\n"
                    f"\n"
                    f"【票据类型】{self.current_card_type or '未知'}\n"
                    f"\n"
                    f"【必须识别的字段列表（共{len(fields_to_extract)}个字段，必须全部识别，一个都不能遗漏）】\n"
                    f"{fields_list_formatted}\n"
                    f"\n"
                    f"【重要要求】\n"
                    f"- **必须识别上述所有{len(fields_to_extract)}个字段，一个都不能遗漏**\n"
                    f"- 如果图片中没有某个字段的值，该字段的值必须填写'无'，但不能跳过该字段\n"
                    f"- 请仔细检查图片中的每一个位置，确保所有字段都被识别和填充\n"
                    f"- 对于组合字段（如'出票人全称'、'出票人账号'等），需要分别识别每个子字段\n"
                    f"- 对于小写金额类型字段，如果为表格形式，需要**仔细核对每一位数字对应的单位**，**必须保证大小写数值相一致**"
                    f"- HTML表格中的小写金额字段**不允许直接用识别结果填充**，**必须用核对过的小写数字填充**，填充内容只能包括**数字和小数点**\n"
                    f"\n"
                    f"【HTML表格模板】\n"
                    f"{html_template}\n"
                    f"\n"
                    f"【输出要求】\n"
                    f"- 只返回填充后的HTML表格（保持原有结构、行列、合并单元格和样式/属性），不要返回任何其他说明文字\n"
                    #f"- 返回填充后的HTML表格（保持原有结构、行列、合并单元格和样式/属性），需要返回解释说明文字\n"
                    f"- 如果HTML表格中存在<repeat>标签，则该标签中包含的内容为表格中某一行的格式，该行可能重复若干次，需要根据识别结果准确判断重复行数并正确填充\n"
                    f"- 不新增或删除字段，不改变表头文案；未识别到的填写'无'\n"
                    f"- 仅在需要填写值的单元格写入文本，避免修改字段名单元格\n"
                    f"- 禁止输出任何猜测或编造的内容\n"
                    f"- 禁止输出未在字段列表中的字段和字段值\n"
                    f"- 不要使用代码块标记符号（例如 ``` ）\n"
                )

                # custom_prompt = (
                #     f"你是专业的票据OCR引擎。请仔细阅读并识别输入图片中的所有内容，并生成一个对应的json识别结果，需要包含解释信息。\n"
                # )
            else:
                # 如果没有HTML模板，使用Markdown表格格式（不应该发生，但作为兜底）
                custom_prompt = (
                    f"你是专业的票据OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                    f"\n"
                    f"任务要求：\n"
                    f"1. 识别票据类型：{self.current_card_type or '未知'}\n"
                    f"2. 提取以下字段（必须全部提取，如果图片中没有该字段则填写'无'）：{fields_list}，禁止提取该列表以外的字段和字段值\n"
                    f"3. 以Markdown表格形式输出，表格包含两列：字段名、字段值\n"
                    f"4. 不要使用代码块标记符号（例如 ``` ）\n"
                    f"5. 输出限制：\n"
                    f"   - 最终输出只包含Markdown表格。\n"
                    f"   - 禁止输出任何猜测或编造的内容。\n"
                    f"   - 禁止输出任何其他文字或解释性内容。\n"
                    f"   - 禁止输出未在字段列表中的字段和字段值。"
                )
            
            # 票据OCR不使用RAG
            use_rag = False
            
            # 票据OCR使用更大的max_tokens，确保能输出完整的HTML表格
            # 根据字段数量动态调整max_tokens（每个字段大约需要50-100 tokens）
            estimated_tokens = len(fields_to_extract) * 100 + 2000  # 基础2000 + 每个字段100
            max_tokens = max(2048, min(estimated_tokens, 8192))  # 最小2048，最大8192
            
            result = self.bill_api.recognize_card(
                image_sr,
                custom_prompt=custom_prompt,
                use_rag=use_rag,
                max_tokens=max_tokens,
                temperature=0.1,  # 降低温度，提高准确性
            )
            
            if not result.get("success"):
                return f"❌ OCR识别失败: {result.get('error', '未知错误')}"
            
            raw_result = (result.get("result") or "").strip()
            # 如果模型按要求直接返回HTML表格，则优先使用HTML（注入可编辑样式）
            if has_html_template and "<table" in raw_result.lower():
                try:
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(raw_result, 'html.parser')
                    table = soup.find('table')
                    if table:
                        # 添加样式使表格更美观且可编辑
                        table['class'] = (table.get('class', []) or []) + ['ocr-result-table']
                        # 获取所有字段名（用于识别哪些单元格是字段名，哪些是值）
                        field_names = set(fields_to_extract)
                        for td in table.find_all('td'):
                            cell_text = td.get_text(strip=True)
                            # 如果单元格文本不是字段名，且不是空，则设置为可编辑（这是值单元格）
                            if cell_text and cell_text not in field_names:
                                td['contenteditable'] = 'true'
                            # 如果单元格为空，也设置为可编辑（可能是待填充的值单元格）
                            elif not cell_text:
                                td['contenteditable'] = 'true'
                        
                        # 优化的表格样式：可调整大小的容器，表格随容器大小变化
                        # 添加JavaScript代码，监听编辑事件并更新隐藏的Textbox
                        styled_html = f"""
                        <style>
                        /* 可调整大小的表格容器 */
                        .ocr-result-table-container {{
                            position: relative;
                            display: inline-block;
                            min-width: 500px;
                            min-height: 300px;
                            max-width: 95vw;
                            max-height: 90vh;
                            width: 100%;
                            height: 600px;
                            resize: both;
                            overflow: auto;  /* 允许滚动，确保表格不超出容器 */
                            border: 2px solid #e0e0e0;
                            border-radius: 8px;
                            padding: 10px;
                            background-color: #f8f9fa;
                            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
                            margin: 20px 0;
                        }}
                        /* 调整大小手柄样式 */
                        .ocr-result-table-container::-webkit-resizer {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            border-radius: 0 0 8px 0;
                            width: 20px;
                            height: 20px;
                        }}
                        /* 调整大小提示 */
                        .ocr-result-table-container::before {{
                            content: '↘ 拖拽调整大小';
                            position: absolute;
                            top: 5px;
                            right: 5px;
                            font-size: 11px;
                            color: #667eea;
                            background: rgba(255, 255, 255, 0.9);
                            padding: 2px 6px;
                            border-radius: 4px;
                            pointer-events: none;
                            opacity: 0.7;
                            z-index: 5;
                            transition: opacity 0.3s ease;
                        }}
                        .ocr-result-table-container:hover::before {{
                            opacity: 1;
                        }}
                        /* 调整大小时的边框高亮 */
                        .ocr-result-table-container:active {{
                            border-color: #667eea;
                            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
                        }}
                        .ocr-result-table {{
                            width: auto;  /* 表格宽度根据内容自适应 */
                            min-width: 100%;  /* 最小宽度为容器宽度 */
                            max-width: 100%;  /* 最大宽度不超过容器 */
                            border-collapse: collapse;
                            margin: 0;
                            font-size: 14px;
                            table-layout: auto;  /* 使用auto，让列宽根据内容自动调整 */
                            box-shadow: none;
                            border-radius: 8px;
                            overflow: visible;  /* 允许内容溢出，不裁剪 */
                            background-color: #ffffff;
                        }}
                        .ocr-result-table th,
                        .ocr-result-table td {{
                            border: 1px solid #e0e0e0;
                            padding: 12px 16px;
                            text-align: left;
                            vertical-align: top;
                            word-break: break-word;
                            word-wrap: break-word;
                            transition: all 0.2s ease;
                            line-height: 1.6;
                            height: auto !important;  /* 行高根据内容自动调整，覆盖HTML中的固定height */
                            min-height: auto !important;
                            overflow: visible;  /* 允许内容显示，不裁剪 */
                            width: auto;  /* 列宽根据内容自动调整 */
                            max-width: none;  /* 不限制最大宽度 */
                        }}
                        /* 字段名列：根据内容自适应宽度 */
                        .ocr-result-table td:not([contenteditable="true"]) {{
                            background-color: #f8f9fa;
                            font-weight: 600;
                            color: #374151;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 120px;  /* 最小宽度 */
                            max-width: 300px;  /* 最大宽度限制，避免过宽 */
                            white-space: nowrap;  /* 字段名不换行 */
                            font-size: 14px;
                            border-right: 2px solid #d1d5db;
                            height: auto !important;
                            overflow: visible;
                        }}
                        /* 值列：根据内容自适应宽度 */
                        .ocr-result-table td[contenteditable="true"] {{
                            background-color: #ffffff;
                            cursor: text;
                            min-height: 20px;
                            height: auto !important;  /* 行高根据内容自动调整 */
                            position: relative;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 200px;  /* 最小宽度 */
                            max-width: none;  /* 不限制最大宽度，允许长文本 */
                            overflow: visible;  /* 允许内容显示 */
                            word-break: break-word;  /* 长文本自动换行 */
                        }}
                        /* 根据文本长度动态调整样式（保持列宽比例） */
                        .ocr-result-table td[contenteditable="true"][data-length="short"] {{
                            font-size: 15px;
                            padding: 10px 14px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="medium"] {{
                            font-size: 14px;
                            padding: 12px 16px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="long"] {{
                            font-size: 13px;
                            padding: 14px 18px;
                            line-height: 1.7;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="very-long"] {{
                            font-size: 12px;
                            padding: 16px 20px;
                            line-height: 1.8;
                            height: auto !important;
                        }}
                        .ocr-result-table th {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: #ffffff;
                            font-weight: 600;
                            font-size: 15px;
                            text-transform: uppercase;
                            letter-spacing: 0.5px;
                            border-color: #5568d3;
                        }}
                        .ocr-result-table tr:nth-child(even) {{
                            background-color: #f8f9fa;
                        }}
                        .ocr-result-table tr:nth-child(odd) {{
                            background-color: #ffffff;
                        }}
                        .ocr-result-table tr:hover {{
                            background-color: #f0f4ff;
                        }}
                        .ocr-result-table tr:hover td:not([contenteditable="true"]) {{
                            background-color: #e5e7eb;
                        }}
                        .ocr-result-table td[contenteditable="true"]:hover {{
                            background-color: #f8f9ff;
                            box-shadow: inset 0 0 0 1px #667eea;
                        }}
                        .ocr-result-table td[contenteditable="true"]:focus {{
                            outline: none;
                            background-color: #eef5ff;
                            box-shadow: inset 0 0 0 2px #667eea, 0 0 0 3px rgba(102, 126, 234, 0.1);
                            border-radius: 4px;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:before {{
                            content: "点击编辑...";
                            color: #999;
                            font-style: italic;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:focus:before {{
                            content: "";
                        }}
                        /* 优化长文本显示 */
                        .ocr-result-table td[contenteditable="true"] {{
                            overflow-wrap: break-word;
                            hyphens: auto;
                        }}
                        /* 响应式设计 */
                        @media (max-width: 768px) {{
                            .ocr-result-table-container {{
                                min-width: 300px;
                                min-height: 200px;
                            }}
                            .ocr-result-table {{
                                font-size: 12px;
                                table-layout: fixed;
                            }}
                            .ocr-result-table th,
                            .ocr-result-table td {{
                                padding: 8px 12px;
                            }}
                            .ocr-result-table td:not([contenteditable="true"]) {{
                                width: 30%;
                                font-size: 12px;
                            }}
                            .ocr-result-table td[contenteditable="true"] {{
                                width: 70%;
                            }}
                        }}
                        </style>
                        <script>
                        (function() {{
                            // 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                            function removeFixedHeights() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    // 移除table的width属性
                                    if (table.hasAttribute('width')) {{
                                        table.removeAttribute('width');
                                    }}
                                    if (table.style.width) {{
                                        table.style.width = '';
                                    }}
                                    
                                    // 移除tr的height和width属性
                                    var rows = table.querySelectorAll('tr');
                                    rows.forEach(function(row) {{
                                        if (row.hasAttribute('height')) {{
                                            row.removeAttribute('height');
                                        }}
                                        if (row.hasAttribute('width')) {{
                                            row.removeAttribute('width');
                                        }}
                                    }});
                                    
                                    // 移除td和th的height和width属性
                                    var cells = table.querySelectorAll('td, th');
                                    cells.forEach(function(cell) {{
                                        if (cell.hasAttribute('height')) {{
                                            cell.removeAttribute('height');
                                        }}
                                        if (cell.hasAttribute('width')) {{
                                            cell.removeAttribute('width');
                                        }}
                                        // 移除内联样式中的height和width
                                        if (cell.style.height) {{
                                            cell.style.height = '';
                                        }}
                                        if (cell.style.width) {{
                                            cell.style.width = '';
                                        }}
                                    }});
                                    
                                    // 移除colgroup中的width属性
                                    var colgroups = table.querySelectorAll('colgroup');
                                    colgroups.forEach(function(colgroup) {{
                                        var cols = colgroup.querySelectorAll('col');
                                        cols.forEach(function(col) {{
                                            if (col.hasAttribute('width')) {{
                                                col.removeAttribute('width');
                                            }}
                                            if (col.style.width) {{
                                                col.style.width = '';
                                            }}
                                        }});
                                    }});
                                }}
                            }}
                            
                            // 根据文本长度动态设置data-length属性
                            function updateCellLength() {{
                                var cells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                cells.forEach(function(cell) {{
                                    var text = cell.textContent || cell.innerText || '';
                                    var length = text.length;
                                    cell.removeAttribute('data-length');
                                    if (length > 0) {{
                                        if (length <= 20) {{
                                            cell.setAttribute('data-length', 'short');
                                        }} else if (length <= 50) {{
                                            cell.setAttribute('data-length', 'medium');
                                        }} else if (length <= 100) {{
                                            cell.setAttribute('data-length', 'long');
                                        }} else {{
                                            cell.setAttribute('data-length', 'very-long');
                                        }}
                                    }}
                                }});
                            }}
                            
                            // 页面加载后执行
                            setTimeout(function() {{
                                removeFixedHeights();
                                updateCellLength();
                            }}, 100);
                            
                            // 监听内容变化
                            var observer = new MutationObserver(function(mutations) {{
                                removeFixedHeights();
                                updateCellLength();
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true,
                                        attributes: true,
                                        attributeFilter: ['height', 'style']
                                    }});
                                }}
                            }}, 200);
                        }})();
                        </script>
                        <div class="ocr-result-table-container">
                            {str(table)}
                        </div>
                        <script>
                        (function() {{
                            var updateTimeout = null;
                            
                            function updateEditedContent() {{
                                // 清除之前的定时器
                                if (updateTimeout) {{
                                    clearTimeout(updateTimeout);
                                }}
                                
                                // 延迟更新，避免频繁触发
                                updateTimeout = setTimeout(function() {{
                                    var table = document.querySelector('.ocr-result-table');
                                    if (!table) return;
                                    
                                    // 获取完整的HTML（包括样式）
                                    var fullHtml = document.querySelector('#bill-ocr-result-html, [id*="bill-ocr-result-html"]');
                                    var htmlContent = '';
                                    
                                    if (fullHtml) {{
                                        // 获取包含表格的完整HTML
                                        var container = fullHtml.querySelector('.ocr-result-table') || fullHtml;
                                        htmlContent = container.innerHTML;
                                    }} else {{
                                        // 如果没有找到容器，直接获取表格的outerHTML
                                        htmlContent = table.outerHTML;
                                    }}
                                    
                                    // 查找隐藏的Textbox - 使用多种方法
                                    var hiddenInput = null;
                                    
                                    // 方法1: 直接通过ID查找
                                    hiddenInput = document.getElementById('bill-ocr-result-html-edited');
                                    
                                    // 方法2: 通过ID包含关键字查找
                                    if (!hiddenInput) {{
                                        var inputs = document.querySelectorAll('input, textarea');
                                        for (var i = 0; i < inputs.length; i++) {{
                                            if (inputs[i].id && inputs[i].id.includes('bill-ocr-result-html-edited')) {{
                                                hiddenInput = inputs[i];
                                                break;
                                            }}
                                        }}
                                    }}
                                    
                                    // 方法3: 通过name属性查找
                                    if (!hiddenInput) {{
                                        hiddenInput = document.querySelector('input[name*="bill-ocr-result-html-edited"], textarea[name*="bill-ocr-result-html-edited"]');
                                    }}
                                    
                                    // 方法4: 通过data属性或class查找
                                    if (!hiddenInput) {{
                                        var allInputs = document.querySelectorAll('input[type="text"], textarea');
                                        for (var i = 0; i < allInputs.length; i++) {{
                                            var input = allInputs[i];
                                            // 检查是否在Gradio的隐藏组件区域
                                            if (input.style.display === 'none' || input.hidden || input.offsetParent === null) {{
                                                // 尝试设置值，看是否能找到正确的输入框
                                                var testValue = input.value;
                                                input.value = 'TEST_' + Date.now();
                                                if (input.value === 'TEST_' + Date.now()) {{
                                                    input.value = testValue; // 恢复原值
                                                    // 这可能是我们要找的输入框，但需要更精确的匹配
                                                }}
                                            }}
                                        }}
                                    }}
                                    
                                    if (hiddenInput) {{
                                        // 获取完整的HTML内容（包括样式）
                                        var styleTag = document.querySelector('style');
                                        var styleContent = styleTag ? styleTag.outerHTML : '';
                                        var fullContent = styleContent + '\\n' + table.outerHTML;
                                        
                                        hiddenInput.value = fullContent;
                                        
                                        // 触发多种事件，确保Gradio捕获到变化
                                        var events = ['input', 'change', 'blur', 'keyup'];
                                        events.forEach(function(eventType) {{
                                            var event = new Event(eventType, {{ bubbles: true, cancelable: true }});
                                            hiddenInput.dispatchEvent(event);
                                        }});
                                        
                                        // 也尝试直接设置属性
                                        if (hiddenInput.setAttribute) {{
                                            hiddenInput.setAttribute('value', fullContent);
                                        }}
                                        
                                        console.log('[DEBUG] 已更新隐藏Textbox，内容长度:', fullContent.length);
                                    }} else {{
                                        console.warn('[DEBUG] 未找到隐藏的Textbox组件');
                                        // 如果找不到，尝试通过window对象存储
                                        if (window.gradioEditedContent === undefined) {{
                                            window.gradioEditedContent = {{}};
                                        }}
                                        window.gradioEditedContent['bill-ocr-result-html-edited'] = htmlContent;
                                    }}
                                }}, 300);
                            }}
                            
                            // 监听所有可编辑单元格的输入事件
                            function attachListeners() {{
                                var editableCells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                editableCells.forEach(function(cell) {{
                                    // 移除旧的监听器（如果存在）
                                    var newCell = cell.cloneNode(true);
                                    cell.parentNode.replaceChild(newCell, cell);
                                    
                                    // 添加新的监听器
                                    newCell.addEventListener('input', updateEditedContent);
                                    newCell.addEventListener('blur', updateEditedContent);
                                    newCell.addEventListener('keyup', updateEditedContent);
                                    newCell.addEventListener('paste', function() {{
                                        setTimeout(updateEditedContent, 100);
                                    }});
                                }});
                                
                                // 初始更新
                                updateEditedContent();
                            }}
                            
                            // 延迟执行，确保DOM已加载
                            setTimeout(attachListeners, 500);
                            
                            // 使用MutationObserver监听表格变化（动态添加的单元格）
                            var observer = new MutationObserver(function(mutations) {{
                                var shouldReattach = false;
                                mutations.forEach(function(mutation) {{
                                    if (mutation.type === 'childList' && mutation.addedNodes.length > 0) {{
                                        shouldReattach = true;
                                    }}
                                }});
                                if (shouldReattach) {{
                                    setTimeout(attachListeners, 100);
                                }}
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true
                                    }});
                                }}
                            }}, 500);
                            
                            // 页面卸载前保存
                            window.addEventListener('beforeunload', updateEditedContent);
                            
                            // 监听导出按钮点击事件，在导出前强制更新内容
                            function setupExportButton() {{
                                var exportBtn = document.getElementById('bill-ocr-export-btn') || 
                                               document.querySelector('button[id*="bill-ocr-export-btn"]') ||
                                               document.querySelector('button:contains("导出结果")');
                                
                                if (exportBtn) {{
                                    exportBtn.addEventListener('click', function(e) {{
                                        console.log('[DEBUG] 导出按钮被点击，强制更新内容...');
                                        // 立即更新内容，不延迟
                                        var table = document.querySelector('.ocr-result-table');
                                        if (table) {{
                                            var styleTag = document.querySelector('style');
                                            var styleContent = styleTag ? styleTag.outerHTML : '';
                                            // 获取编辑后的表格HTML（包含所有用户编辑的内容）
                                            var tableHtml = table.outerHTML;
                                            var fullContent = styleContent + '\\n' + tableHtml;
                                            
                                            console.log('[DEBUG] 获取到的表格HTML长度:', tableHtml.length);
                                            console.log('[DEBUG] 表格内容预览:', tableHtml.substring(0, 200));
                                            
                                            // 查找隐藏的Textbox - 使用多种方法
                                            var hiddenInput = null;
                                            
                                            // 方法1: 直接通过ID查找
                                            hiddenInput = document.getElementById('bill-ocr-result-html-edited');
                                            
                                            // 方法2: 通过ID包含关键字查找
                                            if (!hiddenInput) {{
                                                var inputs = document.querySelectorAll('input, textarea');
                                                for (var i = 0; i < inputs.length; i++) {{
                                                    if (inputs[i].id && inputs[i].id.includes('bill-ocr-result-html-edited')) {{
                                                        hiddenInput = inputs[i];
                                                        break;
                                                    }}
                                                }}
                                            }}
                                            
                                            // 方法3: 查找所有隐藏的输入框
                                            if (!hiddenInput) {{
                                                var allInputs = document.querySelectorAll('input[type="text"], textarea');
                                                for (var i = 0; i < allInputs.length; i++) {{
                                                    var input = allInputs[i];
                                                    // 检查是否是隐藏的组件
                                                    if ((input.style.display === 'none' || input.hidden || input.offsetParent === null) &&
                                                        input.id && input.id.includes('bill')) {{
                                                        hiddenInput = input;
                                                        break;
                                                    }}
                                                }}
                                            }}
                                            
                                            if (hiddenInput) {{
                                                console.log('[DEBUG] 找到隐藏Textbox，ID:', hiddenInput.id);
                                                hiddenInput.value = fullContent;
                                                
                                                // 触发所有可能的事件，确保Gradio捕获到变化
                                                var events = ['input', 'change', 'blur', 'keyup', 'focus'];
                                                events.forEach(function(eventType) {{
                                                    try {{
                                                        var event = new Event(eventType, {{ bubbles: true, cancelable: true }});
                                                        hiddenInput.dispatchEvent(event);
                                                    }} catch(err) {{
                                                        console.error('触发事件失败:', eventType, err);
                                                    }}
                                                }});
                                                
                                                // 也尝试直接设置属性
                                                if (hiddenInput.setAttribute) {{
                                                    hiddenInput.setAttribute('value', fullContent);
                                                }}
                                                
                                                console.log('[DEBUG] 导出前已强制更新，内容长度:', fullContent.length);
                                                console.log('[DEBUG] Textbox当前值长度:', hiddenInput.value.length);
                                            }} else {{
                                                console.error('[DEBUG] 导出前未找到隐藏Textbox，尝试所有输入框...');
                                                var allInputs = document.querySelectorAll('input, textarea');
                                                console.log('[DEBUG] 找到', allInputs.length, '个输入框');
                                                for (var i = 0; i < Math.min(allInputs.length, 10); i++) {{
                                                    console.log('  输入框', i, ':', allInputs[i].id, allInputs[i].name, allInputs[i].className);
                                                }}
                                            }}
                                        }} else {{
                                            console.error('[DEBUG] 未找到表格元素');
                                        }}
                                    }}, true); // 使用捕获阶段，确保先执行
                                }} else {{
                                    // 如果按钮还没加载，延迟重试
                                    setTimeout(setupExportButton, 500);
                                }}
                            }}
                            
                            // 延迟设置导出按钮监听器
                            setTimeout(setupExportButton, 1000);
                        }})();
                        </script>
                        """
                        self.last_ocr_html = styled_html
                        self.last_ocr_markdown = ""  # HTML模式下不生成Markdown
                        return styled_html
                    else:
                        # 如果解析失败，回退到Markdown处理
                        cleaned = self._sanitize_markdown(raw_result)
                        self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                        self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                        return f"🪪 票据OCR识别结果:\n\n{cleaned}"
                except Exception as e:
                    print(f"⚠️ HTML表格解析失败，回退到Markdown格式: {e}")
                    # 解析失败，回退到Markdown处理
                    cleaned = self._sanitize_markdown(raw_result)
                    self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                    self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                    return f"🪪 票据OCR识别结果:\n\n{cleaned}"
            else:
                # 否则按Markdown处理
                cleaned = self._sanitize_markdown(raw_result)
                self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                return f"🪪 票据OCR识别结果:\n\n{cleaned}"
            
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def get_dict_from_html(self, html_content):
        import ast
        if not self.current_parsed_dict:
            prompts = (
                f"根据所给的html表格，生成一个python字典，使用嵌套字典体现表格中的层次关系\n"
                f"【输出要求】\n"
                f"- 只返回填充后的字典，不要返回任何其他说明文字\n"
                f"- 不新增或删除字段，不改变表头文案；多值对应同一个键则使用列表\n"
                f"- 禁止输出任何猜测或编造的内容\n"
                f"- 不要使用代码块标记符号（例如 ``` ）\n"
                f"{html_content}\n"
            )
            
            try:
                self._ensure_card_api_loaded()
                if self.card_api is None:
                    return "❌ 卡证OCR API未初始化"
            except:
                print("导出失败")

            res = self.card_api.general_prompt(prompts)
            print(res.get("result"))
            self.current_parsed_dict = ast.literal_eval(res.get("result"))
        
        return self.current_parsed_dict


    def load_model(self, progress=gr.Progress()):
        """加载模型"""
        if self.is_loaded:
            return "✅ 模型已经加载完成！", gr.update(interactive=True)

        if torch is None:
            return "❌ 模型加载失败: 未检测到PyTorch，请先安装。", gr.update(interactive=False)

        try:
            progress(0.1, desc="检查模型路径...")
            if not os.path.exists(self.model_path):
                return f"❌ 模型路径不存在: {self.model_path}", gr.update(interactive=False)

            progress(0.3, desc="加载模型...")
            self.model = Qwen3VLForConditionalGeneration.from_pretrained(
                self.model_path,
                dtype="auto",
                device_map="cuda",
                load_in_4bit=False,
            )

            progress(0.7, desc="加载处理器...")
            self.processor = AutoProcessor.from_pretrained(self.model_path)
            print("加载处理器")

            # 初始化文档知识库，使用qwen3-vl的文本编码器
            if DOC_KB_AVAILABLE and self.doc_kb_use_qwen3vl:
                try:
                    progress(0.85, desc="初始化文档知识库（使用Qwen3-VL文本编码器）...")
                    self.doc_knowledge_base = DocumentKnowledgeBase(
                        persist_directory="./doc_knowledge_base",
                        embedding_model=None,  # 不使用独立embedding模型
                        qwen3vl_model=self.model,
                        qwen3vl_processor=self.processor
                    )
                    print("✅ 文档知识库初始化成功（使用Qwen3-VL文本编码器）")
                except Exception as e:
                    print(f"⚠️ 文档知识库初始化失败: {e}")
                    import traceback
                    traceback.print_exc()
                    # 如果使用qwen3-vl失败，回退到使用独立embedding模型
                    try:
                        progress(0.85, desc="回退：初始化文档知识库（使用独立embedding模型）...")
                        self.doc_knowledge_base = DocumentKnowledgeBase(
                            persist_directory="./doc_knowledge_base",
                            embedding_model="sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
                        )
                        self.doc_kb_use_qwen3vl = False
                        print("✅ 文档知识库初始化成功（使用独立embedding模型）")
                    except Exception as e2:
                        print(f"⚠️ 文档知识库回退初始化也失败: {e2}")
                        self.doc_knowledge_base = None

            progress(1.0, desc="完成！")
            self.is_loaded = True

            return "✅ 模型加载成功！可以开始使用了。", gr.update(interactive=True)

        except Exception as e:
            return f"❌ 模型加载失败: {str(e)}", gr.update(interactive=False)

    def _prepare_user_message(self, image, prompt):
        prompt_clean = (prompt or "").strip()
        resolved_image = image if image is not None else self.last_image
        if resolved_image is None:
            raise ValueError("❌ 请上传图像！")
        if not prompt_clean:
            raise ValueError("❌ 请输入问题！")
        if image is not None:
            self.last_image = image
        content = [
            {"type": "image", "image": resolved_image},
            {"type": "text", "text": prompt_clean},
        ]
        return prompt_clean, {"role": "user", "content": content}

    def _run_inference(self,
                       image,
                       prompt,
                       max_tokens,
                       temperature,
                       top_p,
                       top_k,
                       repetition_penalty,
                       prepared=None):
        if prepared is None:
            prompt_clean, user_message = self._prepare_user_message(image, prompt)
        else:
            prompt_clean, user_message = prepared
        messages = self.chat_messages + [user_message]

        inputs = self.processor.apply_chat_template(
            messages,
            tokenize=True,
            add_generation_prompt=True,
            return_dict=True,
            return_tensors="pt"
        ).to(self.model.device)

        generation_kwargs = {
            "max_new_tokens": max_tokens,
            "temperature": temperature,
            "top_p": top_p,
            "top_k": top_k,
            "do_sample": True if temperature > 0 else False,
            "repetition_penalty": repetition_penalty
        }

        start_time = time.time()
        with torch.no_grad():
            generated_ids = self.model.generate(**inputs, **generation_kwargs)
        generation_time = time.time() - start_time

        generated_ids_trimmed = [
            out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text = self.processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )
        response = output_text[0]

        assistant_message = {"role": "assistant", "content": [{"type": "text", "text": response}]}
        self.chat_messages.extend([user_message, assistant_message])
        return prompt_clean, response, generation_time

    def _clone_history(self, history):
        return [[turn[0], turn[1]] for turn in history]

    def _chunk_response(self, text, chunk_size=80):
        if not text:
            return []
        return [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]

    def _parse_markdown_sections(self, markdown_text):
        """
        将 Markdown 文本拆分为 table/text 段，支持：
        - 管道表格（| a | b |）
        - HTML <table>（若存在）
        并在解析前对围栏代码块进行去围栏清洗，确保导出与渲染一致。
        """
        sections = []
        if not markdown_text:
            return sections

        # 1) 先去掉围栏，使得“代码块中的表格”也能被识别为可导出的内容
        cleaned_md = self._sanitize_markdown(markdown_text)

        # 2) 先尝试解析 HTML 表格（若模型输出了 <table>）
        html_tables = []
        try:
            from bs4 import BeautifulSoup  # 可选依赖
            soup = BeautifulSoup(cleaned_md, "html.parser")
            for t in soup.find_all("table"):
                headers = []
                header_row = t.find("tr")
                if header_row:
                    # 如果有 <th> 用 th；否则用首行的 td 作为 header
                    ths = header_row.find_all("th")
                    if ths:
                        headers = [th.get_text(strip=True) for th in ths]
                        data_rows = header_row.find_next_siblings("tr")
                    else:
                        tds = header_row.find_all("td")
                        headers = [td.get_text(strip=True) for td in tds]
                        data_rows = header_row.find_next_siblings("tr")
                rows = []
                for r in (data_rows or []):
                    cols = r.find_all(["td", "th"])
                    rows.append([c.get_text(strip=True) for c in cols])
                if headers or rows:
                    html_tables.append({"type": "table", "header": headers, "rows": rows})
        except Exception:
            # 如果 bs4 不在环境中，则略过 HTML 解析
            pass

        # 3) 解析管道表格
        lines = cleaned_md.splitlines()
        i = 0
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()

            # 管道表格判定：当前行和下一行构成 header + 分隔
            is_table = (
                stripped.startswith("|")
                and stripped.count("|") >= 2
                and i + 1 < len(lines)
                and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                and lines[i + 1].strip().startswith("|")
            )

            if is_table:
                header = [cell.strip() for cell in stripped.strip("|").split("|")]
                i += 2  # 跳过 header 与分隔线
                rows = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if not (row_line.startswith("|") and row_line.count("|") >= 2):
                        break
                    row = [cell.strip() for cell in row_line.strip("|").split("|")]
                    rows.append(row)
                    i += 1
                sections.append({"type": "table", "header": header, "rows": rows})
                continue

            # 普通文本块（直到遇到下一个表格或文件结束）
            text_block = []
            while i < len(lines):
                current = lines[i]
                stripped_current = current.strip()
                next_is_table = (
                    stripped_current.startswith("|")
                    and stripped_current.count("|") >= 2
                    and i + 1 < len(lines)
                    and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                    and lines[i + 1].strip().startswith("|")
                )
                if next_is_table:
                    break
                text_block.append(current)
                i += 1
                # 保留空行，改善段落分隔的可读性
                if i < len(lines) and lines[i] == "":
                    text_block.append(lines[i])

            text_content = "\n".join(text_block).strip("\n")
            if text_content:
                sections.append({"type": "text", "text": text_content})

        # 4) 若存在 HTML 表，优先把 HTML 表也加入（放在解析结果前面，避免遗漏）
        if html_tables:
            # 将 HTML 表插在最前面（也可根据需要合并/去重）
            sections = html_tables + sections

        return sections

    @staticmethod
    def _text_to_html_block(text: str) -> str:
        if not text:
            return ""
        escaped = html.escape(text)
        replaced = escaped.replace("\n", "<br>")
        return f'<div class="ocr-text">{replaced}</div>'

    @staticmethod
    def _table_to_html_block(header, rows) -> str:
        header = header or []
        rows = rows or []
        thead = ""
        if header:
            header_cells = "".join(f"<th>{html.escape(str(cell))}</th>" for cell in header)
            thead = f"<thead><tr>{header_cells}</tr></thead>"
        body_rows = []
        for row in rows:
            row_cells = "".join(f"<td>{html.escape(str(cell))}</td>" for cell in row)
            body_rows.append(f"<tr>{row_cells}</tr>")
        tbody = f"<tbody>{''.join(body_rows)}</tbody>" if body_rows else "<tbody></tbody>"
        return f'<table class="ocr-table">{thead}{tbody}</table>'

    def _render_sections_as_html(self, markdown_text: str) -> str:
        if not markdown_text:
            return ""
        sections = self._parse_markdown_sections(markdown_text)
        if not sections:
            escaped = html.escape(markdown_text.strip())
            return f"<pre>{escaped}</pre>" if escaped else ""
        blocks = []
        for section in sections:
            if section.get("type") == "table":
                blocks.append(self._table_to_html_block(section.get("header"), section.get("rows")))
            elif section.get("type") == "text":
                blocks.append(self._text_to_html_block(section.get("text", "")))
        return '<div class="ocr-preview">' + "".join(blocks) + "</div>"

    def chat_with_image(self, image, text, history, max_tokens, temperature, top_p, top_k, repetition_penalty: float = 1.0, presence_penalty: float = 1.5):
        """与图像对话（流式反馈），支持知识库检索增强"""
        original_text = text

        if not self.is_loaded:
            yield history, original_text, "❌ 请先加载模型！"
            return

        # 知识库检索增强
        enhanced_text = text
        kb_context = ""
        if self.doc_knowledge_base and text and text.strip():
            try:
                # 确定搜索范围：如果有最近添加的文档，优先搜索该文档
                doc_ids = [self.last_doc_id] if self.last_doc_id else None
                
                # 从知识库检索相关信息
                search_results = self.doc_knowledge_base.search(
                    query=text.strip(),
                    top_k=3,
                    doc_ids=doc_ids
                )
                
                if search_results:
                    # 构建知识库上下文
                    kb_context_parts = []
                    kb_context_parts.append("\n\n【相关文档内容】\n")
                    for i, result in enumerate(search_results, 1):
                        score = result.get('score', 0.0)
                        result_text = result.get('text', '')
                        metadata = result.get('metadata', {})
                        filename = metadata.get('filename', '未知文档')
                        
                        # 限制每个结果的文本长度，避免上下文过长
                        if len(result_text) > 500:
                            result_text = result_text[:500] + "..."
                        
                        kb_context_parts.append(f"{i}. [来源: {filename}, 相似度: {score:.2f}]\n{result_text}\n")
                    
                    kb_context = "\n".join(kb_context_parts)
                    kb_context += "\n【提示】请基于以上相关文档内容回答用户的问题。如果文档中没有相关信息，请基于图片内容回答。\n"
                    
                    # 将知识库上下文添加到用户问题中
                    enhanced_text = f"{text}{kb_context}"
                    
                    print(f"\n📚 知识库检索结果（共 {len(search_results)} 条）:")
                    for i, result in enumerate(search_results, 1):
                        print(f"  {i}. {result.get('metadata', {}).get('filename', '未知')} (相似度: {result.get('score', 0.0):.3f})")
            except Exception as e:
                print(f"⚠️ 知识库检索失败: {e}")
                # 检索失败时使用原始文本
                enhanced_text = text

        try:
            prepared = self._prepare_user_message(image, enhanced_text)
        except ValueError as exc:
            yield history, original_text, str(exc)
            return

        prompt_clean, _ = prepared
        history_copy = self._clone_history(history)
        # 显示原始问题，不显示知识库上下文（避免界面混乱）
        history_copy.append([f"👤 {original_text}", "🤖 正在思考..."])
        yield self._clone_history(history_copy), original_text, "🤖 正在思考..."

        try:
            _, response, generation_time = self._run_inference(
                image,
                enhanced_text,  # 使用增强后的文本（包含知识库上下文）
                max_tokens,
                temperature,
                top_p,
                top_k,
                repetition_penalty,
                prepared=prepared
            )
        except Exception as e:
            history_copy[-1][1] = f"❌ 生成失败: {str(e)}"
            self.chat_history = self._clone_history(history_copy)
            yield self._clone_history(history_copy), original_text, f"❌ 错误: {str(e)}"
            return

        assembled = ""
        chunks = self._chunk_response(response)
        if not chunks:
            chunks = [""]
        for chunk in chunks:
            assembled += chunk
            history_copy[-1][1] = f"🤖 {assembled}▌"
            yield self._clone_history(history_copy), original_text, f"🤖 {assembled}▌"

        stats = (
            f"⏱️ 生成时间: {generation_time:.2f}秒 | 📝 生成长度: {len(response)}字符"
            f" | ⚙️ 最大长度: {max_tokens}"
        )
        if max_tokens > 1024:
            stats += " | ⏳ 提示: 较大的最大长度可能延长生成时间"
        if kb_context:
            stats += " | 📚 已使用知识库增强"
        history_copy[-1][1] = f"🤖 {response}"
        self.chat_history = self._clone_history(history_copy)
        yield self._clone_history(history_copy), original_text, stats

    def _sanitize_markdown(self, text: str) -> str:
        if not text:
            return ""
        s = text.strip()
        lines = s.splitlines()
        out = []
        in_fence = False
        for line in lines:
            stripped = line.strip()
            if stripped.startswith("```"):
                in_fence = not in_fence
                continue
            if not in_fence:
                out.append(line)
        cleaned = "\n".join(out).strip()
        return cleaned if cleaned else s

    def ocr_analysis(self, image, prompt: str = None):
        """OCR文字识别，可选自定义提示词"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "请识别并提取这张图片中的所有文字内容，尽量还原原本样式，并标注语言类型。"
            " 请确保所有带样式或表格内容使用Markdown表格表示。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1024,
                temperature=0.7,
                top_p=0.8,
                top_k=50,
                repetition_penalty=1.0
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"📝 OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def ocr_card(self, image, prompt: str = None):
        """卡证OCR识别：身份证/银行卡/驾驶证等结构化提取（使用本地模型）"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        
        # 使用与ocr_card_api相同的默认提示词
        default_prompt = (
                "你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                "\n"
                "任务要求如下：\n"
                "\n"
            "1. 识别卡证类型：只允许从以下类别中选择一种：\n"
            "   - 身份证 / 银行卡 / 驾驶证 / 护照 / 工牌 / 其他。\n"
            "   Markdown表格中添加“卡证类型”字段，并用类别选择赋值。\n"
            "   **重要**：如果识别为银行卡，必须严格遵守第3条银行卡特殊要求！\n"
            "\n"
            "2. 输出格式：\n"
            "   - 以Markdown表格形式输出所有识别出的关键字段及其对应的值。\n"
            "   - 若字段中包含“卡号”，请确保该字段的值仅包含数字。\n"
            "   - 不要使用代码块标记符号（例如 ``` ）。\n"
            "\n"
            "3. 银行卡特殊要求（必须严格遵守）：\n"
            "   如果识别的卡证类型是银行卡，必须在Markdown表格的最后额外添加一个字段：\n"
            "   - 字段名：卡面类型（必须添加，不可省略）。\n"
            "   - 基于图片库检索到的相似卡证结果，填充“卡面类型”字段。字段值规则如下：\n"
            "       ① 当出现任何不确定、模糊或不匹配情况时，“卡面类型”字段的值**必须且只能为“其他”**，不得填写相似图片名或其他文本。\n"
            "       ② 若识别出的“发卡行”字段的值与这些相似卡证文件名中`_`前面的银行名称相同，"
            "则“卡面类型”字段的值只能从相似卡证文件名中**严格选择一个**，格式为`银行名称_卡面类型`，去掉文件后缀名，如`中国银行_visa卡`。\n"
            "       ③ 禁止自定义、生成、猜测或编造新的卡面类型值。任何不存在基于图片库检索到的相似卡证文件名的值都视为错误。\n"
            "   **重要提醒**：银行卡的Markdown表格必须包含“卡面类型”字段，这是强制要求，不能省略！\n"
            "   - 如果不是银行卡，则不添加“卡面类型”字段。\n"
            "\n"
            "4. 输出限制：\n"
            "   - 最终输出只包含Markdown表格。\n"
            "   - 禁止输出任何其他文字或解释性内容。\n"
            "   - 如果是银行卡，表格中必须包含“卡面类型”字段，否则输出不完整。\n"
        )

        effective_prompt = (prompt or "").strip() or default_prompt

        # RAG检索（使用与API版本相同的逻辑）
        rag_results = []
        try:
            self._ensure_card_rag_loaded()
            if self.card_rag_store and getattr(self.card_rag_store, "image_embeddings", None):
                rag_results = self._rag_search_card(image, top_k=3)
        except Exception as e:
            print(f"⚠️ RAG检索失败: {str(e)}")
            rag_results = []

        # 在终端输出RAG相似度匹配结果（与API版本一致）
        if rag_results:
            print("\n" + "=" * 60)
            print("📊 RAG相似度匹配结果")
            print("=" * 60)
            print(f"找到 {len(rag_results)} 张相似图片：\n")
            for i, r in enumerate(rag_results, 1):
                filename = r.get("filename", "未知")
                similarity = r.get("similarity", 0.0)
                print(f"  {i}. {filename}")
                print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
            print("=" * 60 + "\n")
        else:
            print("\n⚠️ 未找到相似图片\n")

        # 构建增强提示词（使用与API版本相同的逻辑）
        enhanced_prompt = self._build_enhanced_prompt_card(
            base_prompt=default_prompt,
            rag_results=rag_results,
            custom_prompt=effective_prompt if (prompt or "").strip() else None
        )

        # 在终端输出发送给模型的完整prompt（与API版本一致）
        print("\n" + "=" * 80)
        print("📝 发送给模型的完整Prompt")
        print("=" * 80)
        print(enhanced_prompt)
        print("=" * 80 + "\n")

        try:
            # 使用本地模型进行推理
            prompt_clean, response, _ = self._run_inference(
                image,
                enhanced_prompt,
                max_tokens=1024,
                temperature=0.3,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 卡证OCR识别失败: {str(e)}"

    def ocr_card_api(self, image, prompt: str = None):
        """卡证OCR识别（API调用 + RAG增强）"""
        # 注：如无需强制本地模型加载，可移除此判断
        try:
            self._ensure_card_api_loaded()
            if self.card_api is None:
                return "�?卡证OCR 初始化失败"
            default_prompt = (
                "你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                "\n"
                "任务要求如下：\n"
                "\n"
            "1. 识别卡证类型：只允许从以下类别中选择一种：\n"
            "   - 身份证 / 银行卡 / 驾驶证 / 护照 / 工牌 / 其他。\n"
            "   Markdown表格中添加“卡证类型”字段，并用类别选择赋值。\n"
            "   **重要**：如果识别为银行卡，必须严格遵守第3条银行卡特殊要求！\n"
            "\n"
            "2. 输出格式：\n"
            "   - 以Markdown表格形式输出所有识别出的关键字段及其对应的值。\n"
            "   - 若字段中包含“卡号”，请确保该字段的值仅包含数字。\n"
            "   - 不要使用代码块标记符号（例如 ``` ）。\n"
            "\n"
            "3. 银行卡特殊要求（必须严格遵守）：\n"
            "   如果识别的卡证类型是银行卡，必须在Markdown表格的最后额外添加一个字段：\n"
            "   - 字段名：卡面类型（必须添加，不可省略）。\n"
            "   - 基于图片库检索到的相似卡证结果，填充“卡面类型”字段。字段值规则如下：\n"
            "       ① 当出现任何不确定、模糊或不匹配情况时，“卡面类型”字段的值**必须且只能为“其他”**，不得填写相似图片名或其他文本。\n"
            "       ② 若识别出的“发卡行”字段的值与这些相似卡证文件名中`_`前面的银行名称相同，"
            "则“卡面类型”字段的值只能从相似卡证文件名中**严格选择一个**，格式为`银行名称_卡面类型`，去掉文件后缀名，如`中国银行_visa卡`。\n"
            "       ③ 禁止自定义、生成、猜测或编造新的卡面类型值。任何不存在基于图片库检索到的相似卡证文件名的值都视为错误。\n"
            "   **重要提醒**：银行卡的Markdown表格必须包含“卡面类型”字段，这是强制要求，不能省略！\n"
            "   - 如果不是银行卡，则不添加“卡面类型”字段。\n"
            "\n"
            "4. 输出限制：\n"
            "   - 最终输出只包含Markdown表格。\n"
            "   - 禁止输出任何其他文字或解释性内容。\n"
            "   - 如果是银行卡，表格中必须包含“卡面类型”字段，否则输出不完整。\n"
            )

            effective_prompt = (prompt or "").strip() or default_prompt
            result = self.card_api.recognize_card(
                image,
                custom_prompt=effective_prompt,
                use_rag=True,
            )
            if not result.get("success"):
                return f"�?卡证OCR 调用失败: {result.get('error') or '未知错误'}"

            # 在终端输出RAG相似度匹配结果
            rag_info = result.get("rag_info")
            if rag_info and rag_info.get("enabled") and rag_info.get("results"):
                print("\n" + "=" * 60)
                print("📊 RAG相似度匹配结果")
                print("=" * 60)
                print(f"找到 {len(rag_info['results'])} 张相似图片：\n")
                for i, r in enumerate(rag_info["results"], 1):
                    filename = r.get("filename", "未知")
                    similarity = r.get("similarity", 0.0)
                    print(f"  {i}. {filename}")
                    print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
                print("=" * 60 + "\n")
            elif rag_info and not rag_info.get("enabled"):
                print(f"\n⚠️ RAG未启用: {rag_info.get('reason', '未知原因')}\n")
            else:
                print("\n⚠️ 未找到相似图片\n")

            cleaned = self._sanitize_markdown(result.get("result") or "")
            self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
        except Exception as e:
            return f"�?卡证OCR识别失败: {str(e)}"

    def ocr_receipt(self, image, prompt: str = None):
        """票据OCR识别：发票/小票等表格与关键项解析"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "你是发票/小票OCR专家。请解析图片中的票据并输出：\n"
            "- 以Markdown表格给出关键信息：票据类型、开票日期、发票代码、发票号码、校验码、购买方、销售方、税号、项目、数量、单价、金额、税率、税额、合计金额(含税/不含税)；\n"
            "- 若检测到多行项目，请以表格形式逐行列出；\n"
            "- 表格下方给出识别置信度与可疑项提示；\n"
            "- 不要使用围栏代码块，保持Markdown可渲染。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1536,
                temperature=0.2,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🧾 票据OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 票据OCR识别失败: {str(e)}"

    def ocr_agreement(self, image, prompt: str = None):
        """协议OCR识别：合同/协议段落与条款解析"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "你是合同/协议OCR与条款解析助手。请完成：\n"
            "1) 识别全文，保持段落结构；\n"
            "2) 以Markdown表格提炼关键信息：合同名称、甲方、乙方、签署日期、生效日期、终止日期、金额/币种、违约条款、争议解决、签章情况；\n"
            "3) 如有编号的条款，保留编号并逐条列出；\n"
            "4) 在末尾给出“风险提示”列表（如空白处、涂改处、关键要素缺失等）；\n"
            "5) 不要输出围栏代码块。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=2048,
                temperature=0.3,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 协议OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>协议OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"📄 协议OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 协议OCR识别失败: {str(e)}"

    def _format_ocr_result(self, result):
        """
        格式化 OCR 结果为纯文本字符串（只提取 rec_texts 内容，不包含坐标等信息）
        
        Args:
            result: OCR 结果，可能是字符串、字典或其他类型
            
        Returns:
            格式化后的纯文本字符串（只包含识别到的文本内容）
        """
        import json
        if result is None:
            return ""
        
        if isinstance(result, str):
            # 如果是字符串，直接返回
            return result
        elif isinstance(result, dict):
            # 如果是字典，优先提取 rec_texts（PaddleOCR API 返回的文本字段）
            if "rec_texts" in result:
                rec_texts = result["rec_texts"]
                if isinstance(rec_texts, list):
                    # 如果是列表，合并所有文本行
                    return "\n".join(str(text) for text in rec_texts if text)
                else:
                    return str(rec_texts)
            # 其次查找其他常见的文本字段
            elif "text" in result:
                text = result["text"]
                if isinstance(text, list):
                    return "\n".join(str(t) for t in text if t)
                return str(text)
            elif "content" in result:
                content = result["content"]
                if isinstance(content, list):
                    return "\n".join(str(c) for c in content if c)
                return str(content)
            elif "prunedResult" in result:
                # 递归处理 prunedResult，提取其中的 rec_texts
                return self._format_ocr_result(result["prunedResult"])
            elif "result" in result:
                return self._format_ocr_result(result["result"])
            else:
                # 如果字典中没有明确的文本字段，返回空字符串（不返回坐标等元数据）
                return ""
        elif isinstance(result, list):
            # 如果是列表，提取每个元素的文本内容
            text_parts = []
            for item in result:
                text = self._format_ocr_result(item)
                if text:  # 只添加非空文本
                    text_parts.append(text)
            return "\n".join(text_parts)
        else:
            # 其他类型直接转换为字符串
            return str(result)

    def _chunk_text_for_rag(self, text, chunk_size=500, overlap=50, max_chunks=1000):
        """
        对文本进行切片，用于RAG相似度计算
        
        Args:
            text: 要切片的文本
            chunk_size: 每个切片的最大字符数
            overlap: 切片之间的重叠字符数
            max_chunks: 最大切片数量（防止内存溢出）
            
        Returns:
            文本切片列表
        """
        if not text:
            return []
        
        # 安全检查：确保 overlap < chunk_size，防止无限循环
        if overlap >= chunk_size:
            overlap = max(1, chunk_size // 10)  # 默认使用10%的重叠
        
        chunks = []
        start = 0
        text_length = len(text)
        prev_start = -1  # 用于检测是否卡住
        
        # 如果文本太大，先限制处理范围
        if text_length > 1000000:  # 如果超过1MB，只处理前1MB
            print(f"⚠️ 文本过大（{text_length}字符），仅处理前1MB用于切片")
            text = text[:1000000]
            text_length = len(text)
        
        while start < text_length and len(chunks) < max_chunks:
            end = min(start + chunk_size, text_length)
            chunk = text[start:end].strip()
            
            if chunk:
                chunks.append(chunk)
            
            # 计算下一个起始位置
            next_start = end - overlap
            # 安全检查：确保向前移动
            if next_start <= start:
                next_start = start + 1  # 至少移动1个字符
            
            # 防止卡在同一个位置
            if next_start == prev_start:
                next_start = start + chunk_size  # 强制移动
            
            prev_start = start
            start = next_start
        
        if len(chunks) >= max_chunks:
            print(f"⚠️ 切片数量达到上限（{max_chunks}），已停止切片")
        
        return chunks
    
    def _calculate_text_similarity(self, query, text):
        """
        计算查询文本与目标文本的相似度（使用简单的词重叠度）
        
        Args:
            query: 查询文本（关键字段）
            text: 目标文本（切片内容）
            
        Returns:
            相似度分数（0-1之间）
        """
        if not query or not text:
            return 0.0
        
        # 转换为小写并分词
        query_words = set(query.lower().split())
        text_words = set(text.lower().split())
        
        if not query_words:
            return 0.0
        
        # 计算交集和并集
        intersection = query_words & text_words
        union = query_words | text_words
        
        # Jaccard相似度
        if not union:
            return 0.0
        
        jaccard = len(intersection) / len(union)
        
        # 同时考虑查询词在文本中的出现频率
        text_lower = text.lower()
        query_lower = query.lower()
        if query_lower in text_lower:
            # 如果查询文本完全包含在目标文本中，增加相似度
            jaccard = max(jaccard, 0.5)
        
        return jaccard
    
    def _rag_search_text_chunks(self, key_fields, text_chunks, top_k=3):
        """
        根据关键字段对文本切片进行RAG相似度映射，返回相似度最高的top_k个切片
        
        Args:
            key_fields: 关键字段列表（用户自定义）
            text_chunks: 文本切片列表
            top_k: 返回最相似的切片数量
            
        Returns:
            每个关键字段对应的最相似切片列表
        """
        if not key_fields or not text_chunks:
            return {}
        
        field_chunks = {}
        
        for field in key_fields:
            if not field or not field.strip():
                continue
            
            field = field.strip()
            similarities = []
            
            # 计算每个切片与关键字段的相似度
            for idx, chunk in enumerate(text_chunks):
                similarity = self._calculate_text_similarity(field, chunk)
                similarities.append({
                    'chunk_index': idx,
                    'chunk_text': chunk,
                    'similarity': similarity
                })
            
            # 按相似度排序，取top_k
            similarities.sort(key=lambda x: x['similarity'], reverse=True)
            top_chunks = similarities[:top_k]
            
            field_chunks[field] = top_chunks
        
        return field_chunks

    def _pdf_to_images(self, pdf_path_or_bytes):
        """
        将PDF转换为图像列表
        
        Args:
            pdf_path_or_bytes: PDF文件路径或字节数据
            
        Returns:
            图像列表（PIL Image对象）
        """
        if not PDF_AVAILABLE:
            return None
        
        # 优先尝试 PyMuPDF（不需要外部依赖）
        if PYMUPDF_AVAILABLE:
            try:
                import fitz
                from PIL import Image as PILImage
                if isinstance(pdf_path_or_bytes, bytes):
                    pdf_doc = fitz.open(stream=pdf_path_or_bytes, filetype="pdf")
                else:
                    pdf_doc = fitz.open(pdf_path_or_bytes)
                
                images = []
                for page_num in range(len(pdf_doc)):
                    page = pdf_doc[page_num]
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x缩放提高清晰度
                    img = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    images.append(img)
                pdf_doc.close()
                return images
            except Exception as e:
                print(f"⚠️ PyMuPDF转换失败: {e}，尝试使用pdf2image...")
                # 如果 PyMuPDF 失败，尝试 pdf2image
        
        # 回退到 pdf2image（需要 poppler）
        if PDF2IMAGE_AVAILABLE:
            try:
                if isinstance(pdf_path_or_bytes, bytes):
                    images = convert_from_bytes(pdf_path_or_bytes)
                else:
                    images = convert_from_path(pdf_path_or_bytes)
                return images
            except Exception as e:
                error_msg = str(e)
                if "poppler" in error_msg.lower():
                    print(f"⚠️ PDF转换失败: {error_msg}")
                    print("💡 提示: pdf2image需要poppler，建议安装PyMuPDF: pip install PyMuPDF")
                else:
                    print(f"⚠️ PDF转换失败: {e}")
                return None
        
        return None
    

    PADDLEOCR_API_URL = "https://wdc9jbw9l1f8996b.aistudio-app.com/ocr"
    PADDLEOCR_TOKEN = "61236296494fb5e32ee89aef50d4d6aa99fa2ba7"

    def _parse_pdf_page_spec(self, pages_spec: str, total_pages: int):
        """
        解析用户输入的页码字符串，支持格式：
        - 逗号分隔页码：1,3,5
        - 区间：2-4
        - 混合输入与空格
        返回去重且排序后的页码列表（1-based）
        """
        if not pages_spec or not pages_spec.strip():
            return []

        normalized = (
            pages_spec.replace("，", ",")
            .replace("、", ",")
            .replace(";", ",")
            .strip()
        )
        if not normalized:
            return []

        tokens = re.split(r"[,\s]+", normalized)
        selected_pages = set()

        for token in tokens:
            if not token:
                continue
            token = token.strip()
            if not token:
                continue
            if "-" in token:
                parts = token.split("-", 1)
                if len(parts) != 2:
                    continue
                try:
                    start = int(parts[0])
                    end = int(parts[1])
                except ValueError:
                    continue
                if start > end:
                    start, end = end, start
                for page in range(start, end + 1):
                    if 1 <= page <= total_pages:
                        selected_pages.add(page)
            else:
                try:
                    page = int(token)
                except ValueError:
                    continue
                if 1 <= page <= total_pages:
                    selected_pages.add(page)

        return sorted(selected_pages)

    def _filter_pdf_pages(self, pdf_path: str, pages_spec: str):
        """
        根据页码过滤PDF页面，只返回包含指定页的新PDF路径。
        如果解析失败或依赖缺失则返回None。
        """
        if not os.path.exists(pdf_path):
            return None

        pages_spec = pages_spec or ""
        if pages_spec.strip().lower() == "all":
            return None

        # 优先使用 PyMuPDF，失败后再尝试 PyPDF2
        if PYMUPDF_AVAILABLE:
            try:
                import fitz

                src_doc = fitz.open(pdf_path)
                total_pages = src_doc.page_count
                selected_pages = self._parse_pdf_page_spec(pages_spec, total_pages)

                if not selected_pages:
                    print("⚠️ PDF页码解析为空，继续识别全部页面")
                    return None

                filtered_doc = fitz.open()
                for page_num in selected_pages:
                    filtered_doc.insert_pdf(
                        src_doc, from_page=page_num - 1, to_page=page_num - 1
                    )

                temp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
                temp_pdf_path = temp_pdf.name
                temp_pdf.close()
                filtered_doc.save(temp_pdf_path)
                filtered_doc.close()
                src_doc.close()
                print(f"✅ 使用PyMuPDF 过滤PDF页面: {selected_pages}")
                return temp_pdf_path
            except Exception as e:
                print(f"⚠️ 使用PyMuPDF过滤PDF失败: {e}")

        try:
            from PyPDF2 import PdfReader, PdfWriter
        except ImportError:
            print("⚠️ 未安装PyPDF2，无法按页裁剪PDF，继续识别全部页面")
            return None

        try:
            reader = PdfReader(pdf_path)
            total_pages = len(reader.pages)
            selected_pages = self._parse_pdf_page_spec(pages_spec, total_pages)

            if not selected_pages:
                print("⚠️ PDF页码解析为空，继续识别全部页面")
                return None

            writer = PdfWriter()
            for page_num in selected_pages:
                writer.add_page(reader.pages[page_num - 1])

            temp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            temp_pdf_path = temp_pdf.name
            temp_pdf.close()
            with open(temp_pdf_path, "wb") as output_pdf:
                writer.write(output_pdf)

            print(f"✅ 使用PyPDF2 过滤PDF页面: {selected_pages}")
            return temp_pdf_path
        except Exception as e:
            print(f"⚠️ 使用PyPDF2过滤PDF失败: {e}")
            return None


    def ocr_document(self, image_or_file, prompt: str = None, is_pdf: bool = False, pdf_pages: str = "all", engine: str = "paddle"):
        """
        文档OCR识别：使用 PaddleOCR API 直接调用
        提取到的文字保存下来准备进行切片做RAG知识库
        
        Args:
            image_or_file: 图像（PIL Image）或文件路径/字节数据
            prompt: 自定义提示词（暂不使用，保留兼容性）
            is_pdf: 是否为PDF文件
            pdf_pages: PDF页码，如"all"表示所有页，"1,3,5"表示指定页
            engine: OCR引擎，支持 "paddle" 或 "qwen3"
        """
        import os
        import base64
        import requests
        import json
        import time
        from pathlib import Path
        from PIL import Image
        import tempfile

        engine = (engine or "paddle").lower()
        if engine.startswith("qwen"):
            return self._ocr_document_qwen3(
                image_or_file=image_or_file,
                prompt=prompt,
                pdf_pages=pdf_pages
            )
        
        try:
            # 记录临时文件列表，便于统一清理
            temp_files = []

            # 确定文件路径或处理字节数据
            file_path = None
            
            if isinstance(image_or_file, Image.Image):
                # 如果是 PIL Image，需要先保存为临时文件
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    image_or_file.save(tmp.name, "PNG")
                    file_path = tmp.name
                    temp_files.append(tmp.name)
            elif isinstance(image_or_file, str):
                # 如果是文件路径
                file_path = image_or_file
            elif isinstance(image_or_file, bytes):
                # 如果是字节数据（PDF或图片的字节流）
                # 根据 is_pdf 参数确定文件扩展名
                suffix = ".pdf" if is_pdf else ".png"
                with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                    tmp.write(image_or_file)
                    file_path = tmp.name
                    temp_files.append(tmp.name)
            else:
                return "❌ 不支持的文件类型（需要 PIL Image、文件路径或字节数据）"
            
            if not file_path or not os.path.exists(file_path):
                return f"❌ 文件不存在: {file_path}"

            # 如果是PDF且指定了页码，则尝试裁剪PDF
            if (
                (is_pdf or file_path.lower().endswith(".pdf"))
                and pdf_pages
                and pdf_pages.strip()
                and pdf_pages.strip().lower() != "all"
            ):
                filtered_pdf = self._filter_pdf_pages(file_path, pdf_pages)
                if filtered_pdf:
                    file_path = filtered_pdf
                    temp_files.append(filtered_pdf)
                else:
                    print("⚠️ 未能按页码裁剪PDF，继续识别全部页面")
            
            # 创建输出目录（用于保存OCR结果，准备RAG切片）
            output_dir = "ocr_output"
            os.makedirs(output_dir, exist_ok=True)
            
            # 读取文件并编码
            print(f"正在读取文件: {file_path}")
            with open(file_path, "rb") as file:
                file_bytes = file.read()
            file_data = base64.b64encode(file_bytes).decode("ascii")
            
            file_size_mb = len(file_bytes) / (1024 * 1024)
            print(f"文件大小: {file_size_mb:.2f} MB")
            
            # 准备请求
            headers = {
                "Authorization": f"token {self.PADDLEOCR_TOKEN}",
                "Content-Type": "application/json"
            }
            
            # 根据文件类型设置 fileType
            if is_pdf or file_path.lower().endswith('.pdf'):
                file_type = 0  # PDF
            else:
                file_type = 1  # 图片
            
            payload = {
                "file": file_data,
                "fileType": file_type,
                "useDocOrientationClassify": False,
                "useDocUnwarping": False,
                "useTextlineOrientation": False,
            }
            
            # 发送请求
            print("正在进行PaddleOCR识别...")
            start_time = time.time()
            
            # 根据文件大小设置超时
            timeout = max(600, int(file_size_mb * 60))
            
            response = requests.post(self.PADDLEOCR_API_URL, json=payload, headers=headers, timeout=timeout)
            
            elapsed_time = time.time() - start_time
            print(f"响应时间: {elapsed_time:.2f}秒")
            
            if response.status_code != 200:
                return f"❌ 请求失败，状态码: {response.status_code}\n响应: {response.text[:500]}"
            
            result = response.json()
            
            if "result" not in result:
                return f"❌ 响应格式错误: {result}"
            
            ocr_result = result["result"]
            ocr_results = ocr_result.get("ocrResults", [])
            
            if not ocr_results:
                return "❌ 未识别到任何内容"
            
            # 获取输入文件名（不含扩展名）
            input_filename = os.path.splitext(os.path.basename(file_path))[0]
            
            # 收集所有页面的文本内容（用于RAG切片）
            all_text_parts = []
            all_markdown_parts = []
            
            # 保存文本结果（合并所有页面）
            txt_file = os.path.join(output_dir, f"{input_filename}_ocr.txt")
            md_file = os.path.join(output_dir, f"{input_filename}_ocr.md")
            json_file = os.path.join(output_dir, f"{input_filename}_ocr.json")
            
            with open(txt_file, "w", encoding="utf-8") as f:
                for i, res in enumerate(ocr_results):
                    pruned_result = res.get("prunedResult", "")
                    formatted_result = self._format_ocr_result(pruned_result)
                    all_text_parts.append(formatted_result)
                    
                    f.write(f"\n{'='*60}\n")
                    f.write(f"第 {i + 1} 页\n")
                    f.write(f"{'='*60}\n\n")
                    f.write(formatted_result)
                    f.write("\n\n")
            
            # 保存 Markdown 格式
            with open(md_file, "w", encoding="utf-8") as f:
                f.write(f"# {input_filename} OCR 结果\n\n")
                for i, res in enumerate(ocr_results):
                    pruned_result = res.get("prunedResult", "")
                    formatted_result = self._format_ocr_result(pruned_result)
                    all_markdown_parts.append(formatted_result)
                    
                    f.write(f"## 第 {i + 1} 页\n\n")
                    if "\n" in formatted_result:
                        f.write("```\n")
                        f.write(formatted_result)
                        f.write("\n```\n")
                    else:
                        f.write(formatted_result)
                    f.write("\n\n---\n\n")
            
            # 保存 JSON 结果
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(ocr_result, f, ensure_ascii=False, indent=2)
            
            # 下载并保存图片（如果需要）
            saved_images = 0
            for i, res in enumerate(ocr_results):
                image_url = res.get("ocrImage")
                if image_url:
                    try:
                        img_response = requests.get(image_url, timeout=30)
                        if img_response.status_code == 200:
                            img_filename = os.path.join(output_dir, f"{input_filename}_page_{i + 1}.jpg")
                            with open(img_filename, "wb") as f:
                                f.write(img_response.content)
                            saved_images += 1
                    except:
                        pass
            
            # 合并所有文本（用于RAG切片）
            full_text = "\n\n".join(all_text_parts)
            
            # 保存每页的rec_texts文本（用于分页显示和RAG处理）
            page_texts = []  # 每页的文本列表
            for i, res in enumerate(ocr_results):
                pruned_result = res.get("prunedResult", {})
                if isinstance(pruned_result, dict) and "rec_texts" in pruned_result:
                    rec_texts = pruned_result["rec_texts"]
                    if isinstance(rec_texts, list):
                        page_text = "\n".join(str(text) for text in rec_texts if text)
                    else:
                        page_text = str(rec_texts)
                else:
                    page_text = all_text_parts[i] if i < len(all_text_parts) else ""
                page_texts.append(page_text)
            
            # 先构建并返回结果（不等待文本切片）
            result_markdown = f"## 文档OCR识别结果\n\n"
            result_markdown += f"### 📄 文件信息\n\n"
            result_markdown += f"**文件名：** {input_filename}\n"
            result_markdown += f"**文件类型：** {'PDF' if file_type == 0 else '图片'}\n"
            result_markdown += f"**识别页数：** {len(ocr_results)}\n"
            result_markdown += f"**处理时间：** {elapsed_time:.2f}秒\n\n"
            
            result_markdown += f"### 📝 识别文本（第1页，共{len(page_texts)}页）\n\n"
            if page_texts:
                result_markdown += f"{page_texts[0][:1000]}{'...' if len(page_texts[0]) > 1000 else ''}\n\n"
            
            result_markdown += f"### 💾 保存的文件\n\n"
            result_markdown += f"- **文本文件：** `{txt_file}`\n"
            result_markdown += f"- **Markdown文件：** `{md_file}`\n"
            result_markdown += f"- **JSON文件：** `{json_file}`\n"
            if saved_images > 0:
                result_markdown += f"- **OCR图片：** {saved_images} 张\n"
            result_markdown += f"\n### 📚 知识库\n\n"
            kb_status = f"⏳ 正在对文本进行切片处理，完成后将自动保存到知识库。总字符数：{len(full_text)}\n"
            result_markdown += kb_status
            
            # 先保存基础数据到实例变量
            self.last_ocr_text = full_text
            self.last_ocr_page_texts = page_texts  # 保存每页文本
            self.last_ocr_markdown = result_markdown
            self.last_ocr_output_dir = output_dir
            self.last_ocr_files = {
                'txt': txt_file,
                'md': md_file,
                'json': json_file
            }
            
            self.last_ocr_html = "<h2>文档OCR识别结果</h2>" + self._render_sections_as_html(result_markdown)
            
            # 在后台进行文本切片（不阻塞返回）
            try:
                # 对于大文本，使用更大的切片大小以减少切片数量
                chunk_size = 1000 if len(full_text) > 100000 else 500
                text_chunks = self._chunk_text_for_rag(full_text, chunk_size=chunk_size, max_chunks=500)
                print(f"✅ 文本切片完成，共 {len(text_chunks)} 个切片")
                # 更新实例变量
                self.last_ocr_text_chunks = text_chunks
                
                # 保存到知识库
                if self.doc_knowledge_base and full_text:
                    try:
                        doc_id = self.doc_knowledge_base.add_document(
                            text=full_text,
                            filename=input_filename,
                            chunks=text_chunks,
                            metadata={
                                "file_type": "PDF" if file_type == 0 else "图片",
                                "page_count": len(ocr_results),
                                "ocr_time": elapsed_time
                            }
                        )
                        self.last_doc_id = doc_id
                        print(f"✅ 文档已保存到知识库，ID: {doc_id}")
                    except Exception as e:
                        print(f"⚠️ 保存到知识库失败: {e}")
            except MemoryError as e:
                print(f"⚠️ 内存不足，使用简化切片策略: {e}")
                # 如果内存不足，使用更大的切片，减少切片数量
                try:
                    text_chunks = self._chunk_text_for_rag(full_text, chunk_size=2000, overlap=100, max_chunks=200)
                    self.last_ocr_text_chunks = text_chunks
                    # 保存到知识库
                    if self.doc_knowledge_base and full_text:
                        try:
                            doc_id = self.doc_knowledge_base.add_document(
                                text=full_text,
                                filename=input_filename,
                                chunks=text_chunks
                            )
                            self.last_doc_id = doc_id
                            print(f"✅ 文档已保存到知识库，ID: {doc_id}")
                        except Exception as e:
                            print(f"⚠️ 保存到知识库失败: {e}")
                except:
                    # 如果还是失败，至少保存完整文本作为一个大切片
                    self.last_ocr_text_chunks = [full_text[:50000]] if len(full_text) > 50000 else [full_text]
            except Exception as e:
                print(f"⚠️ 文本切片失败: {e}")
                # 如果切片失败，至少保存完整文本作为一个大切片
                self.last_ocr_text_chunks = [full_text[:50000]] if len(full_text) > 50000 else [full_text]
            
            return result_markdown
            
        except requests.exceptions.Timeout:
            return f"❌ 请求超时（超过 {timeout} 秒）"
        except requests.exceptions.RequestException as e:
            return f"❌ 请求失败: {e}"
        except Exception as e:
            import traceback
            return f"❌ 处理过程中出现错误: {e}\n{traceback.format_exc()}"
        finally:
            for tmp_path in temp_files:
                try:
                    if tmp_path and os.path.exists(tmp_path):
                        os.unlink(tmp_path)
                except:
                    pass
    
    def _ocr_document_qwen3(self, image_or_file, prompt: str = None, pdf_pages: str = "all"):
        """使用Qwen3-VL进行文档OCR，并将结果切片后写入知识库。"""
        import os
        import time
        import json
        from pathlib import Path

        if not self.is_loaded:
            return "❌ 请先加载模型"

        start_time = time.time()
        default_prompt = (
            "请识别这份文档中的全部文字内容，保持原始阅读顺序。"
            "对于表格或列表，请使用Markdown表格或列表输出，不要使用代码块。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt

        output_dir = "ocr_output"
        os.makedirs(output_dir, exist_ok=True)

        input_filename = "document"
        is_pdf = False
        pages = []

        if isinstance(image_or_file, str):
            input_filename = Path(image_or_file).stem
            is_pdf = image_or_file.lower().endswith(".pdf")
            media, err = _load_media(None, image_or_file, need_all_pages=True)
            if err:
                return f"❌ 加载文件失败: {err}"
            pages = media if isinstance(media, list) else [media]
        elif isinstance(image_or_file, list):
            pages = image_or_file
        elif image_or_file is not None:
            pages = [image_or_file]

        if not pages:
            return "❌ 未获取到可识别的页面"

        def _filter_pages_for_spec(all_pages, spec: str):
            if not spec or str(spec).strip().lower() == "all":
                return all_pages
            indices = []
            for part in str(spec).split(","):
                part = part.strip()
                if part.isdigit():
                    idx = int(part) - 1
                    if 0 <= idx < len(all_pages):
                        indices.append(idx)
            return [all_pages[i] for i in indices] if indices else all_pages

        pages = _filter_pages_for_spec(pages, pdf_pages)

        page_texts = []
        for idx, page in enumerate(pages, 1):
            try:
                _, response, _ = self._run_inference(
                    page,
                    effective_prompt,
                    max_tokens=2048,
                    temperature=0.2,
                    top_p=0.8,
                    top_k=40,
                    repetition_penalty=1.0,
                )
                cleaned = self._sanitize_markdown(response)
            except Exception as e:
                cleaned = f"⚠️ 第{idx}页识别失败: {e}"
            page_texts.append(cleaned or "")

        full_text = "\n\n".join(page_texts)

        txt_file = os.path.join(output_dir, f"{input_filename}_ocr.txt")
        md_file = os.path.join(output_dir, f"{input_filename}_ocr.md")
        json_file = os.path.join(output_dir, f"{input_filename}_ocr.json")

        try:
            with open(txt_file, "w", encoding="utf-8") as f:
                for i, text in enumerate(page_texts, 1):
                    f.write(f"\n{'='*60}\n")
                    f.write(f"第{i}页\n")
                    f.write(f"{'='*60}\n\n")
                    f.write(text)
                    f.write("\n\n")
        except Exception as e:
            print(f"⚠️ 写入txt失败: {e}")

        try:
            with open(md_file, "w", encoding="utf-8") as f:
                f.write(f"# {input_filename} OCR 结果（Qwen3）\n\n")
                for i, text in enumerate(page_texts, 1):
                    f.write(f"## 第{i}页\n\n")
                    f.write(text)
                    f.write("\n\n---\n\n")
        except Exception as e:
            print(f"⚠️ 写入md失败: {e}")

        try:
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump({"engine": "qwen3", "pages": page_texts}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"⚠️ 写入json失败: {e}")

        elapsed_time = time.time() - start_time

        result_markdown = f"## 文档OCR识别结果（Qwen3）\n\n"
        result_markdown += f"**文件名：** {input_filename}\n"
        result_markdown += f"**文件类型：** {'PDF' if is_pdf else '图片'}\n"
        result_markdown += f"**识别页数：** {len(page_texts)}\n"
        result_markdown += f"**处理时间：** {elapsed_time:.2f}秒\n\n"

        result_markdown += f"### 📝 识别文本（第1页，共{len(page_texts)}页）\n\n"
        if page_texts:
            result_markdown += f"{page_texts[0][:1000]}{'...' if len(page_texts[0]) > 1000 else ''}\n\n"

        result_markdown += "### 💾 保存的文件\n\n"
        result_markdown += f"- **文本文件：** `{txt_file}`\n"
        result_markdown += f"- **Markdown文件：** `{md_file}`\n"
        result_markdown += f"- **JSON文件：** `{json_file}`\n"
        result_markdown += "\n### 📚 知识库\n\n"
        result_markdown += f"⏳ 正在对文本进行切片处理，完成后将自动保存到知识库。总字符数：{len(full_text)}\n"

        self.last_ocr_text = full_text
        self.last_ocr_page_texts = page_texts
        self.last_ocr_markdown = result_markdown
        self.last_ocr_output_dir = output_dir
        self.last_ocr_files = {
            "txt": txt_file,
            "md": md_file,
            "json": json_file,
        }
        self.last_ocr_html = "<h2>文档OCR识别结果</h2>" + self._render_sections_as_html(result_markdown)

        try:
            chunk_size = 1000 if len(full_text) > 100000 else 500
            text_chunks = self._chunk_text_for_rag(full_text, chunk_size=chunk_size, max_chunks=500)
            self.last_ocr_text_chunks = text_chunks
            if self.doc_knowledge_base and full_text:
                try:
                    doc_id = self.doc_knowledge_base.add_document(
                        text=full_text,
                        filename=input_filename,
                        chunks=text_chunks,
                        metadata={
                            "file_type": "PDF" if is_pdf else "图片",
                            "page_count": len(page_texts),
                            "ocr_time": elapsed_time,
                            "engine": "qwen3",
                        },
                    )
                    self.last_doc_id = doc_id
                    print(f"✅ 文档已保存到知识库，ID: {doc_id}")
                except Exception as e:
                    print(f"⚠️ 保存到知识库失败: {e}")
        except MemoryError as e:
            print(f"⚠️ 内存不足，使用简化切片策略: {e}")
            try:
                text_chunks = self._chunk_text_for_rag(full_text, chunk_size=2000, overlap=100, max_chunks=200)
                self.last_ocr_text_chunks = text_chunks
                if self.doc_knowledge_base and full_text:
                    try:
                        doc_id = self.doc_knowledge_base.add_document(
                            text=full_text,
                            filename=input_filename,
                            chunks=text_chunks,
                            metadata={
                                "file_type": "PDF" if is_pdf else "图片",
                                "page_count": len(page_texts),
                                "ocr_time": elapsed_time,
                                "engine": "qwen3",
                            },
                        )
                        self.last_doc_id = doc_id
                        print(f"✅ 文档已保存到知识库，ID: {doc_id}")
                    except Exception as e:
                        print(f"⚠️ 保存到知识库失败: {e}")
            except Exception:
                self.last_ocr_text_chunks = [full_text[:50000]] if len(full_text) > 50000 else [full_text]
        except Exception as e:
            print(f"⚠️ 文本切片失败: {e}")
            self.last_ocr_text_chunks = [full_text[:50000]] if len(full_text) > 50000 else [full_text]

        return result_markdown
    
    def extract_document_fields_with_rag(self, key_fields, custom_prompt=None):
        """
        识别关键字段：根据关键字段对文本切片进行相似度映射，
        将相似度最高的三段文本提供给大模型进行信息抽取
        
        Args:
            key_fields: 关键字段列表（用户自定义）
            custom_prompt: 自定义提示词
            
        Returns:
            信息抽取结果（表格形式，一个字段一行）
        """
        if not hasattr(self, 'last_ocr_text') or not self.last_ocr_text:
            return "❌ 请先进行文档OCR识别"
        
        if not key_fields or not isinstance(key_fields, list):
            return "❌ 请提供关键字段列表"
        
        # 过滤空字段
        key_fields = [f.strip() for f in key_fields if f and f.strip()]
        if not key_fields:
            return "❌ 关键字段列表为空"
        
        try:
            # 对每个关键字段进行RAG相似度映射（如果文本切片可用）
            rag_context_parts = []
            
            if hasattr(self, 'last_ocr_text_chunks') and self.last_ocr_text_chunks:
                print(f"🔍 开始RAG相似度匹配，共 {len(key_fields)} 个关键字段，{len(self.last_ocr_text_chunks)} 个文本切片")
                
                # 对每个关键字段进行RAG相似度匹配
                field_chunks = self._rag_search_text_chunks(
                    key_fields, 
                    self.last_ocr_text_chunks, 
                    top_k=3
                )
                
                # 为每个字段组织相关文本切片
                for field in key_fields:
                    if field in field_chunks:
                        chunks = field_chunks[field]
                        if chunks:
                            # 获取相似度最高的3个切片
                            top_chunks = chunks[:3]
                            field_texts = []
                            for i, chunk_info in enumerate(top_chunks, 1):
                                similarity = chunk_info['similarity']
                                chunk_text = chunk_info['chunk_text']
                                field_texts.append(f"[相似度: {similarity:.2%}] {chunk_text}")
                            
                            # 为每个字段组织相关文本
                            field_context = f"**关键字段「{field}」的相关文本片段（相似度最高的3段）：**\n"
                            field_context += "\n".join(field_texts)
                            rag_context_parts.append(field_context)
                            print(f"  ✅ {field}: 找到 {len(chunks)} 个相关片段")
                        else:
                            print(f"  ⚠️ {field}: 未找到相关片段")
                    else:
                        print(f"  ⚠️ {field}: 未找到相关片段")
                
                # 合并所有字段的相关文本
                if rag_context_parts:
                    rag_context = "\n\n".join(rag_context_parts)
                else:
                    # 如果没有找到相关片段，使用全文的前5000字符
                    rag_context = f"**文档全文（前5000字符）：**\n{self.last_ocr_text[:5000]}"
            else:
                # 如果没有切片，使用全文的前5000字符
                print("⚠️ 文本切片不可用，使用全文前5000字符")
                rag_context = f"**文档全文（前5000字符）：**\n{self.last_ocr_text[:5000]}"
            
            # 构建提示词（要求大模型提取所有字段）
            fields_list = "、".join([f"「{f}」" for f in key_fields])
            default_prompt = (
                f"你是专业的文档信息抽取专家。请从以下文档文本中提取以下关键字段的信息：{fields_list}\n\n"
                f"**文档相关内容：**\n{rag_context}\n\n"
                "**任务要求：**\n"
                "1. 仔细阅读上述文档内容，特别是每个关键字段的相关文本片段\n"
                "2. 从文档中提取每个关键字段的值\n"
                "3. 只输出Markdown表格，格式如下：\n"
                "| 字段名 | 字段值 |\n"
                "|--------|--------|\n"
                "4. 每个关键字段必须占一行，字段名和字段值都要填写\n"
                "5. 如果某个字段在文档中找不到，字段值填写\"未找到\"\n"
                "6. 只输出表格，不要输出其他说明文字、代码块标记或其他内容\n"
                "7. 确保提取所有字段，不要遗漏\n"
            )
            
            effective_prompt = custom_prompt if custom_prompt and custom_prompt.strip() else default_prompt
            
            # 使用qwen3-vl-plus API进行信息抽取
            self._ensure_doc_api_loaded()
            if self.doc_api is None or not self.doc_api.is_loaded:
                return "❌ 文档OCR未加载"
            
            # 尝试获取第一页图像（如果有保存的OCR图像）
            image_for_api = None
            if hasattr(self, 'last_ocr_files') and self.last_ocr_files:
                # 尝试加载第一页的OCR图像
                try:
                    from PIL import Image
                    # 查找第一页图像
                    output_dir = getattr(self, 'last_ocr_output_dir', 'ocr_output')
                    import glob
                    image_files = glob.glob(os.path.join(output_dir, '*_page_1.jpg'))
                    if image_files:
                        image_for_api = Image.open(image_files[0]).convert("RGB")
                except:
                    pass
            
            # 如果找不到图像，创建一个简单的占位图像或使用文本
            try:
                if image_for_api is None:
                    # 创建一个简单的文本图像占位符
                    from PIL import Image, ImageDraw, ImageFont
                    # 创建一个白色背景的图像
                    img = Image.new('RGB', (800, 600), color='white')
                    draw = ImageDraw.Draw(img)
                    # 在图像上写入提示文本
                    try:
                        text = "文档OCR识别结果\n请根据文本内容提取字段"
                        draw.text((50, 50), text, fill='black')
                    except:
                        pass
                    image_for_api = img
                
                # 调用API进行信息抽取
                api_result = self.doc_api.recognize_card(
                    image=image_for_api,
                    custom_prompt=effective_prompt,
                    max_tokens=2048,
                    temperature=0.2,
                    top_p=0.8,
                    use_rag=False
                )
                
                if api_result.get("success", False):
                    response = api_result.get("result", "")
                    print(f"✅ 大模型调用成功，响应长度: {len(response)}")
                    # 提取表格部分
                    table_html = self._extract_table_from_response(response)
                    print(f"✅ 表格提取完成")
                    return table_html
                else:
                    error_msg = api_result.get("error", "调用失败")
                    print(f"❌ 调用失败: {error_msg}")
                    return f"❌ 调用失败: {error_msg}"
            except Exception as api_error:
                # 如果API调用失败，返回格式化的提示词结果
                print(f"⚠️ 调用异常: {api_error}")
                import traceback
                print(traceback.format_exc())
                return f"❌ 调用失败: {str(api_error)}"
            
        except Exception as e:
            import traceback
            return f"❌ 信息抽取失败: {str(e)}\n{traceback.format_exc()}"
    
    def _extract_table_from_response(self, response):
        """
        从大模型响应中提取表格，格式化为HTML表格（一个字段一行）
        
        Args:
            response: 大模型的响应文本
            
        Returns:
            HTML格式的表格
        """
        import re
        import html
        
        # 清理响应文本，移除代码块标记
        response = response.strip()
        if response.startswith('```'):
            # 移除代码块标记
            lines = response.split('\n')
            start_idx = 0
            end_idx = len(lines)
            for i, line in enumerate(lines):
                if line.strip().startswith('```'):
                    if i == 0:
                        start_idx = 1
                    else:
                        end_idx = i
                        break
            response = '\n'.join(lines[start_idx:end_idx])
        
        # 尝试提取Markdown表格
        # 匹配完整的表格结构：表头行 + 分隔行 + 数据行
        all_lines = response.split('\n')
        table_lines = []
        in_table = False
        header_line_idx = -1
        separator_line_idx = -1
        
        # 查找表格开始位置
        for i, line in enumerate(all_lines):
            line_stripped = line.strip()
            # 检查是否是表格行（包含 | 符号）
            if '|' in line_stripped:
                # 检查是否是分隔行（只包含 -、:、| 和空格）
                if re.match(r'^[\s\|:\-]+$', line_stripped):
                    if header_line_idx >= 0 and separator_line_idx < 0:
                        # 找到分隔行
                        separator_line_idx = i
                        in_table = True
                        continue
                elif not in_table:
                    # 可能是表头行
                    header_line_idx = i
                    table_lines.append(line_stripped)
                elif in_table:
                    # 数据行
                    table_lines.append(line_stripped)
            elif in_table:
                # 表格结束
                break
        
        # 如果找到了表格行，解析它们
        if len(table_lines) > 1:
            html_table = '<div style="margin: 20px 0;">\n'
            html_table += '<table class="ocr-result-table" style="width: 100%; border-collapse: collapse; border: 1px solid #dee2e6; font-size: 14px;">\n'
            html_table += '<thead><tr style="background: #f8f9fa; font-weight: 600;">'
            html_table += '<th style="border: 1px solid #dee2e6; padding: 12px; text-align: left; width: 30%;">字段名</th>'
            html_table += '<th style="border: 1px solid #dee2e6; padding: 12px; text-align: left; width: 70%;">字段值</th>'
            html_table += '</tr></thead>\n<tbody>\n'
            
            # 解析表格行（跳过第一行表头，从数据行开始）
            for line in table_lines[1:]:
                # 跳过分隔行
                if re.match(r'^[\s\|:\-]+$', line):
                    continue
                
                # 分割单元格
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                
                # 如果单元格数量 >= 2，提取字段名和字段值
                if len(cells) >= 2:
                    # 第一个单元格是字段名，第二个是字段值
                    field_name = cells[0].strip()
                    field_value = cells[1].strip() if len(cells) > 1 else ""
                    
                    # 清理字段名和字段值（移除可能的标记符号）
                    field_name = re.sub(r'^[-\*\s]+', '', field_name).strip()
                    field_value = re.sub(r'^[-\*\s]+', '', field_value).strip()
                    
                    # 如果字段名不是"字段名"等表头关键词，才添加
                    if field_name and field_name not in ['字段名', '字段值', 'Field Name', 'Field Value']:
                        html_table += f'<tr>'
                        html_table += f'<td style="border: 1px solid #dee2e6; padding: 12px; font-weight: 500; background: #fafbfc;">{html.escape(field_name)}</td>'
                        html_table += f'<td style="border: 1px solid #dee2e6; padding: 12px; word-wrap: break-word;">{html.escape(field_value)}</td>'
                        html_table += '</tr>\n'
            
            html_table += '</tbody></table>\n</div>'
            return html_table
        
        # 如果没有找到标准表格，尝试提取字段名:字段值格式
        lines = response.split('\n')
        html_table = '<div style="margin: 20px 0;">\n'
        html_table += '<table class="ocr-result-table" style="width: 100%; border-collapse: collapse; border: 1px solid #dee2e6; font-size: 14px;">\n'
        html_table += '<thead><tr style="background: #f8f9fa; font-weight: 600;">'
        html_table += '<th style="border: 1px solid #dee2e6; padding: 12px; text-align: left; width: 30%;">字段名</th>'
        html_table += '<th style="border: 1px solid #dee2e6; padding: 12px; text-align: left; width: 70%;">字段值</th>'
        html_table += '</tr></thead>\n<tbody>\n'
        
        found_rows = False
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 跳过代码块标记和分隔行
            if line.startswith('```') or re.match(r'^[\s\|:\-]+$', line):
                continue
            
            # 尝试匹配字段名:字段值或字段名：字段值
            if ':' in line or '：' in line:
                parts = re.split('[:：]', line, 1)
                if len(parts) == 2:
                    field = parts[0].strip()
                    value = parts[1].strip()
                    # 清理字段名（移除可能的标记符号）
                    field = re.sub(r'^[-\*\s]+', '', field).strip()
                    value = re.sub(r'^[-\*\s]+', '', value).strip()
                    if field and value:
                        html_table += f'<tr>'
                        html_table += f'<td style="border: 1px solid #dee2e6; padding: 12px; font-weight: 500; background: #fafbfc;">{html.escape(field)}</td>'
                        html_table += f'<td style="border: 1px solid #dee2e6; padding: 12px; word-wrap: break-word;">{html.escape(value)}</td>'
                        html_table += '</tr>\n'
                        found_rows = True
        
        if not found_rows:
            # 如果都没有找到，返回原始响应的前500字符
            html_table += f'<tr><td colspan="2" style="border: 1px solid #dee2e6; padding: 12px; color: #dc3545;">⚠️ 无法解析响应格式，原始内容：<br/>{html.escape(response[:500])}</td></tr>\n'
        
        html_table += '</tbody></table>\n</div>'
        return html_table

    def _ensure_card_rag_loaded(self):
        """懒加载卡证RAG图片库（若存在 rag_cards 目录），支持多种RAG实现方式。"""
        if self.card_rag_ready:
            return
        try:
            if not os.path.isdir(self.card_rag_dir):
                self.card_rag_ready = True  # 标记为已尝试，避免重复检查
                return
            
            # 优先尝试使用 multimodal_rag 模块
            try:
                from multimodal_rag import MultiModalDocumentLoader, MultiModalVectorStore
                loader = MultiModalDocumentLoader()
                docs = loader.load_images_from_folder(self.card_rag_dir)
                if not docs:
                    self.card_rag_ready = True
                    return
                store = MultiModalVectorStore(persist_directory="./multimodal_chroma_card")
                store.create_vector_store(docs)
                self.card_rag_store = store
                self.card_rag_ready = True
                print("✅ 使用multimodal_rag加载RAG图片库成功")
                return
            except Exception as e:
                print(f"⚠️ 使用multimodal_rag加载失败: {e}，尝试使用简化版RAG")
            
            # 如果multimodal_rag不可用，尝试使用SimpleRAGStore（从ocr_card_rag_api导入）
            try:
                from ocr_card_rag_api import SimpleRAGStore
                print("使用简化版RAG功能（基于卡面样式特征）...")
                store = SimpleRAGStore(use_style_features=True)
                store.load_images_from_folder(self.card_rag_dir)
                
                if not store.image_embeddings:
                    print("⚠️ RAG图片库为空")
                    self.card_rag_ready = True
                    return False
                
                self.card_rag_store = store
                self.card_rag_ready = True
                print(f"✅ 使用简化版RAG加载成功，共 {len(store.image_embeddings)} 张图片")
                return
            except Exception as e:
                print(f"⚠️ 使用简化版RAG加载失败: {e}")
            
            # 如果都失败了，标记为已尝试
            self.card_rag_ready = True
        except Exception as e:
            print(f"加载RAG图片库失败: {e}")
            self.card_rag_ready = True

    def _ensure_card_api_loaded(self):
        """懒加载卡证OCR API（RAG增强 + Qwen API 客户端）"""
        if self.card_api is not None:
            return
        try:
            api = CardOCRWithRAG(
                api_key=None,
                model="qwen3-vl-plus",
                rag_image_dir=self.card_rag_dir,
                persist_directory="./multimodal_chroma_card",
            )
            api.load_model()
            api.load_rag_library()
            self.card_api = api
        except Exception:
            self.card_api = None
        except Exception:
            # RAG 初始化失败时忽略，走纯模型路径
            self.card_rag_store = None
            self.card_rag_ready = True

    def _rag_search_card(self, image, top_k: int = 3):
        """
        对输入图片进行RAG检索，返回相似图片信息（与ocr_card_rag_api.py中的逻辑一致）
        
        Args:
            image: 输入图片（PIL Image）
            top_k: 返回最相似的k张图片
            
        Returns:
            相似图片列表，每个元素包含 {filename, similarity, metadata}
        """
        if not self.card_rag_store or not hasattr(self.card_rag_store, "image_embeddings"):
            return []
            
        try:
            # 生成查询图片的嵌入向量
            # 兼容两种实现：MultiModalVectorStore 使用 .embeddings.embed_image，SimpleRAGStore 直接使用 .embed_image
            if hasattr(self.card_rag_store, "embeddings") and hasattr(self.card_rag_store.embeddings, "embed_image"):
                # 使用 MultiModalVectorStore
                query_emb = self.card_rag_store.embeddings.embed_image(image)
            elif hasattr(self.card_rag_store, "embed_image"):
                # 使用 SimpleRAGStore
                query_emb = self.card_rag_store.embed_image(image)
            else:
                print("⚠️ RAG存储不支持embed_image方法")
                return []
            
            # 计算与图片库中所有图片的相似度
            similarities = []
            # 如果SimpleRAGStore有compute_similarity方法，使用它（支持样式相似度）
            use_compute_similarity = hasattr(self.card_rag_store, "compute_similarity")
            
            # 确保查询向量的维度
            query_dim = len(query_emb) if hasattr(query_emb, '__len__') else query_emb.shape[0] if hasattr(query_emb, 'shape') else 0
            
            for idx, emb in enumerate(self.card_rag_store.image_embeddings):
                try:
                    # 检查维度是否匹配
                    emb_dim = len(emb) if hasattr(emb, '__len__') else emb.shape[0] if hasattr(emb, 'shape') else 0
                    
                    if query_dim != emb_dim:
                        # 维度不匹配，跳过或使用默认相似度
                        print(f"⚠️ 特征维度不匹配: 查询向量={query_dim}, 图片库向量={emb_dim}，跳过该图片")
                        continue
                    
                    if use_compute_similarity:
                        # 使用样式相似度或CLIP相似度（根据SimpleRAGStore的配置）
                        similarity = self.card_rag_store.compute_similarity(query_emb, emb)
                    else:
                        # 使用余弦相似度（MultiModalVectorStore）
                        dot_product = np.dot(query_emb, emb)
                        norm_query = np.linalg.norm(query_emb)
                        norm_emb = np.linalg.norm(emb)
                        denom = norm_query * norm_emb + 1e-8
                        similarity = float(dot_product / denom) if denom > 0 else 0.0
                    similarities.append((similarity, idx))
                except Exception as e:
                    # 如果计算相似度时出错，跳过该图片
                    print(f"⚠️ 计算相似度失败（图片{idx}）: {str(e)}")
                    continue
            
            # 排序并取Top-K
            similarities.sort(key=lambda x: x[0], reverse=True)
            top_results = []
            
            for sim, idx in similarities[:top_k]:
                if idx < len(self.card_rag_store.image_metadatas):
                    meta = self.card_rag_store.image_metadatas[idx]
                    filename = meta.get("filename") or os.path.basename(meta.get("source", "")) or f"图片{idx+1}"
                    top_results.append({
                        "filename": filename,
                        "similarity": sim,
                        "metadata": meta
                    })
                    
            return top_results
            
        except Exception as e:
            print(f"⚠️ RAG检索失败: {str(e)}")
            return []

    def _build_enhanced_prompt_card(self, base_prompt: str, rag_results: list, custom_prompt: str = None):
        """
        构建增强后的提示词（包含RAG检索结果，与ocr_card_rag_api.py中的逻辑一致）
        
        Args:
            base_prompt: 基础提示词
            rag_results: RAG检索结果
            custom_prompt: 用户自定义提示词
            
        Returns:
            增强后的完整提示词
        """
        if custom_prompt:
            prompt = custom_prompt
        else:
            prompt = base_prompt
            
        # 如果有RAG检索结果，添加到提示词中
        if rag_results:
            rag_context = "\n基于图片库检索到的相似卡证：\n"
            for rank, result in enumerate(rag_results, 1):
                filename = result["filename"]
                similarity = result["similarity"]
                rag_context += f"- 卡面{rank}: {filename} | 相似度={similarity:.3f}\n"
            rag_context += "\n"
            filenames = [result["filename"].split(".")[0] for result in rag_results]
            banks = [filename.split("_")[0] for filename in filenames]
            prompt = rag_context + prompt
            prompt = prompt + (
                f"6. 如果是银行卡且字段列表包含'卡面类型'，则按照以下规则填充：\n"
                f"  - 基于图片库检索到的相似卡证结果{filenames}，填充\"卡面类型\"字段。字段值规则如下：\n"
                f"       -**禁止**自定义、生成、猜测或编造新的卡面类型值。\n"
                f"       -当出现任何不确定、模糊或不匹配情况时，\"卡面类型\"字段的值**必须且只能为\"其他\"**。\n"
                f"       -若识别出的\"发卡行\"字段的值存在与{banks}中银行名称相同的情况，"
                f"则\"卡面类型\"字段的值只能从{filenames}中**严格选择一个**。\n"
            )
            
        return prompt

    def _ensure_bill_api_loaded(self):
        """懒加载票据OCR API（使用qwen-vl-max模型）"""
        if self.bill_api is not None:
            return
        try:
            api = CardOCRWithRAG(
                api_key=None,
                # model="qwen3-vl-plus",
                model="qwen-vl-max", # 都试试
                rag_image_dir=None,  # 票据OCR不使用RAG
                persist_directory=None,
            )
            api.load_model()
            # 票据OCR不使用RAG，跳过RAG库加载
            self.bill_api = api
        except Exception:
            self.bill_api = None

    def _ensure_doc_api_loaded(self):
        """懒加载文档OCR API（使用qwen3-vl-plus模型）"""
        if self.doc_api is not None:
            return
        try:
            api = CardOCRWithRAG(
                api_key=None,
                model="qwen3-vl-plus",  # 文档OCR使用qwen3-vl-plus模型
                rag_image_dir=None,  # 文档OCR不使用RAG
                persist_directory=None,
            )
            api.load_model()
            # 文档OCR不使用RAG，跳过RAG库加载
            self.doc_api = api
        except Exception as e:
            print(f"⚠️ 文档OCR 加载失败: {e}")
            self.doc_api = None

    def _load_field_templates(self):
        """从card_field_templates目录下的md文件加载字段模板"""
        templates = {}
        html_templates = {}  # 存储HTML表格内容
        templates_dir = "card_field_templates"
        
        def parse_html_table(content):
            """解析HTML表格，提取字段名称，正确处理rowspan和子字段组合"""
            try:
                from bs4 import BeautifulSoup
            except ImportError:
                print("⚠️ 需要安装beautifulsoup4来解析HTML表格")
                return []
            
            fields = []
            try:
                soup = BeautifulSoup(content, 'html.parser')
                table = soup.find('table')
                if not table:
                    return []
                
                rows = table.find_all('tr')
                if not rows:
                    return []
                
                # 子字段列表（需要与父类别组合）
                sub_fields = ['名称', '统一社会信用代码/纳税人识别号', '全称', '账号', '开户银行', '开户行行号', '开户行名称', '出票人', '承兑人']
                
                # 用于跟踪每个列位置的活跃rowspan类别
                # 格式: {列位置: {'name': '类别名', 'remaining_rows': 剩余行数}}
                active_rowspans = {}
                
                # 遍历每一行
                for row_idx, row in enumerate(rows):
                    cells = row.find_all(['th', 'td'])
                    if not cells:
                        continue
                    
                    # 第一步：计算每个单元格的实际列位置（考虑colspan和rowspan）
                    current_col = 0
                    row_cells_info = []
                    
                    for cell in cells:
                        text = cell.get_text(strip=True)
                        colspan = int(cell.get('colspan', 1))
                        rowspan = int(cell.get('rowspan', 1))
                        
                        # 跳过被rowspan占用的列
                        while current_col in active_rowspans:
                            current_col += 1
                        
                        cell_info = {
                            'text': text,
                            'col': current_col,
                            'colspan': colspan,
                            'rowspan': rowspan
                        }
                        row_cells_info.append(cell_info)
                        
                        current_col += colspan
                    
                    # 第三步：设置新的rowspan类别（在同一行处理字段提取之前）
                    for cell_info in row_cells_info:
                        text = cell_info['text']
                        col = cell_info['col']
                        colspan = cell_info['colspan']
                        rowspan = cell_info['rowspan']
                        
                        # 如果有rowspan，记录活跃的类别
                        # 注意：即使文本在sub_fields中，如果有rowspan，也应该作为类别处理
                        if rowspan > 1 and text:
                            for c in range(col, col + colspan):
                                active_rowspans[c] = {
                                    'name': text,
                                    'remaining_rows': rowspan - 1
                                }
                    
                    # 第四步：处理当前行的字段提取
                    for cell_info in row_cells_info:
                        text = cell_info['text']
                        col = cell_info['col']
                        colspan = cell_info['colspan']
                        rowspan = cell_info['rowspan']
                        
                        if not text:
                            continue
                        
                        # 如果该单元格有rowspan，说明它是类别，已经在上面设置了active_rowspans，跳过
                        if rowspan > 1:
                            continue
                        
                        # 检查是否是子字段
                        if text in sub_fields:
                            # 查找该列位置的活跃rowspan类别
                            parent_category = None
                            # 检查当前列及其左侧列是否有活跃的rowspan
                            for check_col in range(col, -1, -1):
                                if check_col in active_rowspans:
                                    parent_category = active_rowspans[check_col]['name']
                                    break
                            
                            if parent_category:
                                # 组合字段名：父类别 + 子字段
                                full_field = f"{parent_category}_{text}"
                                if full_field not in fields:
                                    fields.append(full_field)
                            else:
                                # 没有父类别，作为独立字段（如单独的"出票人"、"承兑人"）
                                if text not in fields:
                                    fields.append(text)
                        else:
                            # 独立字段（如"出票日期"、"汇票到期日"、"票据状态"等）
                            if colspan > 1:
                                # 跨列字段，直接添加
                                if text not in fields:
                                    fields.append(text)
                            else:
                                # 单列字段，检查该列是否有活跃的rowspan（且不是当前单元格）
                                if col not in active_rowspans or active_rowspans[col]['name'] != text:
                                    if text not in fields:
                                        fields.append(text)
                    
                    # 第五步：更新rowspan剩余行数，移除已结束的（在字段提取之后）
                    for col in list(active_rowspans.keys()):
                        active_rowspans[col]['remaining_rows'] -= 1
                        if active_rowspans[col]['remaining_rows'] < 0:
                            del active_rowspans[col]
                
                # 去重并保持顺序
                fields = list(dict.fromkeys(fields))
                
                return fields
                
            except Exception as e:
                print(f"⚠️ 解析HTML表格失败: {e}")
                import traceback
                traceback.print_exc()
                return []
        
        def parse_markdown_table(content):
            """解析Markdown表格，提取字段名称列"""
            fields = []
            lines = content.split('\n')
            in_table = False
            header_found = False
            field_name_col_idx = None
            header_col_count = None
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # 检测表格开始（包含 | 的行）
                if '|' in line:
                    if not in_table:
                        in_table = True
                        header_found = False
                    
                    # 分割表格行（保留空字符串以保持列索引）
                    all_cells = [cell.strip() for cell in line.split('|')]
                    # 移除首尾的空字符串（Markdown表格格式：| col1 | col2 |）
                    cells = [c for c in all_cells[1:-1] if c.strip()] if len(all_cells) > 2 else [c.strip() for c in all_cells if c.strip()]
                    
                    # 处理表头
                    if not header_found and len(cells) > 0:
                        # 查找"字段名称"列的索引
                        for idx, cell in enumerate(cells):
                            if '字段名称' in cell or '字段名' in cell:
                                field_name_col_idx = idx
                                header_col_count = len(cells)
                                break
                        header_found = True
                        continue
                    
                    # 跳过分隔行（包含---的行）
                    if '---' in line or all(c in '-: ' for c in line):
                        continue
                    
                    # 提取字段名称
                    if field_name_col_idx is not None and len(cells) > 0:
                        field_name = None
                        # 如果列数与表头相同（3列），使用表头确定的列索引
                        if header_col_count and len(cells) == header_col_count:
                            if len(cells) > field_name_col_idx:
                                # 检查第一列是否是字段类别（如"出票信息"、"收款信息"等）
                                first_col = cells[0].strip() if len(cells) > 0 else ""
                                category_keywords = ['出票信息', '收款信息', '承兑信息', '承兑信息（续）', '保证信息', '保证信息（续）']
                                # 如果第一列是类别，则字段名称在第二列（索引1）
                                if first_col in category_keywords:
                                    if len(cells) > 1:
                                        field_name = cells[1].strip()
                                else:
                                    # 如果第一列不是类别，可能是字段名称在指定列
                                    field_name = cells[field_name_col_idx].strip()
                        # 如果列数不同（通常是2列），假设第一列是字段名称
                        elif len(cells) == 2:
                            field_name = cells[0].strip()
                        
                        # 添加字段名称（排除空值和类别名）
                        if field_name and field_name not in ['出票信息', '收款信息', '承兑信息', '承兑信息（续）', '保证信息', '保证信息（续）']:
                            if field_name not in fields:
                                fields.append(field_name)
                else:
                    # 如果不在表格中，尝试解析列表格式
                    if line.startswith('- '):
                        field = line.replace('- ', '').strip()
                        if field and field not in fields:
                            fields.append(field)
            
            return fields
        
        try:
            if os.path.isdir(templates_dir):
                # 从目录中的md文件加载
                for filename in os.listdir(templates_dir):
                    if filename.endswith('.md'):
                        card_type = filename.replace('.md', '')
                        filepath = os.path.join(templates_dir, filename)
                        fields = []
                        try:
                            with open(filepath, 'r', encoding='utf-8') as f:
                                content = f.read()
                                # 检查是否是HTML表格格式
                                # 兼容带属性的<table ...>，采用更宽松的检测
                                is_html = '<table' in content.lower()
                                
                                # 保存HTML内容（如果是HTML格式）
                                if is_html:
                                    html_templates[card_type] = content
                                
                                # 先尝试解析HTML表格格式
                                if is_html:
                                    fields = parse_html_table(content)
                                else:
                                    # 尝试解析Markdown表格格式
                                    fields = parse_markdown_table(content)
                                
                                # 如果表格解析没有结果，再尝试列表格式
                                if not fields:
                                    for line in content.split('\n'):
                                        line = line.strip()
                                        if line.startswith('- '):
                                            field = line.replace('- ', '').strip()
                                            if field:
                                                fields.append(field)
                            
                            if fields:
                                # 确保第一个字段是"卡证类型"（如果还没有的话）
                                if not fields or fields[0] != "卡证类型":
                                    fields.insert(0, "卡证类型")
                                templates[card_type] = fields
                                print(f"✅ 成功加载 {card_type} 字段模板，共 {len(fields)} 个字段")
                            else:
                                print(f"⚠️ {card_type} 字段模板解析结果为空，将使用默认模板")
                        except Exception as e:
                            print(f"⚠️ 加载模板文件 {filename} 失败: {e}")
                            continue
            else:
                # 如果目录不存在，使用默认模板
                templates = {
                    "身份证": ["卡证类型", "姓名", "性别", "民族", "出生日期", "住址", "公民身份号码", "签发机关", "有效期限"],
                    "银行卡": ["卡证类型", "发卡行", "卡号", "有效期", "姓名", "卡面类型"],
                    "驾驶证": ["卡证类型", "姓名", "性别", "国籍", "住址", "出生日期", "初次领证日期", "准驾车型", "有效期限", "档案编号", "证号"],
                    "护照": ["卡证类型", "姓名", "性别", "出生日期", "出生地点", "护照号码", "签发日期", "有效期至", "签发机关"],
                    "工牌": ["卡证类型", "姓名", "工号", "部门", "职位", "公司名称", "有效期"],
                    "银行承兑汇票": ["卡证类型", "出票人名称", "出票人账号", "出票人开户行", "出票人保证人姓名", "票据金额（大写）", "票据金额（小写）", "收款人名称", "收款人账号", "收款人开户行", "保证人地址", "保证日期", "承兑人名称", "承兑人账号", "承兑人开户行行号", "承兑人开户行名称", "承兑人承诺", "本汇票已承兑，到期无条件付款", "承兑日期", "交易合同号", "能否转让", "保证人姓名", "信用等级", "审查意见"],
                    "其他": ["卡证类型", "姓名", "证件号码", "有效期"]
                }
        except Exception as e:
            print(f"⚠️ 加载字段模板失败: {e}")
            # 使用默认模板
            templates = {
                "身份证": ["卡证类型", "姓名", "性别", "民族", "出生日期", "住址", "公民身份号码", "签发机关", "有效期限"],
                "银行卡": ["卡证类型", "发卡行", "卡号", "有效期", "姓名", "卡面类型"],
                "驾驶证": ["卡证类型", "姓名", "性别", "国籍", "住址", "出生日期", "初次领证日期", "准驾车型", "有效期限", "档案编号", "证号"],
                "护照": ["卡证类型", "姓名", "性别", "出生日期", "出生地点", "护照号码", "签发日期", "有效期至", "签发机关"],
                "工牌": ["卡证类型", "姓名", "工号", "部门", "职位", "公司名称", "有效期"],
                "银行承兑汇票": ["卡证类型", "出票人名称", "出票人账号", "出票人开户行", "出票人保证人姓名", "票据金额（大写）", "票据金额（小写）", "收款人名称", "收款人账号", "收款人开户行", "保证人地址", "保证日期", "承兑人名称", "承兑人账号", "承兑人开户行行号", "承兑人开户行名称", "承兑人承诺", "本汇票已承兑，到期无条件付款", "承兑日期", "交易合同号", "能否转让", "保证人姓名", "信用等级", "审查意见"],
                "其他": ["卡证类型", "姓名", "证件号码", "有效期"]
            }
        # 将HTML模板存储到实例变量中
        self.field_template_htmls = html_templates
        return templates

    def detect_card_type(self, image):
        """第一步：识别卡证类型并加载默认字段模板"""
        if image is None:
            return None, [], "❌ 请先上传图片"
        
        try:
            # 在提供给大模型前先做一次超分辨率预处理
            image_sr = self._super_resolve_image_for_ocr(image)

            self._ensure_card_api_loaded()
            if self.card_api is None:
                return None, [], "❌ 卡证OCR 未初始化"
            
            # 使用简化的提示词只识别卡证类型（不包含银行承兑汇票）
            type_prompt = (
                "请识别这张图片中的卡证类型。\n"
                "只允许从以下类别中选择一种：身份证、银行卡、驾驶证、护照、工牌、其他。\n"
                "只输出卡证类型，不要输出其他内容。"
            )
            
            result = self.card_api.recognize_card(
                image_sr,
                custom_prompt=type_prompt,
                use_rag=False,
                max_tokens=50,
                temperature=0.1
            )
            
            if not result.get("success"):
                return None, [], None, f"❌ 识别失败: {result.get('error', '未知错误')}"
            
            # 从结果中提取卡证类型（不包含银行承兑汇票）
            result_text = result.get("result", "").strip()
            card_types = ["身份证", "银行卡", "驾驶证", "护照", "工牌", "其他"]
            detected_type = None
            
            for ct in card_types:
                if ct in result_text:
                    detected_type = ct
                    break
            
            if not detected_type:
                detected_type = "其他"
            
            # 加载对应的默认字段模板（卡证OCR不使用HTML模板）
            templates = self._load_field_templates()
            default_fields = templates.get(detected_type, templates.get("其他", []))
            
            # 卡证OCR不使用HTML模板，强制设置为None
            html_template = None
            
            # 保存当前状态
            self.current_card_type = detected_type
            self.current_default_fields = default_fields.copy()
            self.current_custom_fields = []
            self.current_field_template_html = None  # 卡证OCR不使用HTML模板
            
            return detected_type, default_fields, html_template, f"✅ 识别成功：{detected_type}"
            
        except Exception as e:
            return None, [], None, f"❌ 识别失败: {str(e)}"

    def detect_bill_type(self, image):
        """票据识别第一步：识别票据类型并加载默认字段模板（使用HTML模板）"""
        fixed_bill_type = ["银行承兑汇票", "商业承兑汇票", "转账支票", "现金支票", "普通支票", "本票", "付款回单", "收款回单", "代发业务回单", "电子发票（铁路电子客票）",]
        extendable_bill_type = ["代发业务清单", "单位活期明细对账单", "电子发票（增值税专用发票）", "电子发票（普通发票）", "中央非税收入统一票据" ]

        supported_bill_type = fixed_bill_type + extendable_bill_type
        if image is None:
            return None, [], None, "❌ 请先上传图片"
        
        try:
            # 在提供给大模型前先做一次超分辨率预处理
            image_sr = self._super_resolve_image_for_ocr(image)

            self._ensure_bill_api_loaded()
            if self.bill_api is None:
                return None, [], None, "❌ 票据OCR 未初始化"
            
            # 票据OCR只识别银行承兑汇票
            type_prompt = (
                "请识别这张图片中的票据类型。\n"
                f"只允许从以下类别中选择一种：{supported_bill_type}。\n"
                "转账支票类型必须有\"转账支票\"关键词，现金支票类型必须有\"现金支票\"关键词，其他支票为普通支票\n"
                "只输出票据类型，不要输出其他内容。"
            )
            
            result = self.bill_api.recognize_card(
                image_sr,
                custom_prompt=type_prompt,
                use_rag=False,
                max_tokens=50,
                temperature=0.1
            )
            
            if not result.get("success"):
                return None, [], None, f"❌ 识别失败: {result.get('error', '未知错误')}"
            
            # 从结果中提取票据类型
            result_text = result.get("result", "").strip()
            detected_type = None
            
            for bt in supported_bill_type:
                if bt in result_text:
                    detected_type = bt
                    break
            
            # No need to set default
            # if not detected_type:
            #     detected_type = "银行承兑汇票"  # 默认使用银行承兑汇票
            
            # 加载对应的默认字段模板（票据OCR使用HTML模板）
            templates = self._load_field_templates()
            #todo: add template of other bills
            default_fields = templates.get(detected_type, templates.get("其他票据", [])) 
            
            # 获取HTML表格内容（票据OCR必须使用HTML模板）
            html_template = getattr(self, 'field_template_htmls', {}).get(detected_type, None)
            
            # 保存当前状态
            self.current_card_type = detected_type
            self.current_default_fields = default_fields.copy()
            self.current_custom_fields = []
            self.current_field_template_html = html_template
            self.current_parsed_dict = None # 清除过去解析结果
            
            return detected_type, default_fields, html_template, f"✅ 识别成功：{detected_type}"
            
        except Exception as e:
            return None, [], None, f"❌ 识别失败: {str(e)}"

    def update_fields(self, card_type, default_fields, custom_fields_text):
        """第二步：合并默认字段和自定义字段"""
        try:
            # 解析自定义字段（每行一个字段）
            custom_fields = []
            if custom_fields_text:
                for line in custom_fields_text.strip().split('\n'):
                    field = line.strip()
                    if field and field not in default_fields:
                        custom_fields.append(field)
            
            # 合并字段
            all_fields = default_fields + custom_fields
            
            # 保存当前状态
            self.current_card_type = card_type
            self.current_default_fields = default_fields
            self.current_custom_fields = custom_fields
            
            return all_fields, f"✅ 字段已更新，共 {len(all_fields)} 个字段"
            
        except Exception as e:
            return [], f"❌ 更新字段失败: {str(e)}"

    def ocr_card_with_fields(self, image, fields_to_extract):
        """第三步：使用指定字段进行OCR识别"""
        if image is None:
            return "❌ 请先上传图片"
        
        if not fields_to_extract:
            return "❌ 请先设置要提取的字段"
        
        try:
            # 在提供给大模型前先做一次超分辨率预处理
            image_sr = self._super_resolve_image_for_ocr(image)

            self._ensure_card_api_loaded()
            if self.card_api is None:
                return "❌ 卡证OCR 未初始化"
            
            # 构建包含字段列表的提示词
            fields_list = "、".join(fields_to_extract)
            
            # 卡证OCR不使用HTML模板，只使用Markdown格式
            has_html_template = False
            
            if False:  # 卡证OCR不使用HTML模板
                # 如果有HTML模板，要求大模型返回填充后的HTML表格
                custom_prompt = (
                    f"你是专业的票据/卡证OCR引擎。请阅读并识别输入图片内容，并在下面提供的HTML表格模板中填充对应字段的值。\n"
                    f"\n"
                    f"【卡证类型】{self.current_card_type or '未知'}\n"
                    f"【字段列表（必须全部覆盖，缺失填写'无'）】{fields_list}\n"
                    f"\n"
                    f"【HTML表格模板】\n"
                    f"{html_template}\n"
                    f"\n"
                    f"要求：\n"
                    f"- 只返回填充后的HTML表格（保持原有结构、行列、合并单元格和样式/属性），不要返回任何其他说明文字。\n"
                    f"- 不新增或删除字段，不改变表头文案；未识别到的填写'无'。\n"
                    f"- 仅在需要填写值的单元格写入文本，避免修改字段名单元格。\n"
                    f"- 禁止输出任何猜测或编造的内容。\n"
                    f"- 禁止输出未在字段列表中的字段和字段值。\n"
                    f"- 不要使用代码块标记符号（例如 ``` ）。"
                )
            else:
                # 如果没有HTML模板，使用原来的Markdown表格格式
                custom_prompt = (
                    f"你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                    f"\n"
                    f"任务要求：\n"
                    f"1. 识别卡证类型：{self.current_card_type or '未知'}\n"
                    f"2. 提取以下字段（必须全部提取，如果图片中没有该字段则填写'无'）：{fields_list}，禁止提取该列表以外的字段和字段值\n"
                    f"3. 以Markdown表格形式输出，表格包含两列：字段名、字段值\n"
                    f"4. 不要使用代码块标记符号（例如 ``` ）\n"
                    f"5. 输出限制：\n"
                    f"   - 最终输出只包含Markdown表格。\n"
                    f"   - 禁止输出任何猜测或编造的内容。\n"
                    f"   - 禁止输出任何其他文字或解释性内容。\n"
                    f"   - 禁止输出未在字段列表中的字段和字段值。"
                )
            
            # 只有银行卡类型才使用RAG
            use_rag = (self.current_card_type == "银行卡")
            
            result = self.card_api.recognize_card(
                image_sr,
                custom_prompt=custom_prompt,
                use_rag=use_rag,
            )
            
            if not result.get("success"):
                return f"❌ OCR识别失败: {result.get('error', '未知错误')}"
            
            # 在终端输出RAG相似度匹配结果
            rag_info = result.get("rag_info")
            if rag_info and rag_info.get("enabled") and rag_info.get("results"):
                print("\n" + "=" * 60)
                print("📊 RAG相似度匹配结果")
                print("=" * 60)
                print(f"找到 {len(rag_info['results'])} 张相似图片：\n")
                for i, r in enumerate(rag_info["results"], 1):
                    filename = r.get("filename", "未知")
                    similarity = r.get("similarity", 0.0)
                    print(f"  {i}. {filename}")
                    print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
                print("=" * 60 + "\n")
            
            raw_result = (result.get("result") or "").strip()
            
            # 如果模型按要求直接返回HTML表格，则优先使用HTML（注入可编辑样式）
            if has_html_template and "<table" in raw_result.lower():
                try:
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(raw_result, 'html.parser')
                    table = soup.find('table')
                    if table:
                        # 添加样式使表格更美观且可编辑
                        table['class'] = (table.get('class', []) or []) + ['ocr-result-table']
                        # 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                        for tr in table.find_all('tr'):
                            if tr.get('height'):
                                del tr['height']
                            if tr.get('width'):
                                del tr['width']
                        for td in table.find_all('td'):
                            if td.get('height'):
                                del td['height']
                            if td.get('width'):
                                del td['width']
                        # 移除table的固定width属性
                        if table.get('width'):
                            del table['width']
                        if table.get('style'):
                            # 移除style中的width和height（使用Python的re模块）
                            import re
                            style = table.get('style', '')
                            style = re.sub(r'width\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                            style = re.sub(r'height\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                            style = style.strip()
                            if style:
                                table['style'] = style
                            else:
                                del table['style']
                        # 移除colgroup中的固定宽度设置
                        for colgroup in table.find_all('colgroup'):
                            for col in colgroup.find_all('col'):
                                if col.get('width'):
                                    del col['width']
                                if col.get('style'):
                                    import re
                                    style = col.get('style', '')
                                    style = re.sub(r'width\s*:\s*[^;]+;?', '', style, flags=re.IGNORECASE)
                                    style = style.strip()
                                    if style:
                                        col['style'] = style
                                    else:
                                        del col['style']
                        # 获取所有字段名（用于识别哪些单元格是字段名，哪些是值）
                        field_names = set(fields_to_extract)
                        for td in table.find_all('td'):
                            cell_text = td.get_text(strip=True)
                            # 如果单元格文本不是字段名，且不是空，则设置为可编辑（这是值单元格）
                            if cell_text and cell_text not in field_names:
                                td['contenteditable'] = 'true'
                            # 如果单元格为空，也设置为可编辑（可能是待填充的值单元格）
                            elif not cell_text:
                                td['contenteditable'] = 'true'
                        
                        # 优化的表格样式：可调整大小的容器，表格随容器大小变化
                        styled_html = f"""
                        <style>
                        /* 可调整大小的表格容器 */
                        .ocr-result-table-container {{
                            position: relative;
                            display: inline-block;
                            min-width: 500px;
                            min-height: 300px;
                            max-width: 95vw;
                            max-height: 90vh;
                            width: 100%;
                            height: 600px;
                            resize: both;
                            overflow: auto;  /* 允许滚动，确保表格不超出容器 */
                            border: 2px solid #e0e0e0;
                            border-radius: 8px;
                            padding: 10px;
                            background-color: #f8f9fa;
                            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
                            margin: 20px 0;
                        }}
                        /* 调整大小手柄样式 */
                        .ocr-result-table-container::-webkit-resizer {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            border-radius: 0 0 8px 0;
                            width: 20px;
                            height: 20px;
                        }}
                        /* 调整大小提示 */
                        .ocr-result-table-container::before {{
                            content: '↘ 拖拽调整大小';
                            position: absolute;
                            top: 5px;
                            right: 5px;
                            font-size: 11px;
                            color: #667eea;
                            background: rgba(255, 255, 255, 0.9);
                            padding: 2px 6px;
                            border-radius: 4px;
                            pointer-events: none;
                            opacity: 0.7;
                            z-index: 5;
                            transition: opacity 0.3s ease;
                        }}
                        .ocr-result-table-container:hover::before {{
                            opacity: 1;
                        }}
                        /* 调整大小时的边框高亮 */
                        .ocr-result-table-container:active {{
                            border-color: #667eea;
                            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
                        }}
                        .ocr-result-table {{
                            width: auto;  /* 表格宽度根据内容自适应 */
                            min-width: 100%;  /* 最小宽度为容器宽度 */
                            max-width: 100%;  /* 最大宽度不超过容器 */
                            border-collapse: collapse;
                            margin: 0;
                            font-size: 14px;
                            table-layout: auto;  /* 使用auto，让列宽根据内容自动调整 */
                            box-shadow: none;
                            border-radius: 8px;
                            overflow: visible;  /* 允许内容溢出，不裁剪 */
                            background-color: #ffffff;
                        }}
                        .ocr-result-table th,
                        .ocr-result-table td {{
                            border: 1px solid #e0e0e0;
                            padding: 12px 16px;
                            text-align: left;
                            vertical-align: top;
                            word-break: break-word;
                            word-wrap: break-word;
                            transition: all 0.2s ease;
                            line-height: 1.6;
                            height: auto !important;  /* 行高根据内容自动调整，覆盖HTML中的固定height */
                            min-height: auto !important;
                            overflow: visible;  /* 允许内容显示，不裁剪 */
                            width: auto;  /* 列宽根据内容自动调整 */
                            max-width: none;  /* 不限制最大宽度 */
                        }}
                        /* 字段名列：根据内容自适应宽度 */
                        .ocr-result-table td:not([contenteditable="true"]) {{
                            background-color: #f8f9fa;
                            font-weight: 600;
                            color: #374151;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 120px;  /* 最小宽度 */
                            max-width: 300px;  /* 最大宽度限制，避免过宽 */
                            white-space: nowrap;  /* 字段名不换行 */
                            font-size: 14px;
                            border-right: 2px solid #d1d5db;
                            height: auto !important;
                            overflow: visible;
                        }}
                        /* 值列：根据内容自适应宽度 */
                        .ocr-result-table td[contenteditable="true"] {{
                            background-color: #ffffff;
                            cursor: text;
                            min-height: 20px;
                            height: auto !important;  /* 行高根据内容自动调整 */
                            position: relative;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 200px;  /* 最小宽度 */
                            max-width: none;  /* 不限制最大宽度，允许长文本 */
                            overflow: visible;  /* 允许内容显示 */
                            word-break: break-word;  /* 长文本自动换行 */
                        }}
                        /* 根据文本长度动态调整样式（保持列宽比例） */
                        .ocr-result-table td[contenteditable="true"][data-length="short"] {{
                            font-size: 15px;
                            padding: 10px 14px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="medium"] {{
                            font-size: 14px;
                            padding: 12px 16px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="long"] {{
                            font-size: 13px;
                            padding: 14px 18px;
                            line-height: 1.7;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="very-long"] {{
                            font-size: 12px;
                            padding: 16px 20px;
                            line-height: 1.8;
                            height: auto !important;
                        }}
                        .ocr-result-table th {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: #ffffff;
                            font-weight: 600;
                            font-size: 15px;
                            text-transform: uppercase;
                            letter-spacing: 0.5px;
                            border-color: #5568d3;
                        }}
                        .ocr-result-table tr:nth-child(even) {{
                            background-color: #f8f9fa;
                        }}
                        .ocr-result-table tr:nth-child(odd) {{
                            background-color: #ffffff;
                        }}
                        .ocr-result-table tr:hover {{
                            background-color: #f0f4ff;
                        }}
                        .ocr-result-table tr:hover td:not([contenteditable="true"]) {{
                            background-color: #e5e7eb;
                        }}
                        .ocr-result-table td[contenteditable="true"]:hover {{
                            background-color: #f8f9ff;
                            box-shadow: inset 0 0 0 1px #667eea;
                        }}
                        .ocr-result-table td[contenteditable="true"]:focus {{
                            outline: none;
                            background-color: #eef5ff;
                            box-shadow: inset 0 0 0 2px #667eea, 0 0 0 3px rgba(102, 126, 234, 0.1);
                            border-radius: 4px;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:before {{
                            content: "点击编辑...";
                            color: #999;
                            font-style: italic;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:focus:before {{
                            content: "";
                        }}
                        /* 优化长文本显示 */
                        .ocr-result-table td[contenteditable="true"] {{
                            overflow-wrap: break-word;
                            hyphens: auto;
                        }}
                        /* 响应式设计 */
                        @media (max-width: 768px) {{
                            .ocr-result-table-container {{
                                min-width: 300px;
                                min-height: 200px;
                            }}
                            .ocr-result-table {{
                                font-size: 12px;
                                table-layout: fixed;
                            }}
                            .ocr-result-table th,
                            .ocr-result-table td {{
                                padding: 8px 12px;
                            }}
                            .ocr-result-table td:not([contenteditable="true"]) {{
                                width: 30%;
                                font-size: 12px;
                            }}
                            .ocr-result-table td[contenteditable="true"] {{
                                width: 70%;
                            }}
                        }}
                        </style>
                        <script>
                        (function() {{
                            // 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                            function removeFixedHeights() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    // 移除table的width属性
                                    if (table.hasAttribute('width')) {{
                                        table.removeAttribute('width');
                                    }}
                                    if (table.style.width) {{
                                        table.style.width = '';
                                    }}
                                    
                                    // 移除tr的height和width属性
                                    var rows = table.querySelectorAll('tr');
                                    rows.forEach(function(row) {{
                                        if (row.hasAttribute('height')) {{
                                            row.removeAttribute('height');
                                        }}
                                        if (row.hasAttribute('width')) {{
                                            row.removeAttribute('width');
                                        }}
                                    }});
                                    
                                    // 移除td和th的height和width属性
                                    var cells = table.querySelectorAll('td, th');
                                    cells.forEach(function(cell) {{
                                        if (cell.hasAttribute('height')) {{
                                            cell.removeAttribute('height');
                                        }}
                                        if (cell.hasAttribute('width')) {{
                                            cell.removeAttribute('width');
                                        }}
                                        // 移除内联样式中的height和width
                                        if (cell.style.height) {{
                                            cell.style.height = '';
                                        }}
                                        if (cell.style.width) {{
                                            cell.style.width = '';
                                        }}
                                    }});
                                    
                                    // 移除colgroup中的width属性
                                    var colgroups = table.querySelectorAll('colgroup');
                                    colgroups.forEach(function(colgroup) {{
                                        var cols = colgroup.querySelectorAll('col');
                                        cols.forEach(function(col) {{
                                            if (col.hasAttribute('width')) {{
                                                col.removeAttribute('width');
                                            }}
                                            if (col.style.width) {{
                                                col.style.width = '';
                                            }}
                                        }});
                                    }});
                                }}
                            }}
                            
                            // 根据文本长度动态设置data-length属性
                            function updateCellLength() {{
                                var cells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                cells.forEach(function(cell) {{
                                    var text = cell.textContent || cell.innerText || '';
                                    var length = text.length;
                                    cell.removeAttribute('data-length');
                                    if (length > 0) {{
                                        if (length <= 20) {{
                                            cell.setAttribute('data-length', 'short');
                                        }} else if (length <= 50) {{
                                            cell.setAttribute('data-length', 'medium');
                                        }} else if (length <= 100) {{
                                            cell.setAttribute('data-length', 'long');
                                        }} else {{
                                            cell.setAttribute('data-length', 'very-long');
                                        }}
                                    }}
                                }});
                            }}
                            
                            // 页面加载后执行
                            setTimeout(function() {{
                                removeFixedHeights();
                                updateCellLength();
                            }}, 100);
                            
                            // 监听内容变化
                            var observer = new MutationObserver(function(mutations) {{
                                removeFixedHeights();
                                updateCellLength();
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true,
                                        attributes: true,
                                        attributeFilter: ['height', 'style']
                                    }});
                                }}
                            }}, 200);
                        }})();
                        </script>
                        <div class="ocr-result-table-container">
                            {str(table)}
                        </div>
                        """
                        self.last_ocr_html = styled_html
                        self.last_ocr_markdown = ""  # HTML模式下不生成Markdown
                        return styled_html
                    else:
                        # 如果解析失败，回退到Markdown处理
                        cleaned = self._sanitize_markdown(raw_result)
                        self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
                        self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                        return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
                except Exception as e:
                    print(f"⚠️ HTML表格解析失败，回退到Markdown格式: {e}")
                    # 解析失败，回退到Markdown处理
                    cleaned = self._sanitize_markdown(raw_result)
                    self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
                    self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                    return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
            else:
                # 否则按Markdown处理
                cleaned = self._sanitize_markdown(raw_result)
            self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
            
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def ocr_bill_with_fields(self, image, fields_to_extract):
        """票据OCR第三步：使用指定字段进行OCR识别（使用HTML模板）"""
        if image is None:
            return "❌ 请先上传图片"
        
        if not fields_to_extract:
            return "❌ 请先设置要提取的字段"
        
        try:
            # 在提供给大模型前先做一次超分辨率预处理
            image_sr = self._super_resolve_image_for_ocr(image)

            self._ensure_bill_api_loaded()
            if self.bill_api is None:
                return "❌ 票据OCR 未初始化"
            
            # 构建包含字段列表的提示词
            fields_list = "、".join(fields_to_extract)

            # 票据OCR使用HTML模板
            html_template = getattr(self, 'current_final_fields_html', None)
            if not html_template:
                html_template = getattr(self, 'current_field_template_html', None)
            has_html_template = html_template is not None and html_template.strip()
            
            if has_html_template:
                # 如果有HTML模板，要求大模型返回填充后的HTML表格
                # 将字段列表格式化为更清晰的格式，确保模型不会遗漏
                fields_list_formatted = "\n".join([f"  {i+1}. {field}" for i, field in enumerate(fields_to_extract)])
                
                custom_prompt = (
                    f"你是专业的票据OCR引擎。请仔细阅读并识别输入图片中的所有内容，并在下面提供的HTML表格模板中填充对应字段的值。\n"
                    f"\n"
                    f"【票据类型】{self.current_card_type or '未知'}\n"
                    f"\n"
                    f"【必须识别的字段列表（共{len(fields_to_extract)}个字段，必须全部识别，一个都不能遗漏）】\n"
                    f"{fields_list_formatted}\n"
                    f"\n"
                    f"【重要要求】\n"
                    f"- **必须识别上述所有{len(fields_to_extract)}个字段，一个都不能遗漏**\n"
                    f"- 如果图片中没有某个字段的值，该字段的值必须填写'无'，但不能跳过该字段\n"
                    f"- 请仔细检查图片中的每一个位置，确保所有字段都被识别和填充\n"
                    f"- 对于组合字段（如'出票人全称'、'出票人账号'等），需要分别识别每个子字段\n"
                    f"- 对于小写金额类型字段，如果为表格形式，需要**仔细核对每一位数字对应的单位**，**必须保证大小写数值相一致**"
                    f"- HTML表格中的小写金额字段**不允许直接用识别结果填充**，**必须用核对过的小写数字填充**，填充内容只能包括**数字和小数点**\n"
                    f"\n"
                    f"【HTML表格模板】\n"
                    f"{html_template}\n"
                    f"\n"
                    f"【输出要求】\n"
                    f"- 只返回填充后的HTML表格（保持原有结构、行列、合并单元格和样式/属性），不要返回任何其他说明文字\n"
                    #f"- 返回填充后的HTML表格（保持原有结构、行列、合并单元格和样式/属性），需要返回解释说明文字\n"
                    f"- 如果HTML表格中存在<repeat>标签，则该标签中包含的内容为表格中某一行的格式，该行可能重复若干次，需要根据识别结果准确判断重复行数并正确填充\n"
                    f"- 不新增或删除字段，不改变表头文案；未识别到的填写'无'\n"
                    f"- 仅在需要填写值的单元格写入文本，避免修改字段名单元格\n"
                    f"- 禁止输出任何猜测或编造的内容\n"
                    f"- 禁止输出未在字段列表中的字段和字段值\n"
                    f"- 不要使用代码块标记符号（例如 ``` ）\n"
                )

                # custom_prompt = (
                #     f"你是专业的票据OCR引擎。请仔细阅读并识别输入图片中的所有内容，并生成一个对应的json识别结果，需要包含解释信息。\n"
                # )
            else:
                # 如果没有HTML模板，使用Markdown表格格式（不应该发生，但作为兜底）
                custom_prompt = (
                    f"你是专业的票据OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                    f"\n"
                    f"任务要求：\n"
                    f"1. 识别票据类型：{self.current_card_type or '未知'}\n"
                    f"2. 提取以下字段（必须全部提取，如果图片中没有该字段则填写'无'）：{fields_list}，禁止提取该列表以外的字段和字段值\n"
                    f"3. 以Markdown表格形式输出，表格包含两列：字段名、字段值\n"
                    f"4. 不要使用代码块标记符号（例如 ``` ）\n"
                    f"5. 输出限制：\n"
                    f"   - 最终输出只包含Markdown表格。\n"
                    f"   - 禁止输出任何猜测或编造的内容。\n"
                    f"   - 禁止输出任何其他文字或解释性内容。\n"
                    f"   - 禁止输出未在字段列表中的字段和字段值。"
                )
            
            # 票据OCR不使用RAG
            use_rag = False
            
            # 票据OCR使用更大的max_tokens，确保能输出完整的HTML表格
            # 根据字段数量动态调整max_tokens（每个字段大约需要50-100 tokens）
            estimated_tokens = len(fields_to_extract) * 100 + 2000  # 基础2000 + 每个字段100
            max_tokens = max(2048, min(estimated_tokens, 8192))  # 最小2048，最大8192
            
            result = self.bill_api.recognize_card(
                image_sr,
                custom_prompt=custom_prompt,
                use_rag=use_rag,
                max_tokens=max_tokens,
                temperature=0.1,  # 降低温度，提高准确性
            )
            
            if not result.get("success"):
                return f"❌ OCR识别失败: {result.get('error', '未知错误')}"
            
            raw_result = (result.get("result") or "").strip()
            # 如果模型按要求直接返回HTML表格，则优先使用HTML（注入可编辑样式）
            if has_html_template and "<table" in raw_result.lower():
                try:
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(raw_result, 'html.parser')
                    table = soup.find('table')
                    if table:
                        # 添加样式使表格更美观且可编辑
                        table['class'] = (table.get('class', []) or []) + ['ocr-result-table']
                        # 获取所有字段名（用于识别哪些单元格是字段名，哪些是值）
                        field_names = set(fields_to_extract)
                        for td in table.find_all('td'):
                            cell_text = td.get_text(strip=True)
                            # 如果单元格文本不是字段名，且不是空，则设置为可编辑（这是值单元格）
                            if cell_text and cell_text not in field_names:
                                td['contenteditable'] = 'true'
                            # 如果单元格为空，也设置为可编辑（可能是待填充的值单元格）
                            elif not cell_text:
                                td['contenteditable'] = 'true'
                        
                        # 优化的表格样式：可调整大小的容器，表格随容器大小变化
                        # 添加JavaScript代码，监听编辑事件并更新隐藏的Textbox
                        styled_html = f"""
                        <style>
                        /* 可调整大小的表格容器 */
                        .ocr-result-table-container {{
                            position: relative;
                            display: inline-block;
                            min-width: 500px;
                            min-height: 300px;
                            max-width: 95vw;
                            max-height: 90vh;
                            width: 100%;
                            height: 600px;
                            resize: both;
                            overflow: auto;  /* 允许滚动，确保表格不超出容器 */
                            border: 2px solid #e0e0e0;
                            border-radius: 8px;
                            padding: 10px;
                            background-color: #f8f9fa;
                            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
                            margin: 20px 0;
                        }}
                        /* 调整大小手柄样式 */
                        .ocr-result-table-container::-webkit-resizer {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            border-radius: 0 0 8px 0;
                            width: 20px;
                            height: 20px;
                        }}
                        /* 调整大小提示 */
                        .ocr-result-table-container::before {{
                            content: '↘ 拖拽调整大小';
                            position: absolute;
                            top: 5px;
                            right: 5px;
                            font-size: 11px;
                            color: #667eea;
                            background: rgba(255, 255, 255, 0.9);
                            padding: 2px 6px;
                            border-radius: 4px;
                            pointer-events: none;
                            opacity: 0.7;
                            z-index: 5;
                            transition: opacity 0.3s ease;
                        }}
                        .ocr-result-table-container:hover::before {{
                            opacity: 1;
                        }}
                        /* 调整大小时的边框高亮 */
                        .ocr-result-table-container:active {{
                            border-color: #667eea;
                            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
                        }}
                        .ocr-result-table {{
                            width: auto;  /* 表格宽度根据内容自适应 */
                            min-width: 100%;  /* 最小宽度为容器宽度 */
                            max-width: 100%;  /* 最大宽度不超过容器 */
                            border-collapse: collapse;
                            margin: 0;
                            font-size: 14px;
                            table-layout: auto;  /* 使用auto，让列宽根据内容自动调整 */
                            box-shadow: none;
                            border-radius: 8px;
                            overflow: visible;  /* 允许内容溢出，不裁剪 */
                            background-color: #ffffff;
                        }}
                        .ocr-result-table th,
                        .ocr-result-table td {{
                            border: 1px solid #e0e0e0;
                            padding: 12px 16px;
                            text-align: left;
                            vertical-align: top;
                            word-break: break-word;
                            word-wrap: break-word;
                            transition: all 0.2s ease;
                            line-height: 1.6;
                            height: auto !important;  /* 行高根据内容自动调整，覆盖HTML中的固定height */
                            min-height: auto !important;
                            overflow: visible;  /* 允许内容显示，不裁剪 */
                            width: auto;  /* 列宽根据内容自动调整 */
                            max-width: none;  /* 不限制最大宽度 */
                        }}
                        /* 字段名列：根据内容自适应宽度 */
                        .ocr-result-table td:not([contenteditable="true"]) {{
                            background-color: #f8f9fa;
                            font-weight: 600;
                            color: #374151;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 120px;  /* 最小宽度 */
                            max-width: 300px;  /* 最大宽度限制，避免过宽 */
                            white-space: nowrap;  /* 字段名不换行 */
                            font-size: 14px;
                            border-right: 2px solid #d1d5db;
                            height: auto !important;
                            overflow: visible;
                        }}
                        /* 值列：根据内容自适应宽度 */
                        .ocr-result-table td[contenteditable="true"] {{
                            background-color: #ffffff;
                            cursor: text;
                            min-height: 20px;
                            height: auto !important;  /* 行高根据内容自动调整 */
                            position: relative;
                            width: auto;  /* 宽度根据内容自适应 */
                            min-width: 200px;  /* 最小宽度 */
                            max-width: none;  /* 不限制最大宽度，允许长文本 */
                            overflow: visible;  /* 允许内容显示 */
                            word-break: break-word;  /* 长文本自动换行 */
                        }}
                        /* 根据文本长度动态调整样式（保持列宽比例） */
                        .ocr-result-table td[contenteditable="true"][data-length="short"] {{
                            font-size: 15px;
                            padding: 10px 14px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="medium"] {{
                            font-size: 14px;
                            padding: 12px 16px;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="long"] {{
                            font-size: 13px;
                            padding: 14px 18px;
                            line-height: 1.7;
                            height: auto !important;
                        }}
                        .ocr-result-table td[contenteditable="true"][data-length="very-long"] {{
                            font-size: 12px;
                            padding: 16px 20px;
                            line-height: 1.8;
                            height: auto !important;
                        }}
                        .ocr-result-table th {{
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                            color: #ffffff;
                            font-weight: 600;
                            font-size: 15px;
                            text-transform: uppercase;
                            letter-spacing: 0.5px;
                            border-color: #5568d3;
                        }}
                        .ocr-result-table tr:nth-child(even) {{
                            background-color: #f8f9fa;
                        }}
                        .ocr-result-table tr:nth-child(odd) {{
                            background-color: #ffffff;
                        }}
                        .ocr-result-table tr:hover {{
                            background-color: #f0f4ff;
                        }}
                        .ocr-result-table tr:hover td:not([contenteditable="true"]) {{
                            background-color: #e5e7eb;
                        }}
                        .ocr-result-table td[contenteditable="true"]:hover {{
                            background-color: #f8f9ff;
                            box-shadow: inset 0 0 0 1px #667eea;
                        }}
                        .ocr-result-table td[contenteditable="true"]:focus {{
                            outline: none;
                            background-color: #eef5ff;
                            box-shadow: inset 0 0 0 2px #667eea, 0 0 0 3px rgba(102, 126, 234, 0.1);
                            border-radius: 4px;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:before {{
                            content: "点击编辑...";
                            color: #999;
                            font-style: italic;
                        }}
                        .ocr-result-table td[contenteditable="true"]:empty:focus:before {{
                            content: "";
                        }}
                        /* 优化长文本显示 */
                        .ocr-result-table td[contenteditable="true"] {{
                            overflow-wrap: break-word;
                            hyphens: auto;
                        }}
                        /* 响应式设计 */
                        @media (max-width: 768px) {{
                            .ocr-result-table-container {{
                                min-width: 300px;
                                min-height: 200px;
                            }}
                            .ocr-result-table {{
                                font-size: 12px;
                                table-layout: fixed;
                            }}
                            .ocr-result-table th,
                            .ocr-result-table td {{
                                padding: 8px 12px;
                            }}
                            .ocr-result-table td:not([contenteditable="true"]) {{
                                width: 30%;
                                font-size: 12px;
                            }}
                            .ocr-result-table td[contenteditable="true"] {{
                                width: 70%;
                            }}
                        }}
                        </style>
                        <script>
                        (function() {{
                            // 移除所有固定的height和width属性，让行高和列宽根据内容自动调整
                            function removeFixedHeights() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    // 移除table的width属性
                                    if (table.hasAttribute('width')) {{
                                        table.removeAttribute('width');
                                    }}
                                    if (table.style.width) {{
                                        table.style.width = '';
                                    }}
                                    
                                    // 移除tr的height和width属性
                                    var rows = table.querySelectorAll('tr');
                                    rows.forEach(function(row) {{
                                        if (row.hasAttribute('height')) {{
                                            row.removeAttribute('height');
                                        }}
                                        if (row.hasAttribute('width')) {{
                                            row.removeAttribute('width');
                                        }}
                                    }});
                                    
                                    // 移除td和th的height和width属性
                                    var cells = table.querySelectorAll('td, th');
                                    cells.forEach(function(cell) {{
                                        if (cell.hasAttribute('height')) {{
                                            cell.removeAttribute('height');
                                        }}
                                        if (cell.hasAttribute('width')) {{
                                            cell.removeAttribute('width');
                                        }}
                                        // 移除内联样式中的height和width
                                        if (cell.style.height) {{
                                            cell.style.height = '';
                                        }}
                                        if (cell.style.width) {{
                                            cell.style.width = '';
                                        }}
                                    }});
                                    
                                    // 移除colgroup中的width属性
                                    var colgroups = table.querySelectorAll('colgroup');
                                    colgroups.forEach(function(colgroup) {{
                                        var cols = colgroup.querySelectorAll('col');
                                        cols.forEach(function(col) {{
                                            if (col.hasAttribute('width')) {{
                                                col.removeAttribute('width');
                                            }}
                                            if (col.style.width) {{
                                                col.style.width = '';
                                            }}
                                        }});
                                    }});
                                }}
                            }}
                            
                            // 根据文本长度动态设置data-length属性
                            function updateCellLength() {{
                                var cells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                cells.forEach(function(cell) {{
                                    var text = cell.textContent || cell.innerText || '';
                                    var length = text.length;
                                    cell.removeAttribute('data-length');
                                    if (length > 0) {{
                                        if (length <= 20) {{
                                            cell.setAttribute('data-length', 'short');
                                        }} else if (length <= 50) {{
                                            cell.setAttribute('data-length', 'medium');
                                        }} else if (length <= 100) {{
                                            cell.setAttribute('data-length', 'long');
                                        }} else {{
                                            cell.setAttribute('data-length', 'very-long');
                                        }}
                                    }}
                                }});
                            }}
                            
                            // 页面加载后执行
                            setTimeout(function() {{
                                removeFixedHeights();
                                updateCellLength();
                            }}, 100);
                            
                            // 监听内容变化
                            var observer = new MutationObserver(function(mutations) {{
                                removeFixedHeights();
                                updateCellLength();
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true,
                                        attributes: true,
                                        attributeFilter: ['height', 'style']
                                    }});
                                }}
                            }}, 200);
                        }})();
                        </script>
                        <div class="ocr-result-table-container">
                            {str(table)}
                        </div>
                        <script>
                        (function() {{
                            var updateTimeout = null;
                            
                            function updateEditedContent() {{
                                // 清除之前的定时器
                                if (updateTimeout) {{
                                    clearTimeout(updateTimeout);
                                }}
                                
                                // 延迟更新，避免频繁触发
                                updateTimeout = setTimeout(function() {{
                                    var table = document.querySelector('.ocr-result-table');
                                    if (!table) return;
                                    
                                    // 获取完整的HTML（包括样式）
                                    var fullHtml = document.querySelector('#bill-ocr-result-html, [id*="bill-ocr-result-html"]');
                                    var htmlContent = '';
                                    
                                    if (fullHtml) {{
                                        // 获取包含表格的完整HTML
                                        var container = fullHtml.querySelector('.ocr-result-table') || fullHtml;
                                        htmlContent = container.innerHTML;
                                    }} else {{
                                        // 如果没有找到容器，直接获取表格的outerHTML
                                        htmlContent = table.outerHTML;
                                    }}
                                    
                                    // 查找隐藏的Textbox - 使用多种方法
                                    var hiddenInput = null;
                                    
                                    // 方法1: 直接通过ID查找
                                    hiddenInput = document.getElementById('bill-ocr-result-html-edited');
                                    
                                    // 方法2: 通过ID包含关键字查找
                                    if (!hiddenInput) {{
                                        var inputs = document.querySelectorAll('input, textarea');
                                        for (var i = 0; i < inputs.length; i++) {{
                                            if (inputs[i].id && inputs[i].id.includes('bill-ocr-result-html-edited')) {{
                                                hiddenInput = inputs[i];
                                                break;
                                            }}
                                        }}
                                    }}
                                    
                                    // 方法3: 通过name属性查找
                                    if (!hiddenInput) {{
                                        hiddenInput = document.querySelector('input[name*="bill-ocr-result-html-edited"], textarea[name*="bill-ocr-result-html-edited"]');
                                    }}
                                    
                                    // 方法4: 通过data属性或class查找
                                    if (!hiddenInput) {{
                                        var allInputs = document.querySelectorAll('input[type="text"], textarea');
                                        for (var i = 0; i < allInputs.length; i++) {{
                                            var input = allInputs[i];
                                            // 检查是否在Gradio的隐藏组件区域
                                            if (input.style.display === 'none' || input.hidden || input.offsetParent === null) {{
                                                // 尝试设置值，看是否能找到正确的输入框
                                                var testValue = input.value;
                                                input.value = 'TEST_' + Date.now();
                                                if (input.value === 'TEST_' + Date.now()) {{
                                                    input.value = testValue; // 恢复原值
                                                    // 这可能是我们要找的输入框，但需要更精确的匹配
                                                }}
                                            }}
                                        }}
                                    }}
                                    
                                    if (hiddenInput) {{
                                        // 获取完整的HTML内容（包括样式）
                                        var styleTag = document.querySelector('style');
                                        var styleContent = styleTag ? styleTag.outerHTML : '';
                                        var fullContent = styleContent + '\\n' + table.outerHTML;
                                        
                                        hiddenInput.value = fullContent;
                                        
                                        // 触发多种事件，确保Gradio捕获到变化
                                        var events = ['input', 'change', 'blur', 'keyup'];
                                        events.forEach(function(eventType) {{
                                            var event = new Event(eventType, {{ bubbles: true, cancelable: true }});
                                            hiddenInput.dispatchEvent(event);
                                        }});
                                        
                                        // 也尝试直接设置属性
                                        if (hiddenInput.setAttribute) {{
                                            hiddenInput.setAttribute('value', fullContent);
                                        }}
                                        
                                        console.log('[DEBUG] 已更新隐藏Textbox，内容长度:', fullContent.length);
                                    }} else {{
                                        console.warn('[DEBUG] 未找到隐藏的Textbox组件');
                                        // 如果找不到，尝试通过window对象存储
                                        if (window.gradioEditedContent === undefined) {{
                                            window.gradioEditedContent = {{}};
                                        }}
                                        window.gradioEditedContent['bill-ocr-result-html-edited'] = htmlContent;
                                    }}
                                }}, 300);
                            }}
                            
                            // 监听所有可编辑单元格的输入事件
                            function attachListeners() {{
                                var editableCells = document.querySelectorAll('.ocr-result-table td[contenteditable="true"]');
                                editableCells.forEach(function(cell) {{
                                    // 移除旧的监听器（如果存在）
                                    var newCell = cell.cloneNode(true);
                                    cell.parentNode.replaceChild(newCell, cell);
                                    
                                    // 添加新的监听器
                                    newCell.addEventListener('input', updateEditedContent);
                                    newCell.addEventListener('blur', updateEditedContent);
                                    newCell.addEventListener('keyup', updateEditedContent);
                                    newCell.addEventListener('paste', function() {{
                                        setTimeout(updateEditedContent, 100);
                                    }});
                                }});
                                
                                // 初始更新
                                updateEditedContent();
                            }}
                            
                            // 延迟执行，确保DOM已加载
                            setTimeout(attachListeners, 500);
                            
                            // 使用MutationObserver监听表格变化（动态添加的单元格）
                            var observer = new MutationObserver(function(mutations) {{
                                var shouldReattach = false;
                                mutations.forEach(function(mutation) {{
                                    if (mutation.type === 'childList' && mutation.addedNodes.length > 0) {{
                                        shouldReattach = true;
                                    }}
                                }});
                                if (shouldReattach) {{
                                    setTimeout(attachListeners, 100);
                                }}
                            }});
                            
                            setTimeout(function() {{
                                var table = document.querySelector('.ocr-result-table');
                                if (table) {{
                                    observer.observe(table, {{
                                        childList: true,
                                        subtree: true,
                                        characterData: true
                                    }});
                                }}
                            }}, 500);
                            
                            // 页面卸载前保存
                            window.addEventListener('beforeunload', updateEditedContent);
                            
                            // 监听导出按钮点击事件，在导出前强制更新内容
                            function setupExportButton() {{
                                var exportBtn = document.getElementById('bill-ocr-export-btn') || 
                                               document.querySelector('button[id*="bill-ocr-export-btn"]') ||
                                               document.querySelector('button:contains("导出结果")');
                                
                                if (exportBtn) {{
                                    exportBtn.addEventListener('click', function(e) {{
                                        console.log('[DEBUG] 导出按钮被点击，强制更新内容...');
                                        // 立即更新内容，不延迟
                                        var table = document.querySelector('.ocr-result-table');
                                        if (table) {{
                                            var styleTag = document.querySelector('style');
                                            var styleContent = styleTag ? styleTag.outerHTML : '';
                                            // 获取编辑后的表格HTML（包含所有用户编辑的内容）
                                            var tableHtml = table.outerHTML;
                                            var fullContent = styleContent + '\\n' + tableHtml;
                                            
                                            console.log('[DEBUG] 获取到的表格HTML长度:', tableHtml.length);
                                            console.log('[DEBUG] 表格内容预览:', tableHtml.substring(0, 200));
                                            
                                            // 查找隐藏的Textbox - 使用多种方法
                                            var hiddenInput = null;
                                            
                                            // 方法1: 直接通过ID查找
                                            hiddenInput = document.getElementById('bill-ocr-result-html-edited');
                                            
                                            // 方法2: 通过ID包含关键字查找
                                            if (!hiddenInput) {{
                                                var inputs = document.querySelectorAll('input, textarea');
                                                for (var i = 0; i < inputs.length; i++) {{
                                                    if (inputs[i].id && inputs[i].id.includes('bill-ocr-result-html-edited')) {{
                                                        hiddenInput = inputs[i];
                                                        break;
                                                    }}
                                                }}
                                            }}
                                            
                                            // 方法3: 查找所有隐藏的输入框
                                            if (!hiddenInput) {{
                                                var allInputs = document.querySelectorAll('input[type="text"], textarea');
                                                for (var i = 0; i < allInputs.length; i++) {{
                                                    var input = allInputs[i];
                                                    // 检查是否是隐藏的组件
                                                    if ((input.style.display === 'none' || input.hidden || input.offsetParent === null) &&
                                                        input.id && input.id.includes('bill')) {{
                                                        hiddenInput = input;
                                                        break;
                                                    }}
                                                }}
                                            }}
                                            
                                            if (hiddenInput) {{
                                                console.log('[DEBUG] 找到隐藏Textbox，ID:', hiddenInput.id);
                                                hiddenInput.value = fullContent;
                                                
                                                // 触发所有可能的事件，确保Gradio捕获到变化
                                                var events = ['input', 'change', 'blur', 'keyup', 'focus'];
                                                events.forEach(function(eventType) {{
                                                    try {{
                                                        var event = new Event(eventType, {{ bubbles: true, cancelable: true }});
                                                        hiddenInput.dispatchEvent(event);
                                                    }} catch(err) {{
                                                        console.error('触发事件失败:', eventType, err);
                                                    }}
                                                }});
                                                
                                                // 也尝试直接设置属性
                                                if (hiddenInput.setAttribute) {{
                                                    hiddenInput.setAttribute('value', fullContent);
                                                }}
                                                
                                                console.log('[DEBUG] 导出前已强制更新，内容长度:', fullContent.length);
                                                console.log('[DEBUG] Textbox当前值长度:', hiddenInput.value.length);
                                            }} else {{
                                                console.error('[DEBUG] 导出前未找到隐藏Textbox，尝试所有输入框...');
                                                var allInputs = document.querySelectorAll('input, textarea');
                                                console.log('[DEBUG] 找到', allInputs.length, '个输入框');
                                                for (var i = 0; i < Math.min(allInputs.length, 10); i++) {{
                                                    console.log('  输入框', i, ':', allInputs[i].id, allInputs[i].name, allInputs[i].className);
                                                }}
                                            }}
                                        }} else {{
                                            console.error('[DEBUG] 未找到表格元素');
                                        }}
                                    }}, true); // 使用捕获阶段，确保先执行
                                }} else {{
                                    // 如果按钮还没加载，延迟重试
                                    setTimeout(setupExportButton, 500);
                                }}
                            }}
                            
                            // 延迟设置导出按钮监听器
                            setTimeout(setupExportButton, 1000);
                        }})();
                        </script>
                        """
                        self.last_ocr_html = styled_html
                        self.last_ocr_markdown = ""  # HTML模式下不生成Markdown
                        return styled_html
                    else:
                        # 如果解析失败，回退到Markdown处理
                        cleaned = self._sanitize_markdown(raw_result)
                        self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                        self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                        return f"🪪 票据OCR识别结果:\n\n{cleaned}"
                except Exception as e:
                    print(f"⚠️ HTML表格解析失败，回退到Markdown格式: {e}")
                    # 解析失败，回退到Markdown处理
                    cleaned = self._sanitize_markdown(raw_result)
                    self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                    self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                    return f"🪪 票据OCR识别结果:\n\n{cleaned}"
            else:
                # 否则按Markdown处理
                cleaned = self._sanitize_markdown(raw_result)
                self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
                self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
                return f"🪪 票据OCR识别结果:\n\n{cleaned}"
            
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def get_dict_from_html(self, html_content):
        import ast
        if not self.current_parsed_dict:
            prompts = (
                f"根据所给的html表格，生成一个python字典，使用嵌套字典体现表格中的层次关系\n"
                f"【输出要求】\n"
                f"- 只返回填充后的字典，不要返回任何其他说明文字\n"
                f"- 不新增或删除字段，不改变表头文案；多值对应同一个键则使用列表\n"
                f"- 禁止输出任何猜测或编造的内容\n"
                f"- 不要使用代码块标记符号（例如 ``` ）\n"
                f"{html_content}\n"
            )
            
            try:
                self._ensure_card_api_loaded()
                if self.card_api is None:
                    return "❌ 卡证OCR API未初始化"
            except:
                print("导出失败")

            res = self.card_api.general_prompt(prompts)
            print(res.get("result"))
            self.current_parsed_dict = ast.literal_eval(res.get("result"))
        
        return self.current_parsed_dict


    def load_model(self, progress=gr.Progress()):
        """加载模型"""
        if self.is_loaded:
            return "✅ 模型已经加载完成！", gr.update(interactive=True)

        if torch is None:
            return "❌ 模型加载失败: 未检测到PyTorch，请先安装。", gr.update(interactive=False)

        try:
            progress(0.1, desc="检查模型路径...")
            if not os.path.exists(self.model_path):
                return f"❌ 模型路径不存在: {self.model_path}", gr.update(interactive=False)

            progress(0.3, desc="加载模型...")
            self.model = Qwen3VLForConditionalGeneration.from_pretrained(
                self.model_path,
                dtype="auto",
                device_map="cuda",
                load_in_4bit=False,
            )

            progress(0.7, desc="加载处理器...")
            self.processor = AutoProcessor.from_pretrained(self.model_path)
            print("加载处理器")

            # 初始化文档知识库，使用qwen3-vl的文本编码器
            if DOC_KB_AVAILABLE and self.doc_kb_use_qwen3vl:
                try:
                    progress(0.85, desc="初始化文档知识库（使用Qwen3-VL文本编码器）...")
                    self.doc_knowledge_base = DocumentKnowledgeBase(
                        persist_directory="./doc_knowledge_base",
                        embedding_model=None,  # 不使用独立embedding模型
                        qwen3vl_model=self.model,
                        qwen3vl_processor=self.processor
                    )
                    print("✅ 文档知识库初始化成功（使用Qwen3-VL文本编码器）")
                except Exception as e:
                    print(f"⚠️ 文档知识库初始化失败: {e}")
                    import traceback
                    traceback.print_exc()
                    # 如果使用qwen3-vl失败，回退到使用独立embedding模型
                    try:
                        progress(0.85, desc="回退：初始化文档知识库（使用独立embedding模型）...")
                        self.doc_knowledge_base = DocumentKnowledgeBase(
                            persist_directory="./doc_knowledge_base",
                            embedding_model="sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
                        )
                        self.doc_kb_use_qwen3vl = False
                        print("✅ 文档知识库初始化成功（使用独立embedding模型）")
                    except Exception as e2:
                        print(f"⚠️ 文档知识库回退初始化也失败: {e2}")
                        self.doc_knowledge_base = None

            progress(1.0, desc="完成！")
            self.is_loaded = True

            return "✅ 模型加载成功！可以开始使用了。", gr.update(interactive=True)

        except Exception as e:
            return f"❌ 模型加载失败: {str(e)}", gr.update(interactive=False)

    def _prepare_user_message(self, image, prompt):
        prompt_clean = (prompt or "").strip()
        resolved_image = image if image is not None else self.last_image
        if resolved_image is None:
            raise ValueError("❌ 请上传图像！")
        if not prompt_clean:
            raise ValueError("❌ 请输入问题！")
        if image is not None:
            self.last_image = image
        content = [
            {"type": "image", "image": resolved_image},
            {"type": "text", "text": prompt_clean},
        ]
        return prompt_clean, {"role": "user", "content": content}

    def _run_inference(self,
                       image,
                       prompt,
                       max_tokens,
                       temperature,
                       top_p,
                       top_k,
                       repetition_penalty,
                       prepared=None):
        if prepared is None:
            prompt_clean, user_message = self._prepare_user_message(image, prompt)
        else:
            prompt_clean, user_message = prepared
        messages = self.chat_messages + [user_message]

        inputs = self.processor.apply_chat_template(
            messages,
            tokenize=True,
            add_generation_prompt=True,
            return_dict=True,
            return_tensors="pt"
        ).to(self.model.device)

        generation_kwargs = {
            "max_new_tokens": max_tokens,
            "temperature": temperature,
            "top_p": top_p,
            "top_k": top_k,
            "do_sample": True if temperature > 0 else False,
            "repetition_penalty": repetition_penalty
        }

        start_time = time.time()
        with torch.no_grad():
            generated_ids = self.model.generate(**inputs, **generation_kwargs)
        generation_time = time.time() - start_time

        generated_ids_trimmed = [
            out_ids[len(in_ids):] for in_ids, out_ids in zip(inputs.input_ids, generated_ids)
        ]
        output_text = self.processor.batch_decode(
            generated_ids_trimmed, skip_special_tokens=True, clean_up_tokenization_spaces=False
        )
        response = output_text[0]

        assistant_message = {"role": "assistant", "content": [{"type": "text", "text": response}]}
        self.chat_messages.extend([user_message, assistant_message])
        return prompt_clean, response, generation_time

    def _clone_history(self, history):
        return [[turn[0], turn[1]] for turn in history]

    def _chunk_response(self, text, chunk_size=80):
        if not text:
            return []
        return [text[i:i + chunk_size] for i in range(0, len(text), chunk_size)]

    def _parse_markdown_sections(self, markdown_text):
        """
        将 Markdown 文本拆分为 table/text 段，支持：
        - 管道表格（| a | b |）
        - HTML <table>（若存在）
        并在解析前对围栏代码块进行去围栏清洗，确保导出与渲染一致。
        """
        sections = []
        if not markdown_text:
            return sections

        # 1) 先去掉围栏，使得“代码块中的表格”也能被识别为可导出的内容
        cleaned_md = self._sanitize_markdown(markdown_text)

        # 2) 先尝试解析 HTML 表格（若模型输出了 <table>）
        html_tables = []
        try:
            from bs4 import BeautifulSoup  # 可选依赖
            soup = BeautifulSoup(cleaned_md, "html.parser")
            for t in soup.find_all("table"):
                headers = []
                header_row = t.find("tr")
                if header_row:
                    # 如果有 <th> 用 th；否则用首行的 td 作为 header
                    ths = header_row.find_all("th")
                    if ths:
                        headers = [th.get_text(strip=True) for th in ths]
                        data_rows = header_row.find_next_siblings("tr")
                    else:
                        tds = header_row.find_all("td")
                        headers = [td.get_text(strip=True) for td in tds]
                        data_rows = header_row.find_next_siblings("tr")
                rows = []
                for r in (data_rows or []):
                    cols = r.find_all(["td", "th"])
                    rows.append([c.get_text(strip=True) for c in cols])
                if headers or rows:
                    html_tables.append({"type": "table", "header": headers, "rows": rows})
        except Exception:
            # 如果 bs4 不在环境中，则略过 HTML 解析
            pass

        # 3) 解析管道表格
        lines = cleaned_md.splitlines()
        i = 0
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()

            # 管道表格判定：当前行和下一行构成 header + 分隔
            is_table = (
                stripped.startswith("|")
                and stripped.count("|") >= 2
                and i + 1 < len(lines)
                and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                and lines[i + 1].strip().startswith("|")
            )

            if is_table:
                header = [cell.strip() for cell in stripped.strip("|").split("|")]
                i += 2  # 跳过 header 与分隔线
                rows = []
                while i < len(lines):
                    row_line = lines[i].strip()
                    if not (row_line.startswith("|") and row_line.count("|") >= 2):
                        break
                    row = [cell.strip() for cell in row_line.strip("|").split("|")]
                    rows.append(row)
                    i += 1
                sections.append({"type": "table", "header": header, "rows": rows})
                continue

            # 普通文本块（直到遇到下一个表格或文件结束）
            text_block = []
            while i < len(lines):
                current = lines[i]
                stripped_current = current.strip()
                next_is_table = (
                    stripped_current.startswith("|")
                    and stripped_current.count("|") >= 2
                    and i + 1 < len(lines)
                    and set(lines[i + 1].replace("|", "").strip()) <= set("-: ")
                    and lines[i + 1].strip().startswith("|")
                )
                if next_is_table:
                    break
                text_block.append(current)
                i += 1
                # 保留空行，改善段落分隔的可读性
                if i < len(lines) and lines[i] == "":
                    text_block.append(lines[i])

            text_content = "\n".join(text_block).strip("\n")
            if text_content:
                sections.append({"type": "text", "text": text_content})

        # 4) 若存在 HTML 表，优先把 HTML 表也加入（放在解析结果前面，避免遗漏）
        if html_tables:
            # 将 HTML 表插在最前面（也可根据需要合并/去重）
            sections = html_tables + sections

        return sections

    @staticmethod
    def _text_to_html_block(text: str) -> str:
        if not text:
            return ""
        escaped = html.escape(text)
        replaced = escaped.replace("\n", "<br>")
        return f'<div class="ocr-text">{replaced}</div>'

    @staticmethod
    def _table_to_html_block(header, rows) -> str:
        header = header or []
        rows = rows or []
        thead = ""
        if header:
            header_cells = "".join(f"<th>{html.escape(str(cell))}</th>" for cell in header)
            thead = f"<thead><tr>{header_cells}</tr></thead>"
        body_rows = []
        for row in rows:
            row_cells = "".join(f"<td>{html.escape(str(cell))}</td>" for cell in row)
            body_rows.append(f"<tr>{row_cells}</tr>")
        tbody = f"<tbody>{''.join(body_rows)}</tbody>" if body_rows else "<tbody></tbody>"
        return f'<table class="ocr-table">{thead}{tbody}</table>'

    def _render_sections_as_html(self, markdown_text: str) -> str:
        if not markdown_text:
            return ""
        sections = self._parse_markdown_sections(markdown_text)
        if not sections:
            escaped = html.escape(markdown_text.strip())
            return f"<pre>{escaped}</pre>" if escaped else ""
        blocks = []
        for section in sections:
            if section.get("type") == "table":
                blocks.append(self._table_to_html_block(section.get("header"), section.get("rows")))
            elif section.get("type") == "text":
                blocks.append(self._text_to_html_block(section.get("text", "")))
        return '<div class="ocr-preview">' + "".join(blocks) + "</div>"

    def chat_with_image(self, image, text, history, max_tokens, temperature, top_p, top_k, repetition_penalty: float = 1.0, presence_penalty: float = 1.5):
        """与图像对话（流式反馈），支持知识库检索增强"""
        original_text = text

        if not self.is_loaded:
            yield history, original_text, "❌ 请先加载模型！"
            return

        # 知识库检索增强
        enhanced_text = text
        kb_context = ""
        if self.doc_knowledge_base and text and text.strip():
            try:
                # 确定搜索范围：如果有最近添加的文档，优先搜索该文档
                doc_ids = [self.last_doc_id] if self.last_doc_id else None
                
                # 从知识库检索相关信息
                search_results = self.doc_knowledge_base.search(
                    query=text.strip(),
                    top_k=3,
                    doc_ids=doc_ids
                )
                
                if search_results:
                    # 构建知识库上下文
                    kb_context_parts = []
                    kb_context_parts.append("\n\n【相关文档内容】\n")
                    for i, result in enumerate(search_results, 1):
                        score = result.get('score', 0.0)
                        result_text = result.get('text', '')
                        metadata = result.get('metadata', {})
                        filename = metadata.get('filename', '未知文档')
                        
                        # 限制每个结果的文本长度，避免上下文过长
                        if len(result_text) > 500:
                            result_text = result_text[:500] + "..."
                        
                        kb_context_parts.append(f"{i}. [来源: {filename}, 相似度: {score:.2f}]\n{result_text}\n")
                    
                    kb_context = "\n".join(kb_context_parts)
                    kb_context += "\n【提示】请基于以上相关文档内容回答用户的问题。如果文档中没有相关信息，请基于图片内容回答。\n"
                    
                    # 将知识库上下文添加到用户问题中
                    enhanced_text = f"{text}{kb_context}"
                    
                    print(f"\n📚 知识库检索结果（共 {len(search_results)} 条）:")
                    for i, result in enumerate(search_results, 1):
                        print(f"  {i}. {result.get('metadata', {}).get('filename', '未知')} (相似度: {result.get('score', 0.0):.3f})")
            except Exception as e:
                print(f"⚠️ 知识库检索失败: {e}")
                # 检索失败时使用原始文本
                enhanced_text = text

        try:
            prepared = self._prepare_user_message(image, enhanced_text)
        except ValueError as exc:
            yield history, original_text, str(exc)
            return

        prompt_clean, _ = prepared
        history_copy = self._clone_history(history)
        # 显示原始问题，不显示知识库上下文（避免界面混乱）
        history_copy.append([f"👤 {original_text}", "🤖 正在思考..."])
        yield self._clone_history(history_copy), original_text, "🤖 正在思考..."

        try:
            _, response, generation_time = self._run_inference(
                image,
                enhanced_text,  # 使用增强后的文本（包含知识库上下文）
                max_tokens,
                temperature,
                top_p,
                top_k,
                repetition_penalty,
                prepared=prepared
            )
        except Exception as e:
            history_copy[-1][1] = f"❌ 生成失败: {str(e)}"
            self.chat_history = self._clone_history(history_copy)
            yield self._clone_history(history_copy), original_text, f"❌ 错误: {str(e)}"
            return

        assembled = ""
        chunks = self._chunk_response(response)
        if not chunks:
            chunks = [""]
        for chunk in chunks:
            assembled += chunk
            history_copy[-1][1] = f"🤖 {assembled}▌"
            yield self._clone_history(history_copy), original_text, f"🤖 {assembled}▌"

        stats = (
            f"⏱️ 生成时间: {generation_time:.2f}秒 | 📝 生成长度: {len(response)}字符"
            f" | ⚙️ 最大长度: {max_tokens}"
        )
        if max_tokens > 1024:
            stats += " | ⏳ 提示: 较大的最大长度可能延长生成时间"
        if kb_context:
            stats += " | 📚 已使用知识库增强"
        history_copy[-1][1] = f"🤖 {response}"
        self.chat_history = self._clone_history(history_copy)
        yield self._clone_history(history_copy), original_text, stats

    def _sanitize_markdown(self, text: str) -> str:
        if not text:
            return ""
        s = text.strip()
        lines = s.splitlines()
        out = []
        in_fence = False
        for line in lines:
            stripped = line.strip()
            if stripped.startswith("```"):
                in_fence = not in_fence
                continue
            if not in_fence:
                out.append(line)
        cleaned = "\n".join(out).strip()
        return cleaned if cleaned else s

    def ocr_analysis(self, image, prompt: str = None):
        """OCR文字识别，可选自定义提示词"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "请识别并提取这张图片中的所有文字内容，尽量还原原本样式，并标注语言类型。"
            " 请确保所有带样式或表格内容使用Markdown表格表示。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1024,
                temperature=0.7,
                top_p=0.8,
                top_k=50,
                repetition_penalty=1.0
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"📝 OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ OCR识别失败: {str(e)}"

    def ocr_card(self, image, prompt: str = None):
        """卡证OCR识别：身份证/银行卡/驾驶证等结构化提取（使用本地模型）"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        
        # 使用与ocr_card_api相同的默认提示词
        default_prompt = (
                "你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                "\n"
                "任务要求如下：\n"
                "\n"
            "1. 识别卡证类型：只允许从以下类别中选择一种：\n"
            "   - 身份证 / 银行卡 / 驾驶证 / 护照 / 工牌 / 其他。\n"
            "   Markdown表格中添加“卡证类型”字段，并用类别选择赋值。\n"
            "   **重要**：如果识别为银行卡，必须严格遵守第3条银行卡特殊要求！\n"
            "\n"
            "2. 输出格式：\n"
            "   - 以Markdown表格形式输出所有识别出的关键字段及其对应的值。\n"
            "   - 若字段中包含“卡号”，请确保该字段的值仅包含数字。\n"
            "   - 不要使用代码块标记符号（例如 ``` ）。\n"
            "\n"
            "3. 银行卡特殊要求（必须严格遵守）：\n"
            "   如果识别的卡证类型是银行卡，必须在Markdown表格的最后额外添加一个字段：\n"
            "   - 字段名：卡面类型（必须添加，不可省略）。\n"
            "   - 基于图片库检索到的相似卡证结果，填充“卡面类型”字段。字段值规则如下：\n"
            "       ① 当出现任何不确定、模糊或不匹配情况时，“卡面类型”字段的值**必须且只能为“其他”**，不得填写相似图片名或其他文本。\n"
            "       ② 若识别出的“发卡行”字段的值与这些相似卡证文件名中`_`前面的银行名称相同，"
            "则“卡面类型”字段的值只能从相似卡证文件名中**严格选择一个**，格式为`银行名称_卡面类型`，去掉文件后缀名，如`中国银行_visa卡`。\n"
            "       ③ 禁止自定义、生成、猜测或编造新的卡面类型值。任何不存在基于图片库检索到的相似卡证文件名的值都视为错误。\n"
            "   **重要提醒**：银行卡的Markdown表格必须包含“卡面类型”字段，这是强制要求，不能省略！\n"
            "   - 如果不是银行卡，则不添加“卡面类型”字段。\n"
            "\n"
            "4. 输出限制：\n"
            "   - 最终输出只包含Markdown表格。\n"
            "   - 禁止输出任何其他文字或解释性内容。\n"
            "   - 如果是银行卡，表格中必须包含“卡面类型”字段，否则输出不完整。\n"
        )

        effective_prompt = (prompt or "").strip() or default_prompt

        # RAG检索（使用与API版本相同的逻辑）
        rag_results = []
        try:
            self._ensure_card_rag_loaded()
            if self.card_rag_store and getattr(self.card_rag_store, "image_embeddings", None):
                rag_results = self._rag_search_card(image, top_k=3)
        except Exception as e:
            print(f"⚠️ RAG检索失败: {str(e)}")
            rag_results = []

        # 在终端输出RAG相似度匹配结果（与API版本一致）
        if rag_results:
            print("\n" + "=" * 60)
            print("📊 RAG相似度匹配结果")
            print("=" * 60)
            print(f"找到 {len(rag_results)} 张相似图片：\n")
            for i, r in enumerate(rag_results, 1):
                filename = r.get("filename", "未知")
                similarity = r.get("similarity", 0.0)
                print(f"  {i}. {filename}")
                print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
            print("=" * 60 + "\n")
        else:
            print("\n⚠️ 未找到相似图片\n")

        # 构建增强提示词（使用与API版本相同的逻辑）
        enhanced_prompt = self._build_enhanced_prompt_card(
            base_prompt=default_prompt,
            rag_results=rag_results,
            custom_prompt=effective_prompt if (prompt or "").strip() else None
        )

        # 在终端输出发送给模型的完整prompt（与API版本一致）
        print("\n" + "=" * 80)
        print("📝 发送给模型的完整Prompt")
        print("=" * 80)
        print(enhanced_prompt)
        print("=" * 80 + "\n")

        try:
            # 使用本地模型进行推理
            prompt_clean, response, _ = self._run_inference(
                image,
                enhanced_prompt,
                max_tokens=1024,
                temperature=0.3,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 卡证OCR识别失败: {str(e)}"

    def ocr_card_api(self, image, prompt: str = None):
        """卡证OCR识别（API调用 + RAG增强）"""
        # 注：如无需强制本地模型加载，可移除此判断
        try:
            self._ensure_card_api_loaded()
            if self.card_api is None:
                return "�?卡证OCR 初始化失败"
            default_prompt = (
                "你是专业的卡证OCR引擎，请对输入图片进行结构化识别，并仅输出Markdown表格。\n"
                "\n"
                "任务要求如下：\n"
                "\n"
            "1. 识别卡证类型：只允许从以下类别中选择一种：\n"
            "   - 身份证 / 银行卡 / 驾驶证 / 护照 / 工牌 / 其他。\n"
            "   Markdown表格中添加“卡证类型”字段，并用类别选择赋值。\n"
            "   **重要**：如果识别为银行卡，必须严格遵守第3条银行卡特殊要求！\n"
            "\n"
            "2. 输出格式：\n"
            "   - 以Markdown表格形式输出所有识别出的关键字段及其对应的值。\n"
            "   - 若字段中包含“卡号”，请确保该字段的值仅包含数字。\n"
            "   - 不要使用代码块标记符号（例如 ``` ）。\n"
            "\n"
            "3. 银行卡特殊要求（必须严格遵守）：\n"
            "   如果识别的卡证类型是银行卡，必须在Markdown表格的最后额外添加一个字段：\n"
            "   - 字段名：卡面类型（必须添加，不可省略）。\n"
            "   - 基于图片库检索到的相似卡证结果，填充“卡面类型”字段。字段值规则如下：\n"
            "       ① 当出现任何不确定、模糊或不匹配情况时，“卡面类型”字段的值**必须且只能为“其他”**，不得填写相似图片名或其他文本。\n"
            "       ② 若识别出的“发卡行”字段的值与这些相似卡证文件名中`_`前面的银行名称相同，"
            "则“卡面类型”字段的值只能从相似卡证文件名中**严格选择一个**，格式为`银行名称_卡面类型`，去掉文件后缀名，如`中国银行_visa卡`。\n"
            "       ③ 禁止自定义、生成、猜测或编造新的卡面类型值。任何不存在基于图片库检索到的相似卡证文件名的值都视为错误。\n"
            "   **重要提醒**：银行卡的Markdown表格必须包含“卡面类型”字段，这是强制要求，不能省略！\n"
            "   - 如果不是银行卡，则不添加“卡面类型”字段。\n"
            "\n"
            "4. 输出限制：\n"
            "   - 最终输出只包含Markdown表格。\n"
            "   - 禁止输出任何其他文字或解释性内容。\n"
            "   - 如果是银行卡，表格中必须包含“卡面类型”字段，否则输出不完整。\n"
            )

            effective_prompt = (prompt or "").strip() or default_prompt
            result = self.card_api.recognize_card(
                image,
                custom_prompt=effective_prompt,
                use_rag=True,
            )
            if not result.get("success"):
                return f"�?卡证OCR 调用失败: {result.get('error') or '未知错误'}"

            # 在终端输出RAG相似度匹配结果
            rag_info = result.get("rag_info")
            if rag_info and rag_info.get("enabled") and rag_info.get("results"):
                print("\n" + "=" * 60)
                print("📊 RAG相似度匹配结果")
                print("=" * 60)
                print(f"找到 {len(rag_info['results'])} 张相似图片：\n")
                for i, r in enumerate(rag_info["results"], 1):
                    filename = r.get("filename", "未知")
                    similarity = r.get("similarity", 0.0)
                    print(f"  {i}. {filename}")
                    print(f"     相似度: {similarity:.4f} ({similarity*100:.2f}%)")
                print("=" * 60 + "\n")
            elif rag_info and not rag_info.get("enabled"):
                print(f"\n⚠️ RAG未启用: {rag_info.get('reason', '未知原因')}\n")
            else:
                print("\n⚠️ 未找到相似图片\n")

            cleaned = self._sanitize_markdown(result.get("result") or "")
            self.last_ocr_markdown = f"## 卡证OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>卡证OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🪪 卡证OCR识别结果:\n\n{cleaned}"
        except Exception as e:
            return f"�?卡证OCR识别失败: {str(e)}"

    def ocr_receipt(self, image, prompt: str = None):
        """票据OCR识别：发票/小票等表格与关键项解析"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "你是发票/小票OCR专家。请解析图片中的票据并输出：\n"
            "- 以Markdown表格给出关键信息：票据类型、开票日期、发票代码、发票号码、校验码、购买方、销售方、税号、项目、数量、单价、金额、税率、税额、合计金额(含税/不含税)；\n"
            "- 若检测到多行项目，请以表格形式逐行列出；\n"
            "- 表格下方给出识别置信度与可疑项提示；\n"
            "- 不要使用围栏代码块，保持Markdown可渲染。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=1536,
                temperature=0.2,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 票据OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>票据OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"🧾 票据OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 票据OCR识别失败: {str(e)}"

    def ocr_agreement(self, image, prompt: str = None):
        """协议OCR识别：合同/协议段落与条款解析"""
        if not self.is_loaded:
            return "❌ 请先加载模型！"
        default_prompt = (
            "你是合同/协议OCR与条款解析助手。请完成：\n"
            "1) 识别全文，保持段落结构；\n"
            "2) 以Markdown表格提炼关键信息：合同名称、甲方、乙方、签署日期、生效日期、终止日期、金额/币种、违约条款、争议解决、签章情况；\n"
            "3) 如有编号的条款，保留编号并逐条列出；\n"
            "4) 在末尾给出“风险提示”列表（如空白处、涂改处、关键要素缺失等）；\n"
            "5) 不要输出围栏代码块。"
        )
        effective_prompt = (prompt or "").strip() or default_prompt
        try:
            prompt_clean, response, _ = self._run_inference(
                image,
                effective_prompt,
                max_tokens=2048,
                temperature=0.3,
                top_p=0.8,
                top_k=40,
                repetition_penalty=1.05
            )
            cleaned = self._sanitize_markdown(response)
            self.chat_history.append([f"👤 {prompt_clean}", f"🤖 {cleaned}"])
            self.last_ocr_markdown = f"## 协议OCR识别结果\n\n{cleaned}"
            self.last_ocr_html = "<h2>协议OCR识别结果</h2>" + self._render_sections_as_html(cleaned)
            return f"📄 协议OCR识别结果:\n\n{cleaned}"
        except ValueError as exc:
            return str(exc)
        except Exception as e:
            return f"❌ 协议OCR识别失败: {str(e)}"

    def _format_ocr_result(self, result):
        """
        格式化 OCR 结果为纯文本字符串（只提取 rec_texts 内容，不包含坐标等信息）
        
        Args:
            result: OCR 结果，可能是字符串、字典或其他类型
            
        Returns:
            格式化后的纯文本字符串（只包含识别到的文本内容）
        """
        import json
        if result is None:
            return ""
        
        if isinstance(result, str):
            # 如果是字符串，直接返回
            return result
        elif isinstance(result, dict):
            # 如果是字典，优先提取 rec_texts（PaddleOCR API 返回的文本字段）
            if "rec_texts" in result:
                rec_texts = result["rec_texts"]
                if isinstance(rec_texts, list):
                    # 如果是列表，合并所有文本行
                    return "\n".join(str(text) for text in rec_texts if text)
                else:
                    return str(rec_texts)
            # 其次查找其他常见的文本字段
            elif "text" in result:
                text = result["text"]
                if isinstance(text, list):
                    return "\n".join(str(t) for t in text if t)
                return str(text)
            elif "content" in result:
                content = result["content"]
                if isinstance(content, list):
                    return "\n".join(str(c) for c in content if c)
                return str(content)
            elif "prunedResult" in result:
                # 递归处理 prunedResult，提取其中的 rec_texts
                return self._format_ocr_result(result["prunedResult"])
            elif "result" in result:
                return self._format_ocr_result(result["result"])
            else:
                # 如果字典中没有明确的文本字段，返回空字符串（不返回坐标等元数据）
                return ""
        elif isinstance(result, list):
            # 如果是列表，提取每个元素的文本内容
            text_parts = []
            for item in result:
                text = self._format_ocr_result(item)
                if text:  # 只添加非空文本
                    text_parts.append(text)
            return "\n".join(text_parts)
        else:
            # 其他类型直接转换为字符串
            return str(result)

    def _chunk_text_for_rag(self, text, chunk_size=500, overlap=50, max_chunks=1000):
        """
        对文本进行切片，用于RAG相似度计算
        
        Args:
            text: 要切片的文本
            chunk_size: 每个切片的最大字符数
            overlap: 切片之间的重叠字符数
            max_chunks: 最大切片数量（防止内存溢出）
            
        Returns:
            文本切片列表
        """
        if not text:
            return []
        
        # 安全检查：确保 overlap < chunk_size，防止无限循环
        if overlap >= chunk_size:
            overlap = max(1, chunk_size // 10)  # 默认使用10%的重叠
        
        chunks = []
        start = 0
        text_length = len(text)
        prev_start = -1  # 用于检测是否卡住
        
        # 如果文本太大，先限制处理范围
        if text_length > 1000000:  # 如果超过1MB，只处理前1MB
            print(f"⚠️ 文本过大（{text_length}字符），仅处理前1MB用于切片")
            text = text[:1000000]
            text_length = len(text)
        
        while start < text_length and len(chunks) < max_chunks:
            end = min(start + chunk_size, text_length)
            chunk = text[start:end].strip()
            
            if chunk:
                chunks.append(chunk)
            
            # 计算下一个起始位置
            next_start = end - overlap
            # 安全检查：确保向前移动
            if next_start <= start:
                next_start = start + 1  # 至少移动1个字符
            
            # 防止卡在同一个位置
            if next_start == prev_start:
                next_start = start + chunk_size  # 强制移动
            
            prev_start = start
            start = next_start
        
        if len(chunks) >= max_chunks:
            print(f"⚠️ 切片数量达到上限（{max_chunks}），已停止切片")
        
        return chunks
    
    def _calculate_text_similarity(self, query, text):
        """
        计算查询文本与目标文本的相似度（使用简单的词重叠度）
        
        Args:
            query: 查询文本（关键字段）
            text: 目标文本（切片内容）
            
        Returns:
            相似度分数（0-1之间）
        """
        if not query or not text:
            return 0.0
        
        # 转换为小写并分词
        query_words = set(query.lower().split())
        text_words = set(text.lower().split())
        
        if not query_words:
            return 0.0
        
        # 计算交集和并集
        intersection = query_words & text_words
        union = query_words | text_words
        
        # Jaccard相似度
        if not union:
            return 0.0
        
        jaccard = len(intersection) / len(union)
        
        # 同时考虑查询词在文本中的出现频率
        text_lower = text.lower()
        query_lower = query.lower()
        if query_lower in text_lower:
            # 如果查询文本完全包含在目标文本中，增加相似度
            jaccard = max(jaccard, 0.5)
        
        return jaccard
    
    def _rag_search_text_chunks(self, key_fields, text_chunks, top_k=3):
        """
        根据关键字段对文本切片进行RAG相似度映射，返回相似度最高的top_k个切片
        
        Args:
            key_fields: 关键字段列表（用户自定义）
            text_chunks: 文本切片列表
            top_k: 返回最相似的切片数量
            
        Returns:
            每个关键字段对应的最相似切片列表
        """
        if not key_fields or not text_chunks:
            return {}
        
        field_chunks = {}
        
        for field in key_fields:
            if not field or not field.strip():
                continue
            
            field = field.strip()
            similarities = []
            
            # 计算每个切片与关键字段的相似度
            for idx, chunk in enumerate(text_chunks):
                similarity = self._calculate_text_similarity(field, chunk)
                similarities.append({
                    'chunk_index': idx,
                    'chunk_text': chunk,
                    'similarity': similarity
                })
            
            # 按相似度排序，取top_k
            similarities.sort(key=lambda x: x['similarity'], reverse=True)
            top_chunks = similarities[:top_k]
            
            field_chunks[field] = top_chunks
        
        return field_chunks

    def _pdf_to_images(self, pdf_path_or_bytes):
        """
        将PDF转换为图像列表
        
        Args:
            pdf_path_or_bytes: PDF文件路径或字节数据
            
        Returns:
            图像列表（PIL Image对象）
        """
        if not PDF_AVAILABLE:
            return None
        
        # 优先尝试 PyMuPDF（不需要外部依赖）
        if PYMUPDF_AVAILABLE:
            try:
                import fitz
                from PIL import Image as PILImage
                if isinstance(pdf_path_or_bytes, bytes):
                    pdf_doc = fitz.open(stream=pdf_path_or_bytes, filetype="pdf")
                else:
                    pdf_doc = fitz.open(pdf_path_or_bytes)
                
                images = []
                for page_num in range(len(pdf_doc)):
                    page = pdf_doc[page_num]
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x缩放提高清晰度
                    img = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
                    images.append(img)
                pdf_doc.close()
                return images
            except Exception as e:
                print(f"⚠️ PyMuPDF转换失败: {e}，尝试使用pdf2image...")
                # 如果 PyMuPDF 失败，尝试 pdf2image
        
        # 回退到 pdf2image（需要 poppler）
        if PDF2IMAGE_AVAILABLE:
            try:
                if isinstance(pdf_path_or_bytes, bytes):
                    images = convert_from_bytes(pdf_path_or_bytes)
                else:
                    images = convert_from_path(pdf_path_or_bytes)
                return images
            except Exception as e:
                error_msg = str(e)
                if "poppler" in error_msg.lower():
                    print(f"⚠️ PDF转换失败: {error_msg}")
                    print("💡 提示: pdf2image需要poppler，建议安装PyMuPDF: pip install PyMuPDF")
                else:
                    print(f"⚠️ PDF转换失败: {e}")
                return None
        
        return None
    

    PADDLEOCR_API_URL = "https://wdc9jbw9l1f8996b.aistudio-app.com/ocr"
    PADDLEOCR_TOKEN = "61236296494fb5e32ee89aef50d4d6aa99fa2ba7"

    def _parse_pdf_page_spec(self, pages_spec: str, total_pages: int):
        """
        解析用户输入的页码字符串，支持格式：
        - 逗号分隔页码：1,3,5
        - 区间：2-4
        - 混合输入与空格
        返回去重且排序后的页码列表（1-based）
        """
        if not pages_spec or not pages_spec.strip():
            return []

        normalized = (
            pages_spec.replace("，", ",")
            .replace("、", ",")
            .replace(";", ",")
            .strip()
        )
        if not normalized:
            return []

        tokens = re.split(r"[,\s]+", normalized)
        selected_pages = set()

        for token in tokens:
            if not token:
                continue
            token = token.strip()
            if not token:
                continue
            if "-" in token:
                parts = token.split("-", 1)
                if len(parts) != 2:
                    continue
                try:
                    start = int(parts[0])
                    end = int(parts[1])
                except ValueError:
                    continue
                if start > end:
                    start, end = end, start
                for page in range(start, end + 1):
                    if 1 <= page <= total_pages:
                        selected_pages.add(page)
            else:
                try:
                    page = int(token)
                except ValueError:
                    continue
                if 1 <= page <= total_pages:
                    selected_pages.add(page)

        return sorted(selected_pages)

    def _filter_pdf_pages(self, pdf_path: str, pages_spec: str):
        """
        根据页码过滤PDF页面，只返回包含指定页的新PDF路径。
        如果解析失败或依赖缺失则返回None。
        """
        if not os.path.exists(pdf_path):
            return None

        pages_spec = pages_spec or ""
        if pages_spec.strip().lower() == "all":
            return None

        # 优先使用 PyMuPDF，失败后再尝试 PyPDF2
        if PYMUPDF_AVAILABLE:
            try:
                import fitz

                src_doc = fitz.open(pdf_path)
                total_pages = src_doc.page_count
                selected_pages = self._parse_pdf_page_spec(pages_spec, total_pages)

                if not selected_pages:
                    print("⚠️ PDF页码解析为空，继续识别全部页面")
                    return None

                filtered_doc = fitz.open()
                for page_num in selected_pages:
                    filtered_doc.insert_pdf(
                        src_doc, from_page=page_num - 1, to_page=page_num - 1
                    )

                temp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
                temp_pdf_path = temp_pdf.name
                temp_pdf.close()
                filtered_doc.save(temp_pdf_path)
                filtered_doc.close()
                src_doc.close()
                print(f"✅ 使用PyMuPDF 过滤PDF页面: {selected_pages}")
                return temp_pdf_path
            except Exception as e:
                print(f"⚠️ 使用PyMuPDF过滤PDF失败: {e}")

        try:
            from PyPDF2 import PdfReader, PdfWriter
        except ImportError:
            print("⚠️ 未安装PyPDF2，无法按页裁剪PDF，继续识别全部页面")
            return None

        try:
            reader = PdfReader(pdf_path)
            total_pages = len(reader.pages)
            selected_pages = self._parse_pdf_page_spec(pages_spec, total_pages)

            if not selected_pages:
                print("⚠️ PDF页码解析为空，继续识别全部页面")
                return None

            writer = PdfWriter()
            for page_num in selected_pages:
                writer.add_page(reader.pages[page_num - 1])

            temp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            temp_pdf_path = temp_pdf.name
            temp_pdf.close()
            with open(temp_pdf_path, "wb") as output_pdf:
                writer.write(output_pdf)

            print(f"✅ 使用PyPDF2 过滤PDF页面: {selected_pages}")
            return temp_pdf_path
        except Exception as e:
            print(f"⚠️ 使用PyPDF2过滤PDF失败: {e}")
            return None


    def ocr_document(self, image_or_file, prompt: str = None, is_pdf: bool = False, pdf_pages: str = "all", engine: str = "paddle"):
        """
        文档OCR识别：使用 PaddleOCR API 直接调用
        提取到的文字保存下来准备进行切片做RAG知识库
        
        Args:
            image_or_file: 图像（PIL Image）或文件路径/字节数据
            prompt: 自定义提示词（暂不使用，保留兼容性）
            is_pdf: 是否为PDF文件
            pdf_pages: PDF页码，如"all"表示所有页，"1,3,5"表示指定页
            engine: OCR引擎，支持 "paddle" 或 "qwen3"
        """
        import os
        import base64
        import requests
        import json
        import time
        from pathlib import Path
        from PIL import Image
        import tempfile

        engine = (engine or "paddle").lower()
        if engine.startswith("qwen"):
            return self._ocr_document_qwen3(
                image_or_file=image_or_file,
                prompt=prompt,
                pdf_pages=pdf_pages
            )
        
        try:
            # 记录临时文件列表，便于统一清理
            temp_files = []

            # 确定文件路径或处理字节数据
            file_path = None
            
            if isinstance(image_or_file, Image.Image):
                # 如果是 PIL Image，需要先保存为临时文件
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    image_or_file.save(tmp.name, "PNG")
                    file_path = tmp.name
                    temp_files.append(tmp.name)
            elif isinstance(image_or_file, str):
                # 如果是文件路径
                file_path = image_or_file
            elif isinstance(image_or_file, bytes):
                # 如果是字节数据（PDF或图片的字节流）
                # 根据 is_pdf 参数确定文件扩展名
                suffix = ".pdf" if is_pdf else ".png"
                with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                    tmp.write(image_or_file)
                    file_path = tmp.name
                    temp_files.append(tmp.name)
            else:
                return "❌ 不支持的文件类型（需要 PIL Image、文件路径或字节数据）"
            
            if not file_path or not os.path.exists(file_path):
                return f"❌ 文件不存在: {file_path}"

            # 如果是PDF且指定了页码，则尝试裁剪PDF
            if (
                (is_pdf or file_path.lower().endswith(".pdf"))
                and pdf_pages
                and pdf_pages.strip()
                and pdf_pages.strip().lower() != "all"
            ):
                filtered_pdf = self._filter_pdf_pages(file_path, pdf_pages)
                if filtered_pdf:
                    file_path = filtered_pdf
                    temp_files.append(filtered_pdf)
                else:
                    print("⚠️ 未能按页码裁剪PDF，继续识别全部页面")
            
            # 创建输出目录（用于保存OCR结果，准备RAG切片）
            output_dir = "ocr_output"
            os.makedirs(output_dir, exist_ok=True)
            
            # 读取文件并编码
            print(f"正在读取文件: {file_path}")
            with open(file_path, "rb") as file:
                file_bytes = file.read()
            file_data = base64.b64encode(file_bytes).decode("ascii")
            
            file_size_mb = len(file_bytes) / (1024 * 1024)
            print(f"文件大小: {file_size_mb:.2f} MB")
            
            # 准备请求
            headers = {
                "Authorization": f"token {self.PADDLEOCR_TOKEN}",
                "Content-Type": "application/json"
            }
            
            # 根据文件类型设置 fileType
            if is_pdf or file_path.lower().endswith('.pdf'):
                file_type = 0  # PDF
            else:
                file_type = 1  # 图片
            
            payload = {
                "file": file_data,
                "fileType": file_type,
                "useDocOrientationClassify": False,
                "useDocUnwarping": False,
                "useTextlineOrientation": False,
            }
            
            # 发送请求
            print("正在进行PaddleOCR识别...")
            start_time = time.time()
            
            # 根据文件大小设置超时
            timeout = max(600, int(file_size_mb * 60))
            
            response = requests.post(self.PADDLEOCR_API_URL, json=payload, headers=headers, timeout=timeout)
            
            elapsed_time = time.time() - start_time
            print(f"响应时间: {elapsed_time:.2f}秒")
            
            if response.status_code != 200:
                return f"❌ 请求失败，状态码: {response.status_code}\n响应: {response.text[:500]}"
            
            result = response.json()
            
            if "result" not in result:
                return f"❌ 响应格式错误: {result}"
            
            ocr_result = result["result"]
            ocr_results = ocr_result.get("ocrResults", [])
            
            if not ocr_results:
                return "❌ 未识别到任何内容"
            
            # 获取输入文件名（不含扩展名）
            input_filename = os.path.splitext(os.path.basename(file_path))[0]
            
            # 收集所有页面的文本内容（用于RAG切片）
            all_text_parts = []
            all_markdown_parts = []
            
            # 保存文本结果（合并所有页面）
            txt_file = os.path.join(output_dir, f"{input_filename}_ocr.txt")
            md_file = os.path.join(output_dir, f"{input_filename}_ocr.md")
            json_file = os.path.join(output_dir, f"{input_filename}_ocr.json")
            
            with open(txt_file, "w", encoding="utf-8") as f:
                for i, res in enumerate(ocr_results):
                    pruned_result = res.get("prunedResult", "")
                    formatted_result = self._format_ocr_result(pruned_result)
                    all_text_parts.append(formatted_result)
                    
                    f.write(f"\n{'='*60}\n")
                    f.write(f"第 {i + 1} 页\n")
                    f.write(f"{'='*60}\n\n")
                    f.write(formatted_result)
                    f.write("\n\n")
            
            # 保存 Markdown 格式
            with open(md_file, "w", encoding="utf-8") as f:
                f.write(f"# {input_filename} OCR 结果\n\n")
                for i, res in enumerate(ocr_results):
                    pruned_result = res.get("prunedResult", "")
                    formatted_result = self._format_ocr_result(pruned_result)
                    all_markdown_parts.append(formatted_result)
                    
                    f.write(f"## 第 {i + 1} 页\n\n")
                    if "\n" in formatted_result:
                        f.write("```\n")
                        f.write(formatted_result)
                        f.write("\n```\n")
                    else:
                        f.write(formatted_result)
                    f.write("\n\n---\n\n")
            
            # 保存 JSON 结果
            with open(json_file, "w", encoding="utf-8") as f:
                json.dump(ocr_result, f, ensure_ascii=False, indent=2)
            
            # 下载并保存图片（如果需要）
            saved_images = 0
            for i, res in enumerate(ocr_results):
                image_url = res.get("ocrImage")
                if image_url:
                    try:
                        img_response = requests.get(image_url, timeout=30)
                        if img_response.status_code == 200:
                            img_filename = os.path.join(output_dir, f"{input_filename}_page_{i + 1}.jpg")
                            with open(img_filename, "wb") as f:
                                f.write(img_response.content)
                            saved_images += 1
                    except:
                        pass
            
            # 合并所有文本（用于RAG切片）
            full_text = "\n\n".join(all_text_parts)
            
            # 保存每页的rec_texts文本（用于分页显示和RAG处理）
            page_texts = []  # 每页的文本列表
            for i, res in enumerate(ocr_results):
                pruned_result = res.get("prunedResult", {})
                if isinstance(pruned_result, dict) and "rec_texts" in pruned_result:
                    rec_texts = pruned_result["rec_texts"]
                    if isinstance(rec_texts, list):
                        page_text = "\n".join(str(text) for text in rec_texts if text)
                    else:
                        page_text = str(rec_texts)
                else:
                    page_text = all_text_parts[i] if i < len(all_text_parts) else ""
                page_texts.append(page_text)
            
            # 先构建并返回结果（不等待文本切片）
            result_markdown = f"## 文档OCR识别结果\n\n"
            result_markdown += f"### 📄 文件信息\n\n"
            result_markdown += f"**文件名：** {input_filename}\n"
            result_markdown += f"**文件类型：** {'PDF' if file_type == 0 else '图片'}\n"
            result_markdown += f"**识别页数：** {len(ocr_results)}\n"
            result_markdown += f"**处理时间：** {elapsed_time:.2f}秒\n\n"
            
            result_markdown += f"### 📝 识别文本（第1页，共{len(page_texts)}页）\n\n"
            if page_texts:
                result_markdown += f"{page_texts[0][:1000]}{'...' if len(page_texts[0]) > 1000 else ''}\n\n"
            
            result_markdown += f"### 💾 保存的文件\n\n"
            result_markdown += f"- **文本文件：** `{txt_file}`\n"
            result_markdown += f"- **Markdown文件：** `{md_file}`\n"
            result_markdown += f"- **JSON文件：** `{json_file}`\n"
            if saved_images > 0:
                result_markdown += f"- **OCR图片：** {saved_images} 张\n"
            result_markdown += f"\n### 📚 知识库\n\n"
            kb_status = f"⏳ 正在对文本进行切片处理，完成后将自动保存到知识库。总字符数：{len(full_text)}\n"
            result_markdown += kb_status
            
            # 先保存基础数据到实例变量
            self.last_ocr_text = full_text
            self.last_ocr_page_texts = page_texts  # 保存每页文本
            self.last_ocr_markdown = result_markdown
            self.last_ocr_output_dir = output_dir
            self.last_ocr_files = {
                'txt': txt_file,
                'md': md_file,
                'json': json_file
            }
            
            self.last_ocr_html = "<h2>文档OCR识别结果</h2>" + self._render_sections_as_html(result_markdown)
            
            # 在后台进行文本切片（不阻塞返回）
            try:
                # 对于大文本，使用更大的切片大小以减少切片数量
                chunk_size = 1000 if len(full_text) > 100000 else 500
                text_chunks = self._chunk_text_for_rag(full_text, chunk_size=chunk_size, max_chunks=500)
                print(f"✅ 文本切片完成，共 {len(text_chunks)} 个切片")
                # 更新实例变量
                self.last_ocr_text_chunks = text_chunks
                
                # 保存到知识库
                if self.doc_knowledge_base and full_text:
                    try:
                        doc_id = self.doc_knowledge_base.add_document(
                            text=full_text,
                            filename=input_filename,
                            chunks=text_chunks,
                            metadata={
                                "file_type": "PDF" if file_type == 0 else "图片",
                                "page_count": len(ocr_results),
                                "ocr_time": elapsed_time
                            }
                        )
                        self.last_doc_id = doc_id
                        print(f"✅ 文档已保存到知识库，ID: {doc_id}")
                    except Exception as e:
                        print(f"⚠️ 保存到知识库失败: {e}")
            except MemoryError as e:
                print(f"⚠️ 内存不足，使用简化切片策略: {e}")
                # 如果内存不足，使用更大的切片，减少切片数量
                try:
                    text_chunks = self._chunk_text_for_rag(full_text, chunk_size=2000, overlap=100, max_chunks=200)
                    self.last_ocr_text_chunks = text_chunks
                    # 保存到知识库
                    if self.doc_knowledge_base and full_text:
                        try:
                            doc_id = self.doc_knowledge_base.add_document(
                            text=full_text,
                            filename=input_filename,
                            chunks=text_chunks
                            )
                            self.last_doc_id = doc_id
                            print(f"✅ 文档已保存到知识库，ID: {doc_id}")
                        except Exception as e:
                            print(f"⚠️ 保存到知识库失败: {e}")
                except:
                    # 如果还是失败，至少保存完整文本作为一个大切片
                    self.last_ocr_text_chunks = [full_text[:50000]] if len(full_text) > 50000 else [full_text]
            except Exception as e:
                print(f"⚠️ 文本切片失败: {e}")
                # 如果切片失败，至少保存完整文本作为一个大切片
                self.last_ocr_text_chunks = [full_text[:50000]] if len(full_text) > 50000 else [full_text]
            
            return result_markdown
            
        except requests.exceptions.Timeout:
            return f"❌ 请求超时（超过 {timeout} 秒）"
        except requests.exceptions.RequestException as e:
            return f"❌ 请求失败: {e}"
        except Exception as e:
            import traceback
            return f"❌ 处理过程中出现错误: {e}\n{traceback.format_exc()}"
        finally:
            for tmp_path in temp_files:
                try:
                    if tmp_path and os.path.exists(tmp_path):
                        os.unlink(tmp_path)
                except:
                    pass
    
    def extract_document_fields_with_rag(self, key_fields, custom_prompt=None):
        """
        识别关键字段：根据关键字段对文本切片进行相似度映射，
        将相似度最高的三段文本提供给大模型进行信息抽取
        
        Args:
            key_fields: 关键字段列表（用户自定义）
            custom_prompt: 自定义提示词
            
        Returns:
            信息抽取结果（表格形式，一个字段一行）
        """
        if not hasattr(self, 'last_ocr_text') or not self.last_ocr_text:
            return "❌ 请先进行文档OCR识别"
        
        if not key_fields or not isinstance(key_fields, list):
            return "❌ 请提供关键字段列表"
        
        # 过滤空字段
        key_fields = [f.strip() for f in key_fields if f and f.strip()]
        if not key_fields:
            return "❌ 关键字段列表为空"
        
        try:
            # 对每个关键字段进行RAG相似度映射（如果文本切片可用）
            rag_context_parts = []
            
            if hasattr(self, 'last_ocr_text_chunks') and self.last_ocr_text_chunks:
                print(f"🔍 开始RAG相似度匹配，共 {len(key_fields)} 个关键字段，{len(self.last_ocr_text_chunks)} 个文本切片")
                
                # 对每个关键字段进行RAG相似度匹配
                field_chunks = self._rag_search_text_chunks(
                    key_fields, 
                    self.last_ocr_text_chunks, 
                    top_k=3
                )
                
                # 为每个字段组织相关文本切片
                for field in key_fields:
                    if field in field_chunks:
                        chunks = field_chunks[field]
                        if chunks:
                            # 获取相似度最高的3个切片
                            top_chunks = chunks[:3]
                            field_texts = []
                            for i, chunk_info in enumerate(top_chunks, 1):
                                similarity = chunk_info['similarity']
                                chunk_text = chunk_info['chunk_text']
                                field_texts.append(f"[相似度: {similarity:.2%}] {chunk_text}")
                            
                            # 为每个字段组织相关文本
                            field_context = f"**关键字段「{field}」的相关文本片段（相似度最高的3段）：**\n"
                            field_context += "\n".join(field_texts)
                            rag_context_parts.append(field_context)
                            print(f"  ✅ {field}: 找到 {len(chunks)} 个相关片段")
                        else:
                            print(f"  ⚠️ {field}: 未找到相关片段")
                    else:
                        print(f"  ⚠️ {field}: 未找到相关片段")
                
                # 合并所有字段的相关文本
                if rag_context_parts:
                    rag_context = "\n\n".join(rag_context_parts)
                else:
                    # 如果没有找到相关片段，使用全文的前5000字符
                    rag_context = f"**文档全文（前5000字符）：**\n{self.last_ocr_text[:5000]}"
            else:
                # 如果没有切片，使用全文的前5000字符
                print("⚠️ 文本切片不可用，使用全文前5000字符")
                rag_context = f"**文档全文（前5000字符）：**\n{self.last_ocr_text[:5000]}"
            
            # 构建提示词（要求大模型提取所有字段）
            fields_list = "、".join([f"「{f}」" for f in key_fields])
            default_prompt = (
                f"你是专业的文档信息抽取专家。请从以下文档文本中提取以下关键字段的信息：{fields_list}\n\n"
                f"**文档相关内容：**\n{rag_context}\n\n"
                "**任务要求：**\n"
                "1. 仔细阅读上述文档内容，特别是每个关键字段的相关文本片段\n"
                "2. 从文档中提取每个关键字段的值\n"
                "3. 只输出Markdown表格，格式如下：\n"
                "| 字段名 | 字段值 |\n"
                "|--------|--------|\n"
                "4. 每个关键字段必须占一行，字段名和字段值都要填写\n"
                "5. 如果某个字段在文档中找不到，字段值填写\"未找到\"\n"
                "6. 只输出表格，不要输出其他说明文字、代码块标记或其他内容\n"
                "7. 确保提取所有字段，不要遗漏\n"
            )
            
            effective_prompt = custom_prompt if custom_prompt and custom_prompt.strip() else default_prompt
            
            # 使用qwen3-vl-plus API进行信息抽取
            self._ensure_doc_api_loaded()
            if self.doc_api is None or not self.doc_api.is_loaded:
                return "❌ 文档OCR未加载"
            
            # 尝试获取第一页图像（如果有保存的OCR图像）
            image_for_api = None
            if hasattr(self, 'last_ocr_files') and self.last_ocr_files:
                # 尝试加载第一页的OCR图像
                try:
                    from PIL import Image
                    # 查找第一页图像
                    output_dir = getattr(self, 'last_ocr_output_dir', 'ocr_output')
                    import glob
                    image_files = glob.glob(os.path.join(output_dir, '*_page_1.jpg'))
                    if image_files:
                        image_for_api = Image.open(image_files[0]).convert("RGB")
                except:
                    pass
            
            # 如果找不到图像，创建一个简单的占位图像或使用文本
            try:
                if image_for_api is None:
                    # 创建一个简单的文本图像占位符
                    from PIL import Image, ImageDraw, ImageFont
                    # 创建一个白色背景的图像
                    img = Image.new('RGB', (800, 600), color='white')
                    draw = ImageDraw.Draw(img)
                    # 在图像上写入提示文本
                    try:
                        text = "文档OCR识别结果\n请根据文本内容提取字段"
                        draw.text((50, 50), text, fill='black')
                    except:
                        pass
                    image_for_api = img
                
                # 调用API进行信息抽取
                api_result = self.doc_api.recognize_card(
                    image=image_for_api,
                    custom_prompt=effective_prompt,
                    max_tokens=2048,
                    temperature=0.2,
                    top_p=0.8,
                    use_rag=False
                )
                
                if api_result.get("success", False):
                    response = api_result.get("result", "")
                    print(f"✅ 大模型调用成功，响应长度: {len(response)}")
                    # 提取表格部分
                    table_html = self._extract_table_from_response(response)
                    print(f"✅ 表格提取完成")
                    return table_html
                else:
                    error_msg = api_result.get("error", "调用失败")
                    print(f"❌ 调用失败: {error_msg}")
                    return f"❌ 调用失败: {error_msg}"
            except Exception as api_error:
                # 如果API调用失败，返回格式化的提示词结果
                print(f"⚠️ 调用异常: {api_error}")
                import traceback
                print(traceback.format_exc())
                return f"❌ 调用失败: {str(api_error)}"
            
        except Exception as e:
            import traceback
            return f"❌ 信息抽取失败: {str(e)}\n{traceback.format_exc()}"
    
    def _extract_table_from_response(self, response):
        """
        从大模型响应中提取表格，格式化为HTML表格（一个字段一行）
        
        Args:
            response: 大模型的响应文本
            
        Returns:
            HTML格式的表格
        """
        import re
        import html
        
        # 清理响应文本，移除代码块标记
        response = response.strip()
        if response.startswith('```'):
            # 移除代码块标记
            lines = response.split('\n')
            start_idx = 0
            end_idx = len(lines)
            for i, line in enumerate(lines):
                if line.strip().startswith('```'):
                    if i == 0:
                        start_idx = 1
                    else:
                        end_idx = i
                        break
            response = '\n'.join(lines[start_idx:end_idx])
        
        # 尝试提取Markdown表格
        # 匹配完整的表格结构：表头行 + 分隔行 + 数据行
        all_lines = response.split('\n')
        table_lines = []
        in_table = False
        header_line_idx = -1
        separator_line_idx = -1
        
        # 查找表格开始位置
        for i, line in enumerate(all_lines):
            line_stripped = line.strip()
            # 检查是否是表格行（包含 | 符号）
            if '|' in line_stripped:
                # 检查是否是分隔行（只包含 -、:、| 和空格）
                if re.match(r'^[\s\|:\-]+$', line_stripped):
                    if header_line_idx >= 0 and separator_line_idx < 0:
                        # 找到分隔行
                        separator_line_idx = i
                        in_table = True
                        continue
                elif not in_table:
                    # 可能是表头行
                    header_line_idx = i
                    table_lines.append(line_stripped)
                elif in_table:
                    # 数据行
                    table_lines.append(line_stripped)
            elif in_table:
                # 表格结束
                break
        
        # 如果找到了表格行，解析它们
        if len(table_lines) > 1:
            html_table = '<div style="margin: 20px 0;">\n'
            html_table += '<table class="ocr-result-table" style="width: 100%; border-collapse: collapse; border: 1px solid #dee2e6; font-size: 14px;">\n'
            html_table += '<thead><tr style="background: #f8f9fa; font-weight: 600;">'
            html_table += '<th style="border: 1px solid #dee2e6; padding: 12px; text-align: left; width: 30%;">字段名</th>'
            html_table += '<th style="border: 1px solid #dee2e6; padding: 12px; text-align: left; width: 70%;">字段值</th>'
            html_table += '</tr></thead>\n<tbody>\n'
            
            # 解析表格行（跳过第一行表头，从数据行开始）
            for line in table_lines[1:]:
                # 跳过分隔行
                if re.match(r'^[\s\|:\-]+$', line):
                    continue
                
                # 分割单元格
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                
                # 如果单元格数量 >= 2，提取字段名和字段值
                if len(cells) >= 2:
                    # 第一个单元格是字段名，第二个是字段值
                    field_name = cells[0].strip()
                    field_value = cells[1].strip() if len(cells) > 1 else ""
                    
                    # 清理字段名和字段值（移除可能的标记符号）
                    field_name = re.sub(r'^[-\*\s]+', '', field_name).strip()
                    field_value = re.sub(r'^[-\*\s]+', '', field_value).strip()
                    
                    # 如果字段名不是"字段名"等表头关键词，才添加
                    if field_name and field_name not in ['字段名', '字段值', 'Field Name', 'Field Value']:
                        html_table += f'<tr>'
                        html_table += f'<td style="border: 1px solid #dee2e6; padding: 12px; font-weight: 500; background: #fafbfc;">{html.escape(field_name)}</td>'
                        html_table += f'<td style="border: 1px solid #dee2e6; padding: 12px; word-wrap: break-word;">{html.escape(field_value)}</td>'
                        html_table += '</tr>\n'
            
            html_table += '</tbody></table>\n</div>'
            return html_table
        
        # 如果没有找到标准表格，尝试提取字段名:字段值格式
        lines = response.split('\n')
        html_table = '<div style="margin: 20px 0;">\n'
        html_table += '<table class="ocr-result-table" style="width: 100%; border-collapse: collapse; border: 1px solid #dee2e6; font-size: 14px;">\n'
        html_table += '<thead><tr style="background: #f8f9fa; font-weight: 600;">'
        html_table += '<th style="border: 1px solid #dee2e6; padding: 12px; text-align: left; width: 30%;">字段名</th>'
        html_table += '<th style="border: 1px solid #dee2e6; padding: 12px; text-align: left; width: 70%;">字段值</th>'
        html_table += '</tr></thead>\n<tbody>\n'
        
        found_rows = False
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 跳过代码块标记和分隔行
            if line.startswith('```') or re.match(r'^[\s\|:\-]+$', line):
                continue
            
            # 尝试匹配字段名:字段值或字段名：字段值
            if ':' in line or '：' in line:
                parts = re.split('[:：]', line, 1)
                if len(parts) == 2:
                    field = parts[0].strip()
                    value = parts[1].strip()
                    # 清理字段名（移除可能的标记符号）
                    field = re.sub(r'^[-\*\s]+', '', field).strip()
                    value = re.sub(r'^[-\*\s]+', '', value).strip()
                    if field and value:
                        html_table += f'<tr>'
                        html_table += f'<td style="border: 1px solid #dee2e6; padding: 12px; font-weight: 500; background: #fafbfc;">{html.escape(field)}</td>'
                        html_table += f'<td style="border: 1px solid #dee2e6; padding: 12px; word-wrap: break-word;">{html.escape(value)}</td>'
                        html_table += '</tr>\n'
                        found_rows = True
        
        if not found_rows:
            # 如果都没有找到，返回原始响应的前500字符
            html_table += f'<tr><td colspan="2" style="border: 1px solid #dee2e6; padding: 12px; color: #dc3545;">⚠️ 无法解析响应格式，原始内容：<br/>{html.escape(response[:500])}</td></tr>\n'
        
        html_table += '</tbody></table>\n</div>'
        return html_table

    def export_chat_history(self):
        """导出对话历史"""
        if not self.chat_history:
            return "❌ 没有对话历史可导出！"

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"chat_history_{timestamp}.json"

            # 保存为JSON格式
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.chat_history, f, ensure_ascii=False, indent=2)

            return f"✅ 对话历史已导出到: {filename}"

        except Exception as e:
            return f"❌ 导出失败: {str(e)}"

    def clear_history(self):
        """清空对话历史"""
        self.chat_history = []
        self.chat_messages = []
        self.last_image = None
        self.last_saved_image_path = None
        self.last_image_digest = None
        self.last_ocr_markdown = None
        self.last_ocr_html = None
        if hasattr(self, "session_turn_image_paths"):
            self.session_turn_image_paths.clear()
        return []

    def export_last_ocr(self):
        if not self.last_ocr_markdown:
            return "❌ 没有可保存的文本样式，请先执行一次OCR识别！"

        export_dir = os.path.join("ocr_exports")
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        sections = self._parse_markdown_sections(self.last_ocr_markdown)

        excel_path = os.path.join(export_dir, f"ocr_{timestamp}.xlsx")
        excel_note = ""
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "表格1" if sections else "OCR文本"
            table_idx = 0
            for section in sections:
                if section["type"] == "table":
                    table_idx += 1
                    if table_idx > 1:
                        ws = wb.create_sheet(title=f"表格{table_idx}")
                    ws.append(section["header"])
                    for row in section["rows"]:
                        ws.append(row)
                elif section["type"] == "text" and section["text"]:
                    if table_idx > 0:
                        ws = wb.create_sheet(title=f"文本{table_idx}")
                    for line in section["text"].splitlines():
                        ws.append([line])
            if not sections:
                for line in self.last_ocr_markdown.splitlines():
                    ws.append([line])
            wb.save(excel_path)
        except Exception as exc:
            excel_path = os.path.join(export_dir, f"ocr_{timestamp}.csv")
            with open(excel_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["OCR Result"])
                for line in self.last_ocr_markdown.splitlines():
                    writer.writerow([line])
            excel_note = f"⚠️ Excel导出失败({exc})，已保存为CSV"

        json_path = os.path.join(export_dir, f"ocr_{timestamp}.json")
        json_content = {
            "markdown": self.last_ocr_markdown,
            "sections": sections,
        }
        with open(json_path, "w", encoding='utf-8') as f:
            json.dump(json_content, f, ensure_ascii=False, indent=2)

        message_lines = [
            "✅ 文本样式已保存：",
            f"- Excel: {excel_path}" + (f" ({excel_note})" if excel_note else ""),
            f"- JSON: {json_path}",
        ]
        return "\n".join(message_lines)


DEFAULT_TASK_PROMPTS = {
    "任务问答": "请根据图片完成指定任务。",
    "OCR识别": "请识别并提取这张图片中的所有文字内容，并标注语言类型。请确保所有带样式或表格内容使用Markdown表格表示。",
    "卡证OCR识别": "请进行卡证类识别并以Markdown表格输出关键字段（如姓名、证件号、有效期、卡号等）",
    "票据OCR识别": "请解析发票/小票等票据，输出关键信息和多行项目表格，并在下方给出置信度与可疑项。",
    "协议OCR识别": "请提取合同/协议关键信息（甲乙方、日期、金额、条款等），保留段落与条款编号，并在末尾给出风险提示。",
}


def _plain_text_to_html(text: str) -> str:
    if not text:
        return ""
    escaped = html.escape(str(text))
    replaced = escaped.replace("\n", "<br>")
    return f'<div class="stats-text">{replaced}</div>'


def _get_default_prompt(task: str, code_format: str = None) -> str:
    return DEFAULT_TASK_PROMPTS.get(task, DEFAULT_TASK_PROMPTS["任务问答"])


def _prefill_prompt(task: str):
    """根据任务类型返回默认提示词，用于输入框预填。"""
    return _get_default_prompt(task, None)


# 单例应用
app = AdvancedQwen3VLApp()

# 会话级图片保存目录与轨迹
IMAGE_SAVE_ROOT = "chat_history/images"
SESSION_IMAGE_DIR = os.path.join(IMAGE_SAVE_ROOT, getattr(app, "session_id", datetime.now().strftime("%Y%m%d_%H%M%S")))
os.makedirs(SESSION_IMAGE_DIR, exist_ok=True)
app.session_turn_image_paths = []  # 与对话轮次对齐的图片路径（无图则为 None）


def _load_media(image, file_path, need_all_pages=False):
    """加载上传的图片或PDF。
    返回 (media, error_message)。media 可能是 PIL.Image 或 list[PIL.Image]（当 need_all_pages=True 且输入为 PDF 时）。
    """
    if file_path:
        try:
            lower = str(file_path).lower()
            if lower.endswith(".pdf"):
                pages = app._pdf_to_images(file_path)
                if not pages:
                    return None, "❌ PDF 转图片失败，请检查文件或安装 PyMuPDF/pdf2image"
                pages = [p.convert("RGB") for p in pages]
                return (pages if need_all_pages else pages[0]), None
            return Image.open(file_path).convert("RGB"), None
        except Exception as e:
            return None, f"❌ 读取文件失败: {e}"
    if image is not None:
        return image, None
    return None, "❌ 请上传图片或PDF文件"


def handle_unified_chat(image,
                        file_path,
                        text,
                        history,
                        max_tokens,
                        temperature,
                        top_p,
                        top_k,
                        pro_task,
                        repetition_penalty,
                        presence_penalty):
    """统一的发送处理：按任务分派到不同方法，默认走任务问答。"""
    user_text = (text or "").strip()

    media, load_err = _load_media(image, file_path, need_all_pages=True)
    if load_err:
        stats_update = gr.update(value=_plain_text_to_html(load_err), visible=True)
        yield history, text, stats_update
        return

    # 如果是PDF，取首页图片用于多模态问答（不重复触发OCR，已在上传时完成）
    image = media[0] if isinstance(media, list) else media
    saved_image_path = None
    image_digest = None
    if image is not None:
        try:
            buffer = io.BytesIO()
            image.save(buffer, format="PNG")
            image_digest = hashlib.md5(buffer.getvalue()).hexdigest()
        except Exception:
            image_digest = None

        should_save = image_digest is None or image_digest != getattr(app, "last_image_digest", None)
        if should_save:
            try:
                ts = datetime.now().strftime("%H%M%S%f")
                saved_image_path = os.path.join(SESSION_IMAGE_DIR, f"img_{ts}.png")
                image.save(saved_image_path)
                if image_digest is not None:
                    app.last_image_digest = image_digest
            except Exception:
                saved_image_path = None

    prev_turns = len(history)
    image_recorded = False

    def record_image_path():
        nonlocal image_recorded
        if image_recorded:
            return
        if saved_image_path and saved_image_path != getattr(app, "last_saved_image_path", None):
            app.session_turn_image_paths.append(saved_image_path)
            app.last_saved_image_path = saved_image_path
            if image_digest is not None:
                app.last_image_digest = image_digest
        else:
            existing = getattr(app, "last_saved_image_path", None)
            app.session_turn_image_paths.append(existing)
        image_recorded = True

    try:
        task = pro_task or "任务问答"
        if task == "OCR识别":
            if image is None:
                stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图片或PDF！"), visible=True)
                yield history, text, stats_update
                return

            result = app.ocr_analysis(image)

            if result.startswith("❌"):
                stats_update = gr.update(value=_plain_text_to_html(result), visible=True)
                yield history, text, stats_update
                return

            prompt_text = user_text if user_text else _get_default_prompt(task, None)
            updated_history = history + [[f"👤 {prompt_text}", result]]
            app.chat_history = updated_history
            if not image_recorded:
                record_image_path()
            ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
            stats_update = gr.update(value=ocr_preview, visible=True)
            yield updated_history, "", stats_update
            return

        if task == "卡证OCR识别（API）":
            if image is None:
                stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图片或PDF！"), visible=True)
                yield history, text, stats_update
                return
            result = app.ocr_card_api(image)
            if result.startswith("❌"):
                stats_update = gr.update(value=_plain_text_to_html(result), visible=True)
                yield history, text, stats_update
                return
            prompt_text = user_text if user_text else _get_default_prompt("卡证OCR识别", None)
            updated_history = history + [[f"👤 {prompt_text}", result]]
            app.chat_history = updated_history
            if not image_recorded:
                record_image_path()
            ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
            stats_update = gr.update(value=ocr_preview, visible=True)
            yield updated_history, "", stats_update
            return

        if task == "卡证OCR识别":
            if image is None:
                stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图片或PDF！"), visible=True)
                yield history, text, stats_update
                return
            result = app.ocr_card(image)
            if result.startswith("❌"):
                stats_update = gr.update(value=_plain_text_to_html(result), visible=True)
                yield history, text, stats_update
                return
            prompt_text = user_text if user_text else _get_default_prompt(task, None)
            updated_history = history + [[f"👤 {prompt_text}", result]]
            app.chat_history = updated_history
            if not image_recorded:
                record_image_path()
            ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
            stats_update = gr.update(value=ocr_preview, visible=True)
            yield updated_history, "", stats_update
            return

        if task == "票据OCR识别":
            if image is None:
                stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图片或PDF！"), visible=True)
                yield history, text, stats_update
                return
            result = app.ocr_receipt(image)
            if result.startswith("❌"):
                stats_update = gr.update(value=_plain_text_to_html(result), visible=True)
                yield history, text, stats_update
                return
            prompt_text = user_text if user_text else _get_default_prompt(task, None)
            updated_history = history + [[f"👤 {prompt_text}", result]]
            app.chat_history = updated_history
            if not image_recorded:
                record_image_path()
            ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
            stats_update = gr.update(value=ocr_preview, visible=True)
            yield updated_history, "", stats_update
            return

        if task == "协议OCR识别":
            if image is None:
                stats_update = gr.update(value=_plain_text_to_html("❌ 请上传图片或PDF！"), visible=True)
                yield history, text, stats_update
                return
            result = app.ocr_agreement(image)
            if result.startswith("❌"):
                stats_update = gr.update(value=_plain_text_to_html(result), visible=True)
                yield history, text, stats_update
                return
            prompt_text = user_text if user_text else _get_default_prompt(task, None)
            updated_history = history + [[f"👤 {prompt_text}", result]]
            app.chat_history = updated_history
            if not image_recorded:
                record_image_path()
            ocr_preview = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
            stats_update = gr.update(value=ocr_preview, visible=True)
            yield updated_history, "", stats_update
            return

        # 默认图文问答
        effective_prompt = user_text if user_text else _get_default_prompt(task, None)
        chat_result = app.chat_with_image(
            image,
            effective_prompt,
            history,
            max_tokens,
            temperature,
            top_p,
            top_k,
            repetition_penalty,
            presence_penalty
        )

        if inspect.isgenerator(chat_result):
            for out_history, cleared, stats in chat_result:
                if not image_recorded and len(out_history) > prev_turns:
                    record_image_path()
                app.chat_history = out_history
                stats_update = gr.update(value=_plain_text_to_html(stats), visible=True)
                yield out_history, cleared, stats_update
        else:
            out_history, cleared, stats = chat_result
            if not image_recorded and len(out_history) > prev_turns:
                record_image_path()
            app.chat_history = out_history
            stats_update = gr.update(value=_plain_text_to_html(stats), visible=True)
            yield out_history, cleared, stats_update

        if not image_recorded and len(app.chat_history) > prev_turns:
            record_image_path()

    except Exception as e:
        history.append(["👤", f"❌ 错误: {str(e)}"])
        app.chat_history = history
        if not image_recorded and len(history) > prev_turns:
            record_image_path()
        stats_update = gr.update(value=_plain_text_to_html(f"❌ 错误: {str(e)}"), visible=True)
        yield history, text, stats_update


def save_chat_to_folder(save_dir, history):
    """将当前聊天历史保存到指定文件夹（JSON）。"""
    try:
        if not save_dir:
            return "❌ 保存失败：未指定保存目录"
        os.makedirs(save_dir, exist_ok=True)
        # 每次保存使用独立导出子目录，避免图片累积到同一目录
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = os.path.join(save_dir, timestamp)
        os.makedirs(export_dir, exist_ok=True)
        images_target_dir = os.path.join(export_dir, "images")
        os.makedirs(images_target_dir, exist_ok=True)

        image_paths = getattr(app, "session_turn_image_paths", [])
        copied_rel_paths = []
        seen_images = set()
        for p in image_paths:
            abs_path = os.path.abspath(p) if p else None
            if not p or not os.path.exists(p) or abs_path in seen_images:
                copied_rel_paths.append(None)
                continue
            try:
                basename = os.path.basename(p)
                target = os.path.join(images_target_dir, basename)
                if os.path.abspath(p) != os.path.abspath(target):
                    shutil.copy2(p, target)
                seen_images.add(abs_path)
                copied_rel_paths.append(os.path.join("images", basename))
            except Exception:
                copied_rel_paths.append(None)
        filename = os.path.join(export_dir, f"chat_history_{timestamp}.json")
        # history 是 [(user, bot), ...]
        data = []
        turns = history or []
        for idx, pair in enumerate(turns):
            try:
                u, b = pair
            except Exception:
                u, b = pair, ""
            img_rel = copied_rel_paths[idx] if idx < len(copied_rel_paths) else None
            data.append({"user": u, "assistant": b, "image_path": img_rel})
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return f"✅ 已保存到: {filename}"
    except Exception as e:
        return f"❌ 保存失败: {str(e)}"


def _legacy_create_unified_interface():
    """创建统一Gradio界面。"""

    touch_css = """
    :root {
        --radius-lg: 22px;
        --radius-md: 14px;
        --surface: #ffffff;
        --surface-muted: #f5f7fb;
        --surface-border: #e2e8f0;
        --text-primary: #0f172a;
        --text-secondary: #64748b;
        --accent: #2563eb;
        --accent-soft: rgba(37, 99, 235, 0.12);
    }
    body {
        background: linear-gradient(135deg, #eef2ff 0%, #f9fafc 55%, #ffffff 100%);
        color: var(--text-primary);
    }
    /* 文档上传卡片 */
    #doc-upload-card {
        background: linear-gradient(140deg, rgba(37,99,235,0.08), rgba(59,130,246,0.05));
        border: 1px solid rgba(37, 99, 235, 0.18);
        border-radius: 20px;
        padding: 16px 18px;
        box-shadow: 0 12px 24px rgba(15, 23, 42, 0.06);
        margin-bottom: 16px;
    }
    #doc-upload-card .gradio-file,
    #doc-upload-card .gradio-image,
    #doc-upload-card .gradio-markdown,
    #doc-upload-card .gradio-html {
        background: transparent;
    }
    #doc-upload-card .gradio-radio {
        margin-top: 8px;
    }
    #unified-header-row {
        display: flex;
        align-items: center;
        gap: 18px;
        margin-bottom: 12px;
    }
    #header-controls {
        display: flex;
        align-items: center;
        gap: 12px;
        justify-content: flex-end;
    }
    #header-status textarea,
    #header-status input {
        height: 60px;
    }
    .gradio-container {
        width: min(1920px, 98vw) !important;
        max-width: 100% !important;
        margin: 0 auto;
        padding: clamp(16px, 2vw, 32px) clamp(18px, 3vw, 36px) 48px;
        font-size: 16px;
        color: var(--text-primary);
    }
    .gradio-container .gr-markdown {
        color: var(--text-primary);
    }
    #unified-header {
        background: linear-gradient(130deg, rgba(37, 99, 235, 0.12), rgba(59, 130, 246, 0.1));
        border: 1px solid rgba(37, 99, 235, 0.18);
        padding: 24px 28px;
        border-radius: 28px;
        box-shadow: 0 18px 36px rgba(15, 23, 42, 0.08);
        margin-bottom: 22px;
    }
    #unified-header h1 {
        margin: 0 0 6px;
        font-size: 26px;
        font-weight: 600;
        letter-spacing: 0.2px;
    }
    #unified-header p {
        margin: 0;
        color: var(--text-secondary);
    }
    #unified-mode-bar {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
        background: var(--surface);
        border-radius: 24px;
        box-shadow: 0 20px 40px rgba(15, 23, 42, 0.06);
        padding: clamp(14px, 2vw, 22px);
        gap: 16px 18px;
        margin-bottom: 20px;
        border: 1px solid var(--surface-border);
    }
    #unified-mode-bar .gradio-button,
    #unified-mode-bar button {
        font-size: 16px !important;
        padding: 12px 18px !important;
        border-radius: 14px !important;
    }
    #unified-mode-bar textarea,
    #unified-mode-bar input[type="text"] {
        background: var(--surface);
        border: 1px solid rgba(148, 163, 184, 0.35);
        border-radius: 14px;
        color: var(--text-primary);
        box-shadow: inset 0 1px 3px rgba(15, 23, 42, 0.04);
    }
    #unified-mode-bar textarea:focus,
    #unified-mode-bar input[type="text"]:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
    }
    .gradio-container .tabs {
        background: transparent;
        border: none;
    }
    .gradio-container .tabitem {
        border-radius: var(--radius-md);
        background: #f8fafc;
        border: 1px solid transparent;
        color: var(--text-secondary);
    }
    .gradio-container .tabitem.selected {
        border-color: rgba(37, 99, 235, 0.25);
        color: var(--text-primary);
        background: #ffffff;
        box-shadow: 0 10px 20px rgba(37, 99, 235, 0.08);
    }
    #unified-input-panel,
    #unified-chat-panel,
    #unified-batch-panel,
    #unified-compare-panel {
        background: var(--surface);
        border-radius: 24px;
        padding: 22px 24px;
        box-shadow: 0 22px 44px rgba(15, 23, 42, 0.06);
        border: 1px solid var(--surface-border);
    }
    #unified-input-panel .gradio-slider > label,
    #unified-input-panel .gradio-dropdown > label {
        color: var(--text-secondary);
    }
    #unified-chat-panel {
        display: flex;
        flex-direction: column;
        gap: 16px;
    }
    #unified-chatbot > .wrap {
        background: #f8fafc;
        border-radius: 20px;
        border: 1px solid rgba(148, 163, 184, 0.25);
        padding: 8px 10px;
    }
    #unified-chatbot .message {
        border-radius: 16px !important;
        padding: 12px 14px !important;
        line-height: 1.6;
        font-size: 15px;
        color: var(--text-primary);
    }
    #unified-chatbot .message.user {
        background: linear-gradient(138deg, rgba(37, 99, 235, 0.16), rgba(96, 165, 250, 0.12));
        border: 1px solid rgba(37, 99, 235, 0.22);
        color: var(--text-primary);
        align-self: flex-end;
    }
    #unified-chatbot .message.bot {
        background: #ffffff;
        border: 1px solid rgba(203, 213, 225, 0.9);
        color: var(--text-primary);
        align-self: flex-start;
    }
    #unified-query textarea {
        border-radius: 16px;
        border: 1px solid rgba(148, 163, 184, 0.35);
        background: var(--surface);
        color: var(--text-primary);
        box-shadow: inset 0 1px 3px rgba(15, 23, 42, 0.05);
    }
    #unified-query textarea:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
    }
    #unified-stats .stats-text {
        background: var(--accent-soft);
        border-radius: 16px;
        border: 1px solid rgba(37, 99, 235, 0.2);
        color: var(--text-primary);
        font-weight: 500;
        padding: 12px 14px;
        line-height: 1.6;
        margin-bottom: 12px;
        word-break: break-word;
    }
    .gradio-container .gradio-button.primary {
        background: linear-gradient(135deg, #2563eb, #1d4ed8);
        border: none;
        color: #ffffff;
        font-weight: 600;
        box-shadow: 0 18px 30px rgba(37, 99, 235, 0.22);
    }
    .gradio-container .gradio-button.primary:hover {
        filter: brightness(1.03);
    }
    .gradio-container .gradio-button.secondary {
        background: rgba(37, 99, 235, 0.1);
        border: 1px solid rgba(37, 99, 235, 0.18);
        color: var(--text-primary);
    }
    .gradio-container textarea,
    .gradio-container input[type="text"],
    .gradio-container input[type="number"] {
        background: var(--surface);
        border: 1px solid rgba(148, 163, 184, 0.35);
        color: var(--text-primary);
        border-radius: 16px;
    }
    .gradio-container textarea:focus,
    .gradio-container input[type="text"]:focus,
    .gradio-container input[type="number"]:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
    }
    #unified-batch-panel textarea,
    #unified-compare-panel textarea {
        min-height: 320px;
    }
    .gradio-container .dropdown span.label,
    .gradio-container .slider > label,
    .gradio-container .dropdown label {
        color: var(--text-secondary);
    }
    .gradio-container .gradio-dropdown .wrap select,
    .gradio-container .gradio-dropdown .wrap button {
        background: var(--surface);
        color: var(--text-primary);
        border-color: rgba(148, 163, 184, 0.4);
    }
    .gradio-container .gradio-dropdown .wrap select:focus,
    .gradio-container .gradio-dropdown .wrap button:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.12);
    }
    /*
    Bigger markdown preview area for unified stats (OCR/table preview)
    */
    #unified-stats {
        max-height: 560px;
        overflow: auto;
        border: 1px solid rgba(148, 163, 184, 0.35);
        padding: 12px 14px;
        border-radius: 14px;
        background: #ffffff;
    }
    #unified-stats table {
        width: 100%;
        border-collapse: collapse;
        margin: 8px 0 14px;
    }
    #unified-stats th,
    #unified-stats td {
        border: 1px solid #e5e7eb;
        padding: 8px 10px;
        text-align: left;
        vertical-align: top;
        font-size: 14px;
        line-height: 1.55;
    }
    #unified-stats thead th {
        background: #f8fafc;
        font-weight: 600;
    }
    #unified-stats code {
        background: #f8fafc;
        border: 1px solid #e5e7eb;
        padding: 1px 4px;
        border-radius: 6px;
        font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
    }
    /* 字段表格样式 */
    .gradio-container .dataframe {
        border-radius: 14px;
        border: 1px solid rgba(148, 163, 184, 0.35);
        overflow: hidden;
    }
    .gradio-container .dataframe table {
        width: 100%;
        border-collapse: collapse;
    }
    .gradio-container .dataframe th {
        background: linear-gradient(135deg, rgba(37, 99, 235, 0.1), rgba(59, 130, 246, 0.08));
        color: var(--text-primary);
        font-weight: 600;
        padding: 10px 12px;
        border-bottom: 2px solid rgba(37, 99, 235, 0.2);
    }
    .gradio-container .dataframe td {
        padding: 8px 12px;
        border-bottom: 1px solid rgba(148, 163, 184, 0.2);
    }
    .gradio-container .dataframe tr:hover {
        background: rgba(37, 99, 235, 0.04);
    }
    .gradio-container .dataframe input[type="text"] {
        border: 1px solid rgba(148, 163, 184, 0.3);
        border-radius: 6px;
        padding: 4px 8px;
        width: 100%;
    }
    .gradio-container .dataframe input[type="text"]:focus {
        border-color: var(--accent);
        box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.1);
    }
    """

    with gr.Blocks(
        title="多模态大模型智能助手",
        theme=gr.themes.Soft(),
        css=touch_css
    ) as interface:

        with gr.Row(elem_id="unified-header-row", equal_height=True):
            with gr.Column(scale=3):
                gr.HTML("""
                <section id="unified-header" style="width: 100%;">
                  <h1 style="margin-bottom: 6px;">🤖 多模态大模型智能助手</h1>
                  <p style="margin: 0;">单一界面完成多模态图文问答、OCR及卡证/票据处理，无需模式切换。</p>
                </section>
                """)
            with gr.Column(scale=1, elem_id="header-controls"):
                load_btn = gr.Button("🔄 加载模型", variant="primary", scale=0, elem_id="header-load-btn")
                status_text = gr.Textbox(
                    label=None,
                    value="⏳ 模型未加载，请点击加载模型按钮",
                    interactive=False,
                    lines=2,
                    elem_id="header-status"
                )

        pro_task_state = gr.State("任务问答")
        media_image_state = gr.State()
        media_file_state = gr.State()

        def on_media_upload(file_path, ocr_engine_choice):
            """上传后立即进行OCR并写入知识库，避免再点发送才识别。"""
            if not file_path:
                return None, None, None, "❌ 请先选择文件", gr.update(value="", visible=False)
            media, err = _load_media(None, file_path, need_all_pages=False)
            if err:
                return None, None, None, err, gr.update(value="", visible=False)
            preview = media[0] if isinstance(media, list) else media
            basename = os.path.basename(file_path) if file_path else ""
            # 记录最近一次上传的文件路径，发送时避免重复OCR
            try:
                app.last_uploaded_file_path = os.path.abspath(file_path)
            except Exception:
                app.last_uploaded_file_path = file_path

            engine_choice = (ocr_engine_choice or "").lower()
            ocr_engine_key = "qwen3" if engine_choice.startswith("qwen") else "paddle"

            # 自动执行文档OCR（图片/PNG/JPG等也走文档路径），并尝试入库
            ocr_msg = ""
            ocr_preview_html = ""
            try:
                is_pdf = str(file_path).lower().endswith(".pdf")
                ocr_res = app.ocr_document(
                    file_path,
                    is_pdf=is_pdf,
                    pdf_pages="all",
                    engine=ocr_engine_key
                )
                # 仅保留简短提示，详细结果仍在左侧“状态/知识库”展示
                snippet = (ocr_res or "").replace("\n", " ")
                snippet = (snippet[:120] + "...") if len(snippet) > 120 else snippet
                engine_label = "Qwen3" if ocr_engine_key == "qwen3" else "PaddleOCR"
                ocr_msg = f" | 已自动OCR（{engine_label}）并写入知识库: {snippet}"
                # 侧边栏预览：优先使用HTML，其次用Markdown转义
                if getattr(app, "last_ocr_html", None):
                    ocr_preview_html = app.last_ocr_html
                else:
                    import html as _html
                    md_text = getattr(app, "last_ocr_markdown", "") or (ocr_res or "")
                    escaped = _html.escape(md_text.strip()[:5000])
                    if escaped:
                        ocr_preview_html = f'<div class="stats-text" style="white-space:pre-wrap;">{escaped}</div>'
            except Exception as e:
                ocr_msg = f" | ⚠️ 自动OCR失败: {e}"

            return (
                preview,
                preview,
                file_path,
                f"✅ 已加载: {basename}{ocr_msg}",
                gr.update(value=ocr_preview_html, visible=bool(ocr_preview_html))
            )

        load_btn.click(app.load_model, outputs=[status_text, load_btn])

        with gr.Tab("📚 文档图文问答"):
            with gr.Group(elem_id="doc-upload-card"):
                gr.Markdown("#### 📂 上传文档并即时预览")
                with gr.Row(equal_height=True):
                    with gr.Column(scale=1):
                        media_file = gr.File(
                            label="选择图片或PDF（取首页预览）",
                            file_types=[".pdf", ".png", ".jpg", ".jpeg", ".webp"],
                            type="filepath",
                            height=90,
                        )
                        ocr_engine_selector = gr.Radio(
                            label="文档OCR引擎",
                            choices=["Qwen3-VL（本地）", "PaddleOCR API"],
                            value="Qwen3-VL（本地）",
                            interactive=True,
                            elem_id="doc-ocr-engine"
                        )
                        media_status = gr.Markdown(
                            value="请上传文件，自动完成预览与OCR",
                            elem_id="shared-status"
                        )
                    with gr.Column(scale=1):
                        media_preview = gr.Image(
                            label="预览（PDF取首页）",
                            type="pil",
                            interactive=False,
                            height=320,
                        )
                        doc_ocr_preview = gr.HTML(
                            label="自动OCR预览（上传即识别）",
                            value="",
                            visible=False,
                            elem_id="doc-ocr-preview"
                        )

            media_file.change(
                on_media_upload,
                inputs=[media_file, ocr_engine_selector],
                outputs=[media_preview, media_image_state, media_file_state, media_status, doc_ocr_preview],
            )

            with gr.Row(equal_height=True):
                with gr.Column(scale=1):
                    with gr.Group(elem_id="unified-input-panel"):
                        with gr.Row(equal_height=True):
                            max_tokens = gr.Slider(minimum=512, maximum=16384, value=4096, label="最大生成长度 (out_seq_length)")
                            temperature = gr.Slider(minimum=0.0, maximum=2.0, value=0.7, label="创造性 (temperature)")
                        gr.Markdown("ℹ️ 上方上传并自动预览/识别，直接在此提问即可。")

                        with gr.Accordion("🎛️ 高级参数", open=False, visible=True):
                            top_p = gr.Slider(minimum=0.0, maximum=1.0, value=0.8, label="top_p")
                            top_k = gr.Slider(minimum=0, maximum=100, value=20, label="top_k")
                            repetition_penalty = gr.Slider(minimum=0.8, maximum=2.0, value=1.0, step=0.05, label="repetition_penalty")
                            presence_penalty = gr.Slider(minimum=0.0, maximum=3.0, value=1.5, step=0.1, label="presence_penalty")

                with gr.Column(scale=2):
                    with gr.Group(elem_id="unified-chat-panel"):
                        gr.Markdown("### 图文问答")
                        doc_chatbot = gr.Chatbot(label=None, height=520, show_label=False, type="tuples", elem_id="unified-chatbot", render_markdown=True)
                        doc_text_input = gr.Textbox(label=None, placeholder="输入你想了解的内容，支持直接就文档/图片发问。", lines=3, elem_id="unified-query")
                        with gr.Row():
                            doc_send_btn = gr.Button("发送", variant="primary", scale=1)
                            doc_clear_btn = gr.Button("🗑️ 清空历史", variant="secondary", scale=1)
                        doc_stats_output = gr.HTML(value="", visible=False, elem_id="unified-stats")

            doc_send_btn.click(
                handle_unified_chat,
                inputs=[media_image_state, media_file_state, doc_text_input, doc_chatbot, max_tokens, temperature, top_p, top_k, pro_task_state, repetition_penalty, presence_penalty],
                outputs=[doc_chatbot, doc_text_input, doc_stats_output],
            )
            doc_text_input.submit(
                handle_unified_chat,
                inputs=[media_image_state, media_file_state, doc_text_input, doc_chatbot, max_tokens, temperature, top_p, top_k, pro_task_state, repetition_penalty, presence_penalty],
                outputs=[doc_chatbot, doc_text_input, doc_stats_output],
            )
            def _clear_doc_session():
                app.clear_history()
                return [], "", gr.update(value="", visible=False)

            doc_clear_btn.click(
                _clear_doc_session,
                outputs=[doc_chatbot, doc_text_input, doc_stats_output],
            )
        with gr.Tab("🪪 卡证OCR（三步流程）"):
            gr.Markdown("### 三步流程：识别类型 → 自定义字段 → OCR识别")
            
            with gr.Row():
                with gr.Column(scale=1):
                    with gr.Row():
                        detect_type_btn = gr.Button("🔍 第一步：识别卡证类型", variant="primary")
                    gr.Markdown("ℹ️ 使用顶部统一上传入口，预览即当前卡证/票据。")
                    
                    card_type_output = gr.Textbox(
                        label="识别的卡证类型",
                        interactive=False,
                        visible=False
                    )
                    
                    default_fields_title = gr.Markdown("### 📋 默认字段模板", visible=False)
                    # HTML表格展示（用于HTML格式的模板）
                    default_fields_html = gr.HTML(
                        label="默认字段模板（HTML表格）",
                        visible=False,
                        elem_id="default-fields-html"
                    )
                    # Dataframe展示（用于非HTML格式的模板）
                    default_fields_output = gr.Dataframe(
                        label="默认字段",
                        headers=["序号", "字段名"],
                        datatype=["number", "str"],
                        interactive=True,
                        visible=False,
                        wrap=True,
                        type="array"  # 明确指定返回格式为2D数组
                    )
                    
                    custom_fields_title = gr.Markdown("### ➕ 自定义字段", visible=False)
                    custom_fields_input = gr.Dataframe(
                        label="添加自定义字段（每行一个字段名）",
                        headers=["字段名"],
                        datatype=["str"],
                        interactive=True,
                        visible=False,
                        wrap=True,
                        row_count=(1, "dynamic"),
                        col_count=(1, "fixed"),
                        type="array",  # 明确指定返回格式为2D数组
                        value=[[""]]  # 初始值：一个空行
                    )
                    
                    with gr.Row():
                        add_custom_field_btn = gr.Button("➕ 添加自定义字段", variant="secondary", visible=False, size="sm")
                    
                    with gr.Row():
                        update_fields_btn = gr.Button("🔗 第二步：合并字段", variant="secondary", visible=False)
                    
                    all_fields_title = gr.Markdown("### ✅ 最终字段列表（将用于OCR识别）", visible=False)
                    # HTML表格展示（用于HTML格式的模板）
                    all_fields_html = gr.HTML(
                        label="最终字段列表（HTML表格）",
                        visible=False,
                        elem_id="all-fields-html"
                    )
                    # Dataframe展示（用于非HTML格式的模板）
                    all_fields_output = gr.Dataframe(
                        label="最终字段列表",
                        headers=["序号", "字段名", "来源"],
                        datatype=["number", "str", "str"],
                        interactive=False,
                        visible=False,
                        wrap=True,
                        type="array"  # 明确指定返回格式为2D数组
                    )
                    
                    fields_status = gr.Textbox(
                        label="状态",
                        interactive=False,
                        visible=False
                    )
                    
                    with gr.Row():
                        ocr_with_fields_btn = gr.Button("🚀 第三步：开始OCR识别", variant="primary", visible=False)
                
                with gr.Column(scale=2):
                    with gr.Row():
                        gr.Markdown("### 📊 OCR识别结果")
                        with gr.Column(scale=1, min_width=200):
                            ocr_export_format = gr.Dropdown(
                                choices=["Markdown (.md)", "Excel (.xlsx)", "CSV (.csv)", "JSON (.json)"],
                                value="Markdown (.md)",
                                label="导出格式",
                                visible=False
                            )
                        ocr_export_btn_3step = gr.Button("💾 导出结果", variant="secondary", visible=False, size="sm")
                    
                    # HTML表格展示（用于HTML格式的模板）
                    ocr_result_html = gr.HTML(
                        label="OCR识别结果（HTML表格）",
                        visible=False,
                        elem_id="ocr-result-html"
                    )
                    # Dataframe展示（用于非HTML格式的模板）
                    ocr_result = gr.Dataframe(
                        label="OCR识别结果（可编辑表格）",
                        headers=["字段名", "字段值"],
                        datatype=["str", "str"],
                        interactive=True,
                        visible=False,
                        wrap=True,
                        type="array"
                    )
                    
                    ocr_export_status_3step = gr.Textbox(
                        label="导出状态",
                        interactive=False,
                        visible=False,
                        lines=3
                    )
            
            # 辅助函数：确保值是标量（非可迭代）
            def ensure_scalar(value):
                """确保值是标量，如果是可迭代对象则转换为字符串"""
                if value is None:
                    return ""
                elif isinstance(value, str):
                    return value
                elif isinstance(value, (list, tuple)):
                    return "".join(str(x) for x in value) if value else ""
                elif hasattr(value, '__iter__'):
                    try:
                        return "".join(str(x) for x in value)
                    except:
                        return str(value)
                else:
                    return str(value)
            
            # 第一步：识别卡证类型
            def step1_detect_type(media_image, media_file_path):
                media = media_image
                if media is None and media_file_path:
                    loaded, err = _load_media(None, media_file_path, need_all_pages=False)
                    if err:
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            err
                        )
                    media = loaded[0] if isinstance(loaded, list) else loaded
                if media is None:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        "❌ 请先上传文件"
                    )

                result = app.detect_card_type(media)
                if len(result) == 4:
                    card_type, default_fields, html_template, status_msg = result
                else:
                    # 兼容旧版本（没有HTML模板）
                    card_type, default_fields, status_msg = result
                    html_template = None
                
                if card_type:
                    # 卡证OCR不使用HTML模板，只使用DataFrame展示
                    # 转换为DataFrame格式：[[序号, 字段名], ...]
                    default_fields_df = []
                    for i, field in enumerate(default_fields, 1):
                        field_str = ensure_scalar(field).strip()
                        if field_str:
                            default_fields_df.append([int(i), field_str])
                    # 清空自定义字段
                    custom_fields_df = [[""]]
                    return (
                        gr.update(value=card_type, visible=True),
                        gr.update(visible=False),  # HTML表格隐藏
                        gr.update(value=default_fields_df, visible=True),  # Dataframe
                        gr.update(visible=True),  # default_fields_title
                        gr.update(value=custom_fields_df, visible=True),  # custom_fields_input
                        gr.update(visible=True),  # custom_fields_title
                        gr.update(visible=True),  # add_custom_field_btn
                        gr.update(visible=True),  # update_fields_btn
                        gr.update(visible=False),  # all_fields_title (初始隐藏)
                        gr.update(value=[], visible=False),  # all_fields_output (初始为空)
                        gr.update(value=status_msg, visible=True),  # fields_status
                    )
                else:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(value=status_msg, visible=True),  # fields_status
                    )
            
            # 第二步：更新字段
            def step2_update_fields(card_type, default_fields_df, custom_fields_df):
                # 检查第一步是否完成：需要card_type存在
                if not card_type:
                    return (
                        gr.update(visible=False),
                        gr.update(value=[], visible=False),
                        gr.update(value="❌ 请先完成第一步：识别卡证类型", visible=True)
                    )
                
                # 优先从应用状态获取字段列表（适用于HTML表格情况）
                default_fields = []
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    default_fields = app.current_default_fields.copy()
                    print(f"[DEBUG] 从app状态获取默认字段: {default_fields}")
                
                # 如果应用状态中没有，则从DataFrame提取字段名
                if not default_fields and default_fields_df is not None:
                    print(f"[DEBUG] default_fields_df原始数据: {default_fields_df}, 类型: {type(default_fields_df)}")
                    
                    # 处理不同的数据格式
                    rows = []
                    if hasattr(default_fields_df, 'values'):
                        # 如果是pandas DataFrame
                        try:
                            rows = default_fields_df.values.tolist()
                        except:
                            rows = list(default_fields_df.values) if hasattr(default_fields_df, 'values') else []
                    elif isinstance(default_fields_df, (list, tuple)):
                        # 如果是列表或元组
                        if len(default_fields_df) == 0:
                            rows = []
                        elif len(default_fields_df) > 0 and isinstance(default_fields_df[0], str):
                            # 第一个元素是字符串，可能是列名列表，跳过
                            print(f"[DEBUG] 警告：default_fields_df似乎是列名列表，跳过: {default_fields_df}")
                            rows = []
                        elif len(default_fields_df) > 0 and isinstance(default_fields_df[0], (list, tuple)):
                            # 第一个元素是列表/元组，这是行数据列表
                            rows = default_fields_df
                        else:
                            # 其他情况，尝试作为行数据处理
                            rows = default_fields_df
                    elif isinstance(default_fields_df, dict):
                        # 如果是字典，尝试提取数据
                        if 'data' in default_fields_df:
                            rows = default_fields_df['data']
                        else:
                            rows = []
                    else:
                        rows = []
                    
                    print(f"[DEBUG] 处理后的rows: {rows}, 类型: {type(rows)}, 长度: {len(rows) if hasattr(rows, '__len__') else 'N/A'}")
                    
                    # 遍历行数据
                    for i, row in enumerate(rows):
                        if not row:
                            continue
                        print(f"[DEBUG] 处理行{i}: {row}, 类型: {type(row)}")
                        
                        # 跳过列名（字符串）
                        if isinstance(row, str):
                            print(f"[DEBUG] 跳过列名: {row}")
                            continue
                        
                        # 处理行格式：应该是列表或元组 [序号, 字段名]
                        if isinstance(row, (list, tuple)):
                            if len(row) >= 2:
                                # 取第二个元素（索引1，字段名列）
                                field_value = row[1]
                            else:
                                continue
                        elif isinstance(row, dict):
                            # 如果是字典，尝试从'字段名'键获取
                            field_value = row.get('字段名') or row.get(1)
                        else:
                            continue
                        
                        print(f"[DEBUG] 提取的field_value: {field_value!r}, 类型: {type(field_value)}")
                        # 使用辅助函数确保字段值是标量
                        field_str = ensure_scalar(field_value).strip()
                        print(f"[DEBUG] 提取后: field_str={field_str!r}, type={type(field_str)}")
                        if field_str:
                            default_fields.append(field_str)
                
                # 如果仍然没有字段，说明第一步未完成
                if not default_fields:
                    return (
                        gr.update(visible=False),
                        gr.update(value=[], visible=False),
                        gr.update(value="❌ 请先完成第一步：识别卡证类型", visible=True)
                    )
                
                # 从自定义字段DataFrame提取
                custom_fields_list = []
                if custom_fields_df is not None:
                    print(f"[DEBUG] custom_fields_df原始数据: {custom_fields_df}, 类型: {type(custom_fields_df)}")
                    
                    # 处理不同的数据格式（与默认字段相同的逻辑）
                    if isinstance(custom_fields_df, dict):
                        if 'data' in custom_fields_df:
                            rows = custom_fields_df['data']
                        elif 'values' in custom_fields_df:
                            rows = custom_fields_df['values']
                        else:
                            rows = [v for v in custom_fields_df.values() if isinstance(v, (list, tuple)) and len(v) > 0]
                            if rows:
                                rows = rows[0]
                            else:
                                rows = []
                    elif hasattr(custom_fields_df, 'values'):
                        try:
                            rows = custom_fields_df.values.tolist()
                        except:
                            rows = list(custom_fields_df.values) if hasattr(custom_fields_df, 'values') else []
                    elif isinstance(custom_fields_df, (list, tuple)):
                        rows = custom_fields_df
                    else:
                        try:
                            rows = list(custom_fields_df)
                        except:
                            rows = []
                    
                    print(f"[DEBUG] 处理后的custom rows: {rows}, 类型: {type(rows)}")
                    
                    # 遍历行数据
                    for row in rows:
                        if not row:
                            continue
                        print(f"[DEBUG] 处理自定义行: {row}, 类型: {type(row)}")
                        
                        # 处理不同的行格式
                        if isinstance(row, dict):
                            # 如果是字典，尝试从'字段名'键获取
                            field_value = row.get('字段名') or row.get(0) or (row.get(list(row.keys())[0]) if len(row) > 0 else None)
                        elif isinstance(row, (list, tuple)):
                            # 如果是列表或元组，取第一个元素（索引0）
                            if len(row) > 0:
                                field_value = row[0]
                            else:
                                continue
                        else:
                            continue
                        
                        print(f"[DEBUG] 提取的自定义field_value: {field_value!r}, 类型: {type(field_value)}")
                        # 使用辅助函数确保字段值是标量
                        field_str = ensure_scalar(field_value).strip()
                        if field_str and field_str not in default_fields:
                            custom_fields_list.append(field_str)
                
                # 转换为DataFrame格式：[[序号, 字段名, 来源], ...]
                all_fields_df = []
                default_count = 0
                custom_count = 0
                
                # 添加默认字段（过滤空字段）
                idx = 1
                for field in default_fields:
                    # 使用辅助函数确保字段名是标量
                    field_str = ensure_scalar(field).strip()
                    if field_str:
                        # 确保每个元素都是标量值，不是可迭代对象
                        # 显式转换为字符串，确保不是其他类型
                        field_name = str(field_str)
                        print(f"[DEBUG] 添加默认字段: idx={idx}, field_str={field_str!r}, field_name={field_name!r}, type={type(field_name)}")
                        all_fields_df.append([int(idx), field_name, "默认"])
                        idx += 1
                        default_count += 1
                
                # 添加自定义字段，序号从当前idx开始
                for field in custom_fields_list:
                    # 使用辅助函数确保字段名是标量
                    field_str = ensure_scalar(field).strip()
                    if field_str:
                        # 确保每个元素都是标量值，不是可迭代对象
                        # 显式转换为字符串，确保不是其他类型
                        field_name = str(field_str)
                        print(f"[DEBUG] 添加自定义字段: idx={idx}, field_str={field_str!r}, field_name={field_name!r}, type={type(field_name)}")
                        all_fields_df.append([int(idx), field_name, "自定义"])
                        idx += 1
                        custom_count += 1
                
                # 调试输出
                print(f"\n[DEBUG] 最终字段列表数据:")
                print(f"  all_fields_df类型: {type(all_fields_df)}")
                print(f"  all_fields_df内容: {all_fields_df}")
                for i, row in enumerate(all_fields_df):
                    print(f"  行{i}: {row}, 类型: {type(row)}, 字段名类型: {type(row[1]) if len(row) > 1 else 'N/A'}")
                
                # 生成状态消息
                total_count = len(all_fields_df)
                if total_count == 0:
                    status_msg = "⚠️ 警告：没有有效字段，请至少添加一个字段"
                else:
                    status_msg = f"✅ 字段已更新，共 {total_count} 个字段（默认：{default_count}，自定义：{custom_count}）"
                
                # 保存到app状态
                app.current_card_type = card_type
                app.current_default_fields = default_fields.copy()
                app.current_custom_fields = custom_fields_list.copy()
                
                # 卡证OCR不使用HTML模板，只使用DataFrame
                app.current_final_fields_html = None
                
                # 直接使用DataFrame展示（卡证OCR不使用HTML模板）
                return (
                    gr.update(visible=True),  # all_fields_title
                    gr.update(visible=False, value=""),  # HTML表格隐藏
                    gr.update(value=all_fields_df, visible=True),  # Dataframe
                    gr.update(value=status_msg, visible=True)
                )
            
            # 辅助函数：将Markdown表格转换为Dataframe格式
            def markdown_table_to_dataframe(markdown_text):
                """将Markdown表格转换为Dataframe格式（2D数组）"""
                if not markdown_text:
                    return []
                
                # 解析Markdown表格
                sections = app._parse_markdown_sections(markdown_text)
                dataframe_data = []
                
                for section in sections:
                    if section["type"] == "table":
                        header = section.get("header", [])
                        rows = section.get("rows", [])
                        
                        # 如果header为空，使用第一行作为header
                        if not header and rows:
                            header = rows[0] if rows else ["字段名", "字段值"]
                            rows = rows[1:] if len(rows) > 1 else []
                        
                        # 确保header至少有两列
                        if len(header) < 2:
                            header = ["字段名", "字段值"]
                        
                        # 转换为Dataframe格式：每行是[字段名, 字段值]
                        for row in rows:
                            if len(row) >= 2:
                                dataframe_data.append([str(row[0]).strip(), str(row[1]).strip()])
                            elif len(row) == 1:
                                dataframe_data.append([str(row[0]).strip(), ""])
                
                # 如果没有找到表格，返回空列表
                return dataframe_data if dataframe_data else []
            
            # 第三步：OCR识别
            def step3_ocr(media_image, media_file_path, all_fields_df):
                media = media_image
                if media is None and media_file_path:
                    loaded, err = _load_media(None, media_file_path, need_all_pages=False)
                    if err:
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=True, value=err)
                        )
                    media = loaded
                if media is None:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=True, value="❌ 请先上传文件")
                    )
                image = media[0] if isinstance(media, list) else media
                
                # 优先从应用状态获取字段列表（适用于HTML表格情况）
                fields_list = []
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    fields_list = app.current_default_fields.copy()
                if hasattr(app, 'current_custom_fields') and app.current_custom_fields:
                    fields_list.extend(app.current_custom_fields)
                
                # 如果应用状态中没有，则从DataFrame提取字段名
                if not fields_list:
                    # 安全地检查DataFrame是否为空
                    has_fields = all_fields_df is not None and (
                        (isinstance(all_fields_df, list) and len(all_fields_df) > 0) or
                        (hasattr(all_fields_df, '__len__') and len(all_fields_df) > 0)
                    )
                    
                    if not has_fields:
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False)
                        )
                    
                    # 从DataFrame提取字段名（排除"来源"列）
                    if all_fields_df is not None:
                        # 确保是列表格式
                        if not isinstance(all_fields_df, list):
                            try:
                                all_fields_df = all_fields_df.tolist() if hasattr(all_fields_df, 'tolist') else list(all_fields_df)
                            except:
                                all_fields_df = []
                        
                        for row in all_fields_df:
                            if row and len(row) >= 2 and row[1] and str(row[1]).strip():
                                fields_list.append(str(row[1]).strip())
                
                if not fields_list:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False)
                    )
                
                result = app.ocr_card_with_fields(image, fields_list)
                
                # 卡证OCR不使用HTML模板，只使用Markdown/DataFrame
                # 提取Markdown文本（去掉可能的图标前缀）
                if result.startswith("🪪"):
                    markdown_text = result.split(":", 1)[1].strip() if ":" in result else result
                else:
                    markdown_text = result
                
                # 保存到app状态以便导出
                app.last_ocr_markdown = f"## 卡证OCR识别结果（三步流程）\n\n{markdown_text}"
                
                # 解析OCR结果，提取字段值字典
                ocr_data = {}
                sections = app._parse_markdown_sections(markdown_text)
                for section in sections:
                    if section["type"] == "table":
                        rows = section.get("rows", [])
                        for row in rows:
                            if len(row) >= 2:
                                field_name = str(row[0]).strip()
                                field_value = str(row[1]).strip()
                                if field_name:
                                    ocr_data[field_name] = field_value
                
                # 卡证OCR不使用HTML模板，只使用DataFrame
                # 将OCR结果转换为DataFrame格式
                ocr_dataframe = []
                for field_name, field_value in ocr_data.items():
                    ocr_dataframe.append([field_name, field_value])
                
                # 如果没有数据，返回空DataFrame
                if not ocr_dataframe:
                    ocr_dataframe = [["", ""]]
                
                return (
                    gr.update(visible=False),  # HTML表格隐藏
                    gr.update(value=ocr_dataframe, visible=True),  # Dataframe
                    gr.update(visible=True),  # 导出格式
                    gr.update(visible=True),  # 导出按钮
                    gr.update(visible=False, value="")  # 导出状态
                )
                
                # 以下代码不再使用（卡证OCR不使用HTML模板）
                if False:
                    # 使用HTML表格展示OCR结果
                    try:
                        from bs4 import BeautifulSoup
                        soup = BeautifulSoup(html_template, 'html.parser')
                        table = soup.find('table')
                        
                        if table:
                            # 填充OCR结果到表格中
                            # 策略：遍历所有行，查找包含字段名的单元格，然后在同一行或下一行填充值
                            for row in table.find_all('tr'):
                                cells = row.find_all(['td', 'th'])
                                for i, cell in enumerate(cells):
                                    cell_text = cell.get_text(strip=True)
                                    # 检查是否是字段名（在OCR结果中存在）
                                    if cell_text and cell_text in ocr_data:
                                        # 查找同一行中的下一个空单元格来填充值
                                        found = False
                                        for j in range(i + 1, len(cells)):
                                            next_cell = cells[j]
                                            next_text = next_cell.get_text(strip=True)
                                            # 如果下一个单元格为空，填充OCR结果
                                            if not next_text:
                                                next_cell.string = ocr_data[cell_text]
                                                found = True
                                                break
                                            # 如果下一个单元格不是字段名，也填充（可能是值单元格）
                                            elif next_text not in ocr_data or next_text == '':
                                                next_cell.string = ocr_data[cell_text]
                                                found = True
                                                break
                                        
                                        # 如果同一行没有找到合适的单元格，在当前单元格后插入
                                        if not found:
                                            value_cell = soup.new_tag('td')
                                            value_cell.string = ocr_data[cell_text]
                                            cell.insert_after(value_cell)
                                        
                                        # 标记已处理，避免重复填充
                                        ocr_data.pop(cell_text, None)
                            
                            # 添加自定义字段（如果有）
                            custom_fields = getattr(app, 'current_custom_fields', [])
                            for custom_field in custom_fields:
                                if custom_field in ocr_data:
                                    new_row = soup.new_tag('tr')
                                    new_row['class'] = 'custom-field-row'
                                    field_cell = soup.new_tag('td')
                                    field_cell.string = custom_field
                                    field_cell['colspan'] = '2'
                                    value_cell = soup.new_tag('td')
                                    value_cell.string = ocr_data.get(custom_field, '')
                                    value_cell['colspan'] = '3'
                                    new_row.append(field_cell)
                                    new_row.append(value_cell)
                                    table.append(new_row)
                            
                            # 添加样式
                            styled_html = f"""
                            <style>
                            .ocr-result-table {{
                                width: 100%;
                                border-collapse: collapse;
                                margin: 10px 0;
                                font-size: 14px;
                            }}
                            .ocr-result-table th,
                            .ocr-result-table td {{
                                border: 1px solid #ddd;
                                padding: 8px;
                                text-align: left;
                            }}
                            .ocr-result-table th {{
                                background-color: #f2f2f2;
                                font-weight: bold;
                            }}
                            .ocr-result-table tr:nth-child(even) {{
                                background-color: #f9f9f9;
                            }}
                            .ocr-result-table .custom-field-row {{
                                background-color: #fff3cd !important;
                            }}
                            .ocr-result-table td[contenteditable="true"] {{
                                background-color: #e7f3ff;
                                cursor: text;
                            }}
                            </style>
                            {str(table)}
                            """
                            
                            return (
                                gr.update(value=styled_html, visible=True),  # HTML表格
                                gr.update(value=[], visible=False),  # Dataframe隐藏，传递空列表避免验证错误
                                gr.update(visible=True),  # ocr_export_format
                                gr.update(visible=True),  # ocr_export_btn_3step
                                gr.update(visible=False, value="")  # ocr_export_status_3step
                            )
                    except Exception as e:
                        print(f"⚠️ 生成OCR结果HTML表格失败: {e}")
                        import traceback
                        traceback.print_exc()
                        # 降级到Dataframe展示
                        dataframe_data = markdown_table_to_dataframe(markdown_text)
                        return (
                            gr.update(visible=False, value=""),  # HTML表格隐藏
                            gr.update(value=dataframe_data, visible=True),  # Dataframe
                            gr.update(visible=True),  # ocr_export_format
                            gr.update(visible=True),  # ocr_export_btn_3step
                            gr.update(visible=False, value="")  # ocr_export_status_3step
                        )
                else:
                    # 没有HTML模板，使用Dataframe展示
                    dataframe_data = markdown_table_to_dataframe(markdown_text)
                    return (
                        gr.update(visible=False, value=""),  # HTML表格隐藏
                        gr.update(value=dataframe_data, visible=True),  # Dataframe
                        gr.update(visible=True),  # ocr_export_format
                        gr.update(visible=True),  # ocr_export_btn_3step
                        gr.update(visible=False, value="")  # ocr_export_status_3step
                    )
            
            detect_type_btn.click(
                step1_detect_type,
                inputs=[media_image_state, media_file_state],
                outputs=[
                    card_type_output,
                    default_fields_html,
                    default_fields_output,
                    default_fields_title,
                    custom_fields_input,
                    custom_fields_title,
                    add_custom_field_btn,
                    update_fields_btn,
                    all_fields_title,
                    all_fields_output,
                    fields_status
                ]
            )
            
            # 添加自定义字段按钮的功能
            def add_custom_field(current_data):
                """在自定义字段Dataframe中添加一个新行"""
                if current_data is None:
                    return [[""]]
                # 确保是列表格式
                if not isinstance(current_data, list):
                    try:
                        current_data = current_data.tolist() if hasattr(current_data, 'tolist') else list(current_data)
                    except:
                        current_data = [[""]]
                # 添加一个新行
                new_data = list(current_data) if current_data else []
                new_data.append([""])
                return new_data
            
            add_custom_field_btn.click(
                add_custom_field,
                inputs=[custom_fields_input],
                outputs=[custom_fields_input]
            )
            
            update_fields_btn.click(
                step2_update_fields,
                inputs=[card_type_output, default_fields_output, custom_fields_input],
                outputs=[all_fields_title, all_fields_html, all_fields_output, fields_status]
            )
            
            # 当字段更新后，显示OCR按钮（用于Dataframe）
            def show_ocr_btn_from_dataframe(all_fields_df):
                # 优先检查应用状态（适用于HTML表格情况）
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    return gr.update(visible=True)
                
                # 安全地检查DataFrame
                if all_fields_df is None:
                    return gr.update(visible=False)
                
                # 确保是列表格式
                if not isinstance(all_fields_df, list):
                    try:
                        all_fields_df = all_fields_df.tolist() if hasattr(all_fields_df, 'tolist') else list(all_fields_df)
                    except:
                        return gr.update(visible=False)
                
                if len(all_fields_df) > 0:
                    # 检查是否有有效字段
                    has_fields = any(
                        row and len(row) >= 2 and row[1] and str(row[1]).strip()
                        for row in all_fields_df
                    )
                    return gr.update(visible=has_fields)
                return gr.update(visible=False)
            
            # 当HTML表格更新后，显示OCR按钮（用于HTML）
            def show_ocr_btn_from_html(html_content):
                # HTML组件变化时，直接检查应用状态
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    return gr.update(visible=True)
                return gr.update(visible=False)
            
            # 监听all_fields_output和all_fields_html的变化
            all_fields_output.change(
                show_ocr_btn_from_dataframe,
                inputs=[all_fields_output],
                outputs=[ocr_with_fields_btn]
            )
            all_fields_html.change(
                show_ocr_btn_from_html,
                inputs=[all_fields_html],
                outputs=[ocr_with_fields_btn]
            )
            
            ocr_with_fields_btn.click(
                step3_ocr,
                inputs=[media_image_state, media_file_state, all_fields_output],  # all_fields_output可能为空（HTML表格情况），但会从app状态获取
                outputs=[ocr_result_html, ocr_result, ocr_export_format, ocr_export_btn_3step, ocr_export_status_3step]
            )
            
            # 辅助函数：将Dataframe转换为Markdown表格
            def dataframe_to_markdown_table(dataframe_data):
                """将Dataframe数据转换为Markdown表格格式"""
                if not dataframe_data:
                    return ""
                
                # 确保是列表格式
                if not isinstance(dataframe_data, list):
                    try:
                        dataframe_data = dataframe_data.tolist() if hasattr(dataframe_data, 'tolist') else list(dataframe_data)
                    except:
                        return ""
                
                # 构建Markdown表格
                lines = ["| 字段名 | 字段值 |", "|--------|--------|"]
                for row in dataframe_data:
                    if row and len(row) >= 2:
                        field_name = str(row[0]).strip() if row[0] else ""
                        field_value = str(row[1]).strip() if row[1] else ""
                        lines.append(f"| {field_name} | {field_value} |")
                
                return "\n".join(lines)
            
            # 导出OCR结果（从Dataframe读取当前编辑后的内容）
            def export_ocr_result_3step(dataframe_data, export_format):
                """导出三步流程的OCR结果（支持多种格式）"""
                if not dataframe_data or (isinstance(dataframe_data, list) and len(dataframe_data) == 0):
                    return gr.update(visible=True, value="❌ 没有可保存的OCR结果，请先执行OCR识别！")
                
                export_dir = os.path.join("ocr_exports")
                os.makedirs(export_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                # 确保是列表格式
                if not isinstance(dataframe_data, list):
                    try:
                        dataframe_data = dataframe_data.tolist() if hasattr(dataframe_data, 'tolist') else list(dataframe_data)
                    except:
                        dataframe_data = []
                
                try:
                    # 根据选择的格式导出
                    if export_format == "Markdown (.md)":
                        # 转换为Markdown表格
                        markdown_table = dataframe_to_markdown_table(dataframe_data)
                        markdown_content = f"## 卡证OCR识别结果（三步流程）\n\n{markdown_table}"
                        
                        file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.md")
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write(markdown_content)
                        
                        app.last_ocr_markdown = markdown_content
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 Markdown文件: {file_path}\n\n已保存当前编辑后的内容。")
                    
                    elif export_format == "Excel (.xlsx)":
                        try:
                            from openpyxl import Workbook
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "OCR结果"
                            
                            # 添加表头
                            ws.append(["字段名", "字段值"])
                            
                            # 添加数据行
                            for row in dataframe_data:
                                if row and len(row) >= 2:
                                    ws.append([str(row[0]).strip() if row[0] else "", str(row[1]).strip() if row[1] else ""])
                            
                            file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.xlsx")
                            wb.save(file_path)
                            return gr.update(visible=True, value=f"✅ 导出成功！\n📊 Excel文件: {file_path}\n\n已保存当前编辑后的内容。")
                        except Exception as e:
                            # 如果Excel导出失败，尝试CSV
                            file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.csv")
                            with open(file_path, "w", encoding="utf-8", newline="") as f:
                                writer = csv.writer(f)
                                writer.writerow(["字段名", "字段值"])
                                for row in dataframe_data:
                                    if row and len(row) >= 2:
                                        writer.writerow([str(row[0]).strip() if row[0] else "", str(row[1]).strip() if row[1] else ""])
                            return gr.update(visible=True, value=f"⚠️ Excel导出失败，已保存为CSV\n📄 CSV文件: {file_path}\n\n错误: {str(e)}")
                    
                    elif export_format == "CSV (.csv)":
                        file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.csv")
                        with open(file_path, "w", encoding="utf-8", newline="") as f:
                            writer = csv.writer(f)
                            writer.writerow(["字段名", "字段值"])
                            for row in dataframe_data:
                                if row and len(row) >= 2:
                                    writer.writerow([str(row[0]).strip() if row[0] else "", str(row[1]).strip() if row[1] else ""])
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 CSV文件: {file_path}\n\n已保存当前编辑后的内容。")
                    
                    elif export_format == "JSON (.json)":
                        # 转换为字典列表
                        json_data = []
                        for row in dataframe_data:
                            if row and len(row) >= 2:
                                json_data.append({
                                    "字段名": str(row[0]).strip() if row[0] else "",
                                    "字段值": str(row[1]).strip() if row[1] else ""
                                })
                        
                        file_path = os.path.join(export_dir, f"ocr_3step_{timestamp}.json")
                        with open(file_path, "w", encoding="utf-8") as f:
                            json.dump(json_data, f, ensure_ascii=False, indent=2)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 JSON文件: {file_path}\n\n已保存当前编辑后的内容。")
                    
                    else:
                        return gr.update(visible=True, value=f"❌ 不支持的导出格式: {export_format}")
                        
                except Exception as e:
                    return gr.update(visible=True, value=f"❌ 导出失败: {str(e)}")
            
            ocr_export_btn_3step.click(
                export_ocr_result_3step,
                inputs=[ocr_result, ocr_export_format],
                outputs=[ocr_export_status_3step]
            )

        with gr.Tab("📄 单据OCR（三步流程）"):
            gr.Markdown("### 三步流程：识别类型 → 自定义字段 → OCR识别（使用HTML表格模板）")
            
            with gr.Row():
                with gr.Column(scale=1):
                    gr.Markdown("使用上方统一上传入口，预览即当前处理文件。")

                    with gr.Row():
                        detect_bill_type_btn = gr.Button("🔍 第一步：识别票据类型", variant="primary")
                    
                    bill_type_output = gr.Textbox(
                        label="识别的票据类型",
                        interactive=False,
                        visible=False
                    )

                    bill_field_list = gr.Textbox(
                        label="字段列表",
                        interactive=False,
                        visible=False
                    )
                    
                    bill_default_fields_title = gr.Markdown("### 📋 默认字段模板", visible=False)
                    # HTML表格展示（票据OCR使用HTML模板）
                    bill_default_fields_html = gr.HTML(
                        label="默认字段模板（HTML表格）",
                        visible=False,
                        elem_id="bill-default-fields-html"
                    )
                    
                    bill_custom_fields_title = gr.Markdown("### ➕ 自定义字段", visible=False)
                    bill_custom_fields_input = gr.Dataframe(
                        label="添加自定义字段（每行一个字段名）",
                        headers=["字段名"],
                        datatype=["str"],
                        interactive=True,
                        visible=False,
                        wrap=True,
                        row_count=(1, "dynamic"),
                        col_count=(1, "fixed"),
                        type="array",
                        value=[[""]]
                    )
                    
                    with gr.Row():
                        bill_add_custom_field_btn = gr.Button("➕ 添加自定义字段", variant="secondary", visible=False, size="sm")
                    
                    with gr.Row():
                        bill_update_fields_btn = gr.Button("🔗 第二步：合并字段", variant="secondary", visible=False)
                    
                    bill_all_fields_title = gr.Markdown("### ✅ 最终字段列表（将用于OCR识别）", visible=False)
                    # HTML表格展示（票据OCR使用HTML模板）
                    bill_all_fields_html = gr.HTML(
                        label="最终字段列表（HTML表格）",
                        visible=False,
                        elem_id="bill-all-fields-html"
                    )
                    
                    bill_fields_status = gr.Textbox(
                        label="状态",
                        interactive=False,
                        visible=False
                    )
                    
                    with gr.Row():
                        bill_ocr_with_fields_btn = gr.Button("🚀 第三步：开始OCR识别", variant="primary", visible=False)
                
                with gr.Column(scale=2):
                    with gr.Row():
                        gr.Markdown("### 📊 OCR识别结果")
                        with gr.Column(scale=1, min_width=200):
                            bill_ocr_export_format = gr.Dropdown(
                                choices=["HTML (.html)", "Markdown (.md)", "Excel (.xlsx)", "CSV (.csv)", "JSON (.json)"],
                                value="HTML (.html)",
                                label="导出格式",
                                visible=False
                            )
                        bill_ocr_export_btn_3step = gr.Button("💾 导出结果", variant="secondary", visible=False, size="sm", elem_id="bill-ocr-export-btn")
                    
                    # HTML表格展示（票据OCR使用HTML模板）
                    bill_ocr_result_html = gr.HTML(
                        label="OCR识别结果（HTML表格）",
                        visible=False,
                        elem_id="bill-ocr-result-html"
                    )
                    
                    # 隐藏的Textbox，用于存储编辑后的HTML内容
                    # 注意：不使用elem_id，让Gradio自动生成ID，然后通过返回值更新
                    bill_ocr_result_html_edited = gr.Textbox(
                        label="编辑后的HTML内容",
                        visible=False
                    )
                    
                    bill_ocr_export_status_3step = gr.Textbox(
                        label="导出状态",
                        interactive=False,
                        visible=False,
                        lines=3
                    )
            
            # 第一步：识别票据类型
            def bill_step1_detect_type(media_image, media_file_path):
                media = media_image
                if media is None and media_file_path:
                    loaded, err = _load_media(None, media_file_path, need_all_pages=False)
                    if err:
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            err
                        )
                    media = loaded[0] if isinstance(loaded, list) else loaded
                if media is None:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        "❌ 请先上传文件"
                    )

                result = app.detect_bill_type(media)
                if len(result) == 4:
                    bill_type, default_fields, html_template, status_msg = result
                else:
                    bill_type, default_fields, status_msg = result
                    html_template = None
                
                if bill_type:
                    has_html_template = html_template is not None and html_template.strip()
                    
                    if has_html_template:
                        styled_html = f"""
                        <style>
                        .field-template-table {{
                            width: 100%;
                            border-collapse: collapse;
                            margin: 10px 0;
                            font-size: 14px;
                        }}
                        .field-template-table th,
                        .field-template-table td {{
                            border: 1px solid #ddd;
                            padding: 8px;
                            text-align: left;
                        }}
                        .field-template-table th {{
                            background-color: #f2f2f2;
                            font-weight: bold;
                        }}
                        .field-template-table tr:nth-child(even) {{
                            background-color: #f9f9f9;
                        }}
                        </style>
                        {html_template}
                        """
                        return (
                            gr.update(value=bill_type, visible=True),
                            gr.update(value=styled_html, visible=True),
                            gr.update(visible=True),
                            gr.update(value=[[""]], visible=True),
                            gr.update(visible=True),
                            gr.update(visible=True),
                            gr.update(visible=True),
                            gr.update(value=status_msg, visible=True),
                            default_fields
                        )
                    else:
                        return (
                            gr.update(value=bill_type, visible=True),
                            gr.update(visible=False),
                            gr.update(visible=True),
                            gr.update(value=[[""]], visible=True),
                            gr.update(visible=True),
                            gr.update(visible=True),
                            gr.update(visible=True),
                            gr.update(value=status_msg, visible=True),
                            default_fields
                        )
                else:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(value=status_msg, visible=True),
                        None
                    )
            
            # 第二步：合并字段（票据OCR使用HTML模板）
            def bill_step2_update_fields(card_type, custom_fields_df):
                if not card_type:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(value="❌ 请先完成第一步：识别票据类型", visible=True)
                    )
                
                default_fields = []
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    default_fields = app.current_default_fields.copy()
                
                custom_fields_list = []
                if custom_fields_df is not None:
                    if isinstance(custom_fields_df, (list, tuple)):
                        rows = custom_fields_df
                    else:
                        rows = []
                    
                    for row in rows:
                        if not row:
                            continue
                        if isinstance(row, (list, tuple)) and len(row) > 0:
                            field_value = row[0]
                        else:
                            continue
                        field_str = str(field_value).strip()
                        if field_str and field_str not in default_fields:
                            custom_fields_list.append(field_str)
                
                app.current_custom_fields = custom_fields_list.copy()
                
                html_template = getattr(app, 'current_field_template_html', None)
                has_html_template = html_template is not None and html_template.strip()
                
                final_fields_html = None
                
                if has_html_template and custom_fields_list:
                    try:
                        from bs4 import BeautifulSoup
                        soup = BeautifulSoup(html_template, 'html.parser')
                        table = soup.find('table')
                        
                        if table:
                            for custom_field in custom_fields_list:
                                new_row = soup.new_tag('tr')
                                field_cell = soup.new_tag('td')
                                field_cell.string = custom_field
                                field_cell['colspan'] = '2'
                                value_cell = soup.new_tag('td')
                                value_cell.string = ''
                                value_cell['colspan'] = '3'
                                new_row.append(field_cell)
                                new_row.append(value_cell)
                                table.append(new_row)
                            
                            styled_html = f"""
                            <style>
                            .all-fields-table {{
                                width: 100%;
                                border-collapse: collapse;
                                margin: 10px 0;
                                font-size: 14px;
                            }}
                            .all-fields-table th,
                            .all-fields-table td {{
                                border: 1px solid #ddd;
                                padding: 8px;
                                text-align: left;
                            }}
                            .all-fields-table th {{
                                background-color: #f2f2f2;
                                font-weight: bold;
                            }}
                            .all-fields-table tr:nth-child(even) {{
                                background-color: #f9f9f9;
                            }}
                            .custom-field-row {{
                                background-color: #fff3cd !important;
                            }}
                            </style>
                            {str(table)}
                            """
                            
                            final_fields_html = str(table)
                            app.current_final_fields_html = final_fields_html
                            
                            total_count = len(default_fields) + len(custom_fields_list)
                            status_msg = f"✅ 字段已更新，共 {total_count} 个字段（默认：{len(default_fields)}，自定义：{len(custom_fields_list)}）"
                            
                            return (
                                gr.update(visible=True),
                                gr.update(value=styled_html, visible=True),
                                gr.update(value=status_msg, visible=True)
                            )
                    except Exception as e:
                        print(f"⚠️ 生成HTML表格失败: {e}")
                        status_msg = f"⚠️ 生成HTML表格失败: {e}"
                        return (
                            gr.update(visible=True),
                            gr.update(visible=False),
                            gr.update(value=status_msg, visible=True)
                        )
                elif has_html_template:
                    try:
                        from bs4 import BeautifulSoup
                        soup = BeautifulSoup(html_template, 'html.parser')
                        table = soup.find('table')
                        if table:
                            final_fields_html = str(table)
                        else:
                            final_fields_html = html_template
                    except:
                        final_fields_html = html_template
                    
                    app.current_final_fields_html = final_fields_html
                    
                    styled_html = f"""
                    <style>
                    .all-fields-table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin: 10px 0;
                        font-size: 14px;
                    }}
                    .all-fields-table th,
                    .all-fields-table td {{
                        border: 1px solid #ddd;
                        padding: 8px;
                        text-align: left;
                    }}
                    .all-fields-table th {{
                        background-color: #f2f2f2;
                        font-weight: bold;
                    }}
                    .all-fields-table tr:nth-child(even) {{
                        background-color: #f9f9f9;
                    }}
                    </style>
                    {html_template}
                    """
                    status_msg = f"✅ 字段已更新，共 {len(default_fields)} 个字段"
                    return (
                        gr.update(visible=True),
                        gr.update(value=styled_html, visible=True),
                        gr.update(value=status_msg, visible=True)
                    )
                else:
                    app.current_final_fields_html = None
                    status_msg = "⚠️ 未找到HTML模板"
                    return (
                        gr.update(visible=True),
                        gr.update(visible=False),
                        gr.update(value=status_msg, visible=True)
                    )
            
            # 第三步：OCR识别
            def bill_step3_ocr(media_image, media_file_path):
                media = media_image
                if media is None and media_file_path:
                    loaded, err = _load_media(None, media_file_path, need_all_pages=False)
                    if err:
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=False),
                            gr.update(visible=True, value=err)
                        )
                    media = loaded

                if media is None:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False, value="❌ 请先上传文件")
                    )

                image = media[0] if isinstance(media, list) else media
                fields_list = []
                if hasattr(app, 'current_default_fields') and app.current_default_fields:
                    fields_list = app.current_default_fields.copy()
                if hasattr(app, 'current_custom_fields') and app.current_custom_fields:
                    fields_list.extend(app.current_custom_fields)
                
                if not fields_list:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False, value="")
                    )
                
                result = app.ocr_bill_with_fields(image, fields_list)
                
                html_template = getattr(app, 'current_field_template_html', None)
                has_html_template = html_template is not None and html_template.strip()
                
                if has_html_template and "<table" in result.lower():
                    app.last_ocr_html = result
                    app.last_ocr_markdown = ""
                    return (
                        gr.update(value=result, visible=True),
                        gr.update(value=result, visible=False),  # 同时更新隐藏的Textbox
                        gr.update(visible=True),
                        gr.update(visible=True),
                        gr.update(visible=False, value="")
                    )
                else:
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False, value=""),
                        gr.update(visible=False),
                        gr.update(visible=False),
                        gr.update(visible=False, value="")
                    )
            
            # 导出票据OCR结果
            def bill_export_ocr_result_3step(html_content, export_format, field_list):
                print("[DEBUG] bill_export_ocr_result_3step called")
                # print("[DEBUG] html content:" + html_content)
                # print("[DEBUG] export format:" + export_format)
                if not html_content or not html_content.strip():
                    return gr.update(visible=True, value="❌ 没有可保存的OCR结果，请先执行OCR识别！")
                
                # 如果接收到的内容看起来像是完整的HTML（包含style或script标签），尝试提取表格
                # 否则直接返回错误
                if '<style>' in html_content or '<script>' in html_content:
                    # 这是完整的HTML，需要提取表格部分
                    pass
                elif '<table' not in html_content.lower():
                    return gr.update(visible=True, value="❌ 未找到表格数据，无法导出！")
                
                export_dir = os.path.join("ocr_exports")
                os.makedirs(export_dir, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                
                html_export_template = f"""
                <!DOCTYPE html>
                <html lang="zh-CN">
                <head>
                    <meta charset="UTF-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>OCR 识别结果预览</title>
                    <style>
                        body {{
                            font-family: "Microsoft YaHei", Arial, sans-serif;
                            padding: 40px;
                            background-color: #f4f4f4;
                            display: flex;
                            justify-content: center;
                        }}
                        
                        /* 表格容器样式 */
                        .table-container {{
                            background-color: white;
                            padding: 20px;
                            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
                            border-radius: 8px;
                            overflow-x: auto; /* 防止表格过宽溢出 */
                        }}

                        /* 核心表格样式 */
                        table.ocr-result-table {{
                            border-collapse: collapse; /* 合并边框，必须有 */
                            margin: 0 auto;
                            /* 如果你想覆盖原始的 width，可以在这里加 !important，否则保留原始宽度 */
                        }}

                        /* 单元格样式 */
                        table.ocr-result-table td, table.ocr-result-table th {{
                            border: 1px solid #333; /* 实线边框 */
                            padding: 8px 12px;
                            text-align: center;
                            vertical-align: middle;
                            font-size: 14px;
                            min-width: 60px; /* 最小宽度防止太挤 */
                        }}

                        /* 针对可编辑区域 (contenteditable="true") 的样式优化 */
                        [contenteditable="true"] {{
                            background-color: #eef7ff; /*以此颜色标识可编辑区域 */
                            color: #0056b3;
                            cursor: text;
                            transition: background-color 0.2s;
                        }}

                        [contenteditable="true"]:focus {{
                            background-color: #fff;
                            outline: 2px solid #2196F3; /* 聚焦时的高亮边框 */
                            box-shadow: 0 0 5px rgba(33, 150, 243, 0.5);
                        }}
                        
                        /* 表头/标签列的样式 (不可编辑部分) */
                        td:not([contenteditable="true"]) {{
                            background-color: #fafafa;
                            font-weight: bold;
                            color: #555;
                        }}
                    </style>
                </head>
                <body>

                    <div class="table-container">
                        <h3>OCR 导出预览</h3>
                        {html_content}
                    </div>

                </body>
                </html>
                """

                try:
                    from bs4 import BeautifulSoup
                    # 解析HTML内容，提取表格数据
                    soup = BeautifulSoup(html_content, 'html.parser')
                    # 移除script和style标签，只保留表格
                    for script in soup(["script", "style"]):
                        script.decompose()
                    
                    table = soup.find('table', class_='ocr-result-table') or soup.find('table')
                    
                    if not table:
                        return gr.update(visible=True, value="❌ 未找到表格数据，无法导出！")
                    
                    # 提取表格数据：处理复杂的表格结构（包含rowspan和colspan）
                    def extract_table_data(table):
                        """提取表格数据，处理rowspan和colspan"""
                        data = []
                        rows = table.find_all('tr')
                        
                        for row in rows:
                            cells = row.find_all(['td', 'th'])
                            if not cells:
                                continue
                            
                            # 提取所有单元格的文本
                            row_data = []
                            for cell in cells:
                                cell_text = cell.get_text(strip=True)
                                # 跳过空单元格或样式类名
                                if cell_text and cell_text not in ['et2', 'et9', 'et11']:
                                    row_data.append(cell_text)
                            
                            # 如果行中有数据，尝试配对字段名和值
                            if len(row_data) >= 2:
                                # 尝试配对：第一个是字段名，后续是值
                                for i in range(0, len(row_data) - 1, 2):
                                    if i + 1 < len(row_data):
                                        field = row_data[i]
                                        value = row_data[i + 1]
                                        # 跳过明显的样式类名
                                        if field not in ['et2', 'et9', 'et11', ''] and value not in ['et2', 'et9', 'et11', '']:
                                            data.append([field, value])
                            elif len(row_data) == 1:
                                # 单列数据，可能是字段名或值（需要与上一行配对）
                                pass  # 暂时跳过单列数据
                        
                        return data
                    
                    # 改进的提取方法：处理复杂的表格结构（rowspan和colspan）
                    def extract_simple_table_data(table):
                        """提取表格数据，处理rowspan和colspan"""
                        data = []
                        rows = table.find_all('tr')
                        
                        # 已知的字段名列表（用于识别字段名）
                        known_fields = [
                            '出票日期', '票据状态', '汇票到期日', '票号', '出票人', '收票人',
                            '全称', '账号', '开户银行', '出票保证信息', '票据金额', '承兑人信息',
                            '开户行行号', '开户行名称', '交易合同号', '能否转让', '承兑信息',
                            '承兑保证信息', '评级信息', '备注'
                        ]
                        
                        for row in rows:
                            cells = row.find_all(['td', 'th'])
                            if not cells:
                                continue
                            
                            # 提取所有非空单元格文本
                            cell_texts = []
                            for cell in cells:
                                text = cell.get_text(strip=True)
                                # 跳过样式类名、空文本和纯数字
                                if text and text not in ['et2', 'et9', 'et11', 'et3', '']:
                                    # 检查是否是样式类名（通常是短字符串且全小写或全大写）
                                    if not (len(text) <= 3 and text.isalnum() and text.islower()):
                                        cell_texts.append(text)
                            
                            if len(cell_texts) < 2:
                                continue
                            
                            # 识别字段名和值
                            # 字段名通常是：1) 在known_fields中 2) 较短且不包含大量数字
                            # 值通常是：1) 较长 2) 包含数字或特殊字符
                            field = None
                            values = []
                            
                            for text in cell_texts:
                                # 检查是否是已知字段名
                                is_field = False
                                for known_field in known_fields:
                                    if known_field in text or text in known_field:
                                        if not field:  # 如果还没有找到字段名
                                            field = text
                                            is_field = True
                                            break
                                
                                if not is_field:
                                    # 判断是否是字段名（较短且不包含大量数字）
                                    if not field and len(text) < 15 and text.count('0') + text.count('1') + text.count('2') + text.count('3') + text.count('4') + text.count('5') + text.count('6') + text.count('7') + text.count('8') + text.count('9') < len(text) * 0.3:
                                        field = text
                                    else:
                                        values.append(text)
                            
                            # 如果有字段名和值，添加到数据中
                            if field and values:
                                # 合并多个值为一个（用空格分隔）
                                value = ' '.join(values)
                                # 避免重复添加相同的字段
                                if not any(d[0] == field for d in data):
                                    data.append([field, value])
                            elif field and not values:
                                # 只有字段名没有值，可能是rowspan的情况，跳过或标记为空
                                pass
                        
                        return data
                    
                    def html_to_excel(html_content, output_path):
                        import pandas as pd
                        from bs4 import BeautifulSoup
                        from openpyxl import Workbook
                        from openpyxl.styles import Alignment
                        from openpyxl.utils import get_column_letter
                        soup = BeautifulSoup(html_content, 'html.parser')
                        table = soup.find('table')
                        
                        if not table:
                            print("未找到表格")
                            return

                        wb = Workbook()
                        ws = wb.active
                        
                        # 1. 初始化一个矩阵来跟踪被占用的单元格 (行, 列)
                        # 这是一个简单的稀疏矩阵逻辑: occupied_cells[(row, col)] = True
                        occupied_cells = set()
                        
                        # 获取所有行
                        rows = table.find_all('tr')
                        
                        # 遍历 HTML 行
                        for r_idx, row in enumerate(rows):
                            # 找到当前行内所有的单元格 (th 和 td)
                            cells = row.find_all(['td', 'th'])
                            
                            c_idx = 0 # 当前行的列指针
                            
                            for cell in cells:
                                # 1.1 跳过已经被上一行 rowspan 占用的位置
                                while (r_idx, c_idx) in occupied_cells:
                                    c_idx += 1
                                
                                # 1.2 获取 HTML 属性
                                rowspan = int(cell.get('rowspan', 1))
                                colspan = int(cell.get('colspan', 1))
                                text_value = cell.get_text(strip=True)
                                
                                # 尝试将数字字符串转为数字（可选，为了Excel格式更好看）
                                try:
                                    if text_value.replace('.', '', 1).isdigit():
                                        if '.' in text_value:
                                            text_value = float(text_value)
                                        else:
                                            text_value = int(text_value)
                                except ValueError:
                                    pass

                                # 1.3 写入数据到 Excel (Openpyxl 是 1-based 索引，所以要 +1)
                                # 我们只把值写入合并区域的左上角第一个单元格
                                excel_row = r_idx + 1
                                excel_col = c_idx + 1
                                cell_obj = ws.cell(row=excel_row, column=excel_col, value=text_value)
                                
                                # 设置居中，美观起见
                                cell_obj.alignment = Alignment(horizontal='center', vertical='center')

                                # 1.4 处理合并
                                if rowspan > 1 or colspan > 1:
                                    # 计算结束坐标
                                    end_row = excel_row + rowspan - 1
                                    end_col = excel_col + colspan - 1
                                    
                                    ws.merge_cells(start_row=excel_row, start_column=excel_col,
                                                end_row=end_row, end_column=end_col)
                                    
                                    # 1.5 标记被占用的格子，以便后续循环跳过
                                    for r in range(rowspan):
                                        for c in range(colspan):
                                            # 标记矩阵中的位置 (0-based)
                                            occupied_cells.add((r_idx + r, c_idx + c))
                                else:
                                    # 如果没有合并，也要标记当前位置已占用
                                    occupied_cells.add((r_idx, c_idx))
                                
                                # 移动列指针 (当前单元格本身可能跨了多列)
                                # 注意：这里不需要手动加 colspan，因为上面的 while 循环和 occupy 逻辑会自动处理
                                # 但为了逻辑简单，我们只简单步进，让 while 循环去判断
                                # 实际上，HTML流式布局中，当前标签处理完，指针应该指向下一个逻辑单元格，
                                # 下一个逻辑单元格的实际物理位置由 occupied_cells 决定。
                                # 这里只需简单 +1 ? 不，如果不考虑 rowspan，由于 colspan 占据了位置，
                                # HTML 下一个 td 对应的应该是 c_idx + colspan。
                                # 但因为我们用了 occupied_cells 机制来全盘控制，
                                # 最稳健的方法是只增加 1 (处理下一个td标签)，但上面的 while 会自动把 c_idx 推到正确位置。
                                # 然而，为了避免逻辑死循环，当前 cell 自身的 colspan 需要被跳过吗？
                                # 不，HTML的 td 是挨个排列的。
                                # 例子：<tr><td colspan=2>A</td><td>B</td></tr>
                                # 处理A: c_idx=0. 占用了(0,0)和(0,1).
                                # 下一个循环处理B: c_idx 初始为 0? 不，我们需要累加器。
                                # 让我们修正一下逻辑：我们不应该在循环里 c_idx += 1，而是由逻辑控制。
                                
                                pass # 这一行实际上不需要做任何事，因为下一次循环开始时的 while 会处理
                                
                            # 这里的逻辑稍微需要调整，上面的 for cell in cells 并没有显式的 c_idx 累加器
                            # 我们需要手动维护 c_idx
                            # --- 修正后的内部循环逻辑 ---
                            
                        # --- 重新编写核心循环逻辑以确保万无一失 ---
                        # 清空之前的写入，重新开始最稳健的逻辑
                        wb = Workbook()
                        ws = wb.active
                        occupied_cells = set()
                        
                        for r_idx, row in enumerate(rows):
                            cells = row.find_all(['td', 'th'])
                            c_idx = 0 # 每一行开始，列指针归零
                            
                            for cell in cells:
                                # 只要当前坐标被之前行的 rowspan 占用了，就向右移动
                                while (r_idx, c_idx) in occupied_cells:
                                    c_idx += 1
                                
                                rowspan = int(cell.get('rowspan', 1))
                                colspan = int(cell.get('colspan', 1))
                                text_value = cell.get_text(strip=True)
                                
                                # 写入值
                                ws.cell(row=r_idx+1, column=c_idx+1, value=text_value).alignment = Alignment(horizontal='center', vertical='center')
                                
                                # 执行合并
                                if rowspan > 1 or colspan > 1:
                                    ws.merge_cells(start_row=r_idx+1, start_column=c_idx+1,
                                                end_row=r_idx+rowspan, end_column=c_idx+colspan)
                                
                                # 标记占用
                                for r in range(rowspan):
                                    for c in range(colspan):
                                        occupied_cells.add((r_idx + r, c_idx + c))
                                
                                # 处理完当前 HTML 标签后，列指针其实只需要向前移动 colspan 的距离
                                # 因为当前标签实际上横向占据了 colspan 个位置
                                # 如果不手动加，下一次循环 while 会检测到 occupied 并自动加，
                                # 但手动加更符合直觉
                                # c_idx += colspan # 这种写法有风险，因为 loop 结束回到 while 可能会重复判断
                                # 最简单的方式：不用手动加，让 while ((r, c) in occupied) c++ 自动处理
                                # 只需要在最后做一次 +1 即可吗？
                                # 不，必须基于 HTML 的流式特性。HTML的一个 cell 处理完，下一个 cell 紧接着有效空位。
                                # 所以在标记完占用后，我们什么都不用做，直接进入下一次 cell 循环？
                                # 不对，当前 cell 在 c_idx。下一个 cell 应该从 c_idx + colspan 开始找空位吗？
                                # 是的。因为当前 cell 占据了横向空间。
                                # 所以：
                                
                                c_idx += colspan 
                                # 此时 c_idx 指向了当前单元格右边的第一个位置（可能是空的，也可能被上一行的 rowspan 占用了）
                                # 下一次循环的 while 会处理那个占用情况。

                        wb.save(output_path)
                        print(f"转换成功！文件已保存至: {output_path}")

                    # 使用简单方法提取数据
                    table_data = extract_simple_table_data(table)
                    
                    # 调试信息
                    print(f"[DEBUG] 提取到的表格数据: {len(table_data)} 条")
                    for i, (field, value) in enumerate(table_data[:5]):  # 只打印前5条
                        print(f"  {i+1}. {field}: {value[:50]}...")
                    
                    if not table_data:
                        # 如果提取失败，尝试更简单的方法
                        print("[DEBUG] 简单提取失败，尝试备用方法...")
                        table_data = []
                        rows = table.find_all('tr')
                        for row in rows:
                            cells = row.find_all(['td', 'th'])
                            texts = [cell.get_text(strip=True) for cell in cells]
                            texts = [t for t in texts if t and t not in ['et2', 'et9', 'et11', 'et3'] and len(t) > 1]
                            if len(texts) >= 2:
                                # 简单配对：第一个是字段名，其余是值
                                field = texts[0]
                                value = ' '.join(texts[1:])
                                if field and value:
                                    table_data.append([field, value])
                        
                        if not table_data:
                            return gr.update(visible=True, value="❌ 表格数据为空，无法导出！请检查表格格式。")
                    

                    if export_format == "Markdown (.md)":
                        markdown_lines = ["## 票据OCR识别结果\n\n| 字段名 | 字段值 |"]
                        markdown_lines.append("|--------|--------|")
                        for field, value in table_data:
                            # 转义Markdown特殊字符
                            field_escaped = field.replace('|', '\\|')
                            value_escaped = value.replace('|', '\\|').replace('\n', ' ')
                            markdown_lines.append(f"| {field_escaped} | {value_escaped} |")
                        markdown_content = "\n".join(markdown_lines)
                        
                        file_name = f"bill_ocr_{timestamp}.md"
                        file_path = os.path.join(export_dir, file_name)
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write(markdown_content)
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 Markdown文件已保存到:\n{abs_file_path}")
                    elif export_format == "Excel (.xlsx)":
                        try:
                            # 读取HTML表格
                            file_name = f"bill_ocr_{timestamp}.xlsx"
                            file_path = os.path.join(export_dir, file_name)
                            html_to_excel(html_content, file_path)
                            abs_file_path = os.path.abspath(file_path)
                            return gr.update(visible=True, value=f"✅ 导出成功！\n📄 Excel文件已保存到:\n{abs_file_path}")
                        except Exception as e:
                            print(f"转换过程中出现错误: {e}")

                        # df = pd.json_normalize(table_data, columns=["字段名", "字段值"])

                        # df.to_excel(file_path, index=False)

                    # elif export_format == "CSV (.csv)":
                    #     import pandas as pd
                    #     df = pd.DataFrame(res.values())
                    #     file_name = f"bill_ocr_{timestamp}.csv"
                    #     file_path = os.path.join(export_dir, file_name)
                    #     df.to_csv(file_path, index=False, encoding='utf-8-sig')
                    #     abs_file_path = os.path.abspath(file_path)
                    #     return gr.update(visible=True, value=f"✅ 导出成功！\n📄 CSV文件已保存到:\n{abs_file_path}")
                    elif export_format == "JSON (.json)":
                        import json
                        file_name = f"bill_ocr_{timestamp}.json"
                        file_path = os.path.join(export_dir, file_name)
                        res = app.get_dict_from_html(html_content)
                        with open(file_path, "w", encoding="utf-8") as f:
                            json.dump(res, f, ensure_ascii=False, indent=2)
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 JSON文件已保存到:\n{abs_file_path}")
                    elif export_format == "HTML (.html)":
                        file_name = f"bill_ocr_{timestamp}.html"
                        file_path = os.path.join(export_dir, file_name)
                        with open(file_name, "w", encoding="utf-8") as f:
                            f.write(html_export_template)
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 HTML文件已保存到:\n{abs_file_path}")
                    else:
                        return gr.update(visible=True, value=f"❌ 不支持的导出格式: {export_format}")
                except Exception as e:
                    import traceback
                    error_msg = f"❌ 导出失败: {str(e)}\n{traceback.format_exc()}"
                    print(error_msg)
                    return gr.update(visible=True, value=f"❌ 导出失败: {str(e)}")
            

            # 绑定事件
            detect_bill_type_btn.click(
                bill_step1_detect_type,
                inputs=[media_image_state, media_file_state],
                outputs=[bill_type_output, bill_default_fields_html, bill_default_fields_title,
                        bill_custom_fields_input, bill_custom_fields_title, bill_add_custom_field_btn,
                        bill_update_fields_btn, bill_fields_status, bill_field_list]
            )
            
            def bill_add_custom_field(current_data):
                if current_data is None:
                    current_data = [[""]]
                elif not isinstance(current_data, list):
                    try:
                        current_data = current_data.tolist() if hasattr(current_data, 'tolist') else list(current_data)
                    except:
                        current_data = [[""]]
                new_data = list(current_data) if current_data else []
                new_data.append([""])
                return new_data
            
            bill_add_custom_field_btn.click(
                bill_add_custom_field,
                inputs=[bill_custom_fields_input],
                outputs=[bill_custom_fields_input]
            )
            
            bill_update_fields_btn.click(
                bill_step2_update_fields,
                inputs=[bill_type_output, bill_custom_fields_input],
                outputs=[bill_all_fields_title, bill_all_fields_html, bill_fields_status]
            )
            
            def bill_show_ocr_btn_from_html(all_fields_html):
                if all_fields_html and all_fields_html.strip():
                    return gr.update(visible=True)
                return gr.update(visible=False)
            
            bill_all_fields_html.change(
                bill_show_ocr_btn_from_html,
                inputs=[bill_all_fields_html],
                outputs=[bill_ocr_with_fields_btn]
            )

            bill_ocr_with_fields_btn.click(
                bill_step3_ocr,
                inputs=[media_image_state, media_file_state],
                outputs=[bill_ocr_result_html, bill_ocr_result_html_edited, bill_ocr_export_format, bill_ocr_export_btn_3step, bill_ocr_export_status_3step]
            )
            
            # 监听HTML组件的change事件，同步更新隐藏的Textbox
            def sync_edited_html(html_content):
                if html_content:
                    return html_content
                return ""
            
            bill_ocr_result_html.change(
                sync_edited_html,
                inputs=[bill_ocr_result_html],
                outputs=[bill_ocr_result_html_edited]
            )
            
            # 导出函数：使用JavaScript更新隐藏的Textbox，然后从Textbox读取
            def export_with_js_content(html_edited, export_format, field_list):
                """导出函数：使用JavaScript更新后的内容"""
                print(f"[DEBUG] export_with_js_content接收到内容:")
                print(f"  - html_edited类型: {type(html_edited)}")
                print(f"  - html_edited长度: {len(html_edited) if html_edited else 0}")
                #print(field_list)
                if html_edited:
                    print(f"  - html_edited预览: {html_edited[:200]}...")
                
                if not html_edited or not html_edited.strip():
                    return gr.update(visible=True, value="❌ 没有可保存的OCR结果，请先执行OCR识别！")
                
                # 调用导出函数
                return bill_export_ocr_result_3step(html_edited, export_format, field_list)
            
            # JavaScript函数：在导出前从DOM读取编辑后的表格内容并更新隐藏的Textbox
            js_code = """
            function() {
                var table = document.querySelector('.ocr-result-table');
                if (!table) {
                    console.error('[DEBUG] 未找到表格元素');
                    return [null];
                }
                
                // 获取编辑后的表格HTML（包含所有用户编辑的内容）
                var styleTag = document.querySelector('style');
                var styleContent = styleTag ? styleTag.outerHTML : '';
                var tableHtml = table.outerHTML;
                var fullContent = styleContent + '\\n' + tableHtml;
                
                console.log('[DEBUG] 从DOM获取的表格HTML长度:', tableHtml.length);
                console.log('[DEBUG] 表格内容预览:', tableHtml.substring(0, 200));
                console.log('[DEBUG] 准备返回编辑后的内容，长度:', fullContent.length);
                
                // 返回编辑后的内容，Gradio会自动更新bill_ocr_result_html_edited组件
                return [fullContent];
            }
            """

            # 使用JavaScript更新隐藏的Textbox，然后导出
            # 第一步：JavaScript更新bill_ocr_result_html_edited组件
            # 第二步：从bill_ocr_result_html_edited读取内容并导出
            bill_ocr_export_btn_3step.click(
                fn=lambda x: x,  # 简单的匿名函数：输入什么，返回什么
                inputs=[bill_ocr_result_html_edited], # 占位，确保参数数量匹配
                outputs=[bill_ocr_result_html_edited], 
                js=js_code
            ).then(
                export_with_js_content,
                inputs=[bill_ocr_result_html_edited, bill_ocr_export_format, bill_field_list],
                outputs=[bill_ocr_export_status_3step]
            )

        with gr.Tab("📚 文档OCR识别"):
            gr.Markdown("### 完整流程：文档输入 → 文本检测 → 文本识别（全文） → 布局分析（Layout） → 字段提取（KIE） → 输出结构化数据")
            gr.Markdown("**支持格式：** 图片（JPG/PNG等）和PDF文档")

            def toggle_content(choice):
                if choice == "全文识别":
                    return gr.update(visible=True), gr.update(visible=False)
                elif choice == "图文问答":
                    return gr.update(visible=False), gr.update(visible=True)
                # 处理空选情况
                else:
                    return gr.update(visible=False), gr.update(visible=False)

            mode_dropdown = gr.Dropdown(
                choices=["全文识别", "图文问答"], 
                label="识别模式", 
                value="全文识别"
            )

            with gr.Row(visible=True) as container_all_context:
                with gr.Column(scale=1):
                    doc_file = gr.File(
                        label="上传文档（支持图片和PDF）",
                        file_types=[".pdf", ".jpg", ".jpeg", ".png", ".bmp", ".webp"],
                        height=400
                    )
                    
                    doc_seal_removal_checkbox = gr.Checkbox(
                        label="🔄 印章淡化（仅对图片有效）",
                        value=False,
                        info="在超分辨率处理后进行印章淡化处理（PDF文件不处理）"
                    )
                    
                    doc_pdf_pages = gr.Textbox(
                        label="PDF页码（可选，留空处理所有页，如：1,3,5）",
                        placeholder="留空处理所有页，或输入页码如：1,3,5",
                        visible=False
                    )
                    
                    with gr.Row():
                        doc_ocr_btn = gr.Button("🚀 开始文档OCR识别", variant="primary")
                
                with gr.Column(scale=2):
                    with gr.Row():
                        gr.Markdown("### 📊 OCR识别结果")
                        with gr.Column(scale=1, min_width=200):
                            doc_ocr_export_format = gr.Dropdown(
                                choices=["Markdown (.md)", "Excel (.xlsx)", "CSV (.csv)", "JSON (.json)"],
                                value="Markdown (.md)",
                                label="导出格式",
                                visible=False
                            )
                        doc_ocr_export_btn = gr.Button("💾 导出结果", variant="secondary", visible=False, size="sm")
                    
                    # 分页控制
                    with gr.Row(visible=False) as doc_page_controls:
                        doc_page_prev_btn = gr.Button("⬅️ 上一页", variant="secondary", size="sm")
                        doc_page_info = gr.Markdown("第 1 页 / 共 1 页", elem_id="doc-page-info")
                        doc_page_next_btn = gr.Button("下一页 ➡️", variant="secondary", size="sm")
                    
                    doc_ocr_result_html = gr.HTML(
                        label="OCR识别结果",
                        visible=False,
                        elem_id="doc-ocr-result-html"
                    )
                    
                    # 关键字段输入和信息抽取（OCR识别完成后显示）
                    with gr.Row(visible=False) as doc_extract_controls:
                        with gr.Column():
                            gr.Markdown("### 🔍 关键字段识别")
                            doc_key_fields = gr.Textbox(
                                label="关键字段（每行一个，用于RAG相似度映射）",
                                placeholder="例如：\n合同编号\n甲方\n乙方\n签订日期",
                                lines=5,
                                interactive=True
                            )
                            doc_extract_btn = gr.Button("🔍 识别关键字段", variant="primary")
                    
                    doc_extract_result = gr.HTML(
                        label="信息抽取结果",
                        visible=False,
                        elem_id="doc-extract-result"
                    )
                    
                    doc_ocr_export_status = gr.Textbox(
                        label="导出状态",
                        interactive=False,
                        visible=False,
                        lines=3
                    )
            
            with gr.Row(visible=False) as container_chatbot:
                pass

            mode_dropdown.change(
                fn=toggle_content,
                inputs=mode_dropdown,
                outputs=[container_all_context, container_chatbot]
            )

            # 文件上传变化时，更新UI显示
            def on_file_change(file):
                if file is None:
                    return (
                        gr.update(visible=True),
                        gr.update(visible=False),
                        gr.update(value="")
                    )
                
                file_path = file.name if hasattr(file, 'name') else file
                file_ext = os.path.splitext(file_path)[1].lower()
                
                if file_ext == '.pdf':
                    return (
                        gr.update(visible=False),
                        gr.update(visible=True),
                        gr.update(value="")
                    )
                else:
                    return (
                        gr.update(visible=True),
                        gr.update(visible=False),
                        gr.update(value="")
                    )
            
            # 文档OCR识别处理函数
            def process_doc_ocr(file, pdf_pages, enable_seal_removal):
                print(f"[DEBUG] process_doc_ocr 被调用")
                print(f"[DEBUG] file: {file}, pdf_pages: {pdf_pages}, enable_seal_removal: {enable_seal_removal}")
                
                try:
                    # 优先使用文件上传
                    if file is not None:
                        file_path = file.name if hasattr(file, 'name') else file
                        print(f"[DEBUG] 文件路径: {file_path}")
                        
                        if not file_path:
                            error_msg = "❌ 请上传文档文件或图片！"
                            print(f"[DEBUG] {error_msg}")
                            return (
                                gr.update(visible=False),
                                gr.update(visible=False),  # 分页控件
                                gr.update(visible=False),  # 导出格式
                                gr.update(visible=False),  # 导出按钮
                                gr.update(visible=False),  # 信息抽取控件
                                gr.update(visible=True, value=error_msg),
                                error_msg
                            )
                        
                        if not os.path.exists(file_path):
                            error_msg = f"❌ 文件不存在: {file_path}"
                            print(f"[DEBUG] {error_msg}")
                            return (
                                gr.update(visible=False),
                                gr.update(visible=False),  # 分页控件
                                gr.update(visible=False),  # 导出格式
                                gr.update(visible=False),  # 导出按钮
                                gr.update(visible=False),  # 信息抽取控件
                                gr.update(visible=True, value=error_msg),
                                error_msg
                            )
                        
                        file_ext = os.path.splitext(file_path)[1].lower()
                        print(f"[DEBUG] 文件扩展名: {file_ext}")
                        
                        if file_ext == '.pdf':
                            # 处理PDF（不进行印章淡化处理）
                            print("[DEBUG] 开始处理PDF...")
                            if enable_seal_removal:
                                print("ℹ️ PDF文件不支持印章淡化处理，将直接进行OCR识别")
                            try:
                                with open(file_path, 'rb') as f:
                                    pdf_bytes = f.read()
                                print(f"[DEBUG] PDF文件大小: {len(pdf_bytes)} 字节")
                                result = app.ocr_document(
                                    pdf_bytes, 
                                    None,
                                    is_pdf=True,
                                    pdf_pages=pdf_pages if pdf_pages and pdf_pages.strip() else "all"
                                )
                                print(f"[DEBUG] PDF处理完成，结果长度: {len(result) if result else 0}")
                            except Exception as e:
                                import traceback
                                error_msg = f"❌ PDF处理失败: {str(e)}"
                                print(f"[DEBUG] PDF处理异常: {traceback.format_exc()}")
                                return (
                                    gr.update(visible=False),
                                    gr.update(visible=False),  # 分页控件
                                    gr.update(visible=False),  # 导出格式
                                    gr.update(visible=False),  # 导出按钮
                                    gr.update(visible=False),  # 信息抽取控件
                                    gr.update(visible=True, value=error_msg),
                                    error_msg
                                )
                        else:
                            # 处理图片文件
                            print("[DEBUG] 开始处理图片文件...")
                            try:
                                from PIL import Image
                                img = Image.open(file_path).convert("RGB")
                                print(f"[DEBUG] 图片尺寸: {img.size}")
                                
                                # 如果启用了印章淡化，先进行超分辨率+印章淡化处理
                                if enable_seal_removal:
                                    print("[DEBUG] 启用印章淡化，进行超分辨率+印章淡化处理...")
                                    img = app._super_resolve_image_for_ocr(img, enable_seal_removal=True)
                                    print("[DEBUG] 图片预处理完成")
                                
                                result = app.ocr_document(img, None)
                                print(f"[DEBUG] 图片处理完成，结果长度: {len(result) if result else 0}")
                            except Exception as e:
                                import traceback
                                error_msg = f"❌ 图片处理失败: {str(e)}"
                                print(f"[DEBUG] 图片处理异常: {traceback.format_exc()}")
                                return (
                                    gr.update(visible=False),
                                    gr.update(visible=False),  # 分页控件
                                    gr.update(visible=False),  # 导出格式
                                    gr.update(visible=False),  # 导出按钮
                                    gr.update(visible=False),  # 信息抽取控件
                                    gr.update(visible=True, value=error_msg),
                                    error_msg
                                )
                    else:
                        error_msg = "❌ 请上传文档文件！"
                        print(f"[DEBUG] {error_msg}")
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),  # 分页控件
                            gr.update(visible=False),  # 导出格式
                            gr.update(visible=False),  # 导出按钮
                            gr.update(visible=False),  # 信息抽取控件
                            gr.update(visible=True, value=error_msg),
                            error_msg
                        )
                    
                    if result is None:
                        error_msg = "❌ 处理失败，未返回结果"
                        print(f"[DEBUG] {error_msg}")
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),  # 分页控件
                            gr.update(visible=False),  # 导出格式
                            gr.update(visible=False),  # 导出按钮
                            gr.update(visible=False),  # 信息抽取控件
                            gr.update(visible=True, value=error_msg),
                            error_msg
                        )
                    
                    if result.startswith("❌"):
                        print(f"[DEBUG] 处理失败: {result}")
                        return (
                            gr.update(visible=False),
                            gr.update(visible=False),  # 分页控件
                            gr.update(visible=False),  # 导出格式
                            gr.update(visible=False),  # 导出按钮
                            gr.update(visible=False),  # 信息抽取控件
                            gr.update(visible=True, value=result),
                            result
                        )
                    
                    print("[DEBUG] 准备返回结果...")
                    # 获取第一页的文本用于显示
                    page_texts = getattr(app, 'last_ocr_page_texts', [])
                    if page_texts:
                        # 显示第一页
                        first_page_text = page_texts[0]
                        page_count = len(page_texts)
                        page_html = f"""
                        <div style="padding: 20px; background: #f8f9fa; border-radius: 8px;">
                            <h3>第 1 页 / 共 {page_count} 页</h3>
                            <div style="background: white; padding: 15px; border-radius: 6px; margin-top: 10px; white-space: pre-wrap; font-family: monospace; max-height: 600px; overflow-y: auto;">
                                {html.escape(first_page_text)}
                            </div>
                        </div>
                        """
                    else:
                        page_html = app.last_ocr_html or _plain_text_to_html(app.last_ocr_markdown or "")
                    
                    print(f"[DEBUG] HTML长度: {len(page_html) if page_html else 0}, 页数: {len(page_texts) if page_texts else 0}")
                    
                    # 如果有多个页面，显示分页控件
                    show_page_controls = len(page_texts) > 1
                    # 关键字段输入框在OCR识别完成后就显示（不依赖页数）
                    show_extract_controls = True
                    
                    return (
                        gr.update(value=page_html, visible=True),
                        gr.update(visible=show_page_controls),  # 分页控件
                        gr.update(visible=True),  # 导出格式
                        gr.update(visible=True),  # 导出按钮
                        gr.update(visible=show_extract_controls),  # 信息抽取控件（关键字段输入）
                        gr.update(visible=True, value="✅ 文档OCR识别完成，可导出结果"),
                        "✅ 文档OCR识别完成，可导出结果"
                    )
                except Exception as e:
                    import traceback
                    error_msg = f"❌ 处理失败: {str(e)}"
                    print(f"[DEBUG] 异常: {traceback.format_exc()}")
                    return (
                        gr.update(visible=False),
                        gr.update(visible=False),  # 分页控件
                        gr.update(visible=False),  # 导出格式
                        gr.update(visible=False),  # 导出按钮
                        gr.update(visible=False),  # 信息抽取控件
                        gr.update(visible=True, value=error_msg),
                        error_msg
                    )
            
            # 文件上传变化时，更新UI显示（文档OCR）
            def on_file_change(file):
                if file is None:
                    return (
                        gr.update(visible=False),
                        gr.update(value="")
                    )
                
                file_path = file.name if hasattr(file, 'name') else file
                file_ext = os.path.splitext(file_path)[1].lower()
                
                if file_ext == '.pdf':
                    return (
                        gr.update(visible=True),
                        gr.update(value="")
                    )
                else:
                    return (
                        gr.update(visible=False),
                        gr.update(value="")
                    )
            
            # 分页切换函数
            def change_doc_page(direction):
                """切换文档页面（direction: 'prev' 或 'next'）"""
                page_texts = getattr(app, 'last_ocr_page_texts', [])
                if not page_texts:
                    return (
                        gr.update(value="❌ 没有可显示的页面", visible=True),
                        gr.update(value="第 0 页 / 共 0 页"),
                        gr.update(),
                        gr.update()
                    )
                
                # 从页面信息中获取当前页码（使用隐藏状态）
                current_page = getattr(app, '_current_doc_page', 0)
                
                if direction == 'prev':
                    current_page = max(0, current_page - 1)
                elif direction == 'next':
                    current_page = min(len(page_texts) - 1, current_page + 1)
                
                app._current_doc_page = current_page
                page_count = len(page_texts)
                page_text = page_texts[current_page]
                
                page_html = f"""
                <div style="padding: 20px; background: #f8f9fa; border-radius: 8px;">
                    <h3>第 {current_page + 1} 页 / 共 {page_count} 页</h3>
                    <div style="background: white; padding: 15px; border-radius: 6px; margin-top: 10px; white-space: pre-wrap; font-family: monospace; max-height: 600px; overflow-y: auto;">
                        {html.escape(page_text)}
                    </div>
                </div>
                """
                
                page_info = f"第 {current_page + 1} 页 / 共 {page_count} 页"
                
                return (
                    gr.update(value=page_html, visible=True),
                    gr.update(value=page_info),
                    gr.update(interactive=current_page > 0),  # 上一页按钮
                    gr.update(interactive=current_page < page_count - 1)  # 下一页按钮
                )
            
            # 信息抽取函数
            def extract_doc_fields(key_fields_text):
                """根据关键字段进行信息抽取，返回表格格式"""
                if not key_fields_text or not key_fields_text.strip():
                    return gr.update(value="❌ 请输入关键字段", visible=True)
                
                # 解析关键字段（每行一个）
                key_fields = [f.strip() for f in key_fields_text.strip().split('\n') if f.strip()]
                
                if not key_fields:
                    return gr.update(value="❌ 关键字段列表为空", visible=True)
                
                try:
                    result = app.extract_document_fields_with_rag(key_fields)
                    # 如果结果是HTML表格，直接使用；否则转换为HTML
                    if result.startswith('<table') or result.startswith('<div'):
                        result_html = result
                    else:
                        # 检查是否包含表格
                        if '<table' in result:
                            result_html = result
                        else:
                            result_html = _plain_text_to_html(result)
                    return gr.update(value=result_html, visible=True)
                except Exception as e:
                    import traceback
                    error_msg = f"❌ 信息抽取失败: {str(e)}\n{traceback.format_exc()}"
                    return gr.update(value=_plain_text_to_html(error_msg), visible=True)
            
            doc_file.change(
                on_file_change,
                inputs=[doc_file],
                outputs=[doc_pdf_pages, doc_pdf_pages]
            )
            
            # 绑定分页按钮事件
            doc_page_prev_btn.click(
                lambda: change_doc_page('prev'),
                outputs=[doc_ocr_result_html, doc_page_info, doc_page_prev_btn, doc_page_next_btn]
            )
            
            doc_page_next_btn.click(
                lambda: change_doc_page('next'),
                outputs=[doc_ocr_result_html, doc_page_info, doc_page_prev_btn, doc_page_next_btn]
            )
            
            # 绑定信息抽取按钮事件
            doc_extract_btn.click(
                extract_doc_fields,
                inputs=[doc_key_fields],
                outputs=[doc_extract_result]
            )
            
            # 导出文档OCR结果
            def export_doc_ocr_result(export_format):
                # 检查是否有OCR结果
                if not hasattr(app, 'last_ocr_text') or not app.last_ocr_text:
                    return gr.update(visible=True, value="❌ 没有可导出的结果！请先进行文档OCR识别。")
                
                try:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    export_dir = "ocr_exports"
                    os.makedirs(export_dir, exist_ok=True)
                    
                    # 获取所有页的文本
                    page_texts = getattr(app, 'last_ocr_page_texts', [])
                    full_text = getattr(app, 'last_ocr_text', '')
                    
                    if export_format == "Markdown (.md)":
                        file_name = f"doc_ocr_{timestamp}.md"
                        file_path = os.path.join(export_dir, file_name)
                        
                        # 生成完整的Markdown内容，包含所有页
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write("# 文档OCR识别结果\n\n")
                            f.write(f"**识别页数：** {len(page_texts)}\n")
                            f.write(f"**总字符数：** {len(full_text)}\n\n")
                            f.write("---\n\n")
                            
                            # 写入每一页的内容
                            for i, page_text in enumerate(page_texts, 1):
                                f.write(f"## 第 {i} 页\n\n")
                                if page_text:
                                    f.write("```\n")
                                    f.write(page_text)
                                    f.write("\n```\n\n")
                                else:
                                    f.write("（本页无内容）\n\n")
                                f.write("---\n\n")
                        
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 Markdown文件已保存到:\n{abs_file_path}\n\n共 {len(page_texts)} 页")
                    
                    elif export_format == "Excel (.xlsx)":
                        import pandas as pd
                        
                        # 创建Excel数据：每页一行
                        excel_data = []
                        for i, page_text in enumerate(page_texts, 1):
                            excel_data.append({
                                "页码": i,
                                "内容": page_text if page_text else "（本页无内容）"
                            })
                        
                        # 如果没有分页数据，使用完整文本
                        if not excel_data:
                            excel_data.append({
                                "页码": 1,
                                "内容": full_text if full_text else "（无内容）"
                            })
                        
                        df = pd.DataFrame(excel_data)
                        file_name = f"doc_ocr_{timestamp}.xlsx"
                        file_path = os.path.join(export_dir, file_name)
                        df.to_excel(file_path, index=False, engine='openpyxl')
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 Excel文件已保存到:\n{abs_file_path}\n\n共 {len(excel_data)} 页")
                    
                    elif export_format == "CSV (.csv)":
                        import pandas as pd
                        import csv
                        
                        # 创建CSV数据：每页一行
                        csv_data = []
                        for i, page_text in enumerate(page_texts, 1):
                            csv_data.append({
                                "页码": i,
                                "内容": page_text if page_text else "（本页无内容）"
                            })
                        
                        # 如果没有分页数据，使用完整文本
                        if not csv_data:
                            csv_data.append({
                                "页码": 1,
                                "内容": full_text if full_text else "（无内容）"
                            })
                        
                        df = pd.DataFrame(csv_data)
                        file_name = f"doc_ocr_{timestamp}.csv"
                        file_path = os.path.join(export_dir, file_name)
                        df.to_csv(file_path, index=False, encoding='utf-8-sig')
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 CSV文件已保存到:\n{abs_file_path}\n\n共 {len(csv_data)} 页")
                    
                    elif export_format == "JSON (.json)":
                        import json
                        
                        # 创建JSON数据结构：包含所有页
                        data = {
                            "总页数": len(page_texts),
                            "总字符数": len(full_text),
                            "页面内容": []
                        }
                        
                        # 添加每一页的内容
                        for i, page_text in enumerate(page_texts, 1):
                            data["页面内容"].append({
                                "页码": i,
                                "内容": page_text if page_text else "（本页无内容）",
                                "字符数": len(page_text) if page_text else 0
                            })
                        
                        # 如果没有分页数据，使用完整文本
                        if not data["页面内容"]:
                            data["页面内容"].append({
                                "页码": 1,
                                "内容": full_text if full_text else "（无内容）",
                                "字符数": len(full_text) if full_text else 0
                            })
                        
                        file_name = f"doc_ocr_{timestamp}.json"
                        file_path = os.path.join(export_dir, file_name)
                        with open(file_path, "w", encoding="utf-8") as f:
                            json.dump(data, f, ensure_ascii=False, indent=2)
                        abs_file_path = os.path.abspath(file_path)
                        return gr.update(visible=True, value=f"✅ 导出成功！\n📄 JSON文件已保存到:\n{abs_file_path}\n\n共 {len(data['页面内容'])} 页")
                    else:
                        return gr.update(visible=True, value=f"❌ 不支持的导出格式: {export_format}")
                except Exception as e:
                    import traceback
                    error_msg = f"❌ 导出失败: {str(e)}\n{traceback.format_exc()}"
                    print(error_msg)
                    return gr.update(visible=True, value=error_msg)
            
            doc_ocr_btn.click(
                process_doc_ocr,
                inputs=[doc_file, doc_pdf_pages, doc_seal_removal_checkbox],
                outputs=[doc_ocr_result_html, doc_page_controls, doc_ocr_export_format, doc_ocr_export_btn, doc_extract_controls, doc_ocr_export_status, doc_ocr_export_status],
                show_progress=True
            )
            
            doc_ocr_export_btn.click(
                export_doc_ocr_result,
                inputs=[doc_ocr_export_format],
                outputs=[doc_ocr_export_status]
            )

        with gr.Tab("ℹ️ 使用说明"):
            gr.Markdown(
                """
                - 先点击「加载模型」后再使用各项功能。
                - 在同一界面直接完成图文问答、通用OCR、卡证/票据/协议OCR等能力，无需额外模式切换。
                - 支持高级生成参数调节，适合触摸屏交互。
                - 已默认优化为更易触摸点击的界面尺寸。
                """
            )

    return interface


def main():
    print("🚀 启动Qwen3-VL-8B-Instruct 统一Web界面...")
    interface = _legacy_create_unified_interface()

    def _cleanup():
        # 清理模型与显存
        try:
            app.model = None
            app.processor = None
            app.is_loaded = False
        except Exception:
            pass
        try:
            if torch is not None and hasattr(torch, "cuda") and torch.cuda.is_available():
                torch.cuda.empty_cache()
                torch.cuda.ipc_collect()
        except Exception:
            pass
        try:
            gc.collect()
        except Exception:
            pass

    # 注册进程退出清理
    atexit.register(_cleanup)
    interface.queue()
    interface.launch(
        server_name="127.0.0.1",
        server_port=7860,  # 自动选择可用端口，避免端口占用错误
        share=False,
        debug=True,
        show_error=True,
    )


if __name__ == "__main__":
    main()
